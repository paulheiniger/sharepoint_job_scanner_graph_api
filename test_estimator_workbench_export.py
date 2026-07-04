from __future__ import annotations

import json
import zipfile
from datetime import UTC, datetime
from decimal import Decimal
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from jobscan.estimator.workbench_export import EXCEL_CELL_LIMIT, export_workbench_review_package


def sample_workbench() -> dict:
    return {
        "estimate_id": "review-test",
        "scope": {
            "division": "Roofing",
            "template_type": "roofing",
            "project_type": "roof coating",
            "net_sqft": 10000,
            "created_at": datetime.now(UTC),
        },
        "historical_filters": {"division": "Roofing", "source_year": None},
        "roofing_coating_template_decisions": [
            {
                "include": True,
                "section": "roofing_coating_template_decisions",
                "decision_id": "roofing_coating_system_row_26",
                "template_bucket": "coating",
                "workbook_row": "26",
                "editable_selector_code": "11",
                "resolved_template_option": "Gaco Silicone",
                "selected_pricing_candidate": "GAF High Solids Silicone 55 Gal",
                "basis_sqft": 10000,
                "gal_per_100_sqft": 1.5,
                "waste_factor_pct": 10,
                "unit_price": 38,
                "estimated_gallons": 166.67,
                "estimated_cost": 6333.46,
                "historical_recommendation": "Historical coating decision from 12 jobs.",
                "decision_evidence_count": 12,
                "decision_confidence": "high",
                "product_id": "gaf_high_solids_silicone",
                "product_manufacturer": "GAF",
                "product_guidance": "Use as silicone roof coating.",
                "product_warning_summary": "Do not apply over wet substrate.",
                "product_source_documents": ["product_documents/GAF Silicone.pdf"],
                "workbook_cell_write_preview": [{"cell": "Estimate!A26", "field": "selector_code", "value": "11"}],
                "notes": "Decision-first coating row.",
            }
        ],
        "roofing_labor_template_decisions": [
            {
                "include": True,
                "section": "roofing_labor_template_decisions",
                "decision_id": "roofing_labor_base",
                "template_bucket": "labor_base",
                "labor_task": "Base Coat",
                "workbook_row": "122",
                "days": 2,
                "crew_size": 4,
                "hourly_rate": 72,
                "total_hours": 64,
                "estimated_cost": 4608,
            }
        ],
        "similar_jobs": [{"job_id": "J1", "customer": "Acme"}],
        "review_flags": ["Verify substrate."],
    }


def test_export_package_creates_zip_with_decision_files_and_workbook(tmp_path) -> None:
    workbook = Workbook()
    workbook.active["A1"] = "estimate"
    workbook_path = tmp_path / "generated.xlsx"
    workbook.save(workbook_path)

    zip_path = export_workbench_review_package(
        workbench=sample_workbench(),
        input_notes="Roof coating notes",
        output_dir=tmp_path,
        workbook_path=workbook_path,
    )

    assert zip_path.exists()
    with zipfile.ZipFile(zip_path) as archive:
        names = set(archive.namelist())
        assert {
            "workbench_summary.json",
            "workbench_summary.xlsx",
            "workbench_debug.json",
            "README.txt",
            "estimator_input.txt",
            "exported_workbook.xlsx",
        }.issubset(names)
        assert "workbook_export_error.txt" not in names
        summary = json.loads(archive.read("workbench_summary.json"))
        assert summary["input_notes"] == "Roof coating notes"
        assert summary["parsed_scope"]["project_type"] == "roof coating"
        assert summary["workbook_decisions"]
        assert "materials_final" not in summary
        assert "labor_final" not in summary
        assert "adders_final" not in summary
        assert any(row["product_id"] == "gaf_high_solids_silicone" for row in summary["product_guidance"])
        readme = archive.read("README.txt").decode("utf-8")
        assert "Decision Trace" in readme
        assert "Debug Decision JSON" in readme
        archive.extract("workbench_summary.xlsx", path=tmp_path)

    summary_workbook = load_workbook(tmp_path / "workbench_summary.xlsx", read_only=True, data_only=True)
    sheetnames = set(summary_workbook.sheetnames)
    assert "Workbook Decisions" in sheetnames
    assert "Decision Trace" in sheetnames
    assert "Product Guidance" in sheetnames
    assert "Materials Compact" not in sheetnames
    assert "Labor Compact" not in sheetnames
    assert "Adders Compact" not in sheetnames
    assert "Debug Materials" not in sheetnames


def test_insulation_review_package_uses_decision_sheets(tmp_path) -> None:
    workbench = {
        "estimate_id": "insulation-review",
        "scope": {
            "division": "Insulation",
            "template_type": "insulation",
            "project_type": "spray foam insulation",
            "gross_insulation_area_sqft": 2460,
            "net_insulation_area_sqft": 2388,
            "opening_area_missing": True,
        },
        "historical_filters": {"division": "Insulation", "template_type": "insulation"},
        "area_calculation_trace": [
            {"component": "walls", "formula": "perimeter * height", "selected_area_sqft": 1260}
        ],
        "insulation_foam_template_decisions": [
            {
                "include": True,
                "decision_id": "insulation_foam_template_selector",
                "template_bucket": "foam",
                "workbook_row": "19-21",
                "editable_selector_code": "11",
                "resolved_template_option": "Gaco 2.0 lb.",
                "basis_sqft": 2388,
                "thickness_inches": 3,
                "yield_or_coverage": 12000,
                "unit_price": 2.4,
                "estimated_units": 597,
                "estimated_cost": 1432.8,
            }
        ],
        "insulation_detail_material_template_decisions": [
            {
                "include": True,
                "decision_id": "insulation_caulk_sealant",
                "template_bucket": "caulk_sealant",
                "template_line": "Caulk / Sealant",
                "workbook_row": "41",
                "editable_selector_code": "1",
                "linear_ft": 120,
                "feet_per_unit": 10,
                "unit_price": 6.05,
                "estimated_units": 12,
                "estimated_cost": 72.6,
                "total_hours": 99,
                "daily_rate": 123,
            }
        ],
        "insulation_labor_template_decisions": [
            {
                "include": True,
                "decision_id": "insulation_labor_foam",
                "template_bucket": "labor_foam",
                "workbook_row": "86",
                "labor_task": "Foam",
                "days": 1.5,
                "crew_size": 3,
                "hourly_rate": 72,
                "total_hours": 36,
                "estimated_cost": 2592,
                "gal_per_100_sqft": 1.5,
                "feet_per_unit": 10,
            }
        ],
        "review_flags": ["Missing rollup door width."],
    }

    zip_path = export_workbench_review_package(
        workbench=workbench,
        input_notes="Foam sprayed in a 30x40 metal building.",
        output_dir=tmp_path,
        workbook_export_error="Insulation workbook export was not attempted in this test.",
    )

    with zipfile.ZipFile(zip_path) as archive:
        names = set(archive.namelist())
        assert "workbench_summary.json" in names
        assert "workbench_summary.xlsx" in names
        assert "workbook_export_error.txt" in names
        assert "exported_workbook.xlsx" not in names
        summary = json.loads(archive.read("workbench_summary.json"))
        assert summary["area_calculation_trace"]
        assert summary["insulation_foam_template_decisions"]
        assert summary["workbook_decisions"]
        archive.extract("workbench_summary.xlsx", path=tmp_path)

    summary_workbook = load_workbook(tmp_path / "workbench_summary.xlsx", read_only=True, data_only=True)
    assert "Area Calculation Trace" in summary_workbook.sheetnames
    assert "Insulation Foam Template" in summary_workbook.sheetnames
    detail_headers = [cell.value for cell in next(summary_workbook["Insulation Details"].iter_rows(min_row=1, max_row=1))]
    assert "linear_ft" in detail_headers
    assert "feet_per_unit" in detail_headers
    assert "total_hours" not in detail_headers
    assert "daily_rate" not in detail_headers
    assert "Insulation Labor Plan" in summary_workbook.sheetnames
    labor_headers = [cell.value for cell in next(summary_workbook["Insulation Labor Plan"].iter_rows(min_row=1, max_row=1))]
    assert "labor_task" in labor_headers
    assert "total_hours" in labor_headers
    assert "gal_per_100_sqft" not in labor_headers
    assert "feet_per_unit" not in labor_headers
    assert "Workbook Decisions" in summary_workbook.sheetnames


def test_export_excel_handles_timezone_datetimes_and_long_decision_text(tmp_path) -> None:
    workbench = sample_workbench()
    long_text = "x" * (EXCEL_CELL_LIMIT + 1000)
    workbench["roofing_coating_template_decisions"][0]["product_guidance"] = long_text
    workbench["roofing_coating_template_decisions"][0]["debug_payload"] = {
        "when": datetime.now(UTC),
        "uuid": uuid4(),
        "amount": Decimal("12.34"),
        "items": [long_text],
    }

    zip_path = export_workbench_review_package(workbench=workbench, input_notes="Notes", output_dir=tmp_path)

    with zipfile.ZipFile(zip_path) as archive:
        archive.extract("workbench_summary.xlsx", path=tmp_path)
    workbook = load_workbook(tmp_path / "workbench_summary.xlsx", read_only=True, data_only=True)
    decisions = workbook["Roof Coating System"]
    headers = [cell.value for cell in next(decisions.iter_rows(min_row=1, max_row=1))]
    notes_index = headers.index("product_guidance")
    row = next(decisions.iter_rows(min_row=2, max_row=2))
    value = row[notes_index].value

    assert isinstance(value, str)
    assert len(value) <= EXCEL_CELL_LIMIT
    assert "truncated for Excel" in value


def test_missing_workbook_export_does_not_crash_package_export(tmp_path) -> None:
    zip_path = export_workbench_review_package(
        workbench=sample_workbench(),
        input_notes="No workbook yet",
        output_dir=tmp_path,
        workbook_export_error="Workbook generation failed in test.",
    )

    with zipfile.ZipFile(zip_path) as archive:
        names = set(archive.namelist())
        assert "workbook_export_error.txt" in names
        assert "exported_workbook.xlsx" not in names
        assert b"Workbook generation failed" in archive.read("workbook_export_error.txt")
