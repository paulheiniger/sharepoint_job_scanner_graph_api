from __future__ import annotations

import hashlib
from pathlib import Path

import pytest

from jobscan.estimator.workbook_template import (
    DEFAULT_TEMPLATE_PATH,
    fill_estimate_workbook,
    load_template,
    read_material_reference_pricing,
)
from test_estimator_prototype import sample_data
from jobscan.estimator.estimate import build_estimate


pytest.importorskip("openpyxl")
import openpyxl


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def sample_estimate_result() -> dict:
    return build_estimate(
        "Metal roof, about 12,000 sqft, rusted fasteners, 20 year silicone coating, many penetrations in Louisville",
        sample_data(),
    )


def test_workbook_template_loads() -> None:
    workbook = load_template(DEFAULT_TEMPLATE_PATH)

    assert {"Estimate", "People", "Materials", "General", "Performance & Payment Bonds"}.issubset(workbook.sheetnames)


def test_materials_tab_reference_pricing_reads() -> None:
    refs = read_material_reference_pricing(DEFAULT_TEMPLATE_PATH)

    assert refs["solvents"]
    assert refs["fabric_widths"]
    assert refs["board_thicknesses"]
    assert refs["fasteners"]
    assert refs["plates"]


def test_fill_workbook_writes_mapped_fields_and_preserves_formulas(tmp_path: Path) -> None:
    original_hash = file_hash(DEFAULT_TEMPLATE_PATH)
    result = sample_estimate_result()

    fill = fill_estimate_workbook(
        DEFAULT_TEMPLATE_PATH,
        result,
        tmp_path,
        metadata={
            "job_name": "Draft Metal Roof",
            "job_type": "Roof Coating",
            "site_address": "123 Main St",
            "city_state_zip": "Louisville, KY 40202",
            "contact": "Quin",
            "email": "quin@example.com",
            "phone": "502-555-0100",
        },
        output_name="draft.xlsx",
    )

    assert fill.output_path.exists()
    assert file_hash(DEFAULT_TEMPLATE_PATH) == original_hash

    workbook = openpyxl.load_workbook(fill.output_path, data_only=False)
    ws = workbook["Estimate"]
    assert ws["C2"].value == "Draft Metal Roof"
    assert ws["C3"].value == "Roof Coating"
    assert ws["C4"].value == "123 Main St"
    assert ws["C5"].value == "Louisville, KY 40202"
    assert ws["C6"].value == "Quin"
    assert ws["C8"].value == "quin@example.com"
    assert ws["C9"].value == "502-555-0100"
    assert ws["C12"].value == 12000
    assert ws["H26"].value.startswith("=")
    assert ws["H163"].value.startswith("=")
    assert ws["H169"].value.startswith("=")


def test_labor_days_and_crew_size_write_correctly(tmp_path: Path) -> None:
    fill = fill_estimate_workbook(DEFAULT_TEMPLATE_PATH, sample_estimate_result(), tmp_path, output_name="labor.xlsx")
    ws = openpyxl.load_workbook(fill.output_path, data_only=False)["Estimate"]

    assert ws["B116"].value > 0
    assert ws["C116"].value >= 1
    assert ws["B122"].value > 0
    assert ws["C122"].value >= 1
    assert ws["H116"].value.startswith("=")


def test_material_quantities_write_correctly(tmp_path: Path) -> None:
    fill = fill_estimate_workbook(DEFAULT_TEMPLATE_PATH, sample_estimate_result(), tmp_path, output_name="materials.xlsx")
    ws = openpyxl.load_workbook(fill.output_path, data_only=False)["Estimate"]

    assert ws["C26"].value == 12000
    assert ws["D26"].value > 0
    assert ws["E26"].value == 38
    assert ws["C39"].value == 12000
    assert ws["C47"].value > 0


def test_generated_workbook_path_is_created(tmp_path: Path) -> None:
    fill = fill_estimate_workbook(DEFAULT_TEMPLATE_PATH, sample_estimate_result(), tmp_path, output_name="created.xlsx")

    assert fill.output_path == tmp_path / "created.xlsx"
    assert fill.output_path.exists()
    assert "subtotal_materials" in fill.formula_cells
