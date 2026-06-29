from __future__ import annotations

import json
import math
from datetime import UTC, datetime
from decimal import Decimal
from types import SimpleNamespace
from uuid import uuid4

import pandas as pd
from openpyxl import load_workbook

from jobscan.estimator.evidence_export import build_estimator_evidence_export, sanitize_for_export, write_estimator_evidence_export
from jobscan.estimator.field_estimator import estimate_from_field_notes
from test_field_estimator import field_data


SAMPLE_NOTE = (
    "Roof coating estimate. Metal roof in Louisville KY. Main roof is 120 ft by 80 ft. "
    "Deduct two skylight areas, each 4 ft by 8 ft. Roof condition is fair with some rusted fasteners. "
    "Customer wants a 10-year silicone coating system. Access is easy. Few penetrations."
)


def test_build_estimator_evidence_export_contains_expected_sheets() -> None:
    data = field_data()
    recommendation = estimate_from_field_notes(SAMPLE_NOTE, {"estimated_sqft": 0}, data=data)

    export = build_estimator_evidence_export(recommendation, data=data, notes=SAMPLE_NOTE)

    assert export["run_summary"]["estimated_sqft"] == 9536
    for sheet in [
        "README",
        "run_integrity",
        "parsed_scope",
        "material_plan",
        "material_evidence",
        "labor_plan",
        "labor_evidence",
        "labor_diagnostics",
        "estimate_rollup",
    ]:
        assert sheet in export["sheets"]
        assert export["sheets"][sheet]
    assert "relationship_rows" not in export["sheets"]
    assert "rejected_evidence" not in export["sheets"]
    assert any(row.get("task") == "labor_prep" for row in export["sheets"]["labor_diagnostics"])
    assert any(row.get("included_in_total") is True for row in export["sheets"]["material_plan"])
    assert export["run_summary"]["run_id"]
    assert export["run_summary"]["input_notes_hash"] == export["run_summary"]["parsed_scope_notes_hash"]
    assert export["sheets"]["run_integrity"][0]["stale_source_text_detected"] is False
    assert "runtime_seconds_by_stage" in export["run_summary"]


def test_debug_evidence_export_restores_detailed_sheets() -> None:
    data = field_data()
    recommendation = estimate_from_field_notes(SAMPLE_NOTE, {"estimated_sqft": 0}, data=data)

    export = build_estimator_evidence_export(recommendation, data=data, notes=SAMPLE_NOTE, debug_evidence=True)

    assert "relationship_rows" in export["sheets"]
    assert "rejected_evidence" in export["sheets"]
    assert "similar_jobs" in export["sheets"]


def test_evidence_export_caps_rows_and_strips_line_item_id_lists() -> None:
    recommendation = SimpleNamespace(
        parsed_fields={"estimated_sqft": 10000, "project_type": "roof coating", "substrate": "metal", "coating_type": "silicone"},
        recommended_scope=[],
        material_plan=[{"item": "Coating", "category": "coating", "estimated_cost": 1000, "selected_price_source": "current_pricing"}],
        labor_plan=[],
        travel_plan={},
        historical_calibration={},
        similar_examples=[],
        estimate_low=0,
        estimate_target=0,
        estimate_high=0,
        review_flags=[],
        human_review_required=False,
        draft_workbook_inputs={"header": {"C12_estimated_sqft": 10000}},
        debug={},
    )
    data = SimpleNamespace(
        source_files_used=[],
        warnings=[],
        template_rows=pd.DataFrame(
            [
                {
                    "template_bucket": "coating",
                    "template_type": "roofing",
                    "selected_item_name": f"Coating {index}",
                    "evidence_line_item_ids": ["a", "b", "c"],
                    "quantity_evidence_diagnostics": [{"row": index}],
                }
                for index in range(6)
            ]
        ),
        job_package_summary=pd.DataFrame(),
        pricing_catalog=pd.DataFrame(),
        pricing=pd.DataFrame(),
        relationship_labor_rates=pd.DataFrame(),
    )

    export = build_estimator_evidence_export(recommendation, data=data, notes="Roof coating", evidence_limit=2)
    rows = [row for row in export["sheets"]["material_evidence"] if row.get("evidence_source_table") == "template_rows"]

    assert len(rows) <= 2
    assert all("evidence_line_item_ids" not in row for row in rows)
    assert all("quantity_evidence_diagnostics" not in row for row in rows)
    assert all(row.get("evidence_line_item_count") == 3 for row in rows)
    assert export["run_summary"]["evidence_rows_exported"] <= 2 + len(export["sheets"]["labor_evidence"]) + 1


def test_normal_roofing_evidence_export_excludes_insulation_template_rows() -> None:
    recommendation = SimpleNamespace(
        parsed_fields={"estimated_sqft": 10000, "project_type": "roof coating", "substrate": "metal", "coating_type": "silicone"},
        recommended_scope=[],
        material_plan=[{"item": "Coating", "category": "coating", "estimated_cost": 1000, "selected_price_source": "current_pricing"}],
        labor_plan=[],
        travel_plan={},
        historical_calibration={},
        similar_examples=[],
        estimate_low=0,
        estimate_target=0,
        estimate_high=0,
        review_flags=[],
        human_review_required=False,
        draft_workbook_inputs={"header": {"C12_estimated_sqft": 10000}},
        debug={},
    )
    data = SimpleNamespace(
        source_files_used=[],
        warnings=[],
        template_rows=pd.DataFrame(
            [
                {"template_bucket": "coating", "template_type": "roofing", "selected_item_name": "Roof coating"},
                {"template_bucket": "coating", "template_type": "insulation", "selected_item_name": "Insulation coating"},
                {"template_bucket": "unknown", "template_type": "roofing", "selected_item_name": "Unknown"},
            ]
        ),
        job_package_summary=pd.DataFrame(),
        pricing_catalog=pd.DataFrame(),
        pricing=pd.DataFrame(),
        relationship_labor_rates=pd.DataFrame(),
    )

    export = build_estimator_evidence_export(recommendation, data=data, notes="Roof coating")
    names = [row.get("selected_item_name") for row in export["sheets"]["material_evidence"]]

    assert "Roof coating" in names
    assert "Insulation coating" not in names
    assert "Unknown" not in names


def test_evidence_export_flags_notes_hash_mismatch_and_stale_source_text() -> None:
    data = field_data()
    recommendation = estimate_from_field_notes("Roof is 160 ft by 150 ft. Silicone coating Louisville KY.", data=data)

    export = build_estimator_evidence_export(
        recommendation,
        data=data,
        notes="Roof is 90 ft by 70 ft. Silicone coating Louisville KY.",
    )
    integrity = export["sheets"]["run_integrity"][0]

    assert integrity["hash_mismatch"] is True
    assert integrity["stale_source_text_detected"] is True
    assert "160 ft by 150 ft" in integrity["stale_fields_detected"]


def test_write_estimator_evidence_export_creates_parseable_json_and_xlsx(tmp_path) -> None:
    data = field_data()
    recommendation = estimate_from_field_notes(SAMPLE_NOTE, {"estimated_sqft": 0}, data=data)

    paths = write_estimator_evidence_export(recommendation, data=data, notes=SAMPLE_NOTE, output_dir=tmp_path, base_filename="sample")

    assert paths["json"].exists()
    assert paths["xlsx"].exists()
    parsed = json.loads(paths["json"].read_text(encoding="utf-8"))
    assert parsed["run_summary"]["material_rows"] >= 1
    workbook = load_workbook(paths["xlsx"], read_only=True)
    assert {
        "README",
        "run_integrity",
        "parsed_scope",
        "material_plan",
        "labor_plan",
        "labor_diagnostics",
        "estimate_rollup",
    }.issubset(set(workbook.sheetnames))


def test_rejected_material_evidence_is_exported() -> None:
    recommendation = SimpleNamespace(
        parsed_fields={"estimated_sqft": 10000, "project_type": "roof coating"},
        recommended_scope=[],
        material_plan=[
            {
                "item": "Primer allowance",
                "category": "primer",
                "quantity": 999,
                "unit": "pail",
                "unit_price": 1,
                "estimated_cost": 999,
                "selected_price_source": "rejected_historical_quantity_ratio",
                "sanity_check_status": "blocked_implausible_quantity",
                "rejected_reason": "Historical cost allowance was not a physical quantity.",
                "needs_review": True,
            }
        ],
        labor_plan=[],
        travel_plan={},
        historical_calibration={},
        similar_examples=[],
        estimate_low=0,
        estimate_target=0,
        estimate_high=0,
        review_flags=[],
        human_review_required=True,
        draft_workbook_inputs={"header": {"C12_estimated_sqft": 10000}},
        debug={},
    )

    export = build_estimator_evidence_export(recommendation, debug_evidence=True)
    material_row = export["sheets"]["material_plan"][0]
    rejected_rows = export["sheets"]["rejected_evidence"]

    assert material_row["included_in_total"] is False
    assert any(row.get("package") == "primer" and row.get("severity") == "blocker" for row in rejected_rows)


def test_export_sanitizes_timezone_datetimes_and_nested_values(tmp_path) -> None:
    aware_datetime = datetime(2026, 6, 26, 12, 30, tzinfo=UTC)
    timestamp = pd.Timestamp("2026-06-26T08:15:00", tz="America/New_York")
    row_uuid = uuid4()
    recommendation = SimpleNamespace(
        parsed_fields={
            "estimated_sqft": 10000,
            "generated_at": aware_datetime,
            "timestamp": timestamp,
            "uuid": row_uuid,
            "bad_number": math.nan,
            "nested": {"date": aware_datetime, "items": [Decimal("12.50"), math.inf]},
        },
        recommended_scope=[],
        material_plan=[
            {
                "item": "Coating",
                "category": "coating",
                "estimated_cost": Decimal("1234.56"),
                "created_at": timestamp,
                "nested": {"uuid": row_uuid, "when": aware_datetime},
            }
        ],
        labor_plan=[],
        travel_plan={},
        historical_calibration={},
        similar_examples=[],
        estimate_low=0,
        estimate_target=0,
        estimate_high=0,
        review_flags=[],
        human_review_required=False,
        draft_workbook_inputs={"header": {"C12_estimated_sqft": 10000, "created_at": aware_datetime}},
        debug={},
    )

    paths = write_estimator_evidence_export(recommendation, output_dir=tmp_path, base_filename="timezone")

    assert paths["json"].exists()
    assert paths["xlsx"].exists()
    parsed = json.loads(paths["json"].read_text(encoding="utf-8"))
    assert parsed["sheets"]["parsed_scope"][0]["generated_at"].endswith("+00:00")
    assert parsed["sheets"]["parsed_scope"][0]["bad_number"] is None
    assert parsed["sheets"]["parsed_scope"][0]["nested"]["items"][1] is None
    workbook = load_workbook(paths["xlsx"], read_only=True, data_only=True)
    parsed_scope = workbook["parsed_scope"]
    headers = [cell.value for cell in next(parsed_scope.iter_rows(min_row=1, max_row=1))]
    generated_at_column = headers.index("generated_at") + 1
    generated_at_value = parsed_scope.cell(row=2, column=generated_at_column).value
    assert isinstance(generated_at_value, str)
    sanitized = sanitize_for_export({"when": aware_datetime, "ts": timestamp, "uuid": row_uuid}, excel=True)
    assert isinstance(sanitized["when"], str)
    assert isinstance(sanitized["ts"], str)
    assert isinstance(sanitized["uuid"], str)
