from __future__ import annotations

import math
import re
from datetime import date
from pathlib import Path
from typing import Any

from .estimator import FlooringEstimateResult

DEFAULT_FLOORING_TEMPLATE_PATH = Path("templates/Estimate Flooring - Lee Sporting Shop.xlsx")
DEFAULT_FLOORING_OUTPUT_DIR = Path("output/flooring_estimator/filled_templates")


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("_")
    return cleaned[:90] or "flooring_estimate"


def number_or_none(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(str(value).replace(",", "").replace("$", ""))
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def text_or_blank(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def resolve_flooring_template_path(template_path: Path | str | None = None) -> Path:
    path = Path(template_path) if template_path else DEFAULT_FLOORING_TEMPLATE_PATH
    if not path.exists():
        raise FileNotFoundError(f"Flooring estimate template not found: {path}")
    return path


def _write(ws: Any, cell: str, value: Any) -> None:
    if value is None or value == "":
        return
    ws[cell] = value


def _as_result_dict(result: FlooringEstimateResult | dict[str, Any]) -> dict[str, Any]:
    return result.to_dict() if isinstance(result, FlooringEstimateResult) else dict(result)


def _decision_by_row(result: dict[str, Any]) -> dict[int, dict[str, Any]]:
    decisions: dict[int, dict[str, Any]] = {}
    for decision in result.get("workbook_decisions") or []:
        row = number_or_none(decision.get("workbook_row"))
        if row is not None:
            decisions[int(row)] = decision
    return decisions


def _labor_trip_count(decisions_by_row: dict[int, dict[str, Any]]) -> int:
    days = 0.0
    for row in (116, 120, 124, 126, 128, 130, 132):
        days += number_or_none((decisions_by_row.get(row) or {}).get("days")) or 0.0
    return max(1, int(math.ceil(days or 1)))


def _apply_formula_recalc_flags(workbook: Any) -> None:
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    except Exception:
        pass


def generate_flooring_estimate_workbook(
    result: FlooringEstimateResult | dict[str, Any],
    *,
    template_path: Path | str | None = None,
    output_dir: Path | str = DEFAULT_FLOORING_OUTPUT_DIR,
    output_filename: str | None = None,
    job_name: str = "",
    site_address: str = "",
    city_state_zip: str = "",
    contact_name: str = "",
    contact_title: str = "",
    contact_email: str = "",
    contact_phone: str = "",
    estimator: str = "",
    estimate_date: date | str | None = None,
    round_trip_miles: float | int | None = None,
) -> Path:
    from openpyxl import load_workbook

    payload = _as_result_dict(result)
    parsed = payload.get("parsed_scope") or {}
    decisions_by_row = _decision_by_row(payload)
    template = resolve_flooring_template_path(template_path)
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = output_filename or f"{safe_filename(job_name or 'flooring_estimate')}.xlsx"
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    output_path = out_dir / filename

    workbook = load_workbook(template)
    estimate = workbook["Estimate"]

    _write(estimate, "C1", estimate_date or date.today())
    _write(estimate, "C2", job_name)
    _write(estimate, "C3", parsed.get("job_type") or "Floor System")
    _write(estimate, "C4", site_address)
    _write(estimate, "C5", city_state_zip)
    _write(estimate, "C6", contact_name)
    _write(estimate, "C7", contact_title or estimator)
    _write(estimate, "C8", contact_email)
    _write(estimate, "C9", contact_phone)

    area = number_or_none(parsed.get("area_sqft"))
    _write(estimate, "C12", area)

    for row in (26, 27):
        decision = decisions_by_row.get(row) or {}
        row_area = number_or_none(decision.get("area_sqft")) or area
        _write(estimate, f"A{row}", decision.get("selector_code"))
        _write(estimate, f"C{row}", row_area)
        _write(estimate, f"D{row}", number_or_none(decision.get("gal_per_100_sqft")))
        _write(estimate, f"E{row}", number_or_none(decision.get("unit_price")))
        _write(estimate, f"F{row}", decision.get("description") or decision.get("item"))

    primer = decisions_by_row.get(39) or {}
    if primer:
        _write(estimate, "A39", primer.get("selector_code") or 1)
        _write(estimate, "C39", number_or_none(primer.get("area_sqft")) or area)
        _write(estimate, "E39", number_or_none(primer.get("unit_price")))

    flake = decisions_by_row.get(177) or {}
    if flake:
        _write(estimate, "A177", "Additional Amount w/o Markup")
        _write(estimate, "F177", number_or_none(flake.get("estimated_cost")))
        _write(estimate, "G177", flake.get("item") or "Flake")

    generator = decisions_by_row.get(99) or {}
    if generator:
        _write(estimate, "C99", number_or_none(generator.get("days")))
        _write(estimate, "E99", number_or_none(generator.get("unit_price")))

    trip_count = _labor_trip_count(decisions_by_row)
    miles = number_or_none(round_trip_miles) or 0
    _write(estimate, "B106", 1)
    _write(estimate, "C106", miles)
    _write(estimate, "B108", trip_count)
    _write(estimate, "C108", miles)

    for row in (116, 120, 124, 126, 128, 130, 132):
        decision = decisions_by_row.get(row) or {}
        if decision:
            _write(estimate, f"B{row}", number_or_none(decision.get("days")))
            _write(estimate, f"C{row}", number_or_none(decision.get("crew_size")))

    loading = decisions_by_row.get(137) or {}
    if loading:
        _write(estimate, "C137", number_or_none(loading.get("hours_per_trip")))
        _write(estimate, "E137", number_or_none(loading.get("crew_size")))
        _write(estimate, "G137", number_or_none(loading.get("hourly_rate")))

    travel = decisions_by_row.get(139) or {}
    if travel:
        _write(estimate, "C139", number_or_none(travel.get("hours_per_trip")))
        _write(estimate, "E139", number_or_none(travel.get("crew_size")))
        _write(estimate, "G139", number_or_none(travel.get("hourly_rate")))

    _apply_formula_recalc_flags(workbook)
    workbook.save(output_path)
    return output_path
