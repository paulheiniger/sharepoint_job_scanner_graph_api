from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
from collections import Counter
from pathlib import Path
from typing import Any

import pandas as pd
from sqlalchemy import bindparam, create_engine, inspect, text
from sqlalchemy.engine import Connection, Engine

PARSER_VERSION = "document-content-template-v4"
TEMPLATE_TYPE_ROOFING = "roofing"
TEMPLATE_TYPE_INSULATION = "insulation"
TEMPLATE_TYPE_FLOORING = "flooring"
TEMPLATE_TYPE_UNKNOWN = "unknown"

ROOFING_HEADER_BUCKETS = {
    1: "estimate_date",
    2: "job_name",
    3: "job_type",
    4: "site_address",
    5: "city_state_zip",
    6: "contact",
    8: "email",
    9: "phone",
    12: "estimated_square_feet",
}
HEADER_BUCKETS = ROOFING_HEADER_BUCKETS

ROOFING_MATERIAL_BUCKETS = {
    19: "foam",
    20: "foam",
    21: "foam",
    26: "coating",
    27: "coating",
    28: "coating",
    33: "thinner",
    36: "granules",
    39: "primer",
    43: "caulk_sealant",
    45: "caulk_sealant",
    47: "seams_misc",
    49: "penetrations",
    51: "hvac_units",
    53: "drains",
    58: "board_stock",
    59: "board_stock",
    60: "board_stock",
    63: "fasteners",
    65: "plates",
    69: "dumpsters",
    73: "lift",
    74: "lift",
    76: "delivery_fee",
    79: "fabric",
    82: "edge_metal",
    84: "gutter",
    86: "downspouts",
    88: "roof_hatch",
    90: "scuppers",
    92: "curbs",
    94: "ladders",
    96: "pitch_pockets",
    99: "generator",
    101: "misc",
    103: "freight",
    106: "sales_inspection_trips",
    108: "truck_expense",
}
MATERIAL_BUCKETS = ROOFING_MATERIAL_BUCKETS

ROOFING_LABOR_BUCKETS = {
    116: "labor_prep",
    118: "labor_prime",
    120: "labor_seam_sealer",
    122: "labor_base",
    124: "labor_top_coat",
    126: "labor_caulk",
    128: "labor_details",
    130: "labor_top_coat_granules",
    132: "labor_cleanup",
    134: "labor_misc",
    137: "labor_loading",
    139: "labor_traveling",
    142: "infrared_scan",
    145: "meals_lodging",
}
LABOR_BUCKETS = ROOFING_LABOR_BUCKETS

ROOFING_TOTAL_BUCKETS = {
    154: "warranty",
    156: "misc_insurance",
    158: "permits",
    163: "total_job_cost",
    165: "overhead",
    167: "profit",
    169: "worksheet_price",
    170: "worksheet_price_adjusted",
}
TOTAL_BUCKETS = ROOFING_TOTAL_BUCKETS

INSULATION_HEADER_BUCKETS = {
    1: "estimate_date",
    2: "job_name",
    3: "job_type",
    4: "site_address",
    5: "city_state_zip",
    6: "contact",
    8: "email",
    9: "phone",
    12: "estimated_square_feet",
}

INSULATION_MATERIAL_BUCKETS = {
    19: "foam",
    20: "foam",
    21: "foam",
    24: "membrane",
    26: "primer",
    30: "thermal_barrier_coating",
    31: "thermal_barrier_coating",
    32: "thermal_barrier_coating",
    37: "thinner",
    41: "caulk_sealant",
    43: "caulk_sealant",
    47: "lift",
    48: "lift",
    50: "delivery_fee",
    53: "generator",
    55: "space_heater",
    57: "misc",
    59: "freight",
    61: "abaa_audit",
    63: "abaa_fee",
    65: "drum_disposal",
    68: "sales_inspection_trips",
    70: "truck_expense",
    72: "subtotal_materials",
    73: "sales_tax",
}

INSULATION_LABOR_BUCKETS = {
    78: "labor_set_up",
    80: "labor_mask",
    82: "labor_prime",
    84: "labor_membrane",
    86: "labor_foam",
    88: "labor_dc_315",
    90: "labor_misc",
    92: "labor_clean_up",
    95: "labor_loading",
    97: "labor_traveling",
    100: "meals_lodging",
    103: "labor_subtotal",
}

INSULATION_TOTAL_BUCKETS = {
    116: "total_job_cost",
    118: "overhead",
    120: "profit",
    122: "worksheet_price",
    123: "worksheet_price_adjusted",
    137: "price_per_sqft_estimated_sets",
}

FLOORING_HEADER_BUCKETS = {
    1: "estimate_date",
    2: "job_name",
    3: "job_type",
    4: "site_address",
    5: "city_state_zip",
    6: "contact",
    7: "contact_title",
    8: "email",
    9: "phone",
    12: "estimated_square_feet",
}

FLOORING_MATERIAL_BUCKETS = {
    19: "foam",
    20: "foam",
    21: "foam",
    26: "floor_base_coat",
    27: "floor_topcoat",
    28: "floor_coating",
    33: "thinner",
    36: "granules",
    39: "floor_primer",
    43: "caulk_sealant",
    45: "caulk_sealant",
    47: "seams_misc",
    49: "penetrations",
    51: "hvac_units",
    53: "drains",
    58: "board_stock",
    59: "board_stock",
    60: "board_stock",
    63: "fasteners",
    65: "plates",
    69: "dumpsters",
    73: "lift",
    74: "lift",
    76: "delivery_fee",
    79: "fabric",
    99: "generator",
    103: "freight",
    106: "sales_inspection_trips",
    108: "truck_expense",
}

FLOORING_LABOR_BUCKETS = {
    116: "labor_floor_grind_patch",
    118: "labor_floor_pop_off",
    120: "labor_floor_prep_base",
    122: "labor_floor_patch_grind",
    124: "labor_floor_primer",
    126: "labor_floor_base_coat",
    128: "labor_floor_details",
    130: "labor_floor_topcoat",
    132: "labor_floor_misc",
    137: "labor_loading",
    139: "labor_traveling",
    142: "infrared_scan",
    145: "meals_lodging",
}

FLOORING_TOTAL_BUCKETS = {
    154: "warranty",
    156: "misc_insurance",
    158: "permits",
    163: "total_job_cost",
    165: "overhead",
    167: "profit",
    169: "worksheet_price",
    170: "worksheet_price_adjusted",
    184: "price_per_sqft_estimated_sets",
}

ADDER_ROWS = set(range(173, 181))
ADDER_BUCKETS = {row_number: "estimate_adder" for row_number in ADDER_ROWS}

TEMPLATE_BUCKET_BY_ROW = {
    **HEADER_BUCKETS,
    **MATERIAL_BUCKETS,
    **LABOR_BUCKETS,
    **TOTAL_BUCKETS,
    **ADDER_BUCKETS,
}

EQUIPMENT_BUCKETS = {"dumpsters", "lift", "generator", "delivery_fee", "space_heater", "drum_disposal"}
TRAVEL_BUCKETS = {"sales_inspection_trips", "truck_expense", "labor_traveling", "meals_lodging", "freight"}
WARRANTY_BUCKETS = {"warranty", "misc_insurance", "permits", "abaa_audit", "abaa_fee"}
TOTAL_LINE_BUCKETS = {"total_job_cost", "worksheet_price", "worksheet_price_adjusted", "subtotal_materials", "sales_tax", "labor_subtotal", "price_per_sqft_estimated_sets"}
ADDER_TEMPLATE_BUCKETS = {"estimate_adder", "estimate_adder_no_markup", "misc_materials", "misc_equipment"}
ADDER_AMOUNT_COLUMNS = ("F", "H", "G", "E")

CELL_FRAGMENT_RE = re.compile(r"^\s*([A-Z]{1,4}\d+)\s*:\s*(.*)\s*$")

ROOFING_LABOR_LABEL_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("labor_loading", ("set up/safety", "setup/safety", "set-up", "setup", "set up", "mobilization", "loading", "load")),
    ("labor_cleanup", ("touch/clean up", "touch/cleanup", "clean up", "cleanup", "final clean", "job clean")),
    ("infrared_scan", ("infrared", "ir scan", "thermal scan", "moisture scan")),
    ("labor_top_coat_granules", ("granules", "granule", "broadcast")),
    ("labor_caulk", ("aldo 399", "caulk", "sealant")),
    ("labor_details", ("flash curbs", "pitch pockets", "expansion joints", "penetrations", "pipe stands", "flashing", "curbs", "skylights", "rtu", "details", "detail work")),
    ("labor_seam_sealer", ("seam sealer", "seam treatment", "vertical seams", "seams", "seam", "laps")),
    ("labor_top_coat", ("top coat", "topcoat", "finish coat", "second coat")),
    ("labor_base", ("to/foam/base", "t.o. foam base", "foam/base", "base coat", "first coat", "base")),
    ("labor_prep", ("pressure wash", "power wash", "pwash", "pw/prep", "pwash/prep", "clean/prep", "prep/clean", "broom clean", "sweep/prep", "preparation", "prep")),
    ("labor_prime", ("prime coat", "primer", "prime")),
]

ROOFING_LABOR_ROW_HINTS = {
    116: "labor_loading",
    118: "labor_prime",
    120: "labor_seam_sealer",
    122: "labor_base",
    124: "labor_top_coat",
    126: "labor_caulk",
    128: "labor_details",
    130: "labor_top_coat_granules",
    132: "labor_cleanup",
    134: "labor_misc",
    137: "labor_loading",
    139: "labor_traveling",
    142: "infrared_scan",
    145: "meals_lodging",
}

ROOFING_LABOR_BUCKET_SET = set(ROOFING_LABOR_BUCKETS.values()) | {"labor_loading", "labor_setup", "labor_mobilization"}


def maps_for_template_type(template_type: str) -> dict[str, dict[int, str]]:
    if template_type == TEMPLATE_TYPE_INSULATION:
        return {
            "header": INSULATION_HEADER_BUCKETS,
            "materials": INSULATION_MATERIAL_BUCKETS,
            "labor": INSULATION_LABOR_BUCKETS,
            "totals": INSULATION_TOTAL_BUCKETS,
        }
    if template_type == TEMPLATE_TYPE_FLOORING:
        return {
            "header": FLOORING_HEADER_BUCKETS,
            "materials": FLOORING_MATERIAL_BUCKETS,
            "labor": FLOORING_LABOR_BUCKETS,
            "totals": FLOORING_TOTAL_BUCKETS,
        }
    return {
        "header": ROOFING_HEADER_BUCKETS,
        "materials": ROOFING_MATERIAL_BUCKETS,
        "labor": ROOFING_LABOR_BUCKETS,
        "totals": ROOFING_TOTAL_BUCKETS,
    }


def template_bucket_by_row(template_type: str) -> dict[int, str]:
    maps = maps_for_template_type(template_type)
    return {
        **maps["header"],
        **maps["materials"],
        **maps["labor"],
        **maps["totals"],
        **ADDER_BUCKETS,
    }


def numeric_or_text(value: str) -> int | float | str:
    text_value = str(value).strip()
    if not text_value:
        return ""
    cleaned = text_value.replace("$", "").replace(",", "").strip()
    pct = cleaned.endswith("%")
    if pct:
        cleaned = cleaned[:-1].strip()
    try:
        number = float(cleaned)
    except ValueError:
        return text_value
    if pct:
        return number
    if number.is_integer():
        return int(number)
    return number


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text_value = str(value).strip()
    if not text_value or text_value.startswith("="):
        return None
    cleaned = text_value.replace("$", "").replace(",", "").replace("%", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return None


def is_present(value: Any) -> bool:
    if value is None:
        return False
    text_value = str(value).strip()
    return bool(text_value) and text_value.lower() not in {"nan", "none", "null"}


def cell_key(column: str, row_number: int) -> str:
    return f"{column}{row_number}"


def parse_cell_labeled_text(text_content: str) -> tuple[dict[str, Any], dict[str, str], int]:
    cell_values: dict[str, Any] = {}
    formula_cells: dict[str, str] = {}
    malformed_count = 0
    for fragment in str(text_content or "").split("|"):
        fragment = fragment.strip()
        if not fragment:
            continue
        match = CELL_FRAGMENT_RE.match(fragment)
        if not match:
            malformed_count += 1
            continue
        cell_ref, raw_value = match.groups()
        raw_value = raw_value.strip()
        if raw_value.startswith("="):
            formula_cells[cell_ref] = raw_value
        else:
            converted = numeric_or_text(raw_value)
            if converted != "":
                cell_values[cell_ref] = converted
    return cell_values, formula_cells, malformed_count


def template_section_for_bucket(bucket: str, template_type: str = TEMPLATE_TYPE_ROOFING) -> str:
    maps = maps_for_template_type(template_type)
    if bucket in maps["header"].values():
        return "job_header"
    if bucket in maps["materials"].values():
        if bucket in TRAVEL_BUCKETS:
            return "travel"
        if bucket in EQUIPMENT_BUCKETS:
            return "materials"
        return "materials"
    if bucket in maps["labor"].values():
        if bucket in TRAVEL_BUCKETS:
            return "travel"
        if bucket == "labor_subtotal":
            return "totals"
        return "labor"
    if bucket in WARRANTY_BUCKETS:
        return "warranty_bonding_insurance"
    if bucket in {"overhead", "profit"}:
        return "overhead_profit"
    if bucket in TOTAL_LINE_BUCKETS:
        return "totals"
    if bucket in ADDER_TEMPLATE_BUCKETS:
        return "estimate_adders"
    return "other"


def line_item_kind_for_bucket(bucket: str, template_type: str = TEMPLATE_TYPE_ROOFING) -> str:
    maps = maps_for_template_type(template_type)
    if bucket in maps["header"].values():
        return "header"
    if bucket in {
        "foam",
        "coating",
        "floor_base_coat",
        "floor_topcoat",
        "floor_coating",
        "floor_primer",
        "floor_flake",
        "thermal_barrier_coating",
        "membrane",
        "thinner",
        "granules",
        "primer",
        "caulk_sealant",
        "seams_misc",
        "penetrations",
        "hvac_units",
        "drains",
        "board_stock",
        "fasteners",
        "plates",
        "fabric",
        "edge_metal",
        "gutter",
        "downspouts",
        "roof_hatch",
        "scuppers",
        "curbs",
        "ladders",
        "pitch_pockets",
        "misc",
        "misc_materials",
    }:
        return "material"
    if bucket in EQUIPMENT_BUCKETS or bucket == "misc_equipment":
        return "equipment"
    if bucket in TRAVEL_BUCKETS:
        return "travel"
    if bucket in maps["labor"].values() and bucket != "labor_subtotal":
        return "labor"
    if bucket == "warranty":
        return "warranty"
    if bucket == "misc_insurance":
        return "insurance"
    if bucket == "permits":
        return "permit"
    if bucket in {"overhead", "profit"}:
        return "overhead_profit"
    if bucket == "sales_tax":
        return "tax"
    if bucket in TOTAL_LINE_BUCKETS:
        return "total"
    return "unknown" if bucket == "unknown" else "other"


def classify_estimate_adder(row_label: Any) -> tuple[str, str]:
    label = str(row_label or "").strip().lower()
    if "flake" in label:
        return "floor_flake", "material"
    if "lift" in label:
        return "lift", "equipment"
    if "insurance" in label:
        return "misc_insurance", "insurance"
    if "material" in label:
        return "misc_materials", "material"
    if "equipment" in label:
        return "misc_equipment", "equipment"
    if "markup" in label or "w/o markup" in label or "without markup" in label:
        return "estimate_adder_no_markup", "other"
    return "estimate_adder", "other"


def is_placeholder_adder_label(row_label: Any) -> bool:
    label = str(row_label or "").strip().lower()
    return "additional amount" in label and ("markup" in label or "w/o" in label or "without" in label)


def adder_label_text(row_label: Any, cell_values: dict[str, Any], raw_text: str) -> str:
    parts = [str(row_label or ""), str(raw_text or "")]
    parts.extend(str(value or "") for value in cell_values.values())
    return " ".join(part for part in parts if part.strip())


def adder_estimated_cost(cell_values: dict[str, Any], row_number: int) -> float | None:
    for column in ADDER_AMOUNT_COLUMNS:
        amount = numeric_at(cell_values, row_number, column)
        if amount is not None:
            return amount
    return None


def is_unused_placeholder_adder_row(row: dict[str, Any] | pd.Series) -> bool:
    record = row.to_dict() if isinstance(row, pd.Series) else dict(row)
    row_number = int(record.get("row_number") or 0)
    if row_number not in ADDER_ROWS:
        return False
    raw_text = str(record.get("text_content") or "")
    if not re.search(r"\b[A-Z]{1,4}\d+\s*:", raw_text):
        return False
    cell_values, _formula_cells, _malformed_count = parse_cell_labeled_text(raw_text)
    row_label = value_at(cell_values, row_number, "A")
    label_text = adder_label_text(row_label, cell_values, raw_text)
    return is_placeholder_adder_label(label_text) and adder_estimated_cost(cell_values, row_number) is None


def template_row_id_for_content_row(row: dict[str, Any] | pd.Series) -> str:
    record = row.to_dict() if isinstance(row, pd.Series) else dict(row)
    return stable_template_row_id(record.get("document_id"), record.get("sheet_name"), record.get("row_number"), record.get("cell_range"))


def stable_template_row_id(document_id: Any, sheet_name: Any, row_number: Any, cell_range: Any) -> str:
    key = "||".join(str(part or "") for part in (document_id, sheet_name, row_number, cell_range))
    return f"templaterow-{hashlib.sha1(key.encode('utf-8')).hexdigest()[:28]}"


def value_at(cell_values: dict[str, Any], row_number: int, column: str) -> Any:
    return cell_values.get(cell_key(column, row_number))


def numeric_at(cell_values: dict[str, Any], row_number: int, column: str) -> float | None:
    return to_float(value_at(cell_values, row_number, column))


def bounded_numeric_at(
    cell_values: dict[str, Any],
    row_number: int,
    column: str,
    *,
    min_value: float | None = None,
    max_value: float | None = None,
) -> float | None:
    value = numeric_at(cell_values, row_number, column)
    if value is None:
        return None
    if min_value is not None and value < min_value:
        return None
    if max_value is not None and value > max_value:
        return None
    return value


def _drop_implausible_number(
    row: dict[str, Any],
    field: str,
    *,
    min_value: float | None = None,
    max_value: float | None = None,
) -> None:
    value = to_float(row.get(field))
    if value is None:
        return
    if (min_value is not None and value < min_value) or (max_value is not None and value > max_value):
        row[field] = None
        row["needs_review"] = True
        row["parsed_confidence"] = min(float(row.get("parsed_confidence") or 0.55), 0.65)


def _sanitize_estimator_template_row_numbers(row: dict[str, Any]) -> None:
    bucket = normalize_label_text(row.get("template_bucket"))
    if bucket in {"foam", "roofing foam", "roofing_foam"}:
        _drop_implausible_number(row, "thickness_inches", min_value=0.01, max_value=24.0)
        _drop_implausible_number(row, "yield_or_coverage", min_value=100.0, max_value=20000.0)
        if row.get("yield_or_coverage") is None:
            row["yield_factor"] = None
        if row.get("thickness_inches") is None or row.get("yield_or_coverage") is None:
            row["units_per_sqft_per_inch"] = None
            row["sets_per_sqft_per_inch"] = None
            row["cost_per_sqft_per_inch"] = None
    _drop_implausible_number(row, "crew_size", min_value=0.01, max_value=20.0)
    _drop_implausible_number(row, "crew_selector_code", min_value=0.01, max_value=20.0)


def normalize_label_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("&", " and ")
    text = re.sub(r"\bt\.?\s*o\.?\b", "to", text)
    text = text.replace("-", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def label_bucket_override(row_label: Any, raw_text: Any = "") -> str | None:
    text = normalize_label_text(f"{row_label or ''} {raw_text or ''}")
    if any(term in text for term in ("sales/inspect", "sales inspect", "sales inspection", "sales trip")):
        return "sales_inspection_trips"
    if any(term in text for term in ("truck exp", "truck expense", "truck mileage")):
        return "truck_expense"
    return None


def first_numeric_at(cell_values: dict[str, Any], row_number: int, *columns: str) -> float | None:
    for column in columns:
        value = numeric_at(cell_values, row_number, column)
        if value is not None:
            return value
    return None


def insulation_foam_uses_shifted_formula_layout(cell_values: dict[str, Any], row_number: int) -> bool:
    if numeric_at(cell_values, row_number, "C") is not None:
        return False
    return any(numeric_at(cell_values, row_number, column) is not None for column in ("E", "F", "G", "H", "J", "K"))


def _label_matches(text: str, phrase: str) -> bool:
    normalized_phrase = normalize_label_text(phrase)
    if "/" in normalized_phrase:
        return normalized_phrase in text
    return re.search(rf"(?<![a-z0-9]){re.escape(normalized_phrase)}(?![a-z0-9])", text) is not None


def normalize_roofing_labor_bucket(label: Any, description: Any = "", row_number: int | None = None) -> dict[str, Any]:
    raw_text = " ".join(str(part or "") for part in (label, description))
    text = normalize_label_text(raw_text)
    slash_text = re.sub(r"\s*/\s*", "/", text)
    search_text = f"{text} {slash_text}"
    matched: list[str] = []
    matched_phrases: list[str] = []
    for bucket, phrases in ROOFING_LABOR_LABEL_RULES:
        bucket_matched = False
        for phrase in phrases:
            if _label_matches(search_text, phrase):
                bucket_matched = True
                matched_phrases.append(phrase)
        if bucket_matched and bucket not in matched:
            matched.append(bucket)

    primary = matched[0] if matched else ""
    reason = "text_match" if primary else ""
    if not primary and row_number in ROOFING_LABOR_ROW_HINTS:
        primary = ROOFING_LABOR_ROW_HINTS[int(row_number)]
        reason = "row_number_hint"

    if "labor_prep" in matched and "labor_prime" in matched:
        primary = "labor_prep"
        reason = "composite_text_match"
    if "labor_base" in matched and ("foam/base" in search_text or "to/foam/base" in search_text):
        primary = "labor_base"
        reason = "composite_text_match"
    if "labor_cleanup" in matched and "touch" in search_text:
        primary = "labor_cleanup"
        reason = "composite_text_match"

    return {
        "template_bucket": primary,
        "primary_bucket": primary,
        "package_tags": matched,
        "secondary_buckets": [bucket for bucket in matched if bucket != primary],
        "matched_phrases": sorted(set(matched_phrases)),
        "mapping_reason": reason,
        "is_composite": len(matched) > 1 or "/" in str(label or ""),
    }


def detect_template_type_from_rows(rows: list[dict[str, Any] | pd.Series]) -> str:
    source_text_parts: list[str] = []
    row_numbers: set[int] = set()
    parsed_by_row: dict[int, dict[str, Any]] = {}
    for row in rows:
        record = row.to_dict() if isinstance(row, pd.Series) else dict(row)
        row_number = int(record.get("row_number") or 0)
        row_numbers.add(row_number)
        source_text_parts.append(str(record.get("source_file") or record.get("file_name") or ""))
        source_text_parts.append(str(record.get("text_content") or ""))
        if row_number in {3, 12, 78, 86, 103, 116, 122, 123, 137}:
            cell_values, _formula_cells, _malformed_count = parse_cell_labeled_text(str(record.get("text_content") or ""))
            parsed_by_row[row_number] = cell_values
    source_text = " ".join(source_text_parts).lower()
    job_type_text = str(value_at(parsed_by_row.get(3, {}), 3, "C") or "").lower()
    insulation_signals = (
        "estimate insulation",
        "spray foam",
        "closed-cell",
        "closed cell",
        "open-cell",
        "open cell",
        "dc315",
        "thermal barrier",
        "sq ft calculation",
        "foam wall",
        "roof deck insulation",
    )
    roofing_signals = (
        "roof",
        "roofing",
        "coating",
        "silicone",
        "hydrostop",
        "gaf",
        "metal roof",
        "roof replacement",
        "roof coating",
        "seams",
        "fasteners",
        "top coat",
        "base coat",
        "infrared scan",
        "warranty",
    )
    flooring_signals = (
        "estimate flooring",
        "floor system",
        "flooring",
        "concrete floor",
        "concrete slab",
        "polyaspartic",
        "polyspartic",
        "epoxy floor",
        "flake broadcast",
        "grind/patch",
    )
    insulation_score = sum(1 for signal in insulation_signals if signal in source_text or signal in job_type_text)
    roofing_score = sum(1 for signal in roofing_signals if signal in source_text or signal in job_type_text)
    flooring_score = sum(1 for signal in flooring_signals if signal in source_text or signal in job_type_text)
    if "insulation" in job_type_text:
        insulation_score += 3
    if "floor" in job_type_text:
        flooring_score += 3
    if "roof" in job_type_text or "coating" in job_type_text:
        roofing_score += 3
    if flooring_score >= 2 and flooring_score > max(insulation_score, roofing_score):
        return TEMPLATE_TYPE_FLOORING
    if insulation_score >= 2 and insulation_score > roofing_score:
        return TEMPLATE_TYPE_INSULATION
    if 103 in row_numbers and 116 in row_numbers:
        row_103_label = str(value_at(parsed_by_row.get(103, {}), 103, "A") or "").lower()
        row_116_label = str(value_at(parsed_by_row.get(116, {}), 116, "A") or "").lower()
        if "total hours" in row_103_label or "total job cost" in row_116_label:
            return TEMPLATE_TYPE_INSULATION
    if insulation_score >= 1 and any(row_number in row_numbers for row_number in (78, 80, 86, 88, 92)):
        return TEMPLATE_TYPE_INSULATION
    return TEMPLATE_TYPE_ROOFING


def detect_workbook_template_type(path: Path) -> str:
    try:
        import openpyxl
    except ImportError:
        return TEMPLATE_TYPE_UNKNOWN
    workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
    if "Estimate" not in workbook.sheetnames:
        return TEMPLATE_TYPE_UNKNOWN
    if "Sq Ft Calculation" in workbook.sheetnames:
        return TEMPLATE_TYPE_INSULATION
    ws = workbook["Estimate"]
    job_type = str(ws["C3"].value or "").lower()
    if "floor" in job_type:
        return TEMPLATE_TYPE_FLOORING
    if "insulation" in job_type:
        return TEMPLATE_TYPE_INSULATION
    flooring_labels = " ".join(str(ws.cell(row=row, column=1).value or "") for row in (116, 120, 126, 130, 137, 139)).lower()
    if any(term in flooring_labels for term in ("grind/patch", "prep & base", "top coat", "traveling")) and "floor" in str(path.name).lower():
        return TEMPLATE_TYPE_FLOORING
    labels = " ".join(str(ws.cell(row=row, column=1).value or "") for row in (78, 86, 103, 116, 122, 123)).lower()
    if "total job cost" in labels and ("foam" in labels or "total hours" in labels):
        return TEMPLATE_TYPE_INSULATION
    return TEMPLATE_TYPE_ROOFING


def parse_document_content_row(row: dict[str, Any] | pd.Series, template_type: str | None = None) -> dict[str, Any] | None:
    record = row.to_dict() if isinstance(row, pd.Series) else dict(row)
    sheet_name = str(record.get("sheet_name") or "")
    row_number = int(record.get("row_number") or 0)
    raw_text = str(record.get("text_content") or "")
    if sheet_name.lower() != "estimate" or row_number <= 0 or not raw_text:
        return None
    if not re.search(r"\b[A-Z]{1,4}\d+\s*:", raw_text):
        return None

    cell_values, formula_cells, malformed_count = parse_cell_labeled_text(raw_text)
    template_type = template_type or record.get("template_type") or detect_template_type_from_rows([record])
    bucket = template_bucket_by_row(template_type).get(row_number, "unknown")
    row_label = value_at(cell_values, row_number, "A")
    selected_item_name = value_at(cell_values, row_number, "B")
    label_override_bucket = label_bucket_override(row_label, raw_text)
    if label_override_bucket:
        bucket = label_override_bucket
    bucket_mapping: dict[str, Any] = {}
    if row_number in ADDER_ROWS:
        label_text = adder_label_text(row_label, cell_values, raw_text)
        estimated_cost = adder_estimated_cost(cell_values, row_number)
        if is_placeholder_adder_label(label_text) and estimated_cost is None:
            return None
        bucket, kind = classify_estimate_adder(label_text)
        section = "estimate_adders"
    else:
        if template_type == TEMPLATE_TYPE_ROOFING and not label_override_bucket and (
            bucket in ROOFING_LABOR_BUCKET_SET
            or bucket == "unknown"
            or row_number in ROOFING_LABOR_ROW_HINTS
        ):
            bucket_mapping = normalize_roofing_labor_bucket(row_label, selected_item_name or raw_text, row_number)
            mapped_bucket = bucket_mapping.get("primary_bucket")
            if mapped_bucket:
                bucket = str(mapped_bucket)
        section = template_section_for_bucket(bucket, template_type)
        kind = line_item_kind_for_bucket(bucket, template_type)
    parsed_confidence = 0.95 if bucket != "unknown" and malformed_count == 0 else 0.55
    needs_review = bucket == "unknown" or malformed_count > 0

    out: dict[str, Any] = {
        "template_row_id": stable_template_row_id(record.get("document_id"), sheet_name, row_number, record.get("cell_range")),
        "document_id": record.get("document_id"),
        "job_id": record.get("job_id"),
        "source_file": record.get("source_file") or record.get("file_name"),
        "template_type": template_type,
        "sheet_name": sheet_name,
        "row_number": row_number,
        "cell_range": record.get("cell_range"),
        "template_bucket": bucket,
        "template_section": section,
        "line_item_kind": kind,
        "row_label": row_label,
        "raw_text": raw_text,
        "cell_values": cell_values,
        "formula_cells": formula_cells,
        "selected_item_name": row_label if bucket in {"sales_inspection_trips", "truck_expense"} else selected_item_name,
        "package_tags": bucket_mapping.get("package_tags", []),
        "secondary_buckets": bucket_mapping.get("secondary_buckets", []),
        "bucket_mapping_reason": bucket_mapping.get("mapping_reason", ""),
        "is_composite_label": bucket_mapping.get("is_composite", False),
        "quantity": None,
        "unit": None,
        "unit_price": None,
        "estimated_units": None,
        "estimated_cost": None,
        "selector_code": None,
        "resolved_item_name": row_label if bucket in {"sales_inspection_trips", "truck_expense"} else selected_item_name,
        "area_sqft": None,
        "thickness_inches": None,
        "yield_or_coverage": None,
        "yield_factor": None,
        "estimated_sets": None,
        "foam_brand": None,
        "foam_density_lb": None,
        "units_per_sqft_per_inch": None,
        "sets_per_sqft_per_inch": None,
        "cost_per_sqft_per_inch": None,
        "gal_per_100_sqft": None,
        "gal_per_sqft": None,
        "estimated_gallons": None,
        "linear_ft": None,
        "ft_per_unit": None,
        "margin_pct": None,
        "waste_margin_cell": None,
        "quantity_cell_role": None,
        "formula_model": None,
        "days": None,
        "crew_size": None,
        "total_hours": None,
        "daily_rate": None,
        "crew_selector_code": None,
        "hourly_rate": None,
        "calculated_cost": None,
        "formula_mode": None,
        "trips": None,
        "round_trip_miles": None,
        "cost_per_mile": None,
        "warranty_years": None,
        "overhead_pct": None,
        "profit_pct": None,
        "parsed_confidence": parsed_confidence,
        "needs_review": needs_review,
        "parser_version": PARSER_VERSION,
    }

    if section == "materials":
        out["row_label"] = row_label if row_label not in (None, "") else selected_item_name
        out["quantity"] = numeric_at(cell_values, row_number, "C")
        out["unit_price"] = numeric_at(cell_values, row_number, "E")
        out["estimated_units"] = numeric_at(cell_values, row_number, "G")
        out["estimated_cost"] = numeric_at(cell_values, row_number, "H")
        if template_type == TEMPLATE_TYPE_INSULATION:
            out["selector_code"] = numeric_at(cell_values, row_number, "A")
            out["resolved_item_name"] = selected_item_name
            if row_number in {19, 20, 21}:
                shifted_foam_layout = insulation_foam_uses_shifted_formula_layout(cell_values, row_number)
                area_column = "E" if shifted_foam_layout else "C"
                thickness_column = "F" if shifted_foam_layout else "D"
                unit_price_column = "G" if shifted_foam_layout else "E"
                yield_column = "H" if shifted_foam_layout else "F"
                estimated_units_column = "J" if shifted_foam_layout else "G"
                estimated_cost_column = "K" if shifted_foam_layout else "H"
                out["area_sqft"] = numeric_at(cell_values, row_number, area_column)
                out["quantity"] = out["area_sqft"]
                out["unit_price"] = numeric_at(cell_values, row_number, unit_price_column)
                out["estimated_cost"] = numeric_at(cell_values, row_number, estimated_cost_column)
                raw_thickness = numeric_at(cell_values, row_number, thickness_column)
                raw_yield = numeric_at(cell_values, row_number, yield_column)
                out["thickness_inches"] = bounded_numeric_at(cell_values, row_number, thickness_column, min_value=0.01, max_value=24.0)
                out["yield_or_coverage"] = bounded_numeric_at(cell_values, row_number, yield_column, min_value=100.0, max_value=20000.0)
                if raw_thickness is not None and out["thickness_inches"] is None:
                    out["needs_review"] = True
                    out["parsed_confidence"] = min(float(out.get("parsed_confidence") or 0.55), 0.65)
                if raw_yield is not None and out["yield_or_coverage"] is None:
                    out["needs_review"] = True
                    out["parsed_confidence"] = min(float(out.get("parsed_confidence") or 0.55), 0.65)
                out["yield_factor"] = out["yield_or_coverage"]
                out["estimated_units"] = numeric_at(cell_values, row_number, estimated_units_column)
                if out["estimated_units"] is not None:
                    out["estimated_sets"] = out["estimated_units"] / 1000
                density_match = re.search(r"(\d+(?:\.\d+)?)\s*lb", str(out["resolved_item_name"] or ""), flags=re.IGNORECASE)
                if density_match:
                    out["foam_density_lb"] = float(density_match.group(1))
                brand = re.split(r"\s+\d+(?:\.\d+)?\s*lb", str(out["resolved_item_name"] or ""), flags=re.IGNORECASE)[0].strip(" .-")
                out["foam_brand"] = brand.split()[0] if brand else out["resolved_item_name"]
                if out["area_sqft"] and out["thickness_inches"] and out["estimated_units"]:
                    out["units_per_sqft_per_inch"] = out["estimated_units"] / (out["area_sqft"] * out["thickness_inches"])
                    out["sets_per_sqft_per_inch"] = out["estimated_sets"] / (out["area_sqft"] * out["thickness_inches"]) if out["estimated_sets"] is not None else None
                    if out["estimated_cost"]:
                        out["cost_per_sqft_per_inch"] = out["estimated_cost"] / (out["area_sqft"] * out["thickness_inches"])
                out["formula_model"] = "foam_sets_from_area_thickness_yield"
                out["quantity_cell_role"] = "area_sqft"
            elif row_number in {30, 31, 32}:
                out["area_sqft"] = numeric_at(cell_values, row_number, "C")
                out["gal_per_100_sqft"] = numeric_at(cell_values, row_number, "D")
                if out["gal_per_100_sqft"] is not None:
                    out["gal_per_sqft"] = out["gal_per_100_sqft"] / 100
                out["estimated_gallons"] = numeric_at(cell_values, row_number, "G")
                out["formula_model"] = "coating_gallons_from_area_rate_waste"
                out["waste_margin_cell"] = "A34"
                out["quantity_cell_role"] = "area_sqft"
            elif row_number in {41, 43}:
                out["linear_ft"] = numeric_at(cell_values, row_number, "C")
                out["ft_per_unit"] = numeric_at(cell_values, row_number, "D")
                out["formula_model"] = "sealant_units_from_linear_feet_coverage"
                out["quantity_cell_role"] = "linear_ft"
            elif row_number in {47, 48}:
                out["margin_pct"] = numeric_at(cell_values, row_number, "F")
                out["formula_model"] = "equipment_cost_with_margin"
                out["quantity_cell_role"] = "period"
            elif row_number == 37:
                out["formula_model"] = "thinner_units_from_coating_gallons"
                out["quantity_cell_role"] = "estimated_units"
        elif template_type == TEMPLATE_TYPE_FLOORING:
            out["selector_code"] = numeric_at(cell_values, row_number, "A")
            out["resolved_item_name"] = selected_item_name
            if row_number in {26, 27, 28}:
                out["area_sqft"] = numeric_at(cell_values, row_number, "C")
                out["gal_per_100_sqft"] = numeric_at(cell_values, row_number, "D")
                if out["gal_per_100_sqft"] is not None:
                    out["gal_per_sqft"] = out["gal_per_100_sqft"] / 100
                out["estimated_gallons"] = numeric_at(cell_values, row_number, "G")
                out["estimated_units"] = out["estimated_gallons"]
                out["formula_model"] = "floor_coating_gallons_from_area_rate_margin"
                out["waste_margin_cell"] = "A30"
                out["quantity_cell_role"] = "area_sqft"
            elif row_number == 39:
                out["area_sqft"] = numeric_at(cell_values, row_number, "C")
                out["estimated_gallons"] = numeric_at(cell_values, row_number, "G")
                out["estimated_units"] = out["estimated_gallons"]
                out["formula_model"] = "floor_primer_gallons_from_area_coverage"
                out["quantity_cell_role"] = "area_sqft"
    if bucket in {"sales_inspection_trips", "truck_expense"}:
        out["trips"] = numeric_at(cell_values, row_number, "B")
        out["round_trip_miles"] = numeric_at(cell_values, row_number, "C")
        out["cost_per_mile"] = first_numeric_at(cell_values, row_number, "E", "D")
        out["estimated_cost"] = first_numeric_at(cell_values, row_number, "H", "K")
        out["calculated_cost"] = out["estimated_cost"]
    if template_type in {TEMPLATE_TYPE_ROOFING, TEMPLATE_TYPE_FLOORING} and 116 <= row_number <= 134:
        out["days"] = numeric_at(cell_values, row_number, "B")
        out["crew_size"] = numeric_at(cell_values, row_number, "C")
        out["crew_selector_code"] = out["crew_size"]
        out["total_hours"] = numeric_at(cell_values, row_number, "D")
        out["hourly_rate"] = numeric_at(cell_values, row_number, "G")
        out["estimated_cost"] = numeric_at(cell_values, row_number, "H")
        out["calculated_cost"] = out["estimated_cost"]
        out["daily_rate"] = numeric_at(cell_values, row_number, "J")
        out["formula_mode"] = "mixed_formula"
    if template_type == TEMPLATE_TYPE_INSULATION and bucket not in {"sales_inspection_trips", "truck_expense"} and row_number in {78, 80, 82, 84, 86, 88, 90, 92}:
        out["days"] = numeric_at(cell_values, row_number, "B")
        out["crew_size"] = numeric_at(cell_values, row_number, "C")
        out["crew_selector_code"] = out["crew_size"]
        out["total_hours"] = numeric_at(cell_values, row_number, "D")
        out["hourly_rate"] = numeric_at(cell_values, row_number, "G")
        out["estimated_cost"] = numeric_at(cell_values, row_number, "H")
        out["calculated_cost"] = out["estimated_cost"]
        out["daily_rate"] = numeric_at(cell_values, row_number, "J")
        out["formula_mode"] = "mixed_formula"
    if (template_type in {TEMPLATE_TYPE_ROOFING, TEMPLATE_TYPE_FLOORING} and row_number in {137, 139}) or (template_type == TEMPLATE_TYPE_INSULATION and row_number in {95, 97}):
        out["total_hours"] = numeric_at(cell_values, row_number, "C")
        out["crew_size"] = numeric_at(cell_values, row_number, "E")
        out["unit_price"] = numeric_at(cell_values, row_number, "G")
        out["hourly_rate"] = out["unit_price"]
        out["estimated_cost"] = numeric_at(cell_values, row_number, "H")
        out["calculated_cost"] = out["estimated_cost"]
        out["formula_mode"] = "hours_based"
    if template_type == TEMPLATE_TYPE_INSULATION and row_number == 100:
        out["days"] = numeric_at(cell_values, row_number, "C")
        out["crew_size"] = numeric_at(cell_values, row_number, "E")
        out["unit_price"] = numeric_at(cell_values, row_number, "G")
        out["daily_rate"] = out["unit_price"]
        out["estimated_cost"] = numeric_at(cell_values, row_number, "H")
        out["calculated_cost"] = out["estimated_cost"]
        out["formula_mode"] = "days_based"
    if row_number == 154:
        out["warranty_years"] = numeric_at(cell_values, row_number, "C")
        out["quantity"] = numeric_at(cell_values, row_number, "E")
        out["estimated_cost"] = numeric_at(cell_values, row_number, "H")
    if (template_type in {TEMPLATE_TYPE_ROOFING, TEMPLATE_TYPE_FLOORING} and row_number == 165) or (template_type == TEMPLATE_TYPE_INSULATION and row_number == 118):
        out["overhead_pct"] = numeric_at(cell_values, row_number, "F")
        out["estimated_cost"] = numeric_at(cell_values, row_number, "H")
    if (template_type in {TEMPLATE_TYPE_ROOFING, TEMPLATE_TYPE_FLOORING} and row_number == 167) or (template_type == TEMPLATE_TYPE_INSULATION and row_number == 120):
        out["profit_pct"] = numeric_at(cell_values, row_number, "F")
        out["estimated_cost"] = numeric_at(cell_values, row_number, "H")
    if (template_type in {TEMPLATE_TYPE_ROOFING, TEMPLATE_TYPE_FLOORING} and row_number in {163, 169}) or (template_type == TEMPLATE_TYPE_INSULATION and row_number in {72, 73, 103, 116, 122}):
        out["estimated_cost"] = numeric_at(cell_values, row_number, "H")
    if template_type in {TEMPLATE_TYPE_ROOFING, TEMPLATE_TYPE_FLOORING} and row_number == 170:
        out["estimated_cost"] = numeric_at(cell_values, row_number, "F") or numeric_at(cell_values, row_number, "H")
    if template_type == TEMPLATE_TYPE_INSULATION and row_number == 123:
        out["estimated_cost"] = numeric_at(cell_values, row_number, "H")
    if template_type == TEMPLATE_TYPE_INSULATION and row_number == 137:
        out["unit_price"] = numeric_at(cell_values, row_number, "B")
        out["estimated_units"] = numeric_at(cell_values, row_number, "D")
    if row_number in ADDER_ROWS:
        out["row_label"] = row_label
        out["selected_item_name"] = row_label
        out["estimated_cost"] = adder_estimated_cost(cell_values, row_number)
        has_label = is_present(row_label)
        has_numeric_amount = out["estimated_cost"] is not None
        has_formula_amount = any(cell_key(column, row_number) in formula_cells for column in ADDER_AMOUNT_COLUMNS)
        if has_label and has_numeric_amount:
            out["needs_review"] = malformed_count > 0
            out["parsed_confidence"] = 0.9 if malformed_count == 0 else 0.65
        elif is_placeholder_adder_label(adder_label_text(row_label, cell_values, raw_text)) and not has_numeric_amount and not has_formula_amount:
            out["needs_review"] = False
            out["parsed_confidence"] = 0.8
        elif has_label and not has_numeric_amount:
            out["needs_review"] = True
            out["parsed_confidence"] = 0.65
        elif has_numeric_amount and not has_label:
            out["needs_review"] = True
            out["parsed_confidence"] = 0.55
    _sanitize_estimator_template_row_numbers(out)
    return out


def parse_document_content_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    parsed: list[dict[str, Any]] = []
    by_document: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        document_id = str(row.get("document_id") or "")
        by_document.setdefault(document_id, []).append(row)
    for document_rows in by_document.values():
        template_type = detect_template_type_from_rows(document_rows)
        for row in document_rows:
            parsed_row = parse_document_content_row(row, template_type=template_type)
            if parsed_row:
                parsed.append(parsed_row)
    return parsed


TEMPLATE_ROW_COLUMNS = [
    "template_row_id",
    "document_id",
    "job_id",
    "source_file",
    "template_type",
    "sheet_name",
    "row_number",
    "cell_range",
    "template_bucket",
    "template_section",
    "line_item_kind",
    "row_label",
    "raw_text",
    "cell_values",
    "formula_cells",
    "selected_item_name",
    "quantity",
    "unit",
    "unit_price",
    "estimated_units",
    "estimated_cost",
    "selector_code",
    "resolved_item_name",
    "area_sqft",
    "thickness_inches",
    "yield_or_coverage",
    "yield_factor",
    "estimated_sets",
    "foam_brand",
    "foam_density_lb",
    "units_per_sqft_per_inch",
    "sets_per_sqft_per_inch",
    "cost_per_sqft_per_inch",
    "gal_per_100_sqft",
    "gal_per_sqft",
    "estimated_gallons",
    "linear_ft",
    "ft_per_unit",
    "margin_pct",
    "waste_margin_cell",
    "quantity_cell_role",
    "formula_model",
    "days",
    "crew_size",
    "total_hours",
    "daily_rate",
    "crew_selector_code",
    "hourly_rate",
    "calculated_cost",
    "formula_mode",
    "trips",
    "round_trip_miles",
    "cost_per_mile",
    "warranty_years",
    "overhead_pct",
    "profit_pct",
    "parsed_confidence",
    "needs_review",
    "parser_version",
]


def db_row(row: dict[str, Any]) -> dict[str, Any]:
    out = {column: row.get(column) for column in TEMPLATE_ROW_COLUMNS}
    out["cell_values"] = json.dumps(out.get("cell_values") or {}, sort_keys=True)
    out["formula_cells"] = json.dumps(out.get("formula_cells") or {}, sort_keys=True)
    return out


_UPSERT_COLUMNS = ", ".join(TEMPLATE_ROW_COLUMNS)
_UPSERT_VALUES = ", ".join(f":{column}" for column in TEMPLATE_ROW_COLUMNS)
_UPSERT_ASSIGNMENTS = ",\n        ".join(
    f"{column} = excluded.{column}" for column in TEMPLATE_ROW_COLUMNS if column != "template_row_id"
)

UPSERT_TEMPLATE_ROW_SQL = text(
    f"""
    INSERT INTO estimate_template_rows ({_UPSERT_COLUMNS})
    VALUES ({_UPSERT_VALUES})
    ON CONFLICT (template_row_id) DO UPDATE SET
        {_UPSERT_ASSIGNMENTS},
        updated_at = CURRENT_TIMESTAMP
    """
)


def upsert_template_rows(conn: Connection, rows: list[dict[str, Any]], batch_size: int = 1000) -> int:
    if not rows:
        return 0
    total = 0
    batch_size = max(batch_size, 1)
    prepared = [db_row(row) for row in rows]
    for start in range(0, len(prepared), batch_size):
        batch = prepared[start : start + batch_size]
        conn.execute(UPSERT_TEMPLATE_ROW_SQL, batch)
        total += len(batch)
    return total


def delete_template_rows_by_id(conn: Connection, template_row_ids: list[str]) -> int:
    clean_ids = [template_row_id for template_row_id in template_row_ids if template_row_id]
    if not clean_ids:
        return 0
    statement = text("DELETE FROM estimate_template_rows WHERE template_row_id IN :template_row_ids").bindparams(
        bindparam("template_row_ids", expanding=True)
    )
    result = conn.execute(statement, {"template_row_ids": clean_ids})
    return int(result.rowcount or 0)


def fetch_template_candidate_documents(
    conn: Connection,
    *,
    document_id: str | None = None,
    limit_documents: int | None = None,
    document_type: str | None = None,
    xlsx_only: bool = True,
    only_unparsed: bool = False,
) -> list[dict[str, Any]]:
    params: dict[str, Any] = {
        "document_id": document_id,
        "document_type": document_type,
        "limit_documents": limit_documents,
        "parser_version": PARSER_VERSION,
    }
    extension_filter = "AND LOWER(COALESCE(d.file_extension, '')) IN ('.xlsx', '.xlsm')" if xlsx_only else ""
    only_unparsed_filter = (
        """
          AND NOT EXISTS (
              SELECT 1
              FROM estimate_template_rows t
              WHERE t.document_id = c.document_id
                AND t.parser_version = :parser_version
          )
        """
        if only_unparsed
        else ""
    )
    limit_sql = "LIMIT :limit_documents" if limit_documents is not None else ""
    statement = text(
        f"""
        SELECT
            c.document_id,
            COALESCE(d.file_name, c.document_id) AS source_file,
            COUNT(*) AS rows_available
        FROM document_content c
        LEFT JOIN documents d ON d.document_id = c.document_id
        WHERE LOWER(COALESCE(c.sheet_name, '')) = 'estimate'
          AND c.row_number IS NOT NULL
          AND c.text_content ~ '[A-Z]{{1,4}}[0-9]+:'
          {extension_filter}
          AND (:document_id IS NULL OR c.document_id = :document_id)
          AND (:document_type IS NULL OR d.document_type = :document_type)
          {only_unparsed_filter}
        GROUP BY c.document_id, COALESCE(d.file_name, c.document_id)
        ORDER BY COALESCE(d.file_name, c.document_id), c.document_id
        {limit_sql}
        """
    )
    try:
        rows = conn.execute(statement, params).mappings().all()
    except Exception:
        sqlite_extension_filter = "AND LOWER(COALESCE(d.file_extension, '')) IN ('.xlsx', '.xlsm')" if xlsx_only else ""
        sqlite_only_unparsed_filter = (
            """
              AND NOT EXISTS (
                  SELECT 1
                  FROM estimate_template_rows t
                  WHERE t.document_id = c.document_id
                    AND t.parser_version = :parser_version
              )
            """
            if only_unparsed
            else ""
        )
        sqlite_statement = text(
            f"""
            SELECT
                c.document_id,
                COALESCE(d.file_name, c.document_id) AS source_file,
                COUNT(*) AS rows_available
            FROM document_content c
            LEFT JOIN documents d ON d.document_id = c.document_id
            WHERE LOWER(COALESCE(c.sheet_name, '')) = 'estimate'
              AND c.row_number IS NOT NULL
              AND c.text_content LIKE '%:%'
              {sqlite_extension_filter}
              AND (:document_id IS NULL OR c.document_id = :document_id)
              AND (:document_type IS NULL OR d.document_type = :document_type)
              {sqlite_only_unparsed_filter}
            GROUP BY c.document_id, COALESCE(d.file_name, c.document_id)
            ORDER BY COALESCE(d.file_name, c.document_id), c.document_id
            {limit_sql}
            """
        )
        rows = conn.execute(sqlite_statement, params).mappings().all()
    return [dict(row) for row in rows]


def fetch_document_content_rows(conn: Connection, document_id: str | None = None) -> list[dict[str, Any]]:
    params: dict[str, Any] = {"document_id": document_id}
    statement = text(
        """
        SELECT
            c.document_id,
            c.job_id,
            COALESCE(d.file_name, c.document_id) AS source_file,
            c.sheet_name,
            c.row_number,
            c.cell_range,
            c.text_content
        FROM document_content c
        LEFT JOIN documents d ON d.document_id = c.document_id
        WHERE LOWER(COALESCE(c.sheet_name, '')) = 'estimate'
          AND c.row_number IS NOT NULL
          AND c.text_content ~ '[A-Z]{1,4}[0-9]+:'
          AND LOWER(COALESCE(d.file_extension, '')) IN ('.xlsx', '.xlsm')
          AND (:document_id IS NULL OR c.document_id = :document_id)
        ORDER BY c.document_id, c.row_number, c.cell_range
        """
    )
    try:
        rows = conn.execute(statement, params).mappings().all()
    except Exception:
        sqlite_statement = text(
            """
            SELECT
                c.document_id,
                c.job_id,
                COALESCE(d.file_name, c.document_id) AS source_file,
                c.sheet_name,
                c.row_number,
                c.cell_range,
                c.text_content
            FROM document_content c
            LEFT JOIN documents d ON d.document_id = c.document_id
            WHERE LOWER(COALESCE(c.sheet_name, '')) = 'estimate'
              AND c.row_number IS NOT NULL
              AND c.text_content LIKE '%:%'
              AND LOWER(COALESCE(d.file_extension, '')) IN ('.xlsx', '.xlsm')
              AND (:document_id IS NULL OR c.document_id = :document_id)
            ORDER BY c.document_id, c.row_number, c.cell_range
            """
        )
        rows = conn.execute(sqlite_statement, params).mappings().all()
    return [dict(row) for row in rows]


def fetch_flooring_repair_candidate_documents(conn: Connection, limit_documents: int | None = None) -> list[dict[str, Any]]:
    params: dict[str, Any] = {"limit_documents": limit_documents}
    limit_sql = "LIMIT :limit_documents" if limit_documents is not None else ""
    has_jobs = inspect(conn).has_table("jobs")
    jobs_join = "LEFT JOIN jobs j ON j.job_id = c.job_id" if has_jobs else ""
    jobs_signal = "OR LOWER(COALESCE(j.division, '')) = 'flooring'" if has_jobs else ""
    statement = text(
        f"""
        SELECT DISTINCT
            c.document_id,
            COALESCE(MAX(d.file_name), MAX(t.source_file), c.document_id) AS source_file
        FROM document_content c
        LEFT JOIN documents d ON d.document_id = c.document_id
        LEFT JOIN estimate_template_rows t ON t.document_id = c.document_id
        {jobs_join}
        WHERE LOWER(COALESCE(c.sheet_name, '')) = 'estimate'
          AND c.row_number IS NOT NULL
          AND c.text_content ~ '[A-Z]{{1,4}}[0-9]+:'
          AND LOWER(COALESCE(d.file_extension, '')) IN ('.xlsx', '.xlsm')
          AND (
                LOWER(COALESCE(d.file_name, t.source_file, '')) LIKE '%floor%'
             {jobs_signal}
             OR LOWER(COALESCE(c.text_content, '')) LIKE '%floor system%'
             OR LOWER(COALESCE(c.text_content, '')) LIKE '%polyaspartic%'
             OR LOWER(COALESCE(c.text_content, '')) LIKE '%polyspartic%'
             OR LOWER(COALESCE(c.text_content, '')) LIKE '%prep & base 707%'
             OR LOWER(COALESCE(c.text_content, '')) LIKE '%grind/patch%'
             OR LOWER(COALESCE(c.text_content, '')) LIKE '%npi epoxy%'
          )
        GROUP BY c.document_id
        ORDER BY source_file, c.document_id
        {limit_sql}
        """
    )
    try:
        rows = conn.execute(statement, params).mappings().all()
    except Exception:
        sqlite_statement = text(
            f"""
            SELECT DISTINCT
                c.document_id,
                COALESCE(MAX(d.file_name), MAX(t.source_file), c.document_id) AS source_file
            FROM document_content c
            LEFT JOIN documents d ON d.document_id = c.document_id
            LEFT JOIN estimate_template_rows t ON t.document_id = c.document_id
            {jobs_join}
            WHERE LOWER(COALESCE(c.sheet_name, '')) = 'estimate'
              AND c.row_number IS NOT NULL
              AND c.text_content LIKE '%:%'
              AND LOWER(COALESCE(d.file_extension, '')) IN ('.xlsx', '.xlsm')
              AND (
                    LOWER(COALESCE(d.file_name, t.source_file, '')) LIKE '%floor%'
                 {jobs_signal}
                 OR LOWER(COALESCE(c.text_content, '')) LIKE '%floor system%'
                 OR LOWER(COALESCE(c.text_content, '')) LIKE '%polyaspartic%'
                 OR LOWER(COALESCE(c.text_content, '')) LIKE '%polyspartic%'
                 OR LOWER(COALESCE(c.text_content, '')) LIKE '%prep & base 707%'
                 OR LOWER(COALESCE(c.text_content, '')) LIKE '%grind/patch%'
                 OR LOWER(COALESCE(c.text_content, '')) LIKE '%npi epoxy%'
              )
            GROUP BY c.document_id
            ORDER BY source_file, c.document_id
            {limit_sql}
            """
        )
        rows = conn.execute(sqlite_statement, params).mappings().all()
    return [dict(row) for row in rows]


def repair_flooring_template_type(
    engine: Engine,
    *,
    batch_size: int = 1000,
    limit_documents: int | None = None,
    progress: bool = False,
) -> dict[str, Any]:
    with engine.connect() as conn:
        candidate_documents = fetch_flooring_repair_candidate_documents(conn, limit_documents=limit_documents)

    documents_considered = len(candidate_documents)
    if progress:
        print(f"Flooring template repair: documents considered: {documents_considered}", flush=True)

    rows_read = 0
    rows_parsed = 0
    rows_upserted = 0
    review_rows = 0
    bucket_counts: Counter[str] = Counter()
    kind_counts: Counter[str] = Counter()

    for index, document in enumerate(candidate_documents, start=1):
        current_document_id = str(document.get("document_id") or "")
        source_file = str(document.get("source_file") or current_document_id)
        with engine.connect() as conn:
            source_rows = fetch_document_content_rows(conn, document_id=current_document_id)
        parsed_rows = [
            parsed
            for row in source_rows
            if (parsed := parse_document_content_row(row, template_type=TEMPLATE_TYPE_FLOORING))
        ]
        with engine.begin() as conn:
            upserted_for_document = upsert_template_rows(conn, parsed_rows, batch_size=batch_size)

        rows_read += len(source_rows)
        rows_parsed += len(parsed_rows)
        rows_upserted += upserted_for_document
        review_rows += sum(1 for row in parsed_rows if row.get("needs_review"))
        bucket_counts.update(str(row.get("template_bucket")) for row in parsed_rows)
        kind_counts.update(str(row.get("line_item_kind")) for row in parsed_rows)
        if progress:
            print(
                f"[{index}/{documents_considered}] {source_file} — "
                f"rows read: {len(source_rows)}, flooring rows upserted: {upserted_for_document}",
                flush=True,
            )

    return {
        "documents_considered": documents_considered,
        "rows_read": rows_read,
        "rows_parsed": rows_parsed,
        "rows_skipped": rows_read - rows_parsed,
        "rows_upserted": rows_upserted,
        "placeholder_rows_deleted": 0,
        "rows_needing_review": review_rows,
        "by_template_bucket": dict(sorted(bucket_counts.items())),
        "by_line_item_kind": dict(sorted(kind_counts.items())),
    }


def parse_existing_document_content(
    engine: Engine,
    document_id: str | None = None,
    batch_size: int = 1000,
    *,
    limit_documents: int | None = None,
    document_type: str | None = None,
    xlsx_only: bool = True,
    only_unparsed: bool = False,
    progress: bool = False,
) -> dict[str, Any]:
    with engine.connect() as conn:
        candidate_documents = fetch_template_candidate_documents(
            conn,
            document_id=document_id,
            limit_documents=limit_documents,
            document_type=document_type,
            xlsx_only=xlsx_only,
            only_unparsed=only_unparsed,
        )

    documents_considered = len(candidate_documents)
    if progress:
        print(f"Template row parse: documents considered: {documents_considered}", flush=True)

    rows_read = 0
    rows_parsed = 0
    rows_upserted = 0
    rows_deleted = 0
    review_rows = 0
    bucket_counts: Counter[str] = Counter()
    kind_counts: Counter[str] = Counter()

    for index, document in enumerate(candidate_documents, start=1):
        current_document_id = str(document.get("document_id") or "")
        source_file = str(document.get("source_file") or current_document_id)
        with engine.connect() as conn:
            source_rows = fetch_document_content_rows(conn, document_id=current_document_id)
        parsed_rows = parse_document_content_rows(source_rows)
        unused_placeholder_ids = [template_row_id_for_content_row(row) for row in source_rows if is_unused_placeholder_adder_row(row)]
        with engine.begin() as conn:
            deleted_for_document = delete_template_rows_by_id(conn, unused_placeholder_ids)
            upserted_for_document = upsert_template_rows(conn, parsed_rows, batch_size=batch_size)

        rows_read += len(source_rows)
        rows_parsed += len(parsed_rows)
        rows_upserted += upserted_for_document
        rows_deleted += deleted_for_document
        review_rows += sum(1 for row in parsed_rows if row.get("needs_review"))
        bucket_counts.update(str(row.get("template_bucket")) for row in parsed_rows)
        kind_counts.update(str(row.get("line_item_kind")) for row in parsed_rows)
        if progress:
            print(
                f"[{index}/{documents_considered}] {source_file} — "
                f"rows read: {len(source_rows)}, parsed/upserted: {len(parsed_rows)}/{upserted_for_document}",
                flush=True,
            )

    skipped = rows_read - rows_parsed
    return {
        "documents_considered": documents_considered,
        "rows_read": rows_read,
        "rows_parsed": rows_parsed,
        "rows_skipped": skipped,
        "rows_upserted": rows_upserted,
        "placeholder_rows_deleted": rows_deleted,
        "rows_needing_review": review_rows,
        "by_template_bucket": dict(sorted(bucket_counts.items())),
        "by_line_item_kind": dict(sorted(kind_counts.items())),
    }


def load_template_rows_for_document(engine: Engine, document_id: str) -> pd.DataFrame:
    with engine.connect() as conn:
        return pd.read_sql_query(
            text("SELECT * FROM estimate_template_rows WHERE document_id = :document_id ORDER BY row_number, cell_range"),
            conn,
            params={"document_id": document_id},
        )


def load_template_rows_for_job(engine: Engine, job_id: str) -> pd.DataFrame:
    with engine.connect() as conn:
        return pd.read_sql_query(
            text("SELECT * FROM estimate_template_rows WHERE job_id = :job_id ORDER BY document_id, row_number, cell_range"),
            conn,
            params={"job_id": job_id},
        )


def load_template_rows_for_jobs(engine: Engine, job_ids: list[str]) -> pd.DataFrame:
    clean_ids = [str(job_id) for job_id in job_ids if str(job_id).strip()]
    if not clean_ids:
        return pd.DataFrame()
    statement = text(
        "SELECT * FROM estimate_template_rows WHERE job_id IN :job_ids ORDER BY job_id, document_id, row_number, cell_range"
    ).bindparams(bindparam("job_ids", expanding=True))
    with engine.connect() as conn:
        return pd.read_sql_query(statement, conn, params={"job_ids": clean_ids})


def bucket_summary(template_rows: pd.DataFrame) -> pd.DataFrame:
    if template_rows.empty:
        return pd.DataFrame()
    df = template_rows.copy()
    df["estimated_cost"] = pd.to_numeric(df.get("estimated_cost"), errors="coerce").fillna(0)
    return (
        df.groupby(["template_bucket", "line_item_kind"], dropna=False, as_index=False)
        .agg(
            rows=("template_row_id", "count"),
            total_estimated_cost=("estimated_cost", "sum"),
            review_rows=("needs_review", "sum"),
        )
        .sort_values(["rows", "total_estimated_cost"], ascending=[False, False])
    )


def labor_task_summary(template_rows: pd.DataFrame) -> pd.DataFrame:
    if template_rows.empty:
        return pd.DataFrame()
    df = template_rows[template_rows["line_item_kind"] == "labor"].copy()
    if df.empty:
        return pd.DataFrame()
    for column in ("days", "crew_size", "total_hours", "estimated_cost"):
        df[column] = pd.to_numeric(df.get(column), errors="coerce")
    return (
        df.groupby("template_bucket", dropna=False, as_index=False)
        .agg(
            rows=("template_row_id", "count"),
            median_days=("days", "median"),
            median_crew_size=("crew_size", "median"),
            median_total_hours=("total_hours", "median"),
            median_estimated_cost=("estimated_cost", "median"),
        )
        .sort_values("rows", ascending=False)
    )


def material_equipment_travel_summary(template_rows: pd.DataFrame) -> pd.DataFrame:
    if template_rows.empty:
        return pd.DataFrame()
    df = template_rows[template_rows["line_item_kind"].isin(["material", "equipment", "travel"])].copy()
    if df.empty:
        return pd.DataFrame()
    for column in ("quantity", "unit_price", "estimated_units", "estimated_cost"):
        df[column] = pd.to_numeric(df.get(column), errors="coerce")
    return (
        df.groupby(["template_bucket", "line_item_kind"], dropna=False, as_index=False)
        .agg(
            rows=("template_row_id", "count"),
            median_quantity=("quantity", "median"),
            median_unit_price=("unit_price", "median"),
            median_estimated_units=("estimated_units", "median"),
            median_estimated_cost=("estimated_cost", "median"),
        )
        .sort_values("rows", ascending=False)
    )


def totals_for_document(template_rows: pd.DataFrame) -> dict[str, float | None]:
    if template_rows.empty:
        return {}
    out: dict[str, float | None] = {}
    for bucket in ("total_job_cost", "overhead", "profit", "worksheet_price", "worksheet_price_adjusted"):
        rows = template_rows[template_rows["template_bucket"] == bucket]
        out[bucket] = to_float(rows.iloc[0].get("estimated_cost")) if not rows.empty else None
    return out


def print_summary(summary: dict[str, Any]) -> None:
    print("Template row parse final summary:")
    print(f"Documents considered: {summary.get('documents_considered', 0)}")
    print(f"Rows read: {summary.get('rows_read', 0)}")
    print(f"Rows parsed: {summary.get('rows_parsed', 0)}")
    print(f"Rows skipped: {summary.get('rows_skipped', 0)}")
    print(f"Rows upserted: {summary.get('rows_upserted', 0)}")
    print(f"Rows needing review: {summary.get('rows_needing_review', 0)}")
    print("Counts by template_bucket:")
    for bucket, count in (summary.get("by_template_bucket") or {}).items():
        print(f"  {bucket}: {count}")
    print("Counts by line_item_kind:")
    for kind, count in (summary.get("by_line_item_kind") or {}).items():
        print(f"  {kind}: {count}")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Parse document_content XLSX Estimate rows into structured template rows.")
    parser.add_argument("--parse-existing", action="store_true", help="Parse all existing document_content Estimate rows.")
    parser.add_argument("--repair-flooring-template-type", action="store_true", help="Force likely flooring estimate documents through the flooring template row map.")
    parser.add_argument("--document-id", help="Parse one document_id from document_content.")
    parser.add_argument("--limit-documents", type=int, help="Maximum number of candidate documents to parse.")
    parser.add_argument("--document-type", help="Only parse documents with this documents.document_type value.")
    parser.add_argument("--xlsx-only", action="store_true", default=True, help="Restrict parsing to .xlsx/.xlsm documents. This is the default.")
    parser.add_argument("--only-unparsed", action="store_true", help="Only parse documents without rows for the current parser version.")
    parser.add_argument("--database-url", default=os.getenv("DATABASE_URL") or os.getenv("NEON_DATABASE_URL"))
    parser.add_argument("--batch-size", type=int, default=1000)
    args = parser.parse_args(argv)
    if not args.parse_existing and not args.document_id and not args.repair_flooring_template_type:
        parser.error("Use --parse-existing, --document-id, or --repair-flooring-template-type.")
    return args


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    if not args.database_url:
        raise SystemExit("Set --database-url, DATABASE_URL, or NEON_DATABASE_URL.")
    engine = create_engine(args.database_url, future=True)
    if args.repair_flooring_template_type:
        summary = repair_flooring_template_type(
            engine,
            batch_size=args.batch_size,
            limit_documents=args.limit_documents,
            progress=True,
        )
    else:
        summary = parse_existing_document_content(
            engine,
            document_id=args.document_id,
            batch_size=args.batch_size,
            limit_documents=args.limit_documents,
            document_type=args.document_type,
            xlsx_only=args.xlsx_only,
            only_unparsed=args.only_unparsed,
            progress=True,
        )
    print_summary(summary)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
