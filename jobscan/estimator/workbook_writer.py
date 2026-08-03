from __future__ import annotations

import math
import re
from datetime import date
from pathlib import Path
from typing import Any

from .rules import first_nonblank, to_float

DEFAULT_ROOFING_ESTIMATE_TEMPLATE_PATH = Path("templates/Estimate + Spec - Roofing.xlsx")
DEFAULT_INSULATION_ESTIMATE_TEMPLATE_PATH = Path(
    "templates/Estimate + Spec - Insulation.xlsx"
)
DEFAULT_FLOORING_ESTIMATE_TEMPLATE_PATH = Path(
    "templates/Estimate + Spec - Flooring.xlsx"
)
DEFAULT_ESTIMATE_TEMPLATE_PATH = Path("templates/Estimate - Full Turnkey.xlsx")
FALLBACK_ESTIMATE_TEMPLATE_PATH = Path("data/estimate_samples/Estimate - Full Turnkey.xlsx")
DEFAULT_ESTIMATE_OUTPUT_DIR = Path("output/estimates")

HEADER_CELLS = {
    "C2_job_name": "C2",
    "C3_job_type": "C3",
    "C4_site_address": "C4",
    "C5_city_state_zip": "C5",
    "C6_contact": "C6",
    "C7_title": "C7",
    "C8_email": "C8",
    "C9_phone": "C9",
    "G1_bill_to_address": "G1",
    "G2_bill_to_city_state_zip": "G2",
    "G3_bill_to_contact": "G3",
    "G4_bill_to_email": "G4",
    "G5_bill_to_phone": "G5",
    "G6_core_sample": "G6",
    "G7_structural_deck": "G7",
    "G8_existing_substrate": "G8",
    "G9_estimator": "G9",
    "C12_estimated_sqft": "C12",
}
INSULATION_HEADER_CELLS = {
    "C2_job_name": "C2",
    "C3_job_type": "C3",
    "C4_site_address": "C4",
    "C5_city_state_zip": "C5",
}

COATING_ROWS = [26, 27, 28]
INSULATION_FOAM_ROWS = [19, 20, 21]
INSULATION_THERMAL_BARRIER_ROWS = [30, 31, 32]
ROOFING_LABOR_LABELS_BY_TASK = {
    "labor_setup_safety": ("setup safety", "full repair"),
    "labor_full_repair": ("full repair",),
    "full_repair": ("full repair",),
    "labor_prep": ("pwash prep", "pw prep", "clean prep", "prep clean", "preparation", "prep"),
    "labor_prime": ("prime", "prime coat", "primer"),
    "labor_tearoff": ("tear off", "tearoff"),
    "labor_tear_off": ("tear off", "tearoff"),
    "labor_seam_sealer": ("seam sealer", "seam treatment"),
    "labor_board": ("board", "board installation"),
    "labor_base": ("base", "base coat", "foam base", "to foam base"),
    "labor_top_coat": ("top coat", "top coat gran"),
    "labor_top_coat_granules": ("top coat gran", "top coat granules"),
    "labor_caulk": ("caulk sf", "caulk"),
    "labor_details": ("details", "detail work"),
    "labor_cleanup": ("touch clean up", "cleanup", "clean up"),
    "labor_loading": ("loading",),
    "labor_traveling": ("traveling", "travel"),
    "labor_floor_grind_patch": ("trip 1 grind", "grind patch"),
    "labor_floor_corner_repair": ("trip 2 corner", "corner repair"),
    "labor_floor_prep_base_flake": ("trip 3 prep", "prep base 707 flake"),
    "labor_floor_patch_grind": ("patch grind",),
    "labor_floor_primer": ("primer",),
    "labor_floor_base_707": ("base coat 707",),
    "labor_floor_details": ("details",),
    "labor_floor_top_coat": ("trip 4 top coat", "top coat"),
    "labor_floor_cleanup": ("touch clean up", "cleanup", "clean up"),
}
ROOFING_DETAIL_LABELS_BY_CATEGORY = {
    "seams_misc": ("misc seams", "miscellaneous seams"),
    "penetrations": ("penetrations",),
    "hvac_units": ("hvac units", "hvac"),
    "drains": ("drains",),
}
ROOFING_ACCESSORY_LABELS_BY_CATEGORY = {
    "edge_metal": ("edge metal",),
    "gutter": ("gutter", "gutters"),
    "downspouts": ("downspouts", "downspout"),
    "roof_hatch": ("roof hatch",),
    "scuppers": ("scuppers", "scupper"),
    "curbs": ("curbs", "curb"),
    "ladders": ("ladders", "ladder"),
    "pitch_pockets": ("pitch pockets", "pitch pocket"),
    "misc": ("misc", "miscellaneous"),
}
INSULATION_LABOR_ROW_BY_TASK = {
    "labor_set_up": 78,
    "set_up": 78,
    "labor_mask": 80,
    "mask": 80,
    "labor_prime": 82,
    "labor_membrane": 84,
    "labor_foam": 86,
    "foam": 86,
    "labor_dc_315": 88,
    "dc_315": 88,
    "labor_misc": 90,
    "labor_clean_up": 92,
    "labor_cleanup": 92,
    "labor_loading": 95,
    "labor_traveling": 97,
    "meals_lodging": 100,
}

MARKUP_PERCENT_CELLS = {
    "roofing": {"overhead_pct": "F165", "profit_pct": "F167"},
    "flooring": {"overhead_pct": "F165", "profit_pct": "F167"},
    "insulation": {"overhead_pct": "F118", "profit_pct": "F120"},
}


def resolve_default_template_path() -> Path:
    return DEFAULT_ESTIMATE_TEMPLATE_PATH if DEFAULT_ESTIMATE_TEMPLATE_PATH.exists() else FALLBACK_ESTIMATE_TEMPLATE_PATH


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("_")
    return cleaned[:90] or "estimate_draft"


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _write_cell(
    ws: Any,
    cell: str,
    value: Any,
    *,
    replace_formula: bool = False,
) -> bool:
    if value is None or value == "":
        return False
    if is_formula(ws[cell].value) and not replace_formula:
        return False
    ws[cell] = value
    return True


def _add_comment(ws: Any, cell: str, text: str) -> None:
    if not text:
        return
    from openpyxl.comments import Comment

    ws[cell].comment = Comment(text[:30000], "Estimator")


def _number(value: Any) -> float | None:
    number = to_float(value)
    if number is None or not math.isfinite(number):
        return None
    return number


def _quantity(row: dict[str, Any]) -> float | int | None:
    value = _number(row.get("quantity"))
    if value is None:
        return None
    return int(value) if float(value).is_integer() else value


def _estimated_sqft(draft_workbook_inputs: dict[str, Any]) -> float | None:
    header = draft_workbook_inputs.get("header") or {}
    return _number(header.get("C12_estimated_sqft") or header.get("estimated_sqft") or header.get("surface_area_sqft"))


def _template_type(draft_workbook_inputs: dict[str, Any], template_path: Path | None = None) -> str:
    explicit = first_nonblank(draft_workbook_inputs.get("template_type"), (draft_workbook_inputs.get("header") or {}).get("template_type")).lower()
    job_type = first_nonblank((draft_workbook_inputs.get("header") or {}).get("C3_job_type")).lower()
    path_text = str(template_path or "").lower()
    if explicit in {"insulation", "roofing", "flooring"}:
        return explicit
    if "insulation" in job_type or "insulation" in path_text:
        return "insulation"
    if "floor" in job_type or "floor" in path_text:
        return "flooring"
    return "roofing"


def _row_text(row: dict[str, Any]) -> str:
    return " ".join(str(row.get(key) or "") for key in ("item", "category", "notes", "task")).lower()


def _normalized_template_label(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _find_labeled_row(
    ws: Any,
    aliases: tuple[str, ...],
    *,
    min_row: int,
    max_row: int,
    columns: tuple[str, ...] = ("A",),
) -> int | None:
    normalized_aliases = {_normalized_template_label(alias) for alias in aliases}
    for row_number in range(min_row, max_row + 1):
        for column in columns:
            label = _normalized_template_label(ws[f"{column}{row_number}"].value)
            if label in normalized_aliases:
                return row_number
    return None


def _first_available_input_row(
    ws: Any,
    rows: tuple[int, ...],
    *,
    input_column: str,
) -> int:
    for row_number in rows:
        if ws[f"{input_column}{row_number}"].value in (None, "", 0):
            return row_number
    return rows[-1]


def _roofing_labor_row(ws: Any, task: str) -> int | None:
    aliases = ROOFING_LABOR_LABELS_BY_TASK.get(task)
    if not aliases:
        return None
    return _find_labeled_row(ws, aliases, min_row=112, max_row=148)


def _roofing_detail_row(ws: Any, category: str) -> int | None:
    aliases = ROOFING_DETAIL_LABELS_BY_CATEGORY.get(category)
    if not aliases:
        return None
    return _find_labeled_row(ws, aliases, min_row=40, max_row=60)


def _roofing_accessory_row(ws: Any, category: str) -> int | None:
    aliases = ROOFING_ACCESSORY_LABELS_BY_CATEGORY.get(category)
    if not aliases:
        return None
    return _find_labeled_row(ws, aliases, min_row=75, max_row=110)


def _manual_adder_rows(ws: Any) -> list[int]:
    rows: list[int] = []
    for row_number in range(1, ws.max_row + 1):
        label = _normalized_template_label(ws[f"A{row_number}"].value)
        if label == "additional amount w o markup":
            rows.append(row_number)
    return rows


def _manual_adder_label(row: dict[str, Any]) -> str:
    label = first_nonblank(row.get("item"), row.get("task"), row.get("flag"), "Review allowance")
    if row.get("needs_review") is True and "review" not in label.lower():
        return f"{label} - REVIEW"
    return label


def _material_basis_note(row: dict[str, Any], fallback: str) -> str:
    measured = _number(row.get("area_sqft"))
    purchase = _number(row.get("basis_sqft"))
    reason = first_nonblank(row.get("quantity_adjustment_reason"))
    parts = [first_nonblank(row.get("notes"), fallback)]
    if measured is not None and purchase is not None and abs(measured - purchase) > 0.5:
        parts.append(
            f"Measured scope: {measured:,.2f} sq ft; formula/purchase basis: {purchase:,.2f} sq ft."
        )
        if reason:
            parts.append(f"Adjustment: {reason}")
    return "\n".join(part for part in parts if part)


def _write_manual_adder(ws: Any, row_number: int, row: dict[str, Any]) -> None:
    label = _manual_adder_label(row)
    amount = _number(row.get("estimated_cost"))
    _write_cell(ws, f"A{row_number}", label)
    if amount is not None:
        _write_cell(ws, f"F{row_number}", round(amount, 2))
    notes = first_nonblank(row.get("notes"), row.get("flag"))
    status = "REVIEW" if row.get("needs_review") or amount is None else ""
    _write_cell(ws, f"G{row_number}", " - ".join(part for part in (status, notes) if part))


def _write_coating_row(ws: Any, row_number: int, row: dict[str, Any], sqft: float | None) -> None:
    item = first_nonblank(row.get("item"), "Roof coating")
    selector_code = _number(row.get("selector_code"))
    gallons = _number(row.get("estimated_gallons") or row.get("quantity"))
    unit_price = _number(row.get("unit_price"))
    area_sqft = _number(row.get("basis_sqft") or row.get("area_sqft")) or sqft
    gal_per_100_sqft = _number(row.get("gal_per_100_sqft"))
    waste_factor_pct = _number(row.get("waste_factor_pct"))
    if selector_code is not None:
        _write_cell(ws, f"A{row_number}", int(selector_code) if float(selector_code).is_integer() else selector_code)
    else:
        _write_cell(ws, f"A{row_number}", item)
    if area_sqft:
        _write_cell(ws, f"C{row_number}", round(area_sqft, 2))
    if gal_per_100_sqft is not None:
        _write_cell(ws, f"D{row_number}", round(gal_per_100_sqft, 4))
    elif area_sqft and gallons:
        _write_cell(ws, f"D{row_number}", round(gallons * 100 / area_sqft, 4))
    elif gallons:
        _write_cell(ws, f"G{row_number}", round(gallons, 2))
    if unit_price is not None:
        _write_cell(ws, f"E{row_number}", round(unit_price, 4))
    if waste_factor_pct is not None:
        _write_cell(ws, "A30", round(waste_factor_pct, 4))
    _add_comment(
        ws,
        f"A{row_number}",
        _material_basis_note(row, "Generated from estimator material plan."),
    )


def _write_primer_row(ws: Any, row: dict[str, Any], sqft: float | None) -> None:
    selector_code = _number(row.get("selector_code"))
    area_sqft = _number(row.get("basis_sqft") or row.get("area_sqft"))
    quantity = _quantity(row)
    unit_price = _number(row.get("unit_price"))
    if selector_code is not None:
        _write_cell(ws, "A39", int(selector_code) if float(selector_code).is_integer() else selector_code)
        if area_sqft is not None:
            _write_cell(ws, "C39", round(area_sqft, 2))
        elif quantity is not None:
            _write_cell(ws, "C39", quantity)
    elif quantity is not None:
        _write_cell(ws, "C39", quantity)
    elif sqft:
        _write_cell(ws, "C39", round(sqft, 2))
    if unit_price is not None:
        _write_cell(ws, "E39", round(unit_price, 4))
    _add_comment(ws, "A39", first_nonblank(row.get("item"), "Primer allowance") + "\n" + first_nonblank(row.get("notes")))


def _write_caulk_sealant_row(ws: Any, row: dict[str, Any]) -> None:
    explicit_row = _number(row.get("workbook_row"))
    row_number = (
        int(explicit_row)
        if explicit_row is not None and int(explicit_row) in {43, 45}
        else _first_available_input_row(ws, (43, 45), input_column="G")
    )
    selector_code = _number(row.get("selector_code"))
    quantity = _number(row.get("estimated_units") or row.get("quantity"))
    unit_price = _number(row.get("unit_price"))
    if selector_code is not None:
        _write_cell(ws, f"A{row_number}", int(selector_code) if float(selector_code).is_integer() else selector_code)
    if unit_price is not None:
        _write_cell(ws, f"E{row_number}", round(unit_price, 4))
    if quantity is not None:
        _write_cell(ws, f"G{row_number}", round(quantity, 4))
    _add_comment(ws, f"A{row_number}", first_nonblank(row.get("item"), "Caulk / sealant allowance") + "\n" + first_nonblank(row.get("notes")))


def _write_fabric_row(ws: Any, row: dict[str, Any]) -> None:
    linear_ft = _number(row.get("linear_ft") or row.get("quantity") or row.get("estimated_units"))
    unit_price = _number(row.get("unit_price"))
    if linear_ft is not None:
        _write_cell(ws, "C79", round(linear_ft, 4))
    if unit_price is not None:
        _write_cell(ws, "E79", round(unit_price, 4))
    _add_comment(ws, "A79", first_nonblank(row.get("item"), "Fabric allowance") + "\n" + first_nonblank(row.get("notes")))


def _write_board_stock_row(ws: Any, row: dict[str, Any]) -> None:
    explicit_row = _number(row.get("workbook_row"))
    row_number = (
        int(explicit_row)
        if explicit_row is not None and int(explicit_row) in {58, 59, 60}
        else _first_available_input_row(ws, (58, 59, 60), input_column="C")
    )
    selector_code = _number(row.get("selector_code"))
    area_sqft = _number(row.get("basis_sqft") or row.get("area_sqft"))
    thickness = _number(row.get("thickness_inches"))
    price_per_square = _number(row.get("price_per_square") or row.get("unit_price"))
    if selector_code is not None:
        _write_cell(ws, f"A{row_number}", int(selector_code) if float(selector_code).is_integer() else selector_code)
    if area_sqft is not None:
        _write_cell(ws, f"C{row_number}", round(area_sqft, 2))
    if thickness is not None:
        _write_cell(ws, f"D{row_number}", round(thickness, 4))
    if price_per_square is not None:
        _write_cell(ws, f"E{row_number}", round(price_per_square, 4))
    _add_comment(
        ws,
        f"A{row_number}",
        first_nonblank(row.get("item"), "Board stock allowance")
        + "\n"
        + _material_basis_note(row, "Generated from estimator material plan."),
    )


def _write_board_fastener_or_plate_row(ws: Any, row: dict[str, Any]) -> None:
    category = str(row.get("category") or row.get("template_bucket") or "").lower()
    explicit_row = _number(row.get("workbook_row"))
    if explicit_row is not None and int(explicit_row) in {63, 65}:
        row_number = int(explicit_row)
    else:
        row_number = 65 if category == "plates" else 63
    unit_price = _number(row.get("unit_price_per_thousand") or row.get("unit_price"))
    quantity = _number(row.get("estimated_units") or row.get("quantity"))
    if unit_price is not None:
        _write_cell(ws, f"E{row_number}", round(unit_price, 4))
    if quantity is not None:
        _write_cell(ws, f"G{row_number}", round(quantity, 4))
    _add_comment(ws, f"A{row_number}", first_nonblank(row.get("item"), "Board fastener/plate allowance") + "\n" + first_nonblank(row.get("notes")))


def _write_granules_row(ws: Any, row: dict[str, Any]) -> None:
    selector_code = _number(row.get("selector_code"))
    area_sqft = _number(row.get("basis_sqft") or row.get("area_sqft"))
    unit_price = _number(row.get("unit_price"))
    quantity = _number(row.get("estimated_units") or row.get("quantity"))
    if selector_code is not None:
        _write_cell(ws, "A36", int(selector_code) if float(selector_code).is_integer() else selector_code)
    if area_sqft is not None:
        _write_cell(ws, "C36", round(area_sqft, 2))
    if unit_price is not None:
        _write_cell(ws, "E36", round(unit_price, 4))
    if quantity is not None:
        _write_cell(ws, "G36", round(quantity, 4))
    _add_comment(ws, "A36", first_nonblank(row.get("item"), "Granules allowance") + "\n" + first_nonblank(row.get("notes")))


def _write_dumpster_row(ws: Any, row: dict[str, Any]) -> None:
    selector_code = _number(row.get("selector_code"))
    area_sqft = _number(row.get("basis_sqft") or row.get("area_sqft"))
    thickness = _number(row.get("debris_thickness_inches") or row.get("thickness_inches"))
    unit_price = _number(row.get("unit_price"))
    margin_pct = _number(row.get("margin_pct"))
    if selector_code is not None:
        _write_cell(ws, "A69", int(selector_code) if float(selector_code).is_integer() else selector_code)
    if area_sqft is not None:
        _write_cell(ws, "C69", round(area_sqft, 2))
    if thickness is not None:
        _write_cell(ws, "D69", round(thickness, 4))
    if unit_price is not None:
        _write_cell(ws, "E69", round(unit_price, 4))
    if margin_pct is not None:
        _write_cell(ws, "F69", round(margin_pct, 4))
    _add_comment(ws, "A69", first_nonblank(row.get("item"), "Dumpster allowance") + "\n" + first_nonblank(row.get("notes")))


def _write_lift_row(ws: Any, row: dict[str, Any]) -> None:
    explicit_row = _number(row.get("workbook_row"))
    row_number = (
        int(explicit_row)
        if explicit_row is not None and int(explicit_row) in {73, 74}
        else _first_available_input_row(ws, (73, 74), input_column="D")
    )
    selector_code = _number(row.get("selector_code"))
    size = first_nonblank(row.get("size"))
    period = _number(row.get("period"))
    unit_price = _number(row.get("unit_price"))
    margin_pct = _number(row.get("margin_pct"))
    if selector_code is not None:
        _write_cell(ws, f"A{row_number}", int(selector_code) if float(selector_code).is_integer() else selector_code)
    if size:
        _write_cell(ws, f"C{row_number}", size)
    if period is not None:
        _write_cell(ws, f"D{row_number}", round(period, 4))
    if unit_price is not None:
        _write_cell(ws, f"E{row_number}", round(unit_price, 4))
    if margin_pct is not None:
        _write_cell(ws, f"F{row_number}", round(margin_pct, 4))
    _add_comment(ws, f"A{row_number}", first_nonblank(row.get("item"), "Lift allowance") + "\n" + first_nonblank(row.get("notes")))


def _write_generator_row(ws: Any, row: dict[str, Any]) -> None:
    days = _number(row.get("days") or row.get("period"))
    unit_price = _number(row.get("unit_price"))
    if days is not None:
        _write_cell(ws, "C99", round(days, 4))
    if unit_price is not None:
        _write_cell(ws, "E99", round(unit_price, 4))
    _add_comment(ws, "A99", first_nonblank(row.get("item"), "Generator allowance") + "\n" + first_nonblank(row.get("notes")))


def _write_delivery_fee_row(ws: Any, row: dict[str, Any]) -> None:
    units = _number(row.get("estimated_units") or row.get("units") or row.get("quantity"))
    unit_price = _number(row.get("unit_price"))
    if unit_price is not None:
        _write_cell(ws, "E76", round(unit_price, 4))
    if units is not None:
        _write_cell(ws, "G76", round(units, 4))
    _add_comment(ws, "A76", first_nonblank(row.get("item"), "Delivery fee") + "\n" + first_nonblank(row.get("notes")))


def _write_freight_row(ws: Any, row: dict[str, Any]) -> None:
    amount = _number(row.get("amount") or row.get("estimated_cost") or row.get("unit_price"))
    if amount is not None:
        _write_cell(ws, "E103", round(amount, 2))
    _add_comment(ws, "A103", first_nonblank(row.get("item"), "Freight") + "\n" + first_nonblank(row.get("notes")))


def _write_roofing_travel_cost_row(ws: Any, row: dict[str, Any]) -> None:
    explicit_row = _number(row.get("workbook_row"))
    category = str(row.get("category") or row.get("template_bucket") or "").lower()
    if category in {"sales_trips", "sales_inspection_trips"}:
        row_number = _find_labeled_row(
            ws,
            ("sales inspect", "sales inspection"),
            min_row=95,
            max_row=115,
        )
    else:
        row_number = _find_labeled_row(
            ws,
            ("truck exp", "truck expense"),
            min_row=95,
            max_row=115,
        )
    if row_number is None and explicit_row is not None and int(explicit_row) in {106, 108}:
        row_number = int(explicit_row)
    if row_number is None:
        return
    trips = _number(row.get("trip_count") or row.get("trips"))
    miles = _number(row.get("round_trip_miles") or row.get("miles"))
    unit_price = _number(row.get("unit_price") or row.get("rate"))
    if trips is not None:
        _write_cell(ws, f"B{row_number}", round(trips, 4))
    if miles is not None:
        _write_cell(ws, f"C{row_number}", round(miles, 4))
    if unit_price is not None:
        _write_cell(ws, f"E{row_number}", round(unit_price, 4))
    _add_comment(ws, f"A{row_number}", first_nonblank(row.get("item"), "Travel / truck expense") + "\n" + first_nonblank(row.get("notes")))


def _write_thinner_row(ws: Any, row: dict[str, Any]) -> None:
    selector_code = _number(row.get("selector_code"))
    unit_price = _number(row.get("unit_price"))
    if selector_code is not None:
        _write_cell(ws, "A33", int(selector_code) if float(selector_code).is_integer() else selector_code)
    if unit_price is not None:
        _write_cell(ws, "E33", round(unit_price, 4))
    _add_comment(ws, "A33", first_nonblank(row.get("item"), "Thinner") + "\n" + first_nonblank(row.get("notes")))


def _write_roofing_accessory_row(ws: Any, row: dict[str, Any]) -> None:
    explicit_row = _number(row.get("workbook_row"))
    category = str(row.get("category") or row.get("template_bucket") or "").lower()
    row_number = _roofing_accessory_row(ws, category)
    if row_number is None and explicit_row is not None:
        row_number = int(explicit_row)
    if row_number is None:
        return
    unit_price = _number(row.get("unit_price"))
    amount = _number(row.get("amount") or row.get("estimated_cost"))
    quantity = _number(row.get("estimated_units") or row.get("units") or row.get("quantity"))
    linear_ft = _number(row.get("linear_ft") or row.get("quantity"))
    if row_number in {82, 84, 86}:
        if linear_ft is not None:
            _write_cell(ws, f"C{row_number}", round(linear_ft, 4))
        if unit_price is not None:
            _write_cell(ws, f"E{row_number}", round(unit_price, 4))
    elif row_number in {88, 90, 92, 94, 96}:
        if unit_price is not None:
            _write_cell(ws, f"E{row_number}", round(unit_price, 4))
        if quantity is not None:
            _write_cell(ws, f"G{row_number}", round(quantity, 4))
    elif row_number == 101:
        if amount is not None:
            _write_cell(ws, "E101", round(amount, 2))
    else:
        if unit_price is not None:
            _write_cell(ws, f"E{row_number}", round(unit_price, 4))
        if quantity is not None:
            _write_cell(ws, f"G{row_number}", round(quantity, 4))
    _add_comment(ws, f"A{row_number}", first_nonblank(row.get("item"), category, "Roof accessory") + "\n" + first_nonblank(row.get("notes")))


def _write_roofing_detail_quantity_row(ws: Any, row: dict[str, Any]) -> None:
    explicit_row = _number(row.get("workbook_row"))
    category = str(row.get("category") or row.get("template_bucket") or "").lower()
    row_number = _roofing_detail_row(ws, category)
    if row_number is None and explicit_row is not None and int(explicit_row) in {47, 49, 51, 53}:
        row_number = int(explicit_row)
    if row_number is None:
        return
    linear_ft = _number(row.get("linear_ft") or row.get("quantity"))
    units = _number(row.get("estimated_units") or row.get("units") or row.get("quantity"))
    amount = _number(row.get("amount") or row.get("estimated_cost"))

    if row_number == 47:
        if linear_ft is not None:
            _write_cell(ws, "C47", round(linear_ft, 4))
    elif units is not None:
        _write_cell(ws, f"D{row_number}", round(units, 4))

    if amount is not None and amount > 0:
        _write_cell(ws, f"H{row_number}", round(amount, 2))

    _add_comment(
        ws,
        f"A{row_number}",
        first_nonblank(row.get("item"), row.get("template_bucket"), "Roof detail quantity") + "\n" + first_nonblank(row.get("notes")),
    )


def _write_roofing_foam_row(ws: Any, row: dict[str, Any]) -> None:
    explicit_row = _number(row.get("workbook_row"))
    row_number = (
        int(explicit_row)
        if explicit_row is not None and int(explicit_row) in {19, 20, 21}
        else _first_available_input_row(ws, (19, 20, 21), input_column="C")
    )
    selector_code = _number(row.get("selector_code") or row.get("editable_selector_code"))
    area_sqft = _number(row.get("basis_sqft") or row.get("area_sqft"))
    thickness = _number(row.get("thickness_inches"))
    unit_price = _number(row.get("unit_price"))
    yield_factor = _number(row.get("yield_factor") or row.get("yield_or_coverage"))

    if selector_code is not None:
        _write_cell(ws, f"A{row_number}", int(selector_code))
    if area_sqft is not None:
        _write_cell(ws, f"C{row_number}", round(area_sqft, 2))
    if thickness is not None:
        _write_cell(ws, f"D{row_number}", round(thickness, 4))
    if unit_price is not None:
        _write_cell(ws, f"E{row_number}", round(unit_price, 4))
    if yield_factor is not None:
        _write_cell(ws, f"F{row_number}", round(yield_factor, 4))

    _add_comment(
        ws,
        f"A{row_number}",
        first_nonblank(row.get("item"), row.get("template_bucket"), "Roofing SPF foam")
        + "\n"
        + _material_basis_note(row, "Generated from estimator material plan."),
    )


def _write_known_material(ws: Any, row: dict[str, Any], sqft: float | None, coating_row_index: int) -> tuple[bool, int]:
    text = _row_text(row)
    category = str(row.get("category") or "").lower()
    explicit_row = _number(row.get("workbook_row"))
    if category in {"roofing_foam", "foam"} or (explicit_row is not None and int(explicit_row) in {19, 20, 21}):
        _write_roofing_foam_row(ws, row)
        return True, coating_row_index
    if category in {"seams_misc", "penetrations", "hvac_units", "drains"} or (
        explicit_row is not None and int(explicit_row) in {47, 49, 51, 53}
    ):
        _write_roofing_detail_quantity_row(ws, row)
        return True, coating_row_index
    if category in {"dumpster", "dumpsters"} or (explicit_row is not None and int(explicit_row) == 69):
        _write_dumpster_row(ws, row)
        return True, coating_row_index
    if category == "lift" or (explicit_row is not None and int(explicit_row) in {73, 74}):
        _write_lift_row(ws, row)
        return True, coating_row_index
    if category == "generator" or (explicit_row is not None and int(explicit_row) == 99):
        _write_generator_row(ws, row)
        return True, coating_row_index
    if category == "delivery_fee" or (explicit_row is not None and int(explicit_row) == 76):
        _write_delivery_fee_row(ws, row)
        return True, coating_row_index
    if category == "freight" or (explicit_row is not None and int(explicit_row) == 103):
        _write_freight_row(ws, row)
        return True, coating_row_index
    if category in {"sales_trips", "sales_inspection_trips", "truck_expense"} or (
        explicit_row is not None and int(explicit_row) in {106, 108}
    ):
        _write_roofing_travel_cost_row(ws, row)
        return True, coating_row_index
    if category == "thinner" or (explicit_row is not None and int(explicit_row) == 33):
        _write_thinner_row(ws, row)
        return True, coating_row_index
    if category in {"edge_metal", "gutter", "downspouts", "roof_hatch", "scuppers", "curbs", "ladders", "pitch_pockets", "misc"} or (
        explicit_row is not None and int(explicit_row) in {82, 84, 86, 88, 90, 92, 94, 96, 101}
    ):
        _write_roofing_accessory_row(ws, row)
        return True, coating_row_index
    if category == "coating":
        if explicit_row is not None and int(explicit_row) in COATING_ROWS:
            target_row = int(explicit_row)
            _write_coating_row(ws, target_row, row, sqft)
            return True, max(coating_row_index, COATING_ROWS.index(target_row) + 1)
        if coating_row_index >= len(COATING_ROWS):
            return False, coating_row_index
        _write_coating_row(ws, COATING_ROWS[coating_row_index], row, sqft)
        return True, coating_row_index + 1
    if "primer" in text:
        _write_primer_row(ws, row, sqft)
        return True, coating_row_index
    if category in {"caulk_detail", "caulk_sealant"} or (row.get("workbook_row") and int(_number(row.get("workbook_row")) or 0) in {43, 45}):
        _write_caulk_sealant_row(ws, row)
        return True, coating_row_index
    if category == "fabric" or (row.get("workbook_row") and int(_number(row.get("workbook_row")) or 0) == 79) or ("fabric" in text and "coating" not in text):
        _write_fabric_row(ws, row)
        return True, coating_row_index
    if category == "board_stock" or (row.get("workbook_row") and int(_number(row.get("workbook_row")) or 0) in {58, 59, 60}):
        _write_board_stock_row(ws, row)
        return True, coating_row_index
    if category in {"fasteners", "fastener_treatment", "plates"} or (row.get("workbook_row") and int(_number(row.get("workbook_row")) or 0) in {63, 65}):
        _write_board_fastener_or_plate_row(ws, row)
        return True, coating_row_index
    if category == "granules" or (row.get("workbook_row") and int(_number(row.get("workbook_row")) or 0) == 36):
        _write_granules_row(ws, row)
        return True, coating_row_index
    if "caulk" in text or "sealant" in text:
        quantity = _quantity(row)
        unit_price = _number(row.get("unit_price"))
        if quantity is not None:
            _write_cell(ws, "G43", quantity)
        if unit_price is not None:
            _write_cell(ws, "E43", round(unit_price, 4))
        _add_comment(ws, "A43", first_nonblank(row.get("item"), "Caulk / sealant allowance") + "\n" + first_nonblank(row.get("notes")))
        return True, coating_row_index
    return False, coating_row_index


def _write_insulation_sqft_calculation(workbook: Any, header: dict[str, Any], sqft: float | None) -> None:
    if "Sq Ft Calculation" not in workbook.sheetnames or sqft is None:
        return
    ws = workbook["Sq Ft Calculation"]
    dimensions = header.get("sqft_calculation_rows") or header.get("dimension_rows") or []
    if isinstance(dimensions, list) and dimensions:
        start_row = 4
        for offset, item in enumerate(dimensions[:12]):
            if not isinstance(item, dict):
                continue
            row_number = start_row + offset
            _write_cell(ws, f"B{row_number}", first_nonblank(item.get("description"), item.get("label"), "Area"))
            signed_area = _number(item.get("area_sqft"))
            if signed_area is not None:
                _write_cell(ws, f"C{row_number}", 1)
                _write_cell(ws, f"D{row_number}", round(signed_area, 2))
            else:
                _write_cell(ws, f"C{row_number}", _number(item.get("height")))
                _write_cell(ws, f"D{row_number}", _number(item.get("width")))
        return
    _write_cell(ws, "B4", "Estimated area from field notes")
    _write_cell(ws, "C4", 1)
    _write_cell(ws, "D4", round(sqft, 2))


def _write_insulation_material(ws: Any, row: dict[str, Any], indexes: dict[str, int]) -> bool:
    text = _row_text(row)
    category = str(row.get("category") or "").lower()
    quantity = _quantity(row)
    unit_price = _number(row.get("unit_price"))
    target_row: int | None = None
    if category == "foam" or "foam" in text:
        if indexes["foam"] >= len(INSULATION_FOAM_ROWS):
            return False
        target_row = INSULATION_FOAM_ROWS[indexes["foam"]]
        indexes["foam"] += 1
    elif "primer" in text:
        target_row = 26
    elif category in {"coating", "thermal_barrier_coating"} or any(term in text for term in ("thermal", "dc 315", "noburn", "coating")):
        if indexes["thermal"] >= len(INSULATION_THERMAL_BARRIER_ROWS):
            return False
        target_row = INSULATION_THERMAL_BARRIER_ROWS[indexes["thermal"]]
        indexes["thermal"] += 1
    elif "membrane" in text:
        target_row = 24
    elif "thinner" in text:
        target_row = 37
    elif "caulk" in text or "sealant" in text:
        target_row = 41 if indexes["caulk"] == 0 else 43
        indexes["caulk"] += 1
    elif "lift" in text:
        target_row = 47 if indexes["lift"] == 0 else 48
        indexes["lift"] += 1
    elif "delivery" in text:
        target_row = 50
    elif "generator" in text:
        target_row = 53
    elif "space heater" in text:
        target_row = 55
    elif "freight" in text:
        target_row = 59
    elif category == "abaa_audits" or "abaa audit" in text:
        target_row = 61
    elif category == "abaa_fee" or "abaa fee" in text:
        target_row = 63
    elif "drum" in text:
        target_row = 65
    elif "sales" in text or "inspection" in text:
        target_row = 68
    elif "truck" in text:
        target_row = 70
    elif "misc" in text:
        target_row = 57
    if target_row is None:
        return False
    if category == "foam" or "foam" in text:
        selector_code = _number(row.get("selector_code"))
        area_sqft = _number(row.get("basis_sqft") or row.get("area_sqft"))
        thickness = _number(row.get("thickness_inches"))
        yield_factor = _number(row.get("yield_factor") or row.get("yield_or_coverage"))
        if selector_code is not None:
            _write_cell(ws, f"A{target_row}", int(selector_code))
        if area_sqft is not None:
            _write_cell(ws, f"C{target_row}", round(area_sqft, 2))
        elif quantity is not None:
            _write_cell(ws, f"C{target_row}", quantity)
        if thickness is not None:
            _write_cell(ws, f"D{target_row}", round(thickness, 4))
        if unit_price is not None:
            _write_cell(ws, f"E{target_row}", round(unit_price, 4))
        if yield_factor is not None:
            _write_cell(ws, f"F{target_row}", round(yield_factor, 4))
    elif category in {"coating", "thermal_barrier_coating"} or any(term in text for term in ("thermal", "dc 315", "noburn", "coating")):
        selector_code = _number(row.get("selector_code"))
        area_sqft = _number(row.get("basis_sqft") or row.get("area_sqft"))
        gal_per_100 = _number(row.get("gal_per_100_sqft"))
        waste_pct = _number(row.get("waste_factor_pct") or row.get("margin_pct"))
        if selector_code is not None:
            _write_cell(ws, f"A{target_row}", int(selector_code))
        if area_sqft is not None:
            _write_cell(ws, f"C{target_row}", round(area_sqft, 2))
        elif quantity is not None:
            _write_cell(ws, f"C{target_row}", quantity)
        if gal_per_100 is not None:
            _write_cell(ws, f"D{target_row}", round(gal_per_100, 4))
        if unit_price is not None:
            _write_cell(ws, f"E{target_row}", round(unit_price, 4))
        if waste_pct is not None:
            _write_cell(ws, "A34", round(waste_pct, 4))
    elif "lift" in text:
        selector_code = _number(row.get("selector_code"))
        size = _number(row.get("size") or row.get("quantity"))
        period = _number(row.get("period") or row.get("days"))
        margin_pct = _number(row.get("margin_pct"))
        if selector_code is not None:
            _write_cell(ws, f"A{target_row}", int(selector_code))
        if size is not None:
            _write_cell(ws, f"C{target_row}", round(size, 2))
        if period is not None:
            _write_cell(ws, f"D{target_row}", round(period, 2))
        if unit_price is not None:
            _write_cell(ws, f"E{target_row}", round(unit_price, 4))
        if margin_pct is not None:
            _write_cell(ws, f"F{target_row}", round(margin_pct, 4))
    elif any(term in text for term in ("generator", "space heater")):
        days = _number(row.get("days") or row.get("period") or row.get("quantity"))
        if days is not None:
            _write_cell(ws, f"C{target_row}", round(days, 2))
        if unit_price is not None:
            _write_cell(ws, f"E{target_row}", round(unit_price, 4))
    elif "sales" in text or "inspection" in text or "truck" in text:
        trips = _number(row.get("trip_count"))
        miles = _number(row.get("round_trip_miles"))
        if trips is not None:
            _write_cell(ws, f"B{target_row}", round(trips, 2))
        if miles is not None:
            _write_cell(ws, f"C{target_row}", round(miles, 2))
        if unit_price is not None:
            _write_cell(ws, f"E{target_row}", round(unit_price, 4))
    elif target_row in {50, 61}:
        if unit_price is not None:
            _write_cell(ws, f"E{target_row}", round(unit_price, 4))
        if quantity is not None:
            _write_cell(ws, f"G{target_row}", quantity)
    elif target_row in {63, 65}:
        if unit_price is not None:
            _write_cell(ws, f"E{target_row}", round(unit_price, 4))
    else:
        selector_code = _number(row.get("selector_code"))
        if selector_code is not None:
            _write_cell(ws, f"A{target_row}", int(selector_code))
        feet_per_unit = _number(row.get("feet_per_unit"))
        if feet_per_unit is not None:
            _write_cell(ws, f"D{target_row}", round(feet_per_unit, 4))
        if quantity is not None:
            _write_cell(ws, f"C{target_row}", quantity)
        if unit_price is not None:
            _write_cell(ws, f"E{target_row}", round(unit_price, 4))
    estimated_cost = _number(row.get("estimated_cost"))
    if estimated_cost is not None:
        _add_comment(ws, f"A{target_row}", f"{first_nonblank(row.get('item'), row.get('category'))}\nEstimator estimated cost: ${estimated_cost:,.2f}\n{first_nonblank(row.get('notes'))}")
    return True


def _write_labor_row(ws: Any, row: dict[str, Any]) -> bool:
    task = first_nonblank(
        row.get("task"),
        row.get("labor_task"),
        row.get("template_bucket"),
    ).strip()
    row_number = _roofing_labor_row(ws, task)
    if row_number is None:
        explicit_row = _number(row.get("workbook_row"))
        if explicit_row is not None and 112 <= int(explicit_row) <= 148:
            row_number = int(explicit_row)
    if row_number is None:
        return False
    if task == "labor_setup_safety":
        _write_cell(
            ws,
            f"A{row_number}",
            first_nonblank(row.get("label"), "Setup/Safety"),
        )
    days = _number(row.get("adjusted_days") or row.get("base_days"))
    crew_size = _number(row.get("crew_size"))
    total_hours = _number(row.get("total_hours"))
    hours_per_trip = _number(row.get("hours_per_trip"))
    hourly_rate = _number(row.get("hourly_rate"))
    estimated_cost = _number(row.get("estimated_cost"))
    if task in {"labor_loading", "labor_traveling"}:
        if hours_per_trip is not None:
            _write_cell(ws, f"C{row_number}", round(hours_per_trip, 2))
        elif total_hours is not None and crew_size:
            _write_cell(ws, f"C{row_number}", round(total_hours / max(crew_size, 1), 2))
        if crew_size is not None:
            _write_cell(ws, f"E{row_number}", int(crew_size))
    else:
        if days is not None:
            _write_cell(ws, f"B{row_number}", round(days, 2))
        if crew_size is not None:
            _write_cell(ws, f"C{row_number}", int(crew_size))
    if hourly_rate is not None:
        _write_cell(ws, f"D{row_number}", round(hourly_rate, 4))
    if total_hours is not None:
        _write_cell(ws, f"G{row_number}", round(total_hours, 4))
    if estimated_cost is not None:
        _add_comment(ws, f"A{row_number}", f"Estimator estimated cost: ${estimated_cost:,.2f}")
    return True


def _write_job_spec(workbook: Any, draft_workbook_inputs: dict[str, Any]) -> None:
    if "Job Spec" not in workbook.sheetnames:
        return
    ws = workbook["Job Spec"]
    scope_lines = draft_workbook_inputs.get("scope_of_work") or []
    if isinstance(scope_lines, str):
        scope_lines = [line.strip() for line in scope_lines.splitlines() if line.strip()]
    if isinstance(scope_lines, list) and scope_lines:
        formatted_lines = []
        for line in scope_lines[:30]:
            text = str(line or "").strip()
            if not text:
                continue
            formatted_lines.append(text if text.startswith(("•", "-")) else f"• {text}")
        if formatted_lines:
            _write_cell(ws, "A13", "\n".join(formatted_lines))

    notes = draft_workbook_inputs.get("spec_notes") or []
    if isinstance(notes, str):
        notes = [line.strip() for line in notes.splitlines() if line.strip()]
    if isinstance(notes, list) and notes:
        notes_label_row = _find_labeled_row(
            ws,
            ("notes",),
            min_row=35,
            max_row=min(ws.max_row, 60),
        )
        if notes_label_row is not None and notes_label_row + 1 <= ws.max_row:
            notes_text = "\n".join(
                str(line or "").strip()
                for line in notes[:20]
                if str(line or "").strip()
            )
            target_cell = f"A{notes_label_row + 1}"
            target_is_merged = any(
                target_cell in merged_range
                for merged_range in ws.merged_cells.ranges
            )
            is_insulation = (
                str(draft_workbook_inputs.get("template_type") or "").lower()
                == "insulation"
            )
            if target_is_merged or not is_insulation:
                _write_cell(ws, target_cell, notes_text)
            else:
                existing_scope = str(ws["A13"].value or "").strip()
                estimator_notes = "Estimator notes:\n" + "\n".join(
                    f"• {line}" for line in notes_text.splitlines() if line
                )
                _write_cell(
                    ws,
                    "A13",
                    "\n\n".join(
                        part for part in (existing_scope, estimator_notes) if part
                    ),
                )

    header = draft_workbook_inputs.get("header") or {}
    mobilizations = _number(header.get("mobilizations"))
    if mobilizations is not None:
        mobilization_row = _find_labeled_row(
            ws,
            ("mobilizations",),
            min_row=1,
            max_row=15,
        )
        if mobilization_row is not None:
            _write_cell(ws, f"B{mobilization_row}", round(mobilizations, 2))

    _write_value_next_to_label(
        ws,
        ("est of days", "estimated days"),
        header.get("estimated_days"),
        min_row=1,
        max_row=15,
    )
    _write_value_next_to_label(
        ws,
        ("est of hours", "estimated hours"),
        header.get("estimated_hours"),
        min_row=1,
        max_row=15,
    )
    _write_value_next_to_label(
        ws,
        ("est crew size", "estimated crew size"),
        header.get("estimated_crew_size"),
        min_row=1,
        max_row=15,
    )
    repair_area_written = _write_value_next_to_label(
        ws,
        ("repair area",),
        header.get("repair_area_description"),
        min_row=1,
        max_row=15,
    )
    if (
        not repair_area_written
        and header.get("repair_area_description")
        and str(draft_workbook_inputs.get("template_type") or "").lower()
        == "roofing"
    ):
        _write_cell(ws, "F9", "Repair Area")
        _write_cell(ws, "G9", header.get("repair_area_description"))
    _write_value_next_to_label(
        ws,
        ("warranty",),
        header.get("warranty_description"),
        min_row=1,
        max_row=10,
        replace_formula=True,
    )


def _write_value_next_to_label(
    ws: Any,
    aliases: tuple[str, ...],
    value: Any,
    *,
    min_row: int,
    max_row: int,
    replace_formula: bool = False,
) -> bool:
    if value is None or value == "":
        return False
    normalized_aliases = {_normalized_template_label(alias) for alias in aliases}
    for row_number in range(max(1, min_row), min(ws.max_row, max_row) + 1):
        for column_number in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_number, column=column_number)
            if _normalized_template_label(cell.value) not in normalized_aliases:
                continue
            target = ws.cell(row=row_number, column=column_number + 1)
            _write_cell(
                ws,
                target.coordinate,
                value,
                replace_formula=replace_formula,
            )
            return True
    return False


def _write_warranty(workbook: Any, draft_workbook_inputs: dict[str, Any]) -> None:
    warranty = draft_workbook_inputs.get("warranty") or {}
    if not warranty or not warranty.get("include", True):
        return
    ws = workbook["Estimate"]
    row_number = _find_labeled_row(
        ws,
        ("warranty",),
        min_row=1,
        max_row=ws.max_row,
    )
    if row_number is None:
        return
    _write_cell(ws, f"B{row_number}", warranty.get("manufacturer"))
    _write_cell(ws, f"C{row_number}", _number(warranty.get("years")))
    _write_cell(ws, f"D{row_number}", warranty.get("warranty_type"))
    _write_cell(ws, f"E{row_number}", _number(warranty.get("area_sqft")))
    _write_cell(ws, f"F{row_number}", _number(warranty.get("unit_cost")))
    _add_comment(
        ws,
        f"A{row_number}",
        first_nonblank(
            warranty.get("notes"),
            "Generated from the approved semantic warranty option.",
        ),
    )


def _prepare_presentation_sheets(workbook: Any) -> None:
    """Remove template-only noise while preserving operational calculations."""
    if "TEST Job Spec" in workbook.sheetnames and "Job Spec" in workbook.sheetnames:
        workbook["TEST Job Spec"].sheet_state = "hidden"

    if "Tracking" in workbook.sheetnames:
        tracking = workbook["Tracking"]
        for row in tracking.iter_rows():
            for cell in row:
                formula = cell.value
                if (
                    isinstance(formula, str)
                    and formula.startswith("=")
                    and not formula.upper().startswith("=IFERROR(")
                ):
                    cell.value = f'=IFERROR({formula[1:]},"")'

    if "Warranty" in workbook.sheetnames and "Estimate" in workbook.sheetnames:
        estimate = workbook["Estimate"]
        warranty_row = _find_labeled_row(
            estimate,
            ("warranty",),
            min_row=1,
            max_row=estimate.max_row,
        )
        warranty_sheet = workbook["Warranty"]
        term_row = _find_labeled_row(
            warranty_sheet,
            ("warranty term",),
            min_row=1,
            max_row=min(warranty_sheet.max_row, 20),
        )
        if warranty_row is not None and term_row is not None:
            formula_cell = next(
                (
                    cell
                    for cell in warranty_sheet[term_row]
                    if isinstance(cell.value, str) and cell.value.startswith("=")
                ),
                None,
            )
            if formula_cell is not None:
                formula_cell.value = (
                    f'=IF(\'Estimate\'!C{warranty_row}>0,'
                    f'\'Estimate\'!C{warranty_row}&" Year "&'
                    f'\'Estimate\'!D{warranty_row},"N/A")'
                )


def _write_insulation_labor_row(ws: Any, row: dict[str, Any]) -> bool:
    task = first_nonblank(row.get("task"), row.get("labor_package")).strip()
    row_number = INSULATION_LABOR_ROW_BY_TASK.get(task)
    if row_number is None:
        return False
    days = _number(row.get("adjusted_days") or row.get("base_days") or row.get("crew_days"))
    crew_size = _number(row.get("crew_size"))
    total_hours = _number(row.get("total_hours") or row.get("labor_hours"))
    hours_per_trip = _number(row.get("hours_per_trip"))
    hourly_rate = _number(row.get("hourly_rate"))
    estimated_cost = _number(row.get("estimated_cost"))
    if row_number in {95, 97}:
        if hours_per_trip is not None:
            _write_cell(ws, f"C{row_number}", round(hours_per_trip, 2))
        elif total_hours is not None and crew_size:
            _write_cell(ws, f"C{row_number}", round(total_hours / max(crew_size, 1), 2))
        if crew_size is not None:
            _write_cell(ws, f"E{row_number}", int(crew_size))
    elif row_number == 100:
        if days is not None:
            _write_cell(ws, "C100", round(days, 2))
        if crew_size is not None:
            _write_cell(ws, "E100", int(crew_size))
    else:
        if days is not None:
            _write_cell(ws, f"B{row_number}", round(days, 2))
        if crew_size is not None:
            _write_cell(ws, f"C{row_number}", int(crew_size))
    if hourly_rate is not None:
        _write_cell(ws, f"D{row_number}", round(hourly_rate, 4))
    if total_hours is not None:
        _write_cell(ws, f"G{row_number}", round(total_hours, 4))
    if estimated_cost is not None:
        _add_comment(ws, f"A{row_number}", f"Estimator estimated cost: ${estimated_cost:,.2f}")
    return True


def _write_travel_row(ws: Any, row: dict[str, Any]) -> dict[str, Any] | None:
    crew_size = _number(row.get("recommended_crew_size") or row.get("crew_size"))
    hours = _number(row.get("travel_labor_hours"))
    if hours is not None:
        if crew_size:
            _write_cell(ws, "C139", round(hours / max(crew_size, 1), 2))
            _write_cell(ws, "E139", int(crew_size))
        else:
            _write_cell(ws, "C139", round(hours, 2))
    vehicle_cost = _number(row.get("travel_vehicle_cost"))
    if vehicle_cost:
        return {
            "item": "Travel / vehicle cost allowance",
            "estimated_cost": vehicle_cost,
            "needs_review": bool(row.get("needs_travel_review")),
            "notes": first_nonblank(row.get("travel_notes"), "Generated from estimator travel plan."),
        }
    return None


def _output_filename(draft_workbook_inputs: dict[str, Any], output_filename: str | None) -> str:
    if output_filename:
        return output_filename if output_filename.lower().endswith(".xlsx") else f"{output_filename}.xlsx"
    header = draft_workbook_inputs.get("header") or {}
    job_name = first_nonblank(header.get("C2_job_name"), "estimate_draft")
    sqft = _estimated_sqft(draft_workbook_inputs)
    suffix = f"_{int(sqft)}sqft" if sqft else ""
    return f"estimate_draft_{safe_filename(job_name)}{suffix}.xlsx"


def _decision_rows_by_type(draft_workbook_inputs: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    """Return workbook decision rows grouped for the existing cell writers.

    The Estimating Assistant now emits decision-native workbook inputs. The
    writer still routes those records through the same low-level cell writers so
    Excel remains the authoritative calculation engine.
    """
    grouped = {"material": [], "labor": [], "travel": [], "adder": []}
    decisions = draft_workbook_inputs.get("workbook_decisions") or []
    for row in decisions:
        if not isinstance(row, dict):
            continue
        row_type = str(row.get("row_type") or "material").lower()
        if row_type not in grouped:
            row_type = "adder" if row_type in {"manual_adder", "review"} else "material"
        grouped[row_type].append(row)
    return grouped


def generate_estimate_workbook(
    draft_workbook_inputs: dict,
    template_path: Path,
    output_dir: Path,
    output_filename: str | None = None,
) -> Path:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("Install openpyxl to generate estimate workbooks.") from exc

    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"Estimate template workbook not found: {template_path}")
    workbook = openpyxl.load_workbook(template_path, data_only=False)
    if "Estimate" not in workbook.sheetnames:
        raise ValueError("Estimate template workbook is missing the 'Estimate' sheet.")
    ws = workbook["Estimate"]
    template_type = _template_type(draft_workbook_inputs, template_path)
    if template_type == "flooring" and "Revised - Patching Only - HJB" in workbook.sheetnames:
        workbook.remove(workbook["Revised - Patching Only - HJB"])

    header = draft_workbook_inputs.get("header") or {}
    _write_cell(ws, "C1", date.today())
    header_cells = INSULATION_HEADER_CELLS if template_type == "insulation" else HEADER_CELLS
    for key, cell in header_cells.items():
        _write_cell(ws, cell, header.get(key))
    pricing = draft_workbook_inputs.get("pricing") or {}
    for key, cell in MARKUP_PERCENT_CELLS.get(template_type, {}).items():
        value = _number(pricing.get(key))
        if value is not None:
            _write_cell(ws, cell, round(value, 4))
    _write_cell(ws, "C12", header.get("C12_estimated_sqft"))
    dimension_lines = [
        f"Gross area: {header.get('gross_area_sqft')}",
        f"Deduction area: {header.get('deduction_area_sqft')}",
        f"Net area: {header.get('net_area_sqft')}",
    ]
    dimension_notes = header.get("dimension_notes") or []
    if isinstance(dimension_notes, str):
        dimension_notes = [dimension_notes]
    _add_comment(ws, "C12", "\n".join([line for line in dimension_lines if not line.endswith("None")] + list(dimension_notes)))

    sqft = _estimated_sqft(draft_workbook_inputs)
    if template_type == "insulation":
        _write_insulation_sqft_calculation(workbook, header, sqft)
    coating_row_index = 0
    insulation_indexes = {"foam": 0, "thermal": 0, "caulk": 0, "lift": 0}
    manual_adders: list[dict[str, Any]] = []
    decision_rows = _decision_rows_by_type(draft_workbook_inputs)
    for row in decision_rows["material"]:
        if not isinstance(row, dict):
            continue
        if template_type == "insulation":
            placed = _write_insulation_material(ws, row, insulation_indexes)
        else:
            placed, coating_row_index = _write_known_material(ws, row, sqft, coating_row_index)
        if not placed:
            manual_adders.append(row)

    for row in decision_rows["labor"]:
        if not isinstance(row, dict):
            continue
        placed = _write_insulation_labor_row(ws, row) if template_type == "insulation" else _write_labor_row(ws, row)
        if not placed:
            manual_adders.append(row)

    for row in decision_rows["travel"]:
        if isinstance(row, dict):
            vehicle_row = _write_travel_row(ws, row)
            if vehicle_row:
                manual_adders.append(vehicle_row)

    for row in decision_rows["adder"]:
        if isinstance(row, dict):
            manual_adders.append(row)

    for row_number, row in zip(_manual_adder_rows(ws), manual_adders):
        _write_manual_adder(ws, row_number, row)

    _write_warranty(workbook, draft_workbook_inputs)
    _write_job_spec(workbook, draft_workbook_inputs)
    _prepare_presentation_sheets(workbook)

    if workbook.calculation is None:
        from openpyxl.workbook.properties import CalcProperties

        workbook.calculation = CalcProperties()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / _output_filename(draft_workbook_inputs, output_filename)
    workbook.save(output_path)
    return output_path
