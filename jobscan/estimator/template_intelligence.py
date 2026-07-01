from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from jobscan.estimator.template_rows import (
    ROOFING_HEADER_BUCKETS,
    ROOFING_LABOR_BUCKETS,
    ROOFING_MATERIAL_BUCKETS,
    ROOFING_TOTAL_BUCKETS,
)


PROJECT_ROOT = Path(__file__).resolve().parents[2]


FOAM_SELECTOR_MAP = {
    "11": "Gaco 2.0 lb.",
    "12": "Gaco 0.5 lb.",
    "21": "NCFI 2.0 lb.",
    "22": "NCFI 0.5 lb.",
    "31": "BASF 2.0 lb.",
    "32": "BASF 0.5 lb.",
    "41": "PSI 2.0 lb.",
    "42": "PSI 0.5 lb.",
    "51": "Demilec 2.0 lb.",
    "52": "Demilec 0.5 lb.",
}


def resolve_template_path(template_path: str | Path) -> Path:
    """Resolve workbook paths used by tests, CLI runs, and template fixtures."""
    path = Path(template_path).expanduser()
    if path.exists():
        return path
    if not path.is_absolute():
        for candidate in (Path.cwd() / path, PROJECT_ROOT / "templates" / path.name):
            if candidate.exists():
                return candidate
    return path


MATERIAL_ROW_MODELS: dict[int, dict[str, Any]] = {
    19: {"bucket": "foam", "model": "foam_sets_from_area_thickness_yield"},
    20: {"bucket": "foam", "model": "foam_sets_from_area_thickness_yield"},
    21: {"bucket": "foam", "model": "foam_sets_from_area_thickness_yield"},
    24: {"bucket": "membrane", "model": "material_cost"},
    26: {"bucket": "primer", "model": "material_cost"},
    30: {"bucket": "thermal_barrier_coating", "model": "coating_gallons_from_area_rate_waste"},
    31: {"bucket": "thermal_barrier_coating", "model": "coating_gallons_from_area_rate_waste"},
    32: {"bucket": "thermal_barrier_coating", "model": "coating_gallons_from_area_rate_waste"},
    37: {"bucket": "thinner", "model": "thinner_units_from_coating_gallons"},
    41: {"bucket": "caulk_sealant", "model": "sealant_units_from_linear_feet_coverage"},
    43: {"bucket": "caulk_sealant", "model": "sealant_units_from_linear_feet_coverage"},
    47: {"bucket": "lift", "model": "equipment_cost_with_margin"},
    48: {"bucket": "lift", "model": "equipment_cost_with_margin"},
    50: {"bucket": "delivery_fee", "model": "fixed_cost"},
    53: {"bucket": "generator", "model": "fixed_cost"},
    55: {"bucket": "space_heater", "model": "fixed_cost"},
    57: {"bucket": "misc", "model": "fixed_cost"},
    59: {"bucket": "freight", "model": "fixed_cost"},
    61: {"bucket": "abaa_audit", "model": "fixed_cost"},
    63: {"bucket": "abaa_fee", "model": "fixed_cost"},
    65: {"bucket": "drum_disposal", "model": "fixed_cost"},
    68: {"bucket": "sales_inspection_trips", "model": "travel_cost"},
    70: {"bucket": "truck_expense", "model": "truck_cost"},
}

LABOR_ROW_MODELS: dict[int, str] = {
    78: "labor_set_up",
    80: "labor_mask",
    82: "labor_prime",
    84: "labor_membrane",
    86: "labor_foam",
    88: "labor_dc_315",
    90: "labor_misc",
    92: "labor_clean_up",
    95: "labor_loading",
    97: "labor_traveling",
    100: "meals_lodging",
}

FOAM_ROLE_BY_COLUMN = {
    "A": "selector_code",
    "B": "resolved_item_name",
    "C": "area_sqft",
    "D": "thickness_inches",
    "E": "unit_price",
    "F": "yield_or_coverage",
    "G": "estimated_units",
    "H": "estimated_cost",
}

COATING_ROLE_BY_COLUMN = {
    "A": "selector_code",
    "B": "resolved_item_name",
    "C": "area_sqft",
    "D": "gal_per_100_sqft",
    "E": "unit_price",
    "G": "estimated_gallons",
    "H": "estimated_cost",
}

THINNER_ROLE_BY_COLUMN = {
    "A": "selector_code",
    "B": "resolved_item_name",
    "E": "unit_price",
    "G": "estimated_units",
    "H": "estimated_cost",
}

SEALANT_ROLE_BY_COLUMN = {
    "A": "selector_code",
    "B": "resolved_item_name",
    "C": "linear_ft",
    "D": "ft_per_unit",
    "E": "unit_price",
    "G": "estimated_units",
    "H": "estimated_cost",
}

LIFT_ROLE_BY_COLUMN = {
    "A": "selector_code",
    "B": "resolved_item_name",
    "C": "size",
    "D": "period",
    "E": "unit_price",
    "F": "margin_pct",
    "H": "estimated_cost",
}

INSULATION_LABOR_ROLE_BY_COLUMN = {
    "A": "labor_task",
    "B": "days",
    "C": "crew_size",
    "D": "total_hours",
    "E": "prevailing_wage",
    "F": "fringe",
    "G": "blended_rate",
    "H": "estimated_cost",
    "J": "daily_rate",
}

INSULATION_SUPPORT_LABOR_ROLE_BY_COLUMN = {
    "A": "labor_task",
    "C": "days_or_hours",
    "E": "crew_size",
    "G": "hourly_or_daily_rate",
    "H": "estimated_cost",
}


ROOFING_LABOR_ROLE_BY_COLUMN = {
    "A": "labor_task",
    "B": "days",
    "C": "crew_size",
    "D": "total_hours",
    "E": "prevailing_wage",
    "F": "fringe",
    "G": "blended_rate",
    "H": "estimated_cost",
    "J": "daily_rate",
}

ROOFING_TOTAL_MARKUP_ROWS = {
    110: "subtotal_materials",
    111: "sales_tax",
    148: "subtotal_labor",
    154: "warranty",
    156: "misc_insurance",
    158: "permits",
    161: "subtotal_warranty_bonding_insurance",
    163: "total_job_cost",
    165: "overhead",
    167: "profit",
    169: "worksheet_price",
    170: "worksheet_price_adjusted",
    184: "price_per_sqft_estimated_sets",
}


PEOPLE_DAILY_RATE_COLUMNS = {
    1: "D",
    2: "E",
    3: "F",
    4: "G",
    5: "H",
    6: "I",
    7: "J",
    8: "K",
}


def _roofing_formula_model_for_bucket(bucket: str, row_number: int) -> str:
    if bucket == "foam":
        return "foam_sets_from_area_thickness_yield"
    if bucket == "coating":
        return "coating_gallons_from_area_rate_waste"
    if bucket == "thinner":
        return "thinner_units_from_coating_gallons"
    if bucket == "granules":
        return "granules_units_from_area_rate"
    if bucket == "primer":
        return "primer_units_from_area_coverage"
    if bucket == "caulk_sealant":
        return "manual_sealant_units_cost"
    if bucket == "board_stock":
        return "board_cost_from_squares"
    if bucket in {"fasteners", "plates"}:
        return "fastener_units_from_board_area"
    if bucket == "dumpsters":
        return "dumpster_count_from_area_thickness_margin"
    if bucket == "lift":
        return "equipment_cost_with_margin"
    if bucket == "fabric":
        return "fabric_cost_from_linear_feet"
    if bucket in {"edge_metal", "gutter", "downspouts"}:
        return "linear_feet_unit_cost"
    if bucket in {"sales_inspection_trips", "truck_expense"}:
        return "travel_cost_from_trips_miles_rate"
    if row_number in {99, 106, 108}:
        return "days_or_trips_rate_cost"
    return "units_rate_cost"


def _roofing_role_map_for_material_row(row_number: int, bucket: str) -> dict[str, str]:
    if bucket == "foam":
        return FOAM_ROLE_BY_COLUMN
    if bucket == "coating":
        return {
            "A": "selector_code",
            "B": "resolved_item_name",
            "C": "area_sqft",
            "D": "gal_per_100_sqft",
            "E": "unit_price",
            "F": "product_reference",
            "G": "estimated_gallons",
            "H": "estimated_cost",
        }
    if bucket == "thinner":
        return THINNER_ROLE_BY_COLUMN
    if bucket == "granules":
        return {
            "A": "selector_code",
            "B": "resolved_item_name",
            "C": "area_sqft",
            "E": "unit_price",
            "F": "color_reference",
            "G": "estimated_units",
            "H": "estimated_cost",
        }
    if bucket == "primer":
        return {
            "A": "selector_code",
            "B": "resolved_item_name",
            "C": "area_sqft",
            "E": "unit_price",
            "G": "estimated_units",
            "H": "estimated_cost",
        }
    if bucket == "caulk_sealant":
        return {
            "A": "selector_code",
            "B": "resolved_item_name",
            "E": "unit_price",
            "G": "estimated_units",
            "H": "estimated_cost",
        }
    if bucket == "board_stock":
        return {
            "A": "selector_code",
            "B": "resolved_item_name",
            "C": "area_sqft",
            "D": "thickness_inches",
            "E": "price_per_square",
            "H": "estimated_cost",
        }
    if bucket in {"fasteners", "plates"}:
        return {"A": "item_name", "E": "unit_price", "G": "estimated_units", "H": "estimated_cost"}
    if bucket == "dumpsters":
        return {
            "A": "selector_code",
            "B": "resolved_item_name",
            "C": "area_sqft",
            "D": "thickness_inches",
            "E": "unit_price",
            "F": "margin_pct",
            "G": "estimated_units",
            "H": "estimated_cost",
        }
    if bucket == "lift":
        return LIFT_ROLE_BY_COLUMN
    if bucket == "fabric":
        return {"A": "item_name", "C": "linear_ft", "D": "width", "E": "unit_price", "H": "estimated_cost"}
    if bucket in {"seams_misc", "penetrations", "hvac_units", "drains"}:
        return {"A": "item_name", "C": "linear_ft", "D": "estimated_units", "H": "estimated_cost"}
    if bucket in {"edge_metal", "gutter", "downspouts"}:
        return {"A": "item_name", "C": "linear_ft", "E": "unit_price", "H": "estimated_cost"}
    if bucket in {"roof_hatch", "scuppers", "curbs", "ladders", "pitch_pockets", "delivery_fee"}:
        return {"A": "item_name", "E": "unit_price", "G": "estimated_units", "H": "estimated_cost"}
    if bucket == "generator":
        return {"A": "item_name", "C": "days", "E": "unit_price", "H": "estimated_cost"}
    if bucket in {"misc", "freight"}:
        return {"A": "item_name", "E": "estimated_cost", "H": "estimated_cost"}
    if bucket in {"sales_inspection_trips", "truck_expense"}:
        return {"A": "item_name", "B": "trip_count", "C": "round_trip_miles", "E": "unit_price", "H": "estimated_cost"}
    return {"A": "item_name", "C": "quantity_or_basis", "E": "unit_price", "G": "estimated_units", "H": "estimated_cost"}


def cell_value(ws: Any, address: str) -> Any:
    return ws[address].value


def extract_if_selector_map(formula: Any) -> dict[str, str]:
    if not isinstance(formula, str) or not formula.startswith("="):
        return {}
    matches = re.findall(r"\$?[A-Z]{1,3}\$?\d+\s*=\s*([0-9]+)\s*,\s*\"([^\"]*)\"", formula)
    return {str(code): label for code, label in matches}


def _formula_without_strings(formula: Any) -> str:
    if not isinstance(formula, str):
        return ""
    return re.sub(r'"[^"]*"', '""', formula)


def formula_dependencies(formula: Any) -> list[str]:
    scrubbed = _formula_without_strings(formula)
    if not scrubbed.startswith("="):
        return []
    refs: list[str] = []
    pattern = re.compile(
        r"(?:(?:'([^']+)'|([A-Za-z0-9_ &]+))!)?\$?([A-Z]{1,3})\$?(\d+)"
        r"(?::\$?([A-Z]{1,3})\$?(\d+))?"
    )
    for match in pattern.finditer(scrubbed):
        sheet = (match.group(1) or match.group(2) or "").strip()
        cell = f"{match.group(3)}{match.group(4)}"
        if match.group(5) and match.group(6):
            cell = f"{cell}:{match.group(5)}{match.group(6)}"
        ref = f"{sheet}!{cell}" if sheet else cell
        if ref not in refs:
            refs.append(ref)
    return refs


def formula_kind(formula: Any) -> str:
    if not isinstance(formula, str) or not formula.startswith("="):
        return ""
    upper = formula.upper()
    if extract_if_selector_map(formula):
        return "if_selector_map"
    if "XLOOKUP(" in upper:
        return "xlookup"
    if "VLOOKUP(" in upper:
        return "vlookup"
    if "INDEX(" in upper and "MATCH(" in upper:
        return "index_match"
    if "IF(" in upper:
        return "conditional_formula"
    if "SUM(" in upper:
        return "sum_formula"
    return "formula"


def _json(value: Any) -> str:
    return json.dumps(value, default=str, sort_keys=True)


def _selector_key(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _as_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        numeric = float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None
    return numeric


def _positive_float(value: Any) -> float | None:
    numeric = _as_float(value)
    if numeric is None or numeric <= 0:
        return None
    return numeric


def foam_product_parts(name: Any) -> dict[str, Any]:
    text = str(name or "").strip()
    density = None
    match = re.search(r"(\d+(?:\.\d+)?)\s*lb", text, flags=re.IGNORECASE)
    if match:
        density = float(match.group(1))
    brand = text
    if text:
        brand = re.split(r"\s+\d+(?:\.\d+)?\s*lb", text, flags=re.IGNORECASE)[0].strip(" .-")
        brand = brand.split()[0] if brand else text
    return {"foam_brand": brand or "", "foam_density_lb": density}


def _derived_area_diagnostics(area: Any, amount: Any, thickness: Any, cost: Any = None) -> dict[str, Any]:
    area_value = _positive_float(area)
    amount_value = _positive_float(amount)
    thickness_value = _positive_float(thickness)
    cost_value = _positive_float(cost)
    if area_value is None or amount_value is None or thickness_value is None:
        return {}
    diagnostics = {
        "units_per_sqft_per_inch": amount_value / (area_value * thickness_value),
        "sets_per_sqft_per_inch": (amount_value / 1000) / (area_value * thickness_value),
    }
    if cost_value is not None:
        diagnostics["cost_per_sqft_per_inch"] = cost_value / (area_value * thickness_value)
    return diagnostics


def _role_map_for_material_row(row_number: int) -> dict[str, str]:
    if row_number in {19, 20, 21}:
        return FOAM_ROLE_BY_COLUMN
    if row_number in {30, 31, 32}:
        return COATING_ROLE_BY_COLUMN
    if row_number == 37:
        return THINNER_ROLE_BY_COLUMN
    if row_number in {41, 43}:
        return SEALANT_ROLE_BY_COLUMN
    if row_number in {47, 48}:
        return LIFT_ROLE_BY_COLUMN
    return {"A": "label_or_selector", "B": "resolved_item_name", "C": "quantity_or_basis", "E": "unit_price", "G": "estimated_units", "H": "estimated_cost"}


def _row_values(ws: Any, row_number: int, max_col: int = 10) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row_number, column=col_idx)
        if cell.value is not None:
            values[cell.column_letter] = cell.value
    return values


def _selector_rows(ws: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_number in [19, 20, 21, 30, 31, 32, 37, 41, 43, 47, 48]:
        formula = cell_value(ws, f"B{row_number}")
        selector_map = extract_if_selector_map(formula)
        if row_number in {19, 20, 21} and len(selector_map) < len(FOAM_SELECTOR_MAP):
            selector_map = FOAM_SELECTOR_MAP
        for code, resolved in selector_map.items():
            rows.append(
                {
                    "sheet_name": ws.title,
                    "row_number": row_number,
                    "selector_cell": f"A{row_number}",
                    "resolved_cell": f"B{row_number}",
                    "selector_code": code,
                    "resolved_item_name": resolved,
                    "formula": formula,
                }
            )
    return rows


def _formula_model_rows(ws: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_number, spec in MATERIAL_ROW_MODELS.items():
        formulas = {col: cell_value(ws, f"{col}{row_number}") for col in ("B", "G", "H") if isinstance(cell_value(ws, f"{col}{row_number}"), str)}
        model = spec["model"]
        basis = ""
        if model == "foam_sets_from_area_thickness_yield":
            basis = (
                f"G{row_number}=((C{row_number}/F{row_number})*D{row_number})*1000 estimated_units; "
                f"estimated_sets=G{row_number}/1000; H{row_number}=E{row_number}*G{row_number}"
            )
        elif model == "coating_gallons_from_area_rate_waste":
            basis = f"G{row_number}=(((C{row_number}/100)*D{row_number}))/((100-A$34)/100); H{row_number}=E{row_number}*G{row_number}"
        rows.append(
            {
                "sheet_name": ws.title,
                "row_number": row_number,
                "template_bucket": spec["bucket"],
                "formula_model": model,
                "role_map": json.dumps(_role_map_for_material_row(row_number), sort_keys=True),
                "formula_basis": basis,
                "formula_cells": json.dumps(formulas, default=str, sort_keys=True),
                "waste_margin_cell": "A34" if row_number in {30, 31, 32} else "",
            }
        )
    for row_number, bucket in LABOR_ROW_MODELS.items():
        formulas = {col: cell_value(ws, f"{col}{row_number}") for col in ("D", "G", "H", "J") if isinstance(cell_value(ws, f"{col}{row_number}"), str)}
        role_map = _insulation_labor_role_map(row_number)
        formula_model = _insulation_labor_formula_model(row_number, formulas)
        regular_labor = row_number in {78, 80, 82, 84, 86, 88, 90, 92}
        if regular_labor:
            basis = (
                f"D{row_number}=(B{row_number}*People!B11)*C{row_number}; "
                f"J{row_number}=People daily rate selected by C{row_number}; "
                f"H{row_number}=IF(G{row_number}=0, B{row_number}*J{row_number}, D{row_number}*G{row_number})"
            )
        elif row_number == 95:
            basis = f"H{row_number}=loading hours/people/rate formula, adjusted by truck expense multiplier B70"
        elif row_number == 97:
            basis = f"H{row_number}=C{row_number}*E{row_number}*G{row_number}*B70"
        elif row_number == 100:
            basis = f"H{row_number}=C{row_number}*E{row_number}*G{row_number}"
        else:
            basis = ""
        rows.append(
            {
                "sheet_name": ws.title,
                "row_number": row_number,
                "template_bucket": bucket,
                "formula_model": formula_model,
                "formula_mode": _insulation_labor_formula_mode(row_number, formulas),
                "role_map": json.dumps(role_map, sort_keys=True),
                "formula_basis": basis,
                "formula_cells": json.dumps(formulas, default=str, sort_keys=True),
                "waste_margin_cell": "",
            }
        )
    return rows


def _material_rows(ws: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_number, spec in MATERIAL_ROW_MODELS.items():
        values = _row_values(ws, row_number, 10)
        role_map = _role_map_for_material_row(row_number)
        selector_map = extract_if_selector_map(values.get("B"))
        if row_number in {19, 20, 21} and len(selector_map) < len(FOAM_SELECTOR_MAP):
            selector_map = FOAM_SELECTOR_MAP
        selector_code = values.get("A") if role_map.get("A") == "selector_code" else None
        resolved_from_selector = selector_map.get(_selector_key(selector_code))
        resolved_item = resolved_from_selector or values.get("B")
        row = {
            "sheet_name": ws.title,
            "row_number": row_number,
            "template_bucket": spec["bucket"],
            "line_item_kind": "material" if spec["bucket"] not in {"lift", "generator", "space_heater"} else "equipment",
            "formula_model": spec["model"],
            "selector_code": selector_code,
            "resolved_item_name": resolved_item,
            "resolved_item_formula": values.get("B") if isinstance(values.get("B"), str) and values.get("B").startswith("=") else "",
            "cell_roles": json.dumps(role_map, sort_keys=True),
            "cell_values": json.dumps(values, default=str, sort_keys=True),
            "selector_map": json.dumps(selector_map, default=str, sort_keys=True),
            "selector_cell_role": "selector_code" if role_map.get("A") == "selector_code" else "label_or_selector",
        }
        for column, role in role_map.items():
            row[role] = values.get(column)
        if resolved_from_selector:
            row["resolved_item_name"] = resolved_from_selector
        if row_number in {19, 20, 21}:
            area = row.get("area_sqft")
            thickness = row.get("thickness_inches")
            yield_factor = row.get("yield_or_coverage")
            estimated_units = _positive_float(row.get("estimated_units"))
            area_value = _positive_float(area)
            thickness_value = _positive_float(thickness)
            yield_value = _positive_float(yield_factor)
            if estimated_units is None and area_value is not None and thickness_value is not None and yield_value is not None:
                estimated_units = ((area_value / yield_value) * thickness_value) * 1000
                row["estimated_units"] = estimated_units
            unit_price = _positive_float(row.get("unit_price"))
            if _positive_float(row.get("estimated_cost")) is None and estimated_units is not None and unit_price is not None:
                row["estimated_cost"] = unit_price * estimated_units
            row["yield_factor"] = yield_factor
            row["estimated_sets"] = estimated_units / 1000 if estimated_units is not None else None
            row.update(foam_product_parts(row.get("resolved_item_name")))
            row.update(_derived_area_diagnostics(area, estimated_units, thickness, row.get("estimated_cost")))
            row["decision_fields"] = "selector_code,resolved_item_name,foam_brand,foam_density_lb,area_sqft,thickness_inches,yield_factor,unit_price"
            row["calculated_output_fields"] = "estimated_units,estimated_sets,estimated_cost,units_per_sqft_per_inch,sets_per_sqft_per_inch,cost_per_sqft_per_inch"
        elif row_number in {30, 31, 32}:
            row["yield_factor"] = None
            row["decision_fields"] = "selector_code,resolved_item_name,area_sqft,gal_per_100_sqft,waste_margin_pct,unit_price"
            row["calculated_output_fields"] = "estimated_gallons,estimated_cost,cost_per_sqft"
        else:
            row["decision_fields"] = ",".join(role for role in role_map.values() if role not in {"estimated_units", "estimated_sets", "estimated_gallons", "estimated_cost"})
            row["calculated_output_fields"] = ",".join(role for role in role_map.values() if role in {"estimated_units", "estimated_sets", "estimated_gallons", "estimated_cost"})
        rows.append(row)
    return rows


def _insulation_labor_role_map(row_number: int) -> dict[str, str]:
    if row_number in {78, 80, 82, 84, 86, 88, 90, 92}:
        return INSULATION_LABOR_ROLE_BY_COLUMN
    return INSULATION_SUPPORT_LABOR_ROLE_BY_COLUMN


def _insulation_labor_formula_model(row_number: int, formulas: dict[str, Any]) -> str:
    if row_number in {78, 80, 82, 84, 86, 88, 90, 92}:
        return "labor_cost_from_days_crew_rate"
    cost_formula = str(formulas.get("H") or "").replace(" ", "").upper()
    if row_number == 95:
        return "loading_cost_from_hours_people_rate_trip_count"
    if row_number == 97:
        return "travel_labor_cost_from_hours_people_rate_trip_count"
    if row_number == 100:
        return "meals_lodging_cost_from_days_people_daily_amount"
    if "C" in cost_formula and "E" in cost_formula and "G" in cost_formula:
        return "labor_cost_from_quantity_people_rate"
    return "labor_cost"


def _insulation_labor_formula_mode(row_number: int, formulas: dict[str, Any]) -> str:
    if row_number in {78, 80, 82, 84, 86, 88, 90, 92}:
        return _labor_formula_mode(row_number, formulas)
    if row_number in {95, 97}:
        return "hours_based"
    if row_number == 100:
        return "days_based"
    return _labor_formula_mode(row_number, formulas)


def _insulation_labor_rows(ws: Any, people_rate_table: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    selector_options = people_rate_table or []
    for row_number, bucket in LABOR_ROW_MODELS.items():
        role_map = _insulation_labor_role_map(row_number)
        values = _row_values(ws, row_number, 11)
        inputs, outputs = _input_and_output_cells(ws, row_number, role_map, 11)
        formulas = _formula_cells_for_row(ws, row_number, 11)
        formula_mode = _insulation_labor_formula_mode(row_number, formulas)
        formula_model = _insulation_labor_formula_model(row_number, formulas)
        regular_labor = row_number in {78, 80, 82, 84, 86, 88, 90, 92}
        crew_selector = values.get("C") if regular_labor else values.get("E")
        selected_daily_rate_cell = _daily_rate_cell_for_selector(crew_selector) if regular_labor else ""
        row = {
            "sheet_name": ws.title,
            "row_number": row_number,
            "template_bucket": bucket,
            "line_item_kind": "labor",
            "labor_task": values.get("A") or bucket,
            "cell_roles": _json(role_map),
            "cell_values": _json(values),
            "formula_model": formula_model,
            "formula_mode": formula_mode,
            "formula_cells": _json(formulas),
            "formula_dependencies": _json(sorted({dep for formula in formulas.values() for dep in formula_dependencies(formula)})),
            "input_cells": _json(inputs),
            "calculated_output_cells": _json(outputs),
            "days_cell": f"B{row_number}" if regular_labor else f"C{row_number}",
            "crew_selector_cell": f"C{row_number}" if regular_labor else f"E{row_number}",
            "crew_person_selector_code": crew_selector,
            "crew_people_selection": crew_selector,
            "crew_selector_options": _json(selector_options),
            "selected_daily_rate_cell": selected_daily_rate_cell,
            "daily_rate_cell": f"J{row_number}" if regular_labor else "",
            "daily_rate_formula": formulas.get("J", ""),
            "total_hours_cell": f"D{row_number}" if regular_labor else f"C{row_number}",
            "total_hours_formula": formulas.get("D", "") if regular_labor else "",
            "hourly_rate_cell": f"G{row_number}",
            "cost_cell": f"H{row_number}",
            "cost_formula": formulas.get("H", ""),
            "decision_fields": "labor_task,days,crew_person_selector_code,selected_daily_rate_cell,daily_rate,hourly_rate,formula_mode",
            "calculated_output_fields": "total_hours,calculated_cost",
        }
        for column, role in role_map.items():
            row[role] = values.get(column)
        if regular_labor:
            row["hourly_rate"] = row.get("blended_rate")
            row["calculated_cost"] = row.get("estimated_cost")
        elif row_number == 100:
            row["days"] = values.get("C")
            row["total_hours"] = None
            row["hourly_rate"] = None
            row["daily_rate"] = values.get("G")
            row["calculated_cost"] = values.get("H")
        else:
            row["days"] = values.get("C")
            row["total_hours"] = values.get("C")
            row["hourly_rate"] = values.get("G")
            row["daily_rate"] = None
            row["calculated_cost"] = values.get("H")
        rows.append(row)
    return rows


def _sqft_calculation_rows(ws: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_number in range(4, 16):
        description = cell_value(ws, f"B{row_number}")
        formula = cell_value(ws, f"E{row_number}")
        if row_number in range(4, 9):
            area_type = "wall"
            model = "height_width_area"
        elif row_number in range(9, 13):
            area_type = "ceiling"
            model = "height_width_area"
        elif row_number in range(13, 15):
            area_type = "gable"
            model = "triangle_height_width_area"
        else:
            area_type = "total"
            model = "sum_area_rows"
        rows.append(
            {
                "sheet_name": ws.title,
                "row_number": row_number,
                "area_type": area_type,
                "description_cell": f"B{row_number}",
                "description": description,
                "height_cell": f"C{row_number}",
                "width_cell": f"D{row_number}",
                "total_cell": f"E{row_number}",
                "total_formula": formula,
                "model": model,
                "final_total_cell": "F15" if row_number == 15 else "",
                "final_total_formula": cell_value(ws, "F15") if row_number == 15 else "",
            }
        )
    return rows


def _formula_cells_for_row(ws: Any, row_number: int, max_col: int = 11) -> dict[str, Any]:
    formulas: dict[str, Any] = {}
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row_number, column=col_idx)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            formulas[cell.column_letter] = cell.value
    return formulas


def _input_and_output_cells(ws: Any, row_number: int, role_map: dict[str, str], max_col: int = 11) -> tuple[dict[str, Any], dict[str, Any]]:
    inputs: dict[str, Any] = {}
    outputs: dict[str, Any] = {}
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row_number, column=col_idx)
        if cell.value is None:
            continue
        role = role_map.get(cell.column_letter, "")
        item = {"role": role, "value": cell.value}
        if isinstance(cell.value, str) and cell.value.startswith("="):
            outputs[cell.coordinate] = item
        else:
            inputs[cell.coordinate] = item
    return inputs, outputs


def _workbook_sheet_rows(workbook: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for ws in workbook.worksheets:
        formula_count = 0
        hidden_rows = 0
        hidden_columns = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
        for dimension in ws.row_dimensions.values():
            if getattr(dimension, "hidden", False):
                hidden_rows += 1
        for dimension in ws.column_dimensions.values():
            if getattr(dimension, "hidden", False):
                hidden_columns += 1
        validation_count = len(list(ws.data_validations.dataValidation)) if ws.data_validations else 0
        rows.append(
            {
                "sheet_name": ws.title,
                "sheet_state": ws.sheet_state,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "formula_count": formula_count,
                "data_validation_count": validation_count,
                "table_count": len(ws.tables),
                "hidden_row_count": hidden_rows,
                "hidden_column_count": hidden_columns,
            }
        )
    return rows


def _named_range_rows(workbook: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    defined_names = getattr(workbook, "defined_names", None)
    if defined_names is None:
        return rows
    try:
        iterable = defined_names.values()
    except AttributeError:
        iterable = defined_names.definedName
    for name in iterable:
        rows.append(
            {
                "name": getattr(name, "name", ""),
                "scope": getattr(name, "localSheetId", None),
                "attr_text": getattr(name, "attr_text", ""),
                "hidden": bool(getattr(name, "hidden", False)),
                "comment": getattr(name, "comment", ""),
            }
        )
    return rows


def _data_validation_rows(workbook: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for ws in workbook.worksheets:
        validations = list(ws.data_validations.dataValidation) if ws.data_validations else []
        for validation in validations:
            rows.append(
                {
                    "sheet_name": ws.title,
                    "type": validation.type,
                    "operator": validation.operator,
                    "formula1": validation.formula1,
                    "formula2": validation.formula2,
                    "allow_blank": validation.allowBlank,
                    "sqref": str(validation.sqref),
                }
            )
    return rows


def _hidden_table_rows(workbook: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for ws in workbook.worksheets:
        if ws.sheet_state != "visible":
            rows.append(
                {
                    "sheet_name": ws.title,
                    "hidden_type": "sheet",
                    "address": f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}",
                    "description": "Hidden worksheet that may contain lookup/reference data.",
                }
            )
        for row_number, dimension in ws.row_dimensions.items():
            if getattr(dimension, "hidden", False):
                rows.append(
                    {
                        "sheet_name": ws.title,
                        "hidden_type": "row",
                        "address": str(row_number),
                        "description": "Hidden row in visible worksheet.",
                    }
                )
        for column_letter, dimension in ws.column_dimensions.items():
            if getattr(dimension, "hidden", False):
                rows.append(
                    {
                        "sheet_name": ws.title,
                        "hidden_type": "column",
                        "address": column_letter,
                        "description": "Hidden column in visible worksheet.",
                    }
                )
    return rows


def _estimate_row_bucket(row_number: int) -> tuple[str, str]:
    if row_number in ROOFING_HEADER_BUCKETS:
        return "header", ROOFING_HEADER_BUCKETS[row_number]
    if row_number in ROOFING_MATERIAL_BUCKETS:
        return "material", ROOFING_MATERIAL_BUCKETS[row_number]
    if row_number in ROOFING_LABOR_BUCKETS:
        return "labor", ROOFING_LABOR_BUCKETS[row_number]
    if row_number in ROOFING_TOTAL_BUCKETS:
        return "total", ROOFING_TOTAL_BUCKETS[row_number]
    if row_number in ROOFING_TOTAL_MARKUP_ROWS:
        return "total", ROOFING_TOTAL_MARKUP_ROWS[row_number]
    return "", ""


def _all_formula_rows(workbook: Any, template_type: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                section = ""
                bucket = ""
                if template_type == "roofing" and ws.title == "Estimate":
                    section, bucket = _estimate_row_bucket(cell.row)
                selector_map = extract_if_selector_map(value)
                rows.append(
                    {
                        "sheet_name": ws.title,
                        "cell": cell.coordinate,
                        "row_number": cell.row,
                        "column_letter": cell.column_letter,
                        "section": section,
                        "template_bucket": bucket,
                        "formula": value,
                        "formula_kind": formula_kind(value),
                        "dependencies": _json(formula_dependencies(value)),
                        "selector_map": _json(selector_map),
                    }
                )
    return rows


def _all_selector_rows(workbook: Any, template_type: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for formula_row in _all_formula_rows(workbook, template_type):
        selector_map = json.loads(formula_row.get("selector_map") or "{}")
        if not selector_map:
            continue
        for code, resolved in selector_map.items():
            selector_cell = ""
            match = re.search(r"([A-Z]{1,3}\d+)\s*=", formula_row["formula"])
            if match:
                selector_cell = match.group(1)
            rows.append(
                {
                    "sheet_name": formula_row["sheet_name"],
                    "row_number": formula_row["row_number"],
                    "formula_cell": formula_row["cell"],
                    "selector_cell": selector_cell,
                    "resolved_cell": formula_row["cell"],
                    "section": formula_row.get("section", ""),
                    "template_bucket": formula_row.get("template_bucket", ""),
                    "selector_code": code,
                    "resolved_item_name": resolved,
                    "formula": formula_row["formula"],
                }
            )
    return rows


def _lookup_table_rows(workbook: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    def add_range(sheet_name: str, table_name: str, start_row: int, end_row: int, columns: list[str]) -> None:
        if sheet_name not in workbook.sheetnames:
            return
        ws = workbook[sheet_name]
        headers = {col: ws[f"{col}{start_row}"].value for col in columns}
        last_label = None
        for row_number in range(start_row + 1, end_row + 1):
            values = {col: ws[f"{col}{row_number}"].value for col in columns}
            if all(value is None for value in values.values()):
                continue
            first_col = columns[0]
            if values.get(first_col) is not None:
                last_label = values[first_col]
            elif last_label is not None:
                values[first_col] = last_label
            rows.append(
                {
                    "sheet_name": sheet_name,
                    "table_name": table_name,
                    "row_number": row_number,
                    "headers": _json(headers),
                    "values": _json(values),
                    "lookup_key": values.get(first_col),
                }
            )

    add_range("Materials", "solvents", 1, 3, ["A", "C", "D", "E"])
    add_range("Materials", "fabric", 5, 12, ["A", "B", "C", "D"])
    add_range("Materials", "board", 14, 21, ["A", "B", "C", "D", "E"])
    add_range("Materials", "fasteners", 24, 32, ["A", "B", "C", "D"])
    add_range("Materials", "plates", 34, 35, ["A", "B", "C", "D"])
    add_range("People", "crew_rate_matrix", 2, 12, ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"])
    add_range("People", "crew_leaders", 15, 19, ["A", "B", "C"])
    add_range("People", "workers", 21, 34, ["A", "B", "C"])
    add_range("People", "temp_workers", 37, 43, ["A", "B", "C"])
    add_range("Performance & Payment Bonds", "performance_payment_bonds", 1, 3, ["A", "B", "C", "D"])
    return rows


def _people_daily_rate_selector_rows(workbook: Any) -> list[dict[str, Any]]:
    if "People" not in workbook.sheetnames:
        return []
    ws = workbook["People"]
    rows: list[dict[str, Any]] = []
    for selector_code, column in PEOPLE_DAILY_RATE_COLUMNS.items():
        daily_rate_cell = f"{column}12"
        daily_rate_formula = ws[daily_rate_cell].value
        crew_components: list[dict[str, Any]] = []
        for row_number in range(3, 11):
            role = ws[f"A{row_number}"].value
            hourly_wage = ws[f"B{row_number}"].value
            burden_rate = ws[f"C{row_number}"].value
            component_formula = ws[f"{column}{row_number}"].value
            if role is None and hourly_wage is None and component_formula is None:
                continue
            crew_components.append(
                {
                    "row_number": row_number,
                    "role": role,
                    "hourly_wage": hourly_wage,
                    "burden_rate": burden_rate,
                    "component_cell": f"People!{column}{row_number}",
                    "component_formula": component_formula,
                }
            )
        rows.append(
            {
                "sheet_name": "People",
                "table_name": "people_daily_rate_selector",
                "selector_code": selector_code,
                "crew_size": selector_code,
                "people_sheet_column": column,
                "daily_rate_cell": f"People!{daily_rate_cell}",
                "daily_rate_formula": daily_rate_formula,
                "hours_per_day_cell": "People!B11",
                "hours_per_day": ws["B11"].value,
                "formula_dependencies": _json(formula_dependencies(daily_rate_formula)),
                "crew_components": _json(crew_components),
            }
        )
    return rows


def _daily_rate_cell_for_selector(selector_code: Any) -> str:
    code = int(_positive_float(selector_code) or 0)
    column = PEOPLE_DAILY_RATE_COLUMNS.get(code)
    return f"People!{column}12" if column else ""


def _labor_formula_mode(row_number: int, formulas: dict[str, Any]) -> str:
    cost_formula = str(formulas.get("H") or "").replace(" ", "").upper()
    if not cost_formula:
        return "manual_or_missing_formula"
    if f"IF(G{row_number}=0" in cost_formula and f"B{row_number}*J{row_number}" in cost_formula and f"D{row_number}*G{row_number}" in cost_formula:
        return "mixed_formula"
    if f"B{row_number}*J{row_number}" in cost_formula:
        return "days_based"
    if f"D{row_number}*G{row_number}" in cost_formula or f"C{row_number}*E{row_number}*G{row_number}" in cost_formula:
        return "hours_based"
    return "formula_based"


def _roofing_material_rows(ws: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    waste_pct = cell_value(ws, "A30")
    for row_number, bucket in ROOFING_MATERIAL_BUCKETS.items():
        values = _row_values(ws, row_number, 11)
        role_map = _roofing_role_map_for_material_row(row_number, bucket)
        inputs, outputs = _input_and_output_cells(ws, row_number, role_map, 11)
        formulas = _formula_cells_for_row(ws, row_number, 11)
        selector_map = extract_if_selector_map(values.get("B"))
        selector_code = values.get("A") if role_map.get("A") == "selector_code" else None
        resolved_from_selector = selector_map.get(_selector_key(selector_code))
        resolved_item_name = resolved_from_selector if resolved_from_selector else values.get("B") if values.get("B") is not None else values.get("A")
        row = {
            "sheet_name": ws.title,
            "row_number": row_number,
            "template_bucket": bucket,
            "line_item_kind": "equipment" if bucket in {"dumpsters", "lift", "generator"} else "material",
            "formula_model": _roofing_formula_model_for_bucket(bucket, row_number),
            "selector_code": selector_code,
            "resolved_item_name": resolved_item_name,
            "resolved_item_formula": values.get("B") if isinstance(values.get("B"), str) and values.get("B").startswith("=") else "",
            "item_name": values.get("A"),
            "cell_roles": _json(role_map),
            "cell_values": _json(values),
            "selector_map": _json(selector_map),
            "formula_cells": _json(formulas),
            "formula_dependencies": _json(sorted({dep for formula in formulas.values() for dep in formula_dependencies(formula)})),
            "input_cells": _json(inputs),
            "calculated_output_cells": _json(outputs),
            "selector_cell_role": "selector_code" if role_map.get("A") == "selector_code" else "item_or_label",
            "waste_factor_pct": waste_pct if bucket == "coating" else None,
            "waste_factor_cell": "A30" if bucket == "coating" else "",
        }
        for column, role in role_map.items():
            row[role] = values.get(column)
        if resolved_from_selector:
            row["resolved_item_name"] = resolved_from_selector
        if bucket == "coating":
            gal_per_100 = values.get("D")
            row["wet_mils_estimate"] = round(float(gal_per_100) * 16, 4) if isinstance(gal_per_100, (int, float)) else None
            row["gal_per_sqft"] = round(float(gal_per_100) / 100, 6) if isinstance(gal_per_100, (int, float)) else None
            row["decision_fields"] = "selector_code,resolved_item_name,area_sqft,gal_per_100_sqft,gal_per_sqft,wet_mils_estimate,waste_factor_pct,unit_price,product_reference"
            row["calculated_output_fields"] = "estimated_gallons,estimated_cost"
        elif bucket == "primer":
            row["decision_fields"] = "selector_code,resolved_item_name,area_sqft,unit_price"
            row["calculated_output_fields"] = "estimated_units,estimated_cost"
        elif bucket in {"seams_misc", "fasteners", "plates", "caulk_sealant", "fabric"}:
            row["decision_fields"] = ",".join(
                role
                for role in role_map.values()
                if role not in {"estimated_units", "estimated_gallons", "estimated_cost"}
            )
            row["calculated_output_fields"] = ",".join(
                role for role in role_map.values() if role in {"estimated_units", "estimated_gallons", "estimated_cost"}
            )
        else:
            row["decision_fields"] = ",".join(
                role
                for role in role_map.values()
                if role not in {"estimated_units", "estimated_gallons", "estimated_cost"}
            )
            row["calculated_output_fields"] = ",".join(
                role for role in role_map.values() if role in {"estimated_units", "estimated_gallons", "estimated_cost"}
            )
        rows.append(row)
    return rows


def _roofing_labor_rows(ws: Any, people_rate_table: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    selector_options = people_rate_table or []
    for row_number, bucket in ROOFING_LABOR_BUCKETS.items():
        values = _row_values(ws, row_number, 11)
        inputs, outputs = _input_and_output_cells(ws, row_number, ROOFING_LABOR_ROLE_BY_COLUMN, 11)
        formulas = _formula_cells_for_row(ws, row_number, 11)
        formula_mode = _labor_formula_mode(row_number, formulas)
        crew_selector = values.get("C") if row_number in range(116, 135) else values.get("E")
        selected_daily_rate_cell = _daily_rate_cell_for_selector(crew_selector) if row_number in range(116, 135) else ""
        row = {
            "sheet_name": ws.title,
            "row_number": row_number,
            "template_bucket": bucket,
            "line_item_kind": "labor",
            "labor_task": values.get("A") or bucket,
            "cell_roles": _json(ROOFING_LABOR_ROLE_BY_COLUMN),
            "cell_values": _json(values),
            "formula_model": "labor_cost_from_days_crew_rate",
            "formula_mode": formula_mode,
            "formula_cells": _json(formulas),
            "formula_dependencies": _json(sorted({dep for formula in formulas.values() for dep in formula_dependencies(formula)})),
            "input_cells": _json(inputs),
            "calculated_output_cells": _json(outputs),
            "days_cell": f"B{row_number}" if row_number in range(116, 135) else f"C{row_number}",
            "crew_selector_cell": f"C{row_number}" if row_number in range(116, 135) else f"E{row_number}",
            "crew_person_selector_code": crew_selector,
            "crew_people_selection": crew_selector,
            "crew_selector_options": _json(selector_options),
            "selected_daily_rate_cell": selected_daily_rate_cell,
            "daily_rate_cell": f"J{row_number}" if row_number in range(116, 135) else "",
            "daily_rate_formula": formulas.get("J", ""),
            "total_hours_cell": f"D{row_number}" if row_number in range(116, 135) else f"C{row_number}",
            "total_hours_formula": formulas.get("D", "") if row_number in range(116, 135) else "",
            "hourly_rate_cell": f"G{row_number}",
            "cost_cell": f"H{row_number}",
            "cost_formula": formulas.get("H", ""),
            "decision_fields": "labor_task,days,crew_person_selector_code,selected_daily_rate_cell,daily_rate,hourly_rate,formula_mode",
            "calculated_output_fields": "total_hours,calculated_cost",
        }
        for column, role in ROOFING_LABOR_ROLE_BY_COLUMN.items():
            row[role] = values.get(column)
        row["hourly_rate"] = row.get("blended_rate")
        row["calculated_cost"] = row.get("estimated_cost")
        rows.append(row)
    return rows


def _roofing_totals_markups(ws: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_number, bucket in ROOFING_TOTAL_MARKUP_ROWS.items():
        values = _row_values(ws, row_number, 11)
        formulas = _formula_cells_for_row(ws, row_number, 11)
        rows.append(
            {
                "sheet_name": ws.title,
                "row_number": row_number,
                "template_bucket": bucket,
                "label": values.get("A") or values.get("G") or values.get("C"),
                "percentage": values.get("F") if row_number in {165, 167} else values.get("G") if row_number == 111 else None,
                "value_cell": "H" + str(row_number) if values.get("H") is not None else "",
                "formula_cells": _json(formulas),
                "formula_dependencies": _json(sorted({dep for formula in formulas.values() for dep in formula_dependencies(formula)})),
                "cell_values": _json(values),
            }
        )
    return rows


def _roofing_row_catalog(material_rows: list[dict[str, Any]], labor_rows: list[dict[str, Any]], totals: list[dict[str, Any]]) -> list[dict[str, Any]]:
    catalog: list[dict[str, Any]] = []
    for row_number, bucket in ROOFING_HEADER_BUCKETS.items():
        catalog.append(
            {
                "sheet_name": "Estimate",
                "row_number": row_number,
                "section": "header",
                "template_bucket": bucket,
                "line_item_kind": "header",
                "formula_model": "",
                "cell_roles": _json({"B": "label", "C": "value"}),
            }
        )
    for row in material_rows:
        catalog.append(
            {
                "sheet_name": row["sheet_name"],
                "row_number": row["row_number"],
                "section": "material",
                "template_bucket": row["template_bucket"],
                "line_item_kind": row["line_item_kind"],
                "formula_model": row["formula_model"],
                "cell_roles": row.get("cell_roles"),
            }
        )
    for row in labor_rows:
        catalog.append(
            {
                "sheet_name": row["sheet_name"],
                "row_number": row["row_number"],
                "section": "labor",
                "template_bucket": row["template_bucket"],
                "line_item_kind": "labor",
                "formula_model": row["formula_model"],
                "cell_roles": row.get("cell_roles"),
            }
        )
    for row in totals:
        catalog.append(
            {
                "sheet_name": row["sheet_name"],
                "row_number": row["row_number"],
                "section": "totals_markups",
                "template_bucket": row["template_bucket"],
                "line_item_kind": "total",
                "formula_model": "markup_or_total_formula",
                "cell_roles": "",
            }
        )
    return sorted(catalog, key=lambda item: (item["sheet_name"], int(item["row_number"])))


def _pricing_product_references(selector_maps: list[dict[str, Any]], lookup_tables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    material_buckets = set(ROOFING_MATERIAL_BUCKETS.values())
    for row in selector_maps:
        if row.get("template_bucket") in material_buckets:
            rows.append(
                {
                    "source_type": "selector_map",
                    "source_table": "Estimate",
                    "template_bucket": row.get("template_bucket"),
                    "row_number": row.get("row_number"),
                    "selector_code": row.get("selector_code"),
                    "product_name": row.get("resolved_item_name"),
                    "formula": row.get("formula"),
                }
            )
    for row in lookup_tables:
        if row.get("sheet_name") == "Materials":
            rows.append(
                {
                    "source_type": "lookup_table",
                    "source_table": row.get("table_name"),
                    "template_bucket": "",
                    "row_number": row.get("row_number"),
                    "selector_code": "",
                    "product_name": row.get("lookup_key"),
                    "formula": "",
                    "values": row.get("values"),
                }
            )
    return rows


def _people_labor_references(
    labor_rows: list[dict[str, Any]],
    lookup_tables: list[dict[str, Any]],
    people_rate_table: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in people_rate_table or []:
        rows.append(
            {
                "source_type": "people_daily_rate_selector",
                "source_table": row.get("table_name"),
                "row_number": 12,
                "labor_package": "",
                "lookup_key": row.get("selector_code"),
                "selector_code": row.get("selector_code"),
                "daily_rate_cell": row.get("daily_rate_cell"),
                "daily_rate_formula": row.get("daily_rate_formula"),
                "values": row.get("crew_components"),
            }
        )
    for row in lookup_tables:
        if row.get("sheet_name") == "People":
            rows.append(
                {
                    "source_type": "people_lookup_table",
                    "source_table": row.get("table_name"),
                    "row_number": row.get("row_number"),
                    "labor_package": "",
                    "lookup_key": row.get("lookup_key"),
                    "values": row.get("values"),
                }
            )
    for row in labor_rows:
        rows.append(
            {
                "source_type": "labor_row_formula_reference",
                "source_table": "Estimate",
                "row_number": row.get("row_number"),
                "labor_package": row.get("template_bucket"),
                "lookup_key": row.get("labor_task"),
                "values": row.get("formula_dependencies"),
                "formula_mode": row.get("formula_mode"),
                "crew_selector_cell": row.get("crew_selector_cell"),
                "crew_person_selector_code": row.get("crew_person_selector_code"),
                "selected_daily_rate_cell": row.get("selected_daily_rate_cell"),
                "cost_formula": row.get("cost_formula"),
            }
        )
    return rows


def _row_catalog(material_rows: list[dict[str, Any]], labor_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    catalog: list[dict[str, Any]] = []
    for row in material_rows:
        catalog.append(
            {
                "sheet_name": row["sheet_name"],
                "row_number": row["row_number"],
                "template_bucket": row["template_bucket"],
                "line_item_kind": row["line_item_kind"],
                "formula_model": row["formula_model"],
                "selector_cell_role": row.get("selector_cell_role", ""),
                "resolved_item_name": row.get("resolved_item_name"),
                "cell_roles": row.get("cell_roles"),
            }
        )
    for row in labor_rows:
        catalog.append(
            {
                "sheet_name": row["sheet_name"],
                "row_number": row["row_number"],
                "template_bucket": row["template_bucket"],
                "line_item_kind": row["line_item_kind"],
                "formula_model": row.get("formula_model") or "labor_cost_from_days_crew_rate",
                "selector_cell_role": "",
                "resolved_item_name": row.get("labor_task"),
                "cell_roles": row.get("cell_roles"),
            }
        )
    return sorted(catalog, key=lambda item: int(item["row_number"]))


def _median(values: list[float]) -> float | None:
    clean = sorted(value for value in values if value is not None)
    if not clean:
        return None
    mid = len(clean) // 2
    if len(clean) % 2:
        return clean[mid]
    return (clean[mid - 1] + clean[mid]) / 2


def _percentile(values: list[float], q: float) -> float | None:
    clean = sorted(value for value in values if value is not None)
    if not clean:
        return None
    if len(clean) == 1:
        return clean[0]
    pos = (len(clean) - 1) * q
    lower = int(pos)
    upper = min(lower + 1, len(clean) - 1)
    fraction = pos - lower
    return clean[lower] + (clean[upper] - clean[lower]) * fraction


def _mode_text(values: list[Any]) -> str:
    counts: dict[str, int] = {}
    for value in values:
        text = str(value or "").strip()
        if text:
            counts[text] = counts.get(text, 0) + 1
    if not counts:
        return ""
    return sorted(counts.items(), key=lambda item: (-item[1], item[0]))[0][0]


def _wide_range_flag(p25: float | None, p75: float | None, median: float | None, label: str) -> str:
    if p25 is None or p75 is None or median in {None, 0}:
        return ""
    if (p75 - p25) / abs(median) > 0.75:
        return f"Wide historical range for {label}; estimator review recommended."
    return ""


def _insulation_decision_history(material_rows: list[dict[str, Any]], labor_rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    foam_rows = [row for row in material_rows if row.get("template_bucket") == "foam"]
    coating_rows = [row for row in material_rows if row.get("template_bucket") == "thermal_barrier_coating"]

    foam_groups: dict[tuple[str, float | None], list[dict[str, Any]]] = {}
    for row in foam_rows:
        key = (str(row.get("resolved_item_name") or ""), _as_float(row.get("foam_density_lb")))
        foam_groups.setdefault(key, []).append(row)
    foam_history: list[dict[str, Any]] = []
    for (product, density), rows in foam_groups.items():
        thickness = [_positive_float(row.get("thickness_inches")) for row in rows]
        yield_values = [_positive_float(row.get("yield_factor") or row.get("yield_or_coverage")) for row in rows]
        unit_prices = [_positive_float(row.get("unit_price")) for row in rows]
        p25_thickness = _percentile([v for v in thickness if v is not None], 0.25)
        p75_thickness = _percentile([v for v in thickness if v is not None], 0.75)
        median_thickness = _median([v for v in thickness if v is not None])
        p25_yield = _percentile([v for v in yield_values if v is not None], 0.25)
        p75_yield = _percentile([v for v in yield_values if v is not None], 0.75)
        median_yield = _median([v for v in yield_values if v is not None])
        foam_history.append(
            {
                "resolved_item_name": product,
                "foam_brand": _mode_text([row.get("foam_brand") for row in rows]),
                "foam_density_lb": density,
                "median_thickness_inches": median_thickness,
                "p25_thickness_inches": p25_thickness,
                "p75_thickness_inches": p75_thickness,
                "median_yield_factor": median_yield,
                "p25_yield_factor": p25_yield,
                "p75_yield_factor": p75_yield,
                "median_unit_price": _median([v for v in unit_prices if v is not None]),
                "evidence_count": len(rows),
                "review_flags": " ".join(
                    flag
                    for flag in (
                        _wide_range_flag(p25_thickness, p75_thickness, median_thickness, "foam thickness"),
                        _wide_range_flag(p25_yield, p75_yield, median_yield, "foam yield"),
                    )
                    if flag
                ),
            }
        )

    coating_history: list[dict[str, Any]] = []
    for product in sorted({str(row.get("resolved_item_name") or "") for row in coating_rows if row.get("resolved_item_name")}):
        rows = [row for row in coating_rows if str(row.get("resolved_item_name") or "") == product]
        gal_rates = [_positive_float(row.get("gal_per_100_sqft")) for row in rows]
        unit_prices = [_positive_float(row.get("unit_price")) for row in rows]
        coating_history.append(
            {
                "resolved_item_name": product,
                "median_gal_per_100_sqft": _median([v for v in gal_rates if v is not None]),
                "median_gal_per_sqft": (_median([v for v in gal_rates if v is not None]) or 0) / 100 if gal_rates else None,
                "median_unit_price": _median([v for v in unit_prices if v is not None]),
                "evidence_count": len(rows),
            }
        )

    labor_history: list[dict[str, Any]] = []
    for package in sorted({str(row.get("template_bucket") or "") for row in labor_rows if row.get("template_bucket")}):
        rows = [row for row in labor_rows if str(row.get("template_bucket") or "") == package]
        labor_history.append(
            {
                "template_bucket": package,
                "task": _mode_text([row.get("labor_task") for row in rows]),
                "median_days": _median(
                    [v for v in [_positive_float(row.get("days") or row.get("days_or_hours")) for row in rows] if v is not None]
                ),
                "median_crew_people_selection": _median(
                    [v for v in [_positive_float(row.get("crew_person_selector_code") or row.get("crew_size")) for row in rows] if v is not None]
                ),
                "median_daily_rate": _median([v for v in [_positive_float(row.get("daily_rate")) for row in rows] if v is not None]),
                "median_hourly_rate": _median([v for v in [_positive_float(row.get("hourly_rate") or row.get("blended_rate")) for row in rows] if v is not None]),
                "median_total_hours": _median([v for v in [_positive_float(row.get("total_hours")) for row in rows] if v is not None]),
                "formula_mode": _mode_text([row.get("formula_mode") for row in rows]),
                "evidence_count": len(rows),
            }
        )
    return {
        "insulation_foam_decision_history": foam_history,
        "insulation_coating_decision_history": coating_history,
        "insulation_labor_decision_history": labor_history,
    }


def _roofing_decision_history(material_rows: list[dict[str, Any]], labor_rows: list[dict[str, Any]], totals: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    coating_rows = [row for row in material_rows if row.get("template_bucket") == "coating"]
    coating_history: list[dict[str, Any]] = []
    for product in sorted({str(row.get("resolved_item_name") or "") for row in coating_rows if row.get("resolved_item_name")}):
        rows = [row for row in coating_rows if str(row.get("resolved_item_name") or "") == product]
        gal_rates = [_positive_float(row.get("gal_per_100_sqft")) for row in rows]
        waste = [_positive_float(row.get("waste_factor_pct")) for row in rows]
        coating_history.append(
            {
                "resolved_item_name": product,
                "median_gal_per_100_sqft": _median([v for v in gal_rates if v is not None]),
                "median_gal_per_sqft": (_median([v for v in gal_rates if v is not None]) or 0) / 100 if gal_rates else None,
                "median_wet_mils": _median([v * 16 for v in gal_rates if v is not None]),
                "median_waste_factor_pct": _median([v for v in waste if v is not None]),
                "evidence_count": len(rows),
            }
        )
    labor_history: list[dict[str, Any]] = []
    for package in sorted({str(row.get("template_bucket") or "") for row in labor_rows if row.get("template_bucket")}):
        rows = [row for row in labor_rows if str(row.get("template_bucket") or "") == package]
        labor_history.append(
            {
                "template_bucket": package,
                "task": _mode_text([row.get("labor_task") for row in rows]),
                "median_days": _median([v for v in [_positive_float(row.get("days")) for row in rows] if v is not None]),
                "median_crew_people_selection": _median(
                    [v for v in [_positive_float(row.get("crew_person_selector_code") or row.get("crew_size")) for row in rows] if v is not None]
                ),
                "median_daily_rate": _median([v for v in [_positive_float(row.get("daily_rate")) for row in rows] if v is not None]),
                "median_hourly_rate": _median([v for v in [_positive_float(row.get("hourly_rate") or row.get("blended_rate")) for row in rows] if v is not None]),
                "median_total_hours": _median([v for v in [_positive_float(row.get("total_hours")) for row in rows] if v is not None]),
                "formula_mode": _mode_text([row.get("formula_mode") for row in rows]),
                "evidence_count": len(rows),
            }
        )
    markup_history = [
        {
            "template_bucket": row.get("template_bucket"),
            "percentage": row.get("percentage"),
            "formula_cells": row.get("formula_cells"),
        }
        for row in totals
        if row.get("template_bucket") in {"overhead", "profit", "sales_tax", "warranty"}
    ]
    return {
        "roofing_coating_decision_history": coating_history,
        "roofing_labor_decision_history": labor_history,
        "roofing_markup_decision_history": markup_history,
    }


def _extract_insulation_template_intelligence(path: Path, workbook: Any) -> dict[str, Any]:
    estimate = workbook["Estimate"]
    sqft = workbook["Sq Ft Calculation"]
    material_rows = _material_rows(estimate)
    people_rate_table = _people_daily_rate_selector_rows(workbook)
    labor_rows = _insulation_labor_rows(estimate, people_rate_table)
    decision_history = _insulation_decision_history(material_rows, labor_rows)
    selector_maps = _selector_rows(estimate)
    formula_models = _formula_model_rows(estimate)
    sqft_rows = _sqft_calculation_rows(sqft)
    return {
        "template_type": "insulation",
        "template_path": str(path),
        "template_name": path.name,
        "sheets": workbook.sheetnames,
        "selector_maps": selector_maps,
        "row_catalog": _row_catalog(material_rows, labor_rows),
        "formula_models": formula_models,
        "sq_ft_calculation": sqft_rows,
        "labor_rows": labor_rows,
        "people_labor_references": _people_labor_references(labor_rows, [], people_rate_table),
        "people_rate_table": people_rate_table,
        "materials_rows": material_rows,
        **decision_history,
        "derived_defaults": {
            "foam_quantity_model": "foam_sets_from_area_thickness_yield",
            "foam_estimated_units_formula": "estimated_units=((area_sqft/yield_factor)*thickness_inches)*1000",
            "foam_estimated_sets_formula": "estimated_sets=estimated_units/1000",
            "foam_cost_formula": "estimated_cost=unit_price*estimated_units",
            "coating_quantity_model": "coating_gallons_from_area_rate_waste",
            "coating_gallons_formula": "estimated_gallons=((area_sqft/100)*gal_per_100_sqft)/((100-waste_margin_pct)/100)",
            "coating_waste_margin_cell": "Estimate!A34",
        },
    }


def _extract_roofing_template_intelligence(path: Path, workbook: Any) -> dict[str, Any]:
    estimate = workbook["Estimate"]
    material_rows = _roofing_material_rows(estimate)
    people_rate_table = _people_daily_rate_selector_rows(workbook)
    labor_rows = _roofing_labor_rows(estimate, people_rate_table)
    totals = _roofing_totals_markups(estimate)
    decision_history = _roofing_decision_history(material_rows, labor_rows, totals)
    selector_maps = _all_selector_rows(workbook, "roofing")
    lookup_tables = _lookup_table_rows(workbook)
    formula_models = _all_formula_rows(workbook, "roofing")
    hidden_tables = _hidden_table_rows(workbook)
    named_ranges = _named_range_rows(workbook)
    data_validations = _data_validation_rows(workbook)
    return {
        "template_type": "roofing",
        "template_path": str(path),
        "template_name": path.name,
        "workbook_sheets": _workbook_sheet_rows(workbook),
        "sheets": workbook.sheetnames,
        "selector_maps": selector_maps,
        "lookup_tables": lookup_tables,
        "material_rows": material_rows,
        "materials_rows": material_rows,
        "labor_rows": labor_rows,
        "formula_models": formula_models,
        "hidden_tables": hidden_tables,
        "named_ranges": named_ranges,
        "data_validations": data_validations,
        "workbook_row_catalog": _roofing_row_catalog(material_rows, labor_rows, totals),
        "row_catalog": _roofing_row_catalog(material_rows, labor_rows, totals),
        "pricing_product_references": _pricing_product_references(selector_maps, lookup_tables),
        "people_labor_references": _people_labor_references(labor_rows, lookup_tables, people_rate_table),
        "people_rate_table": people_rate_table,
        "totals_markups": totals,
        **decision_history,
        "derived_defaults": {
            "foam_quantity_model": "foam_sets_from_area_thickness_yield",
            "coating_quantity_model": "coating_gallons_from_area_rate_waste",
            "coating_waste_factor_cell": "Estimate!A30",
            "coating_wet_mils_rule": "wet_mils = gal_per_100_sqft * 16",
            "warranty_years_cell": "Estimate!C154",
            "warranty_unit_cost_cell": "Estimate!F154",
            "overhead_pct_cell": "Estimate!F165",
            "profit_pct_cell": "Estimate!F167",
            "labor_hours_per_day_cell": "People!B11",
        },
    }


def _detect_template_type(path: Path, workbook: Any, explicit_template_type: str | None = None) -> str:
    explicit = (explicit_template_type or "").strip().lower()
    if explicit in {"roofing", "insulation"}:
        return explicit
    if "Sq Ft Calculation" in workbook.sheetnames:
        return "insulation"
    name = path.name.lower()
    if "insulation" in name:
        return "insulation"
    return "roofing"


def extract_template_intelligence(template_path: str | Path, template_type: str | None = None) -> dict[str, Any]:
    path = resolve_template_path(template_path)
    workbook = load_workbook(path, data_only=False, read_only=False)
    resolved_type = _detect_template_type(path, workbook, template_type)
    if resolved_type == "roofing":
        return _extract_roofing_template_intelligence(path, workbook)
    return _extract_insulation_template_intelligence(path, workbook)


def _frames_from_intelligence(intelligence: dict[str, Any]) -> dict[str, pd.DataFrame]:
    if intelligence.get("template_type") == "roofing":
        return {
            "Workbook Sheets": pd.DataFrame(intelligence.get("workbook_sheets") or []),
            "Selector Maps": pd.DataFrame(intelligence.get("selector_maps") or []),
            "Lookup Tables": pd.DataFrame(intelligence.get("lookup_tables") or []),
            "Material Rows": pd.DataFrame(intelligence.get("material_rows") or intelligence.get("materials_rows") or []),
            "Labor Rows": pd.DataFrame(intelligence.get("labor_rows") or []),
            "Formula Models": pd.DataFrame(intelligence.get("formula_models") or []),
            "Hidden Tables": pd.DataFrame(intelligence.get("hidden_tables") or []),
            "Named Ranges": pd.DataFrame(intelligence.get("named_ranges") or []),
            "Workbook Row Catalog": pd.DataFrame(intelligence.get("workbook_row_catalog") or intelligence.get("row_catalog") or []),
            "Pricing Product References": pd.DataFrame(intelligence.get("pricing_product_references") or []),
            "People Labor References": pd.DataFrame(intelligence.get("people_labor_references") or []),
            "People Rate Table": pd.DataFrame(intelligence.get("people_rate_table") or []),
            "Totals Markups": pd.DataFrame(intelligence.get("totals_markups") or []),
            "Coating Decisions": pd.DataFrame(intelligence.get("roofing_coating_decision_history") or []),
            "Labor Decisions": pd.DataFrame(intelligence.get("roofing_labor_decision_history") or []),
            "Markup Decisions": pd.DataFrame(intelligence.get("roofing_markup_decision_history") or []),
        }
    return {
        "Selector Maps": pd.DataFrame(intelligence.get("selector_maps") or []),
        "Row Catalog": pd.DataFrame(intelligence.get("row_catalog") or []),
        "Formula Models": pd.DataFrame(intelligence.get("formula_models") or []),
        "Sq Ft Calculation": pd.DataFrame(intelligence.get("sq_ft_calculation") or []),
        "Labor Rows": pd.DataFrame(intelligence.get("labor_rows") or []),
        "Materials Rows": pd.DataFrame(intelligence.get("materials_rows") or []),
        "People Labor References": pd.DataFrame(intelligence.get("people_labor_references") or []),
        "People Rate Table": pd.DataFrame(intelligence.get("people_rate_table") or []),
        "Foam Decisions": pd.DataFrame(intelligence.get("insulation_foam_decision_history") or []),
        "Coating Decisions": pd.DataFrame(intelligence.get("insulation_coating_decision_history") or []),
        "Labor Decisions": pd.DataFrame(intelligence.get("insulation_labor_decision_history") or []),
    }


def write_template_intelligence(intelligence: dict[str, Any], json_out: str | Path) -> tuple[Path, Path]:
    json_path = Path(json_out)
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(intelligence, indent=2, default=str), encoding="utf-8")
    xlsx_path = json_path.with_suffix(".xlsx")
    with pd.ExcelWriter(xlsx_path) as writer:
        for sheet_name, frame in _frames_from_intelligence(intelligence).items():
            frame.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    return json_path, xlsx_path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Extract estimator template intelligence from an estimate workbook.")
    parser.add_argument("--template", required=True, help="Path to the estimate workbook template.")
    parser.add_argument("--template-type", choices=["auto", "insulation", "roofing"], default="auto", help="Template type to extract.")
    parser.add_argument("--out", default="output/insulation_template_intelligence.json", help="JSON output path.")
    args = parser.parse_args(argv)
    intelligence = extract_template_intelligence(args.template, None if args.template_type == "auto" else args.template_type)
    json_path, xlsx_path = write_template_intelligence(intelligence, args.out)
    print(f"Wrote template intelligence JSON: {json_path}")
    print(f"Wrote template intelligence workbook: {xlsx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
