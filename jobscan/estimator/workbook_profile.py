from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any


def normalize_label(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


@dataclass(frozen=True)
class FormulaCostRow:
    label: str
    row: int
    trips_cell: str
    miles_cell: str
    rate_cell: str
    cost_cell: str


@dataclass(frozen=True)
class LaborCostRow:
    label: str
    row: int
    days_cell: str
    crew_size_cell: str
    daily_rate_cell: str
    cost_cell: str


@dataclass(frozen=True)
class PerTripLaborCostRow:
    label: str
    row: int
    hours_cell: str
    crew_size_cell: str
    rate_cell: str
    cost_cell: str


@dataclass(frozen=True)
class WarrantyCostRow:
    label: str
    row: int
    manufacturer_cell: str
    years_cell: str
    type_cell: str
    area_cell: str
    unit_cost_cell: str
    cost_cell: str


@dataclass(frozen=True)
class EstimateWorkbookProfile:
    template_type: str
    template_sha256: str
    estimate_sheet: str
    people_sheet: str
    job_spec_sheet: str | None
    sales_inspection: FormulaCostRow
    truck_expense: FormulaCostRow
    labor_rows: dict[str, LaborCostRow]
    per_trip_labor_rows: dict[str, PerTripLaborCostRow]
    crew_daily_rate_cells: dict[int, str]
    warranty: WarrantyCostRow | None
    material_subtotal_cell: str
    labor_subtotal_cell: str
    total_job_cost_cell: str
    final_price_cell: str

    def action_summary(self) -> dict[str, Any]:
        capabilities = [
            "sales_inspection_mileage_cost",
            "truck_mileage_cost",
            "crew_daily_rate_lookup",
            "labor_cost_by_activity",
            "material_subtotal",
            "labor_subtotal",
            "total_job_cost",
            "final_worksheet_price",
        ]
        if self.warranty is not None:
            capabilities.append("warranty_cost")
        return {
            "profile_version": f"spraytec.{self.template_type}_template_profile.v1",
            "template_sha256": self.template_sha256,
            "sheets": {
                "estimate": self.estimate_sheet,
                "people": self.people_sheet,
                "job_spec": self.job_spec_sheet,
            },
            "detected_capabilities": capabilities,
            "supported_crew_sizes": sorted(self.crew_daily_rate_cells),
            "labor_tasks": sorted(
                [*self.labor_rows, *self.per_trip_labor_rows]
            ),
        }


RoofingWorkbookProfile = EstimateWorkbookProfile
InsulationWorkbookProfile = EstimateWorkbookProfile


def discover_roofing_workbook_profile(template_path: Path) -> RoofingWorkbookProfile:
    return _discover_workbook_profile(
        template_path,
        template_type="roofing",
        labor_aliases={
            "labor_setup_safety": ("setup safety", "full repair"),
            "labor_full_repair": ("full repair",),
            "labor_prep": ("pw prep", "pwash prep", "prep"),
            "labor_tearoff": ("tear off", "tearoff"),
            "labor_board": ("board",),
            "labor_base": ("base", "base coat"),
            "labor_caulk": ("caulk sf", "caulk"),
            "labor_details": ("details",),
            "labor_top_coat": ("top coat gran", "top coat"),
            "labor_cleanup": ("touch clean up", "cleanup", "clean up"),
        },
    )


def discover_insulation_workbook_profile(
    template_path: Path,
) -> InsulationWorkbookProfile:
    return _discover_workbook_profile(
        template_path,
        template_type="insulation",
        labor_aliases={
            "labor_set_up": ("set up", "setup"),
            "labor_mask": ("mask",),
            "labor_prime": ("prime",),
            "labor_membrane": ("membrane",),
            "labor_foam": ("foam",),
            "labor_dc_315": ("noburn", "no burn", "dc 315"),
            "labor_misc": ("misc", "miscellaneous"),
            "labor_cleanup": ("clean up", "cleanup"),
        },
    )


def discover_flooring_workbook_profile(
    template_path: Path,
) -> EstimateWorkbookProfile:
    return _discover_workbook_profile(
        template_path,
        template_type="flooring",
        labor_aliases={
            "labor_floor_grind_patch": ("trip 1 grind",),
            "labor_floor_corner_repair": ("trip 2 corner",),
            "labor_floor_prep_base_flake": ("trip 3 prep",),
            "labor_floor_patch_grind": ("patch grind",),
            "labor_floor_primer": ("primer",),
            "labor_floor_base_707": ("base coat 707",),
            "labor_floor_details": ("details",),
            "labor_floor_top_coat": ("trip 4 top coat",),
            "labor_floor_cleanup": ("touch clean up",),
        },
    )


def _discover_workbook_profile(
    template_path: Path,
    *,
    template_type: str,
    labor_aliases: dict[str, tuple[str, ...]],
) -> EstimateWorkbookProfile:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("Install openpyxl to inspect estimate templates.") from exc

    path = Path(template_path)
    workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
    required_sheets = ("Estimate", "People")
    missing_sheets = [name for name in required_sheets if name not in workbook.sheetnames]
    if missing_sheets:
        raise ValueError(
            f"{template_type.title()} template is missing required sheets: "
            + ", ".join(missing_sheets)
        )
    estimate = workbook["Estimate"]
    people = workbook["People"]

    sales_row = _find_label_row(estimate, ("sales inspect", "sales inspection"))
    truck_row = _find_label_row(estimate, ("truck exp", "truck expense"))
    sales_profile = _travel_cost_row(estimate, sales_row, "Sales/inspection")
    truck_profile = _travel_cost_row(estimate, truck_row, "Truck expense")

    labor_rows: dict[str, LaborCostRow] = {}
    for task, aliases in labor_aliases.items():
        row_number = _find_labor_label_row(estimate, aliases)
        cost_cell = _formula_cell_on_row(estimate, row_number, preferred_column="H")
        daily_rate_cell = _formula_cell_on_row(
            estimate,
            row_number,
            preferred_column="J",
        )
        labor_rows[task] = LaborCostRow(
            label=str(estimate[f"A{row_number}"].value or task).strip(),
            row=row_number,
            days_cell=f"B{row_number}",
            crew_size_cell=f"C{row_number}",
            daily_rate_cell=daily_rate_cell,
            cost_cell=cost_cell,
        )
    per_trip_labor_rows = {
        "labor_loading": _per_trip_labor_row(
            estimate,
            _find_label_row(estimate, ("loading",)),
            "Loading",
        ),
        "labor_traveling": _per_trip_labor_row(
            estimate,
            _find_label_row(estimate, ("traveling", "travel")),
            "Traveling",
        ),
    }

    daily_rate_row = _find_label_row(people, ("estimated cost day", "estimated cost per day"))
    crew_daily_rate_cells: dict[int, str] = {}
    for column in range(4, people.max_column + 1):
        crew_size = people.cell(2, column).value
        formula = people.cell(daily_rate_row, column).value
        if isinstance(crew_size, (int, float)) and isinstance(formula, str) and formula.startswith("="):
            crew_daily_rate_cells[int(crew_size)] = people.cell(daily_rate_row, column).coordinate
    if not crew_daily_rate_cells:
        raise ValueError(
            f"{template_type.title()} template has no formula-driven People daily crew rates."
        )

    warranty_row = _find_optional_label_row(estimate, ("warranty",))
    warranty = None
    if warranty_row is not None:
        warranty = WarrantyCostRow(
            label=str(estimate[f"A{warranty_row}"].value or "Warranty").strip(),
            row=warranty_row,
            manufacturer_cell=f"B{warranty_row}",
            years_cell=f"C{warranty_row}",
            type_cell=f"D{warranty_row}",
            area_cell=f"E{warranty_row}",
            unit_cost_cell=f"F{warranty_row}",
            cost_cell=_formula_cell_on_row(
                estimate,
                warranty_row,
                preferred_column="H",
            ),
        )

    return EstimateWorkbookProfile(
        template_type=template_type,
        template_sha256=hashlib.sha256(path.read_bytes()).hexdigest(),
        estimate_sheet="Estimate",
        people_sheet="People",
        job_spec_sheet="Job Spec" if "Job Spec" in workbook.sheetnames else None,
        sales_inspection=sales_profile,
        truck_expense=truck_profile,
        labor_rows=labor_rows,
        per_trip_labor_rows=per_trip_labor_rows,
        crew_daily_rate_cells=crew_daily_rate_cells,
        warranty=warranty,
        material_subtotal_cell=_output_cell_for_label(
            estimate,
            ("subtotal materials",),
        ),
        labor_subtotal_cell=_output_cell_for_label(
            estimate,
            ("subtotal labor subcontractor",),
        ),
        total_job_cost_cell=_output_cell_for_label(estimate, ("total job cost",)),
        final_price_cell=_output_cell_for_label(
            estimate,
            ("work sheet price additional amount w o markup",),
        ),
    )


def _find_label_row(ws: Any, aliases: tuple[str, ...]) -> int:
    normalized_aliases = {normalize_label(alias) for alias in aliases}
    for row in ws.iter_rows():
        for cell in row:
            if normalize_label(cell.value) in normalized_aliases:
                return cell.row
    raise ValueError(
        f"Template sheet {ws.title!r} is missing label: {aliases[0]}."
    )


def _find_optional_label_row(ws: Any, aliases: tuple[str, ...]) -> int | None:
    try:
        return _find_label_row(ws, aliases)
    except ValueError:
        return None


def _find_labor_label_row(ws: Any, aliases: tuple[str, ...]) -> int:
    normalized_aliases = {normalize_label(alias) for alias in aliases}
    for row in ws.iter_rows():
        if not any(normalize_label(cell.value) in normalized_aliases for cell in row):
            continue
        if any(
            isinstance(cell.value, str) and cell.value.startswith("=")
            for cell in row
        ) and isinstance(ws[f"J{row[0].row}"].value, str):
            return row[0].row
    raise ValueError(
        f"Template sheet {ws.title!r} is missing labor label: {aliases[0]}."
    )


def _formula_cell_on_row(
    ws: Any,
    row_number: int,
    *,
    preferred_column: str,
) -> str:
    preferred = ws[f"{preferred_column}{row_number}"]
    if isinstance(preferred.value, str) and preferred.value.startswith("="):
        return preferred.coordinate
    for cell in ws[row_number]:
        if isinstance(cell.value, str) and cell.value.startswith("="):
            return cell.coordinate
    raise ValueError(
        f"Template row {ws.title}!{row_number} has no calculation formula."
    )


def _travel_cost_row(ws: Any, row_number: int, label: str) -> FormulaCostRow:
    cost_cell = _formula_cell_on_row(ws, row_number, preferred_column="H")
    formula = str(ws[cost_cell].value or "")
    referenced_columns = []
    for column, reference_row in re.findall(r"\$?([A-Z]{1,3})\$?(\d+)", formula.upper()):
        if int(reference_row) == row_number and column not in referenced_columns:
            referenced_columns.append(column)
    trips_column = ""
    miles_column = ""
    for header_row in range(row_number - 1, max(row_number - 6, 0), -1):
        trips_column = _optional_header_column(
            ws,
            header_row,
            ("no of trips", "number of trips"),
        )
        miles_column = _optional_header_column(
            ws,
            header_row,
            ("r t miles", "rt miles", "round trip miles"),
        )
        if trips_column and miles_column:
            break
    if not trips_column or not miles_column:
        raise ValueError(f"Template {label} row has no discoverable travel headers.")
    rate_candidates = [
        column
        for column in referenced_columns
        if column not in {trips_column, miles_column, ws[cost_cell].column_letter}
    ]
    if not rate_candidates:
        raise ValueError(f"Template {label} formula has no discoverable rate input.")
    return FormulaCostRow(
        label=label,
        row=row_number,
        trips_cell=f"{trips_column}{row_number}",
        miles_cell=f"{miles_column}{row_number}",
        rate_cell=f"{rate_candidates[0]}{row_number}",
        cost_cell=cost_cell,
    )


def _per_trip_labor_row(
    ws: Any,
    row_number: int,
    label: str,
) -> PerTripLaborCostRow:
    header_row = _nearest_header_row(
        ws,
        row_number,
        ("est hours day", "no of people", "est rate"),
    )
    return PerTripLaborCostRow(
        label=label,
        row=row_number,
        hours_cell=(
            f"{_header_column(ws, header_row, ('est hours day', 'estimated hours day'))}"
            f"{row_number}"
        ),
        crew_size_cell=(
            f"{_header_column(ws, header_row, ('no of people', 'number of people'))}"
            f"{row_number}"
        ),
        rate_cell=(
            f"{_header_column(ws, header_row, ('est rate', 'estimated rate'))}"
            f"{row_number}"
        ),
        cost_cell=_formula_cell_on_row(
            ws,
            row_number,
            preferred_column="H",
        ),
    )


def _nearest_header_row(
    ws: Any,
    row_number: int,
    required_labels: tuple[str, ...],
) -> int:
    normalized_required = {normalize_label(label) for label in required_labels}
    for candidate_row in range(row_number - 1, max(row_number - 6, 0), -1):
        labels = {normalize_label(cell.value) for cell in ws[candidate_row]}
        if normalized_required.issubset(labels):
            return candidate_row
    raise ValueError(f"Template row {ws.title}!{row_number} has no labor input headers.")


def _header_column(ws: Any, row_number: int, aliases: tuple[str, ...]) -> str:
    column = _optional_header_column(ws, row_number, aliases)
    if column:
        return column
    raise ValueError(
        f"Template sheet {ws.title!r} is missing travel header: {aliases[0]}."
    )


def _optional_header_column(
    ws: Any,
    row_number: int,
    aliases: tuple[str, ...],
) -> str:
    normalized_aliases = {normalize_label(alias) for alias in aliases}
    for cell in ws[row_number]:
        if normalize_label(cell.value) in normalized_aliases:
            return cell.column_letter
    return ""


def _output_cell_for_label(ws: Any, aliases: tuple[str, ...]) -> str:
    normalized_aliases = {normalize_label(alias) for alias in aliases}
    for row in ws.iter_rows():
        for cell in row:
            if normalize_label(cell.value) not in normalized_aliases:
                continue
            row_cells = list(ws[cell.row])
            candidates = [
                candidate for candidate in row_cells if candidate.column > cell.column
            ] + [candidate for candidate in row_cells if candidate.column <= cell.column]
            for candidate in candidates:
                if isinstance(candidate.value, str) and candidate.value.startswith("="):
                    return candidate.coordinate
    raise ValueError(
        f"Template sheet {ws.title!r} is missing calculated output: {aliases[0]}."
    )
