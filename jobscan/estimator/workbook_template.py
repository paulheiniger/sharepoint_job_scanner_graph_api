from __future__ import annotations

import re
import shutil
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

from .rules import first_nonblank, to_float

DEFAULT_TEMPLATE_PATH = Path("data/estimate_samples/Estimate - Full Turnkey.xlsx")
DEFAULT_OUTPUT_DIR = Path("output/estimates")
REQUIRED_SHEETS = {"Estimate", "People", "Materials", "General", "Performance & Payment Bonds"}

HEADER_CELL_MAP = {
    "today": "C1",
    "job_name": "C2",
    "job_type": "C3",
    "site_address": "C4",
    "city_state_zip": "C5",
    "contact": "C6",
    "email": "C8",
    "phone": "C9",
    "estimated_square_feet": "C12",
}

MATERIAL_ROW_MAPPINGS: dict[str, dict[str, str]] = {
    "foam_1": {"type_code": "A19", "sqft": "C19", "thickness": "D19", "unit_price": "E19", "yield": "F19"},
    "foam_2": {"type_code": "A20", "sqft": "C20", "thickness": "D20", "unit_price": "E20", "yield": "F20"},
    "foam_3": {"type_code": "A21", "sqft": "C21", "thickness": "D21", "unit_price": "E21", "yield": "F21"},
    "coating_1": {"type_code": "A26", "sqft": "C26", "gal_per_square": "D26", "unit_price": "E26"},
    "coating_2": {"type_code": "A27", "sqft": "C27", "gal_per_square": "D27", "unit_price": "E27"},
    "coating_3": {"type_code": "A28", "sqft": "C28", "gal_per_square": "D28", "unit_price": "E28"},
    "thinner": {"type_code": "A33", "unit_price": "E33", "units": "G33"},
    "granules": {"type_code": "A36", "sqft": "C36", "unit_price": "E36"},
    "primer": {"type_code": "A39", "sqft": "C39", "unit_price": "E39"},
    "caulk_sealant_1": {"type_code": "A43", "unit_price": "E43", "units": "G43"},
    "caulk_sealant_2": {"type_code": "A45", "unit_price": "E45", "units": "G45"},
    "misc_seams": {"linear_feet": "C47"},
    "penetrations": {"units": "D49"},
    "hvac_units": {"units": "D51"},
    "drains": {"units": "D53"},
    "board_stock_1": {"type_code": "A58", "sqft": "C58", "thickness": "D58", "price_per_square": "E58"},
    "board_stock_2": {"type_code": "A59", "sqft": "C59", "thickness": "D59", "price_per_square": "E59"},
    "board_stock_3": {"type_code": "A60", "sqft": "C60", "thickness": "D60", "price_per_square": "E60"},
    "fasteners": {"unit_price_per_thousand": "E63"},
    "plates": {"unit_price_per_thousand": "E65"},
    "dumpsters": {"type_code": "A69", "sqft": "C69", "thickness": "D69", "unit_price": "E69", "margin_pct": "F69"},
    "lift_1": {"type_code": "A73", "size": "C73", "period": "D73", "unit_price": "E73", "margin_pct": "F73"},
    "lift_2": {"type_code": "A74", "size": "C74", "period": "D74", "unit_price": "E74", "margin_pct": "F74"},
    "delivery_fee": {"unit_price": "E76", "units": "G76"},
    "fabric": {"linear_feet": "C79", "width": "D79", "unit_price": "E79"},
    "edge_metal": {"linear_feet": "C82", "unit_price": "E82"},
    "gutter": {"linear_feet": "C84", "unit_price": "E84"},
    "downspouts": {"linear_feet": "C86", "unit_price": "E86"},
    "roof_hatch": {"unit_price": "E88", "units": "G88"},
    "scuppers": {"unit_price": "E90", "units": "G90"},
    "curbs": {"unit_price": "E92", "units": "G92"},
    "ladders": {"unit_price": "E94", "units": "G94"},
    "pitch_pockets": {"unit_price": "E96", "units": "G96"},
    "generator": {"days": "C99", "unit_price": "E99"},
    "misc": {"unit_price": "E101"},
    "freight": {"unit_price": "E103"},
    "sales_inspection_travel": {"trips": "B106", "round_trip_miles": "C106", "cost_per_mile": "E106"},
    "truck_expense": {"trips": "B108", "round_trip_miles": "C108", "cost_per_mile": "E108"},
}

LABOR_ROW_MAPPINGS: dict[str, dict[str, str]] = {
    "prep": {"days": "B116", "crew_size": "C116"},
    "prime": {"days": "B118", "crew_size": "C118"},
    "seam_sealer": {"days": "B120", "crew_size": "C120"},
    "base": {"days": "B122", "crew_size": "C122"},
    "top_coat": {"days": "B124", "crew_size": "C124"},
    "caulk_sf": {"days": "B126", "crew_size": "C126"},
    "details": {"days": "B128", "crew_size": "C128"},
    "top_coat_granules": {"days": "B130", "crew_size": "C130"},
    "touch_cleanup": {"days": "B132", "crew_size": "C132"},
    "misc": {"days": "B134", "crew_size": "C134"},
    "loading": {"hours_per_day": "C137", "crew_size": "E137", "rate": "G137"},
    "traveling": {"hours_per_day": "C139", "crew_size": "E139", "rate": "G139"},
    "infrared_scan": {"hours": "C142", "crew_size": "E142", "rate": "G142"},
    "meals_lodging": {"days": "C145", "crew_size": "E145", "rate": "G145"},
}

WARRANTY_CELL_MAP = {
    "warranty_years": "C154",
    "warranty_type": "D154",
    "warranty_sqft": "E154",
    "warranty_unit_cost": "F154",
    "insurance_amount": "F156",
    "permits_amount": "F158",
    "overhead_pct": "F165",
    "profit_pct": "F167",
}

TOTAL_FORMULA_CELLS = {
    "subtotal_materials": "H110",
    "subtotal_labor": "H148",
    "total_job_cost": "H163",
    "worksheet_price": "H169",
}


@dataclass
class WorkbookFillResult:
    output_path: Path
    written_cells: dict[str, Any] = field(default_factory=dict)
    formula_cells: dict[str, str] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("_")
    return cleaned[:80] or "estimate_draft"


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def load_template(template_path: Path) -> Any:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("Install openpyxl to fill estimate workbook templates.") from exc
    workbook = openpyxl.load_workbook(template_path, data_only=False)
    missing = REQUIRED_SHEETS.difference(workbook.sheetnames)
    if missing:
        raise ValueError(f"Estimate template is missing required sheets: {', '.join(sorted(missing))}")
    return workbook


def write_input_cell(ws: Any, cell: str, value: Any, written: dict[str, Any], *, allow_formula_overwrite: bool = False) -> None:
    if value is None or value == "":
        return
    existing = ws[cell].value
    if is_formula(existing) and not allow_formula_overwrite:
        return
    ws[cell] = value
    written[cell] = value


def write_mapping(ws: Any, mapping: dict[str, str], values: dict[str, Any], written: dict[str, Any]) -> None:
    for field, cell in mapping.items():
        if field in values:
            write_input_cell(ws, cell, values.get(field), written)


def read_material_reference_pricing(template_path: Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_template(template_path)
    ws = workbook["Materials"]
    return {
        "solvents": _read_material_rows(ws, range(2, 4), {"name": "A", "cost": "C", "unit": "D", "per_gal": "E"}),
        "fabric_widths": _read_material_rows(ws, range(6, 13), {"width": "B", "cost": "C", "unit": "D"}),
        "board_thicknesses": _read_material_rows(ws, range(15, 22), {"name": "A", "thickness": "B", "cost": "C", "truckload": "D", "unit": "E"}),
        "fasteners": _read_material_rows(ws, range(25, 33), {"name": "A", "length": "B", "cost": "C", "unit": "D"}),
        "plates": _read_material_rows(ws, range(35, 36), {"name": "A", "cost": "C", "unit": "D"}),
    }


def _read_material_rows(ws: Any, rows: range, columns: dict[str, str]) -> list[dict[str, Any]]:
    out = []
    for row_num in rows:
        record = {field: ws[f"{column}{row_num}"].value for field, column in columns.items()}
        if any(value not in (None, "") for value in record.values()):
            out.append(record)
    return out


def estimate_to_workbook_inputs(result: dict[str, Any], metadata: dict[str, Any] | None = None) -> dict[str, Any]:
    metadata = metadata or {}
    scope = result.get("scope", {})
    materials = result.get("materials", {})
    labor = result.get("labor", {})
    travel = result.get("travel", {})
    decision = result.get("decision_tree", {})
    material_assumptions = decision.get("material_assumptions", {})
    area = to_float(scope.get("surface_area_sqft")) or to_float(scope.get("wall_area_sqft")) or 0
    crew_size = int(to_float(labor.get("recommended_crew_size")) or 1)
    duration_days = to_float(labor.get("estimated_duration_days_high")) or 1
    prep_days = max(round(duration_days * 0.2, 2), 0.25)
    detail_days = max(round(duration_days * 0.2, 2), 0.25) if material_assumptions.get("seam_treatment_recommended") or material_assumptions.get("fastener_treatment_recommended") else 0
    top_days = max(round(duration_days * 0.25, 2), 0.25)
    base_days = max(round(duration_days * 0.25, 2), 0.25)
    cleanup_days = max(round(duration_days * 0.1, 2), 0.25)
    coating_item = next((item for item in materials.get("material_items", []) if str(item.get("unit") or "").lower() == "gal"), {})
    coating_gallons = to_float(coating_item.get("quantity")) or 0
    gal_per_square = coating_gallons * 100 / area if area else None
    foam_item = next((item for item in materials.get("material_items", []) if "board" in str(item.get("unit") or "").lower()), {})
    return {
        "header": {
            "today": date.today(),
            "job_name": first_nonblank(metadata.get("job_name"), scope.get("job_name"), "Estimator Prototype Draft"),
            "job_type": first_nonblank(metadata.get("job_type"), scope.get("project_type")),
            "site_address": first_nonblank(metadata.get("site_address"), scope.get("site_address")),
            "city_state_zip": first_nonblank(metadata.get("city_state_zip"), scope.get("location")),
            "contact": first_nonblank(metadata.get("contact"), scope.get("contact")),
            "email": first_nonblank(metadata.get("email"), scope.get("email")),
            "phone": first_nonblank(metadata.get("phone"), scope.get("phone")),
            "estimated_square_feet": area or None,
        },
        "materials": {
            "foam_1": {
                "sqft": area if foam_item else None,
                "thickness": material_assumptions.get("foam_thickness_inches"),
                "unit_price": foam_item.get("unit_price"),
            },
            "coating_1": {
                "sqft": area if coating_item else None,
                "gal_per_square": gal_per_square,
                "unit_price": coating_item.get("unit_price"),
            },
            "primer": {"sqft": area if material_assumptions.get("primer_allowance_recommended") else None},
            "misc_seams": {"linear_feet": round(area / 100, 1) if material_assumptions.get("seam_treatment_recommended") and area else None},
            "penetrations": {"units": 10 if decision.get("condition_flags", {}).get("penetrations_complexity") == "high" else None},
            "truck_expense": {
                "trips": max(int(duration_days), 1),
                "round_trip_miles": travel.get("estimated_round_trip_miles"),
                "cost_per_mile": 1,
            },
            "sales_inspection_travel": {
                "trips": 1,
                "round_trip_miles": travel.get("estimated_round_trip_miles"),
                "cost_per_mile": 0.75,
            },
        },
        "labor": {
            "prep": {"days": prep_days, "crew_size": crew_size},
            "prime": {"days": 0.5 if material_assumptions.get("primer_allowance_recommended") else 0, "crew_size": crew_size},
            "seam_sealer": {"days": detail_days, "crew_size": crew_size},
            "base": {"days": base_days, "crew_size": crew_size},
            "top_coat": {"days": top_days, "crew_size": crew_size},
            "details": {"days": detail_days, "crew_size": crew_size},
            "touch_cleanup": {"days": cleanup_days, "crew_size": crew_size},
            "traveling": {"hours_per_day": round((to_float(travel.get("estimated_drive_time_minutes_one_way")) or 0) * 2 / 60, 2), "crew_size": crew_size},
            "meals_lodging": {"days": duration_days if travel.get("lodging_required_possible") else 0, "crew_size": crew_size},
        },
        "warranty": {
            "warranty_years": scope.get("warranty_target"),
            "warranty_sqft": area or None,
            "overhead_pct": metadata.get("overhead_pct"),
            "profit_pct": metadata.get("profit_pct"),
        },
    }


def fill_estimate_workbook(
    template_path: Path,
    result: dict[str, Any],
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    *,
    metadata: dict[str, Any] | None = None,
    output_name: str | None = None,
    workbook_inputs: dict[str, Any] | None = None,
) -> WorkbookFillResult:
    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"Estimate workbook template not found: {template_path}")
    output_dir.mkdir(parents=True, exist_ok=True)
    inputs = workbook_inputs or estimate_to_workbook_inputs(result, metadata)
    job_name = first_nonblank(inputs.get("header", {}).get("job_name"), "estimate_draft")
    destination = output_dir / (output_name or f"{safe_filename(job_name)}.xlsx")
    shutil.copy2(template_path, destination)

    workbook = load_template(destination)
    ws = workbook["Estimate"]
    written: dict[str, Any] = {}
    warnings: list[str] = []

    write_mapping(ws, HEADER_CELL_MAP, inputs.get("header", {}), written)
    for row_key, values in (inputs.get("materials") or {}).items():
        mapping = MATERIAL_ROW_MAPPINGS.get(row_key)
        if mapping:
            write_mapping(ws, mapping, values or {}, written)
        else:
            warnings.append(f"Unknown material mapping: {row_key}")
    for row_key, values in (inputs.get("labor") or {}).items():
        mapping = LABOR_ROW_MAPPINGS.get(row_key)
        if mapping:
            write_mapping(ws, mapping, values or {}, written)
        else:
            warnings.append(f"Unknown labor mapping: {row_key}")
    write_mapping(ws, WARRANTY_CELL_MAP, inputs.get("warranty", {}), written)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(destination)

    formula_cells = {name: ws[cell].value for name, cell in TOTAL_FORMULA_CELLS.items() if is_formula(ws[cell].value)}
    formula_cells["price_per_sqft"] = "=H169/C12"
    return WorkbookFillResult(output_path=destination, written_cells=written, formula_cells=formula_cells, warnings=warnings)
