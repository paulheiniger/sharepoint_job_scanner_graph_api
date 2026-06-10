from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Iterable

from .extractors import find_job_folders, scan_job_folder
from .models import JobRecord


FIELD_ORDER = [
    "job_id", "status", "division", "pipeline_status", "scan_root", "source_year",
    "customer", "job_name", "job_type", "site_address", "city", "state", "zip_code",
    "contact_name", "contact_title", "contact_email", "contact_phone",
    "estimate_date", "estimated_sqft", "material_subtotal", "labor_subtotal",
    "warranty_bonding_insurance_subtotal", "total_job_cost", "overhead_pct", "overhead_amount",
    "profit_pct", "profit_amount", "worksheet_price", "final_price", "price_per_sqft",
    "invoice_number", "invoice_amount", "invoice_date",
    "has_signed_contract", "has_invoice", "has_warranty", "has_proposal", "has_job_spec", "has_aerial", "has_notes",
    "photo_count", "duplicate_photo_count", "image_files_cached", "skipped_image_count",
    "crew_leader", "crew_type", "scheduled_sequence", "estimated_start_date", "estimated_duration_days",
    "estimated_end_date", "schedule_status", "ready_to_schedule", "blocking_issue", "schedule_notes",
    "schedule_source_file", "schedule_confidence",
    "folder_name", "folder_path", "folder_url", "estimate_file", "invoice_file", "warnings",
]


def scan_root(root: Path, scan_context: str = "") -> list[JobRecord]:
    root = root.resolve()
    folders = find_job_folders(root)
    records = [scan_job_folder(folder, root=root, scan_context=scan_context) for folder in folders]
    with_estimate = sum(1 for record in records if record.estimate_file)
    without_estimate = len(records) - with_estimate
    immediate_child_count = getattr(find_job_folders, "last_immediate_child_count", 0)
    skipped_admin_count = getattr(find_job_folders, "last_skipped_admin_count", 0)
    print(f"Immediate child job folders found: {immediate_child_count}")
    print(f"Records emitted from estimate workbooks: {with_estimate}")
    print(f"Records emitted without estimate workbook: {without_estimate}")
    print(f"Skipped administrative/template folders: {skipped_admin_count}")
    return records


def records_as_dicts(records: Iterable[JobRecord]) -> list[dict]:
    rows = []
    for record in records:
        row = record.to_dict()
        row["warnings"] = "; ".join(row.get("warnings") or [])
        rows.append({field: row.get(field) for field in FIELD_ORDER})
    return rows


def write_csv(records: Iterable[JobRecord], path: Path) -> None:
    rows = records_as_dicts(records)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELD_ORDER)
        writer.writeheader()
        writer.writerows(rows)


def write_json(records: Iterable[JobRecord], path: Path) -> None:
    rows = records_as_dicts(records)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8")


def write_excel(records: Iterable[JobRecord], path: Path) -> None:
    """Write a simple Excel workbook for local use.

    This is intentionally plain; Power BI/SharePoint Lists should be the longer-term dashboard layer.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise RuntimeError("Install openpyxl to write Excel output: pip install openpyxl") from exc

    rows = records_as_dicts(records)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Index"
    ws.append(FIELD_ORDER)
    for row in rows:
        ws.append([row.get(field) for field in FIELD_ORDER])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 28, "B": 14, "C": 16, "D": 20, "E": 18, "F": 24, "G": 18, "J": 20,
        "Y": 14, "Z": 14, "AA": 18, "AL": 40, "AM": 50, "AN": 50, "AO": 50,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for col in range(1, ws.max_column + 1):
        letter = openpyxl.utils.get_column_letter(col)
        if ws.column_dimensions[letter].width == 13:
            ws.column_dimensions[letter].width = 16
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Format money and percentages.
    money_headers = {"material_subtotal", "labor_subtotal", "total_job_cost", "overhead_amount", "profit_amount", "worksheet_price", "final_price", "price_per_sqft", "invoice_amount"}
    pct_headers = {"overhead_pct", "profit_pct"}
    header_idx = {cell.value: cell.column for cell in ws[1]}
    for header in money_headers:
        col = header_idx.get(header)
        if col:
            for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                for c in cell:
                    c.number_format = '$#,##0.00'
    for header in pct_headers:
        col = header_idx.get(header)
        if col:
            for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                for c in cell:
                    c.number_format = '0.0'

    if ws.max_row > 1:
        table = Table(displayName="JobIndex", ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
