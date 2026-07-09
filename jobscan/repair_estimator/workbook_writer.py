from __future__ import annotations

import math
import re
from datetime import date
from pathlib import Path
from typing import Any

from .estimator import RepairEstimateResult

DEFAULT_REPAIR_TEMPLATE_PATH = Path("templates/Estimate + Spec Form - Contracted Repairs.xlsx")
DEFAULT_REPAIR_OUTPUT_DIR = Path("output/repair_estimator/filled_templates")


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("_")
    return cleaned[:90] or "repair_estimate"


def number_or_none(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(str(value).replace(",", "").replace("$", ""))
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def text_or_blank(value: Any) -> str:
    if value is None:
        return ""
    try:
        import pandas as pd

        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def ceil_or_none(value: Any) -> int | None:
    number = number_or_none(value)
    if number is None or number <= 0:
        return None
    return int(math.ceil(number))


def resolve_repair_template_path(template_path: Path | str | None = None) -> Path:
    path = Path(template_path) if template_path else DEFAULT_REPAIR_TEMPLATE_PATH
    if not path.exists():
        raise FileNotFoundError(f"Repair estimate template not found: {path}")
    return path


def _write(ws: Any, cell: str, value: Any) -> None:
    if value is None or value == "":
        return
    ws[cell] = value


def _as_result_dict(result: RepairEstimateResult | dict[str, Any]) -> dict[str, Any]:
    return result.to_dict() if isinstance(result, RepairEstimateResult) else dict(result)


def _package_lookup(result: dict[str, Any]) -> dict[str, dict[str, Any]]:
    packages: dict[str, dict[str, Any]] = {}
    for package in result.get("selected_repair_packages") or []:
        name = text_or_blank(package.get("material_package"))
        if name:
            packages[name] = package
    return packages


def _package_cost(package: dict[str, Any] | None) -> float | None:
    if not package:
        return None
    return number_or_none(package.get("median_total_cost"))


def _quantity_from_cost(package: dict[str, Any] | None, unit_price: float, default: float = 1.0) -> float:
    cost = _package_cost(package)
    if cost is None or cost <= 0 or unit_price <= 0:
        return default
    return max(default, math.ceil(cost / unit_price))


def _repair_scope_summary(parsed: dict[str, Any]) -> str:
    parts = [
        text_or_blank(parsed.get("roof_type")).replace("_", " "),
        text_or_blank(parsed.get("issue_type")).replace("_", " "),
        text_or_blank(parsed.get("affected_area")),
    ]
    text = " - ".join(part for part in parts if part and part != "unknown")
    return text or "Roof repair"


def _scope_bullets(parsed: dict[str, Any]) -> list[str]:
    bullets = ["Set-up jobsite safety and protect work area."]
    issue = text_or_blank(parsed.get("issue_type"))
    actions = [text_or_blank(action).replace("_", " ") for action in parsed.get("actions_requested") or []]
    materials = [text_or_blank(material).replace("_", " ") for material in parsed.get("materials_mentioned") or []]
    if parsed.get("leak_present"):
        bullets.append("Locate and repair active leak condition.")
    if issue and issue != "unknown":
        bullets.append(f"Repair {issue.replace('_', ' ')} condition.")
    for action in actions:
        if action:
            bullets.append(action.capitalize() + " affected area.")
    if materials:
        bullets.append("Use " + ", ".join(materials) + " as required for the repair.")
    if text_or_blank(parsed.get("access_complexity")) == "high":
        bullets.append("Coordinate access constraints before mobilization.")
    return list(dict.fromkeys(bullets))


def _tech_notes(result: dict[str, Any]) -> str:
    flags = [text_or_blank(flag) for flag in result.get("review_flags") or [] if text_or_blank(flag)]
    evidence = result.get("evidence_summary") or {}
    summary = [
        f"Generated from {evidence.get('similar_repair_count', 0)} similar historical repair(s).",
        f"Confidence: {text_or_blank(result.get('confidence')).title() or 'Review'}",
    ]
    if flags:
        summary.append("Review: " + " ".join(flags))
    return " ".join(summary)


def _labor_plan(result: dict[str, Any]) -> dict[str, Any]:
    total_hours = number_or_none(result.get("estimated_labor_hours_target")) or 4.0
    parsed = result.get("parsed_scope") or {}
    crew_size = 1 if total_hours <= 4 else 2 if total_hours <= 16 else 3
    days = max(1, math.ceil(total_hours / max(crew_size * 8, 1)))
    travel_per_day = 1.0
    per_tech_total = total_hours / crew_size
    onsite_per_day = max((per_tech_total - travel_per_day * days) / days, 0.5)
    return {
        "days": days,
        "crew_size": crew_size,
        "per_tech_total": per_tech_total,
        "onsite_per_day": onsite_per_day,
        "travel_per_day": travel_per_day,
        "emergency": text_or_blank(parsed.get("emergency_or_standard")) == "emergency",
    }


def _material_quantities(result: dict[str, Any]) -> dict[str, float]:
    parsed = result.get("parsed_scope") or {}
    packages = _package_lookup(result)
    sqft = number_or_none(parsed.get("affected_area_sqft")) or 0
    linear_ft = number_or_none(parsed.get("affected_linear_feet")) or 0
    penetrations = number_or_none(parsed.get("penetration_count")) or 0
    issue = text_or_blank(parsed.get("issue_type"))
    materials = set(parsed.get("materials_mentioned") or [])
    actions = set(parsed.get("actions_requested") or [])

    coating_needed = "coating" in materials or "coat" in actions or issue == "small_coating_touch_up"
    fabric_needed = "fabric" in materials or "reinforce_with_fabric" in actions or issue in {
        "pipe_boot_leak",
        "open_seam",
        "curb_leak",
        "drain_leak",
        "puncture_or_patch",
        "skylight_curb_leak",
    }
    caulk_needed = "sealant" in materials or "seal" in actions or issue in {
        "pipe_boot_leak",
        "open_seam",
        "curb_leak",
        "flashing_leak",
        "unknown_leak",
        "skylight_curb_leak",
    }
    fastener_needed = "fasteners" in materials or issue == "exposed_fasteners"
    membrane_needed = "membrane" in materials or issue == "puncture_or_patch"

    silicone_gallons = max(math.ceil(sqft / 100), 1) if coating_needed else 0
    reinforced_gallons = max(math.ceil(sqft / 100), 1) if coating_needed and fabric_needed else 0
    dow_caulk = max(math.ceil(penetrations), 1) if caulk_needed else 0
    aldo_caulk = 1 if text_or_blank(parsed.get("emergency_or_standard")) == "emergency" and caulk_needed else 0
    fabric = max(linear_ft, sqft ** 0.5 if sqft > 0 else 25, 1) if fabric_needed else 0
    fasteners = max(math.ceil(penetrations * 4), 10) if fastener_needed else 0
    misc = 1 if membrane_needed or issue in {"gutter_downspout", "flashing_leak"} else 0

    if not any([silicone_gallons, reinforced_gallons, dow_caulk, aldo_caulk, fabric, fasteners, misc]):
        dow_caulk = _quantity_from_cost(packages.get("caulk_sealant"), 16, default=1)

    return {
        "silicone_coating": silicone_gallons,
        "reinforced_silicone": reinforced_gallons,
        "dow_caulk": dow_caulk,
        "aldo_caulk": aldo_caulk,
        "granules": 0,
        "fabric": round(fabric, 2) if fabric else 0,
        "brushes": 2 if any([silicone_gallons, reinforced_gallons, dow_caulk, aldo_caulk, fabric]) else 0,
        "solvent": 1 if any([dow_caulk, aldo_caulk]) else 0,
        "fasteners": fasteners,
        "misc": misc,
        "boots": max(math.ceil(penetrations), 0) if issue == "pipe_boot_leak" else 0,
        "mileage": 0,
    }


def _apply_formula_recalc_flags(workbook: Any) -> None:
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    except Exception:
        pass


def generate_repair_estimate_workbook(
    result: RepairEstimateResult | dict[str, Any],
    *,
    template_path: Path | str | None = None,
    output_dir: Path | str = DEFAULT_REPAIR_OUTPUT_DIR,
    output_filename: str | None = None,
    job_name: str = "",
    site_address: str = "",
    contact_name: str = "",
    contact_phone: str = "",
    contact_email: str = "",
    estimator: str = "",
    purchase_order: str = "",
    repair_date: date | str | None = None,
) -> Path:
    from openpyxl import load_workbook

    payload = _as_result_dict(result)
    parsed = payload.get("parsed_scope") or {}
    template = resolve_repair_template_path(template_path)
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = output_filename or f"{safe_filename(job_name or parsed.get('issue_type') or 'repair_estimate')}.xlsx"
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    output_path = out_dir / filename

    workbook = load_workbook(template)
    general = workbook["General Estimate"]
    spec = workbook["Job Spec"]
    tracking = workbook["Tracking"]

    _write(general, "G2", job_name or _repair_scope_summary(parsed))
    _write(general, "G3", site_address)
    _write(general, "G4", contact_name)
    _write(general, "G5", contact_phone)
    _write(general, "G6", contact_email)
    _write(general, "G7", estimator)
    _write(general, "B9", _labor_plan(payload)["days"])
    _write(general, "D9", purchase_order)
    _write(general, "G9", repair_date or date.today())
    _write(general, "D8", _repair_scope_summary(parsed))

    quantities = _material_quantities(payload)
    for cell, key in {
        "A11": "silicone_coating",
        "A12": "reinforced_silicone",
        "A13": "dow_caulk",
        "A14": "aldo_caulk",
        "A15": "granules",
        "A16": "fabric",
        "A17": "brushes",
        "A18": "solvent",
        "A19": "fasteners",
        "A20": "misc",
        "A21": "boots",
        "A22": "mileage",
    }.items():
        value = quantities[key]
        if value:
            _write(general, cell, value)

    labor = _labor_plan(payload)
    _write(general, "D23", round(labor["onsite_per_day"], 2))
    _write(general, "F23", round(labor["travel_per_day"], 2))
    active_rows = [24, 25, 26, 27][: int(labor["crew_size"])]
    for row in active_rows:
        _write(general, f"A{row}", round(labor["per_tech_total"], 2))
    _write(general, "A28", "Tech Notes:")
    _write(general, "B28", _tech_notes(payload))

    _write(spec, "A11", "\n".join("• " + bullet for bullet in _scope_bullets(parsed)[:10]))
    _write(spec, "F8", "Total sq. ft.")
    if parsed.get("affected_area_sqft"):
        _write(spec, "G8", parsed.get("affected_area_sqft"))
    if payload.get("similar_repairs"):
        first = payload["similar_repairs"][0]
        _write(
            spec,
            "B51",
            "Historical evidence:\n"
            f"Closest repair: {text_or_blank(first.get('job_name'))}; "
            f"hours={text_or_blank(first.get('historical_labor_hours'))}; "
            f"invoice={text_or_blank(first.get('invoice_amount'))}.",
        )

    _write(tracking, "R2", int(labor["crew_size"]))
    _write(tracking, "S2", "Estimated from repair history; update actuals daily.")

    _apply_formula_recalc_flags(workbook)
    workbook.save(output_path)
    return output_path
