from __future__ import annotations

import csv
import json
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from .models import JobRecord, rel

JOB_TRACKING_SUMMARY_FIELDS = [
    "job_id",
    "tracking_file",
    "customer",
    "job_name",
    "division",
    "pipeline_status",
    "actual_first_work_date",
    "actual_last_work_date",
    "actual_work_day_count",
    "actual_labor_hours",
    "actual_travel_hours",
    "actual_load_hours",
    "actual_os_hours",
    "actual_mileage",
    "actual_os_mileage",
    "actual_foam_strokes",
    "actual_foam_thickness_inches",
    "actual_foam_sqft",
    "actual_foam_yield",
    "actual_base_coat_1",
    "actual_base_coat_2",
    "actual_granules",
    "actual_af_buttergrade",
    "actual_caulk",
    "actual_primer",
    "actual_sf",
    "estimated_labor_hours",
    "estimated_travel_hours",
    "estimated_load_hours",
    "estimated_overhead",
    "estimated_mileage",
    "estimated_os_mileage",
    "estimated_foam_strokes",
    "estimated_foam_thickness_inches",
    "estimated_foam_sqft",
    "estimated_foam_yield",
    "estimated_base_coat_1",
    "estimated_base_coat_2",
    "estimated_granules",
    "estimated_af_buttergrade",
    "estimated_caulk",
    "estimated_primer",
    "estimated_sf",
    "labor_hours_variance",
    "travel_hours_variance",
    "load_hours_variance",
    "foam_strokes_variance",
    "foam_sqft_variance",
    "base_coat_1_variance",
    "base_coat_2_variance",
    "granules_variance",
    "af_buttergrade_variance",
    "caulk_variance",
    "primer_variance",
    "sf_variance",
    "tracking_notes",
    "tracking_warnings",
    "source_file",
    "source_path",
]

JOB_TRACKING_DAILY_FIELDS = [
    "job_id",
    "tracking_file",
    "work_date",
    "labor_hours",
    "travel_hours",
    "load_hours",
    "os_hours",
    "mileage",
    "os_mileage",
    "foam_strokes",
    "foam_thickness_inches",
    "foam_sqft",
    "foam_yield",
    "a_side_lot",
    "b_side_lot",
    "base_coat_1",
    "base_sqft",
    "base_gal_per_sq",
    "base_coat_2",
    "top_sqft",
    "top_gal_per_sq",
    "granules",
    "af_buttergrade",
    "caulk",
    "primer",
    "sf",
    "crew",
    "notes",
    "source_sheet",
    "source_row",
]

STOP_MARKERS = ("daily totals", "insert additional lines here")


def scan_job_tracking_for_records(root: Path, records: list[JobRecord]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    summaries: list[dict[str, Any]] = []
    daily_entries: list[dict[str, Any]] = []
    root = root.resolve()
    for record in records:
        folder = root / record.folder_path
        if not folder.exists():
            continue
        files = [path for path in folder.rglob("*") if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}]
        for path in sorted(files, key=lambda item: str(item).lower()):
            file_summaries, file_daily = extract_job_tracking_file(path, root, record)
            summaries.extend(file_summaries)
            daily_entries.extend(file_daily)
    return summaries, daily_entries


def extract_job_tracking_file(path: Path, root: Path, record: JobRecord) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return [], []

    summaries: list[dict[str, Any]] = []
    daily_entries: list[dict[str, Any]] = []
    for ws in tracking_sheets(wb, path):
        try:
            summary, daily = extract_tracking_sheet(ws, path, root, record)
        except Exception:
            continue
        summaries.append(summary)
        daily_entries.extend(daily)
    if len(summaries) <= 1:
        return summaries, daily_entries
    return [combine_tracking_summaries(summaries, daily_entries)], daily_entries


def combine_tracking_summaries(summaries: list[dict[str, Any]], daily_entries: list[dict[str, Any]]) -> dict[str, Any]:
    combined: dict[str, Any] = {field: None for field in JOB_TRACKING_SUMMARY_FIELDS}
    first = summaries[0]
    for field in [
        "job_id",
        "tracking_file",
        "customer",
        "job_name",
        "division",
        "pipeline_status",
        "source_file",
        "source_path",
    ]:
        combined[field] = next((summary.get(field) for summary in summaries if summary.get(field)), first.get(field))

    dates = sorted({entry.get("work_date") for entry in daily_entries if entry.get("work_date")})
    combined["actual_first_work_date"] = dates[0] if dates else first_nonempty_summary_value(summaries, "actual_first_work_date")
    combined["actual_last_work_date"] = dates[-1] if dates else first_nonempty_summary_value(summaries, "actual_last_work_date")
    combined["actual_work_day_count"] = len(dates) if dates else first_nonempty_summary_value(summaries, "actual_work_day_count")

    additive_actual_fields = [
        "actual_labor_hours",
        "actual_travel_hours",
        "actual_load_hours",
        "actual_os_hours",
        "actual_mileage",
        "actual_os_mileage",
        "actual_foam_strokes",
        "actual_foam_sqft",
        "actual_base_coat_1",
        "actual_base_coat_2",
        "actual_granules",
        "actual_af_buttergrade",
        "actual_caulk",
        "actual_primer",
        "actual_sf",
    ]
    for field in additive_actual_fields:
        combined[field] = sum_summary_values(summaries, field)

    for field in ["actual_foam_thickness_inches", "actual_foam_yield"]:
        combined[field] = average_summary_values(summaries, field)

    for field in [
        "estimated_labor_hours",
        "estimated_travel_hours",
        "estimated_load_hours",
        "estimated_overhead",
        "estimated_mileage",
        "estimated_os_mileage",
        "estimated_foam_strokes",
        "estimated_foam_thickness_inches",
        "estimated_foam_sqft",
        "estimated_foam_yield",
        "estimated_base_coat_1",
        "estimated_base_coat_2",
        "estimated_granules",
        "estimated_af_buttergrade",
        "estimated_caulk",
        "estimated_primer",
        "estimated_sf",
    ]:
        combined[field] = first_nonempty_summary_value(summaries, field)

    for source_key, variance_key in VARIANCE_KEYS.items():
        actual = combined.get(f"actual_{source_key}")
        estimated = combined.get(f"estimated_{source_key}")
        if isinstance(actual, (int, float)) and isinstance(estimated, (int, float)):
            combined[variance_key] = clean_number(float(estimated) - float(actual))
        else:
            combined[variance_key] = first_nonempty_summary_value(summaries, variance_key)

    notes = []
    warnings = []
    for summary in summaries:
        notes.extend(part.strip() for part in str(summary.get("tracking_notes") or "").split(";") if part.strip())
        warnings.extend(part.strip() for part in str(summary.get("tracking_warnings") or "").split(";") if part.strip())
    combined["tracking_notes"] = "; ".join(dict.fromkeys(notes))
    combined["tracking_warnings"] = "; ".join(dict.fromkeys(warnings))
    return {field: combined.get(field) for field in JOB_TRACKING_SUMMARY_FIELDS}


def first_nonempty_summary_value(summaries: list[dict[str, Any]], field: str) -> Any:
    for summary in summaries:
        value = summary.get(field)
        if value is not None and value != "":
            return value
    return None


def sum_summary_values(summaries: list[dict[str, Any]], field: str) -> int | float | None:
    values = [summary.get(field) for summary in summaries if isinstance(summary.get(field), (int, float))]
    if not values:
        return None
    return clean_number(float(sum(values)))


def average_summary_values(summaries: list[dict[str, Any]], field: str) -> int | float | None:
    values = [float(summary.get(field)) for summary in summaries if isinstance(summary.get(field), (int, float))]
    if not values:
        return None
    return clean_number(sum(values) / len(values))


def apply_job_tracking_to_record(record: JobRecord, folder: Path, root: Path, tracking_files: list[Path]) -> None:
    summaries: list[dict[str, Any]] = []
    for path in tracking_files:
        file_summaries, _daily = extract_job_tracking_file(path, root, record)
        summaries.extend(file_summaries)
    if not summaries:
        return
    summary = summaries[0]
    record.has_job_tracking_form = True
    record.job_tracking_file = summary.get("tracking_file")
    for key in (
        "actual_first_work_date",
        "actual_last_work_date",
        "actual_work_day_count",
        "actual_labor_hours",
        "actual_travel_hours",
        "actual_load_hours",
        "actual_mileage",
        "actual_base_coat_1",
        "actual_base_coat_2",
        "actual_af_buttergrade",
        "actual_caulk",
        "labor_hours_variance",
        "tracking_warnings",
    ):
        if hasattr(record, key):
            setattr(record, key, summary.get(key))


def is_job_tracking_file(path: Path) -> bool:
    if "job tracking" in path.name.lower():
        return True
    summaries, _daily = extract_job_tracking_file(path, path.parent, JobRecord(job_id="", folder_name="", folder_path=""))
    return bool(summaries)


def tracking_sheets(wb: Any, path: Path) -> list[Any]:
    filename_match = "job tracking" in path.name.lower()
    matches = []
    for ws in wb.worksheets:
        if not sheet_has_any_value(ws):
            continue
        if filename_match or "job tracking" in ws.title.lower() or sheet_has_tracking_markers(ws):
            matches.append(ws)
    return matches


def sheet_has_any_value(ws: Any) -> bool:
    return any(cell.value is not None for row in ws.iter_rows() for cell in row)


def sheet_has_tracking_markers(ws: Any) -> bool:
    labels = {norm_label(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None}
    return {"actual amounts", "estimated amounts", "over/under"}.issubset(labels)


def extract_tracking_sheet(ws: Any, path: Path, root: Path, record: JobRecord) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    warnings: list[str] = []
    source_path = rel(path, root)
    actual_marker = find_cell(ws, "Actual Amounts")
    estimated_marker = find_cell(ws, "Estimated Amounts")
    over_under_marker = find_cell(ws, "Over/Under")

    daily_entries: list[dict[str, Any]] = []
    actual_totals: dict[str, Any] = {}
    if actual_marker:
        header_row, actual_columns = find_tracking_header(ws, actual_marker.row + 1, actual_marker.row + 5)
        if header_row:
            daily_entries, actual_totals = extract_daily_entries(ws, header_row, actual_columns, source_path, record)
        else:
            warnings.append("Missing Actual Amounts header row")
    else:
        header_row, actual_columns = find_tracking_header(ws, 1, min(ws.max_row, 6))
        if header_row:
            daily_entries, actual_totals = extract_daily_entries(ws, header_row, actual_columns, source_path, record)
        else:
            warnings.append("Missing Actual Amounts section")

    estimated = extract_values_section(ws, estimated_marker.row if estimated_marker else None, prefix="estimated")
    variance = extract_values_section(ws, over_under_marker.row if over_under_marker else None, prefix="", variance=True)
    actual = actual_totals or summed_actuals(daily_entries)
    dates = [entry["work_date"] for entry in daily_entries if entry.get("work_date")]
    notes = [str(entry.get("notes")).strip() for entry in daily_entries if entry.get("notes")]

    summary = {
        "job_id": record.job_id,
        "tracking_file": source_path,
        "customer": record.customer,
        "job_name": record.job_name or tracking_job_name(ws, header_row if header_row else actual_marker.row + 1 if actual_marker else None),
        "division": record.division,
        "pipeline_status": record.pipeline_status,
        "actual_first_work_date": min(dates) if dates else None,
        "actual_last_work_date": max(dates) if dates else None,
        "actual_work_day_count": len(set(dates)),
        "tracking_notes": "; ".join(dict.fromkeys(notes)),
        "tracking_warnings": "; ".join(dict.fromkeys(warnings)),
        "source_file": path.name,
        "source_path": source_path,
    }
    for source_key, target_key in ACTUAL_SUMMARY_KEYS.items():
        summary[target_key] = actual.get(source_key)
    summary.update(estimated)
    summary.update(variance)
    return {field: summary.get(field) for field in JOB_TRACKING_SUMMARY_FIELDS}, daily_entries


def extract_daily_entries(ws: Any, header_row: int, columns: dict[str, int], tracking_file: str, record: JobRecord) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    totals: dict[str, Any] = {}
    blank_streak = 0
    for row_num in range(header_row + 1, ws.max_row + 1):
        row_text = " ".join(str(ws.cell(row=row_num, column=col).value or "") for col in range(1, min(ws.max_column, 20) + 1)).strip()
        marker = norm_label(row_text)
        if any(stop in marker for stop in STOP_MARKERS):
            if "daily totals" in marker:
                totals = row_values_by_columns(ws, row_num, columns)
            break
        if not row_text:
            blank_streak += 1
            if blank_streak >= 3:
                break
            continue
        blank_streak = 0
        work_date = parse_work_date(ws.cell(row=row_num, column=1).value)
        if not work_date:
            continue
        values = row_values_by_columns(ws, row_num, columns)
        entry = {
            "job_id": record.job_id,
            "tracking_file": tracking_file,
            "work_date": work_date,
            "source_sheet": ws.title,
            "source_row": row_num,
        }
        for field in JOB_TRACKING_DAILY_FIELDS:
            if field not in entry:
                entry[field] = values.get(field)
        entries.append({field: entry.get(field) for field in JOB_TRACKING_DAILY_FIELDS})
    return entries, totals


def extract_values_section(ws: Any, marker_row: int | None, *, prefix: str, variance: bool = False) -> dict[str, Any]:
    if marker_row is None:
        return {}
    header_row, columns = find_tracking_header(ws, marker_row + 1, marker_row + 4)
    if not header_row:
        return {}
    value_row = first_numeric_row(ws, header_row + 1, min(ws.max_row, header_row + 4), columns)
    if not value_row:
        return {}
    values = row_values_by_columns(ws, value_row, columns)
    out: dict[str, Any] = {}
    for key, value in values.items():
        if value is None:
            continue
        if variance:
            variance_key = VARIANCE_KEYS.get(key)
            if variance_key:
                out[variance_key] = value
        else:
            out[f"{prefix}_{key}"] = value
    return out


def find_tracking_header(ws: Any, start_row: int, end_row: int) -> tuple[int | None, dict[str, int]]:
    max_row = ws.max_row or 0
    if max_row <= 0:
        return None, {}
    first_row = max(1, start_row)
    last_row = min(end_row, max_row)
    if last_row < first_row:
        return None, {}
    for row_num, row in enumerate(ws.iter_rows(min_row=first_row, max_row=last_row), start=first_row):
        columns: dict[str, int] = {}
        for cell in row:
            key = tracking_key(cell.value)
            if key:
                column = getattr(cell, "column", None)
                if column is not None:
                    columns[key] = column
        if {"labor_hours", "travel_hours", "load_hours"}.issubset(columns):
            return row_num, columns
    return None, {}


def first_numeric_row(ws: Any, start_row: int, end_row: int, columns: dict[str, int]) -> int | None:
    for row_num in range(start_row, min(end_row, ws.max_row) + 1):
        if any(number(ws.cell(row=row_num, column=column).value) is not None for column in columns.values()):
            return row_num
    return None


def row_values_by_columns(ws: Any, row_num: int, columns: dict[str, int]) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for key, column in columns.items():
        raw = ws.cell(row=row_num, column=column).value
        values[key] = text(raw) if key in {"crew", "notes", "a_side_lot", "b_side_lot"} else number(raw)
    return values


def summed_actuals(entries: list[dict[str, Any]]) -> dict[str, Any]:
    totals: dict[str, Any] = {}
    for key in ACTUAL_SUMMARY_KEYS:
        values = [entry.get(key) for entry in entries if isinstance(entry.get(key), (int, float))]
        if values:
            totals[key] = clean_number(sum(values))
    return totals


def parse_work_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and 30000 < value < 60000:
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
    text_value = str(value or "").strip()
    match = re.match(r"^(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})$", text_value)
    if not match:
        return None
    month, day, year = match.groups()
    year_int = int(year)
    if year_int < 100:
        year_int += 2000
    try:
        return date(year_int, int(month), int(day)).isoformat()
    except ValueError:
        return None


def tracking_job_name(ws: Any, header_row: int | None) -> str | None:
    if not header_row:
        return None
    value = ws.cell(row=header_row, column=1).value
    return text(value)


def find_cell(ws: Any, label: str) -> Any | None:
    target = norm_label(label)
    for row in ws.iter_rows():
        for cell in row:
            if norm_label(cell.value) == target:
                return cell
    return None


def tracking_key(value: Any) -> str | None:
    label = norm_label(value)
    aliases = {
        "labor hours": "labor_hours",
        "travel hours": "travel_hours",
        "load hours": "load_hours",
        "os hours": "os_hours",
        "overhead": "overhead",
        "mileage": "mileage",
        "os mileage": "os_mileage",
        "foam strokes": "foam_strokes",
        "thickness in": "foam_thickness_inches",
        "foam thickness": "foam_thickness_inches",
        "sq ft foam": "foam_sqft",
        "sq ft foam": "foam_sqft",
        "foam sqft": "foam_sqft",
        "foam sf": "foam_sqft",
        "foam yield": "foam_yield",
        "a-side lot #": "a_side_lot",
        "a-side lot": "a_side_lot",
        "a side lot #": "a_side_lot",
        "a side lot": "a_side_lot",
        "b-side lot #": "b_side_lot",
        "b-side lot": "b_side_lot",
        "b side lot #": "b_side_lot",
        "b side lot": "b_side_lot",
        "base coat #1": "base_coat_1",
        "base coat": "base_coat_1",
        "sq ft base": "base_sqft",
        "sq. ft. base": "base_sqft",
        "gal sq base": "base_gal_per_sq",
        "gal/sq base": "base_gal_per_sq",
        "base coat #2": "base_coat_2",
        "top coat": "base_coat_2",
        "sq ft top": "top_sqft",
        "sq. ft. top": "top_sqft",
        "gal sq top": "top_gal_per_sq",
        "gal/sq top": "top_gal_per_sq",
        "granules": "granules",
        "af buttergrade": "af_buttergrade",
        "caulk": "caulk",
        "primer": "primer",
        "sf": "sf",
        "crew": "crew",
        "notes": "notes",
    }
    compact = label.replace("(", "").replace(")", "")
    return aliases.get(label) or aliases.get(compact)


def norm_label(value: Any) -> str:
    text_value = str(value or "").strip().lower()
    text_value = re.sub(r"[\r\n]+", " ", text_value)
    text_value = text_value.replace(".", "")
    text_value = re.sub(r"\s+", " ", text_value)
    return text_value


def number(value: Any) -> float | int | None:
    if isinstance(value, (int, float)):
        return clean_number(float(value))
    if value is None:
        return None
    raw = str(value).replace(",", "").replace("$", "").strip()
    if not raw:
        return None
    try:
        return clean_number(float(raw))
    except ValueError:
        return None


def clean_number(value: float) -> float | int:
    return int(value) if value.is_integer() else round(value, 2)


def text(value: Any) -> str | None:
    if value is None:
        return None
    out = str(value).strip()
    return out or None


def write_dataset_csv(rows: list[dict[str, Any]], fields: list[str], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows([{field: row.get(field) for field in fields} for row in rows])


def write_dataset_json(rows: list[dict[str, Any]], fields: list[str], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    normalized = [{field: row.get(field) for field in fields} for row in rows]
    path.write_text(json.dumps(normalized, indent=2, ensure_ascii=False), encoding="utf-8")


ACTUAL_SUMMARY_KEYS = {
    "labor_hours": "actual_labor_hours",
    "travel_hours": "actual_travel_hours",
    "load_hours": "actual_load_hours",
    "os_hours": "actual_os_hours",
    "mileage": "actual_mileage",
    "os_mileage": "actual_os_mileage",
    "foam_strokes": "actual_foam_strokes",
    "foam_thickness_inches": "actual_foam_thickness_inches",
    "foam_sqft": "actual_foam_sqft",
    "foam_yield": "actual_foam_yield",
    "base_coat_1": "actual_base_coat_1",
    "base_coat_2": "actual_base_coat_2",
    "granules": "actual_granules",
    "af_buttergrade": "actual_af_buttergrade",
    "caulk": "actual_caulk",
    "primer": "actual_primer",
    "sf": "actual_sf",
}

VARIANCE_KEYS = {
    "labor_hours": "labor_hours_variance",
    "travel_hours": "travel_hours_variance",
    "load_hours": "load_hours_variance",
    "foam_strokes": "foam_strokes_variance",
    "foam_sqft": "foam_sqft_variance",
    "base_coat_1": "base_coat_1_variance",
    "base_coat_2": "base_coat_2_variance",
    "granules": "granules_variance",
    "af_buttergrade": "af_buttergrade_variance",
    "caulk": "caulk_variance",
    "primer": "primer_variance",
    "sf": "sf_variance",
}
