from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

from .estimate_selection import estimate_id, infer_estimate_scope_type, select_primary_estimate
from .extractors import SPREADSHEET_EXTS, classify_files, extract_estimate_xlsx, money, rel
from .models import JobRecord

ESTIMATE_SUMMARY_FIELDS = [
    "job_id",
    "estimate_id",
    "division",
    "pipeline_status",
    "customer",
    "estimate_role",
    "estimate_scope_type",
    "job_name",
    "job_type",
    "estimate_file",
    "estimate_date",
    "estimate_version",
    "site_address",
    "city",
    "state",
    "zip_code",
    "estimated_sqft",
    "wall_area_sqft",
    "roof_deck_area_sqft",
    "linear_feet",
    "material_subtotal",
    "labor_subtotal",
    "equipment_subtotal",
    "subcontractor_subtotal",
    "travel_lodging",
    "adders_subtotal",
    "warranty_amount",
    "insurance_amount",
    "rental_amount",
    "subcontractor_amount",
    "misc_materials_amount",
    "total_job_cost",
    "overhead_pct",
    "overhead_amount",
    "profit_pct",
    "profit_amount",
    "worksheet_price",
    "final_price",
    "price_per_sqft",
    "estimated_labor_hours",
    "estimated_duration_days",
    "estimated_crew_size",
    "estimated_hours_per_day",
    "coating_required",
    "coating_type",
    "warranty_years",
    "labor_duration_source",
    "source_file",
    "source_path",
    "source_sheet",
    "folder_url",
    "extraction_warnings",
]

ESTIMATE_LINE_ITEM_FIELDS = [
    "estimate_id",
    "job_id",
    "estimate_file",
    "division",
    "pipeline_status",
    "customer",
    "job_name",
    "section",
    "line_item_name",
    "line_item_category",
    "description",
    "quantity",
    "unit",
    "unit_cost",
    "unit_price",
    "extended_cost",
    "markup_pct",
    "labor_days",
    "crew_size",
    "labor_hours",
    "vendor",
    "notes",
    "source_sheet",
    "source_row",
]

SECTION_ALIASES = {
    "materials": "Materials",
    "material": "Materials",
    "labor / subcontractor": "Labor / Subcontractor",
    "labor": "Labor",
    "equipment": "Equipment",
    "travel": "Travel",
    "lodging": "Lodging",
    "warranty": "Warranty",
    "coating": "Coating",
    "foam": "Foam",
    "fasteners": "Fasteners",
    "prep": "Prep",
    "details": "Details",
    "misc": "Misc",
    "subcontractor": "Subcontractor",
}

SUMMARY_LABELS = {
    "wall_area_sqft": ["wall area", "wall sq ft", "wall sqft"],
    "roof_deck_area_sqft": ["roof deck area", "deck area", "roof deck sq ft", "roof deck sqft"],
    "linear_feet": ["linear feet", "lineal feet", "lf"],
    "equipment_subtotal": ["subtotal equipment", "equipment subtotal"],
    "subcontractor_subtotal": ["subtotal subcontractor", "subcontractor subtotal"],
    "travel_lodging": ["travel/lodging", "travel lodging", "lodging", "travel"],
    "coating_type": ["coating type", "coating"],
    "warranty_years": ["warranty years", "warranty"],
}

ADDER_SKIP_PATTERNS = [
    r"^total$",
    r"^subtotal$",
    r"\btotal job cost\b",
    r"\boverhead\b",
    r"\bprofit\b",
    r"\bfinal price\b",
    r"\bprice\s*/?\s*sq\.?\s*ft\b",
    r"\btotal hours\b",
    r"\btotal days\b",
    r"\bworksheet price\b",
]


def estimate_version_from_name(path: Path) -> str | None:
    match = re.search(r"\b(?:rev|revision|version|v)\s*[-_. ]?([A-Za-z0-9]+)\b", path.stem, flags=re.I)
    return match.group(0) if match else None


def scan_estimate_datasets_for_records(root: Path, records: list[JobRecord]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    summaries: list[dict[str, Any]] = []
    line_items: list[dict[str, Any]] = []
    root = root.resolve()

    for record in records:
        folder = root / record.folder_path
        if not folder.exists():
            continue
        info = classify_files(folder)
        estimate_files = sorted(info.get("estimate_files") or [
            path for path in info.get("files", []) if path.suffix.lower() in SPREADSHEET_EXTS
        ], key=lambda path: path.name.lower())
        parsed_estimates = []
        for estimate_file in estimate_files:
            high_level = safe_extract_estimate_xlsx(estimate_file)
            parsed_estimates.append(
                {
                    "path": estimate_file,
                    "estimate_file": rel(estimate_file, root),
                    **high_level,
                }
            )
        primary, _reason = select_primary_estimate(parsed_estimates)
        primary_path = primary.get("path") if primary else None
        for parsed in parsed_estimates:
            estimate_file = parsed["path"]
            role = "primary" if estimate_file == primary_path else "supporting"
            summary, items = extract_estimate_dataset(estimate_file, root, record, estimate_role=role, high_level=parsed)
            summaries.append(summary)
            line_items.extend(items)
    return summaries, line_items


def safe_extract_estimate_xlsx(path: Path) -> dict[str, Any]:
    try:
        return extract_estimate_xlsx(path)
    except Exception as exc:
        return {"warnings": [f"estimate parse failed: {type(exc).__name__}: {exc}"]}


def extract_estimate_dataset(
    path: Path,
    root: Path,
    record: JobRecord,
    *,
    estimate_role: str | None = None,
    high_level: dict[str, Any] | None = None,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    warnings: list[str] = []
    high_level = high_level or safe_extract_estimate_xlsx(path)
    warnings.extend(high_level.get("warnings") or [])
    source_path = rel(path, root)
    eid = estimate_id(record.job_id, source_path)

    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        workbook_summary, line_items, workbook_warnings = extract_workbook_details(wb, path, root, record, eid)
        warnings.extend(workbook_warnings)
    except Exception as exc:
        workbook_summary = {}
        line_items = []
        warnings.append(f"detail extraction failed: {type(exc).__name__}: {exc}")

    scope_type = infer_estimate_scope_type(path, " ".join(str(high_level.get(key) or "") for key in ("job_name", "job_type")))
    role = estimate_role or ("primary" if source_path == record.primary_estimate_file or source_path == record.estimate_file else "supporting")
    summary = {
        "job_id": record.job_id,
        "estimate_id": eid,
        "division": record.division,
        "pipeline_status": record.pipeline_status,
        "customer": record.customer,
        "estimate_role": role,
        "estimate_scope_type": scope_type,
        "job_name": high_level.get("job_name") or record.job_name,
        "job_type": high_level.get("job_type") or record.job_type,
        "estimate_file": source_path,
        "estimate_date": high_level.get("estimate_date") or record.estimate_date,
        "estimate_version": estimate_version_from_name(path),
        "site_address": high_level.get("site_address") or record.site_address,
        "city": high_level.get("city") or record.city,
        "state": high_level.get("state") or record.state,
        "zip_code": high_level.get("zip_code") or record.zip_code,
        "estimated_sqft": high_level.get("estimated_sqft") or record.estimated_sqft,
        "wall_area_sqft": workbook_summary.get("wall_area_sqft"),
        "roof_deck_area_sqft": workbook_summary.get("roof_deck_area_sqft"),
        "linear_feet": workbook_summary.get("linear_feet"),
        "material_subtotal": high_level.get("material_subtotal") or record.material_subtotal,
        "labor_subtotal": high_level.get("labor_subtotal") or record.labor_subtotal,
        "equipment_subtotal": workbook_summary.get("equipment_subtotal"),
        "subcontractor_subtotal": workbook_summary.get("subcontractor_subtotal"),
        "travel_lodging": workbook_summary.get("travel_lodging"),
        "adders_subtotal": workbook_summary.get("adders_subtotal"),
        "warranty_amount": workbook_summary.get("warranty_amount"),
        "insurance_amount": workbook_summary.get("insurance_amount"),
        "rental_amount": workbook_summary.get("rental_amount"),
        "subcontractor_amount": workbook_summary.get("subcontractor_amount"),
        "misc_materials_amount": workbook_summary.get("misc_materials_amount"),
        "total_job_cost": high_level.get("total_job_cost") or record.total_job_cost,
        "overhead_pct": high_level.get("overhead_pct") or record.overhead_pct,
        "overhead_amount": high_level.get("overhead_amount") or record.overhead_amount,
        "profit_pct": high_level.get("profit_pct") or record.profit_pct,
        "profit_amount": high_level.get("profit_amount") or record.profit_amount,
        "worksheet_price": high_level.get("worksheet_price") or record.worksheet_price,
        "final_price": high_level.get("final_price") or record.final_price,
        "price_per_sqft": high_level.get("price_per_sqft") or record.price_per_sqft,
        "estimated_labor_hours": high_level.get("estimated_labor_hours") or record.estimated_labor_hours,
        "estimated_duration_days": high_level.get("estimated_duration_days") or record.estimated_duration_days,
        "estimated_crew_size": high_level.get("estimated_crew_size") or record.estimated_crew_size,
        "estimated_hours_per_day": high_level.get("estimated_hours_per_day") or record.estimated_hours_per_day,
        "coating_required": workbook_summary.get("coating_required"),
        "coating_type": workbook_summary.get("coating_type"),
        "warranty_years": workbook_summary.get("warranty_years"),
        "labor_duration_source": high_level.get("labor_duration_source") or record.labor_duration_source,
        "source_file": path.name,
        "source_path": source_path,
        "source_sheet": "Estimate",
        "folder_url": record.folder_url,
        "extraction_warnings": "; ".join(dict.fromkeys(warnings)),
    }
    return {field: summary.get(field) for field in ESTIMATE_SUMMARY_FIELDS}, line_items


def extract_workbook_details(wb: Any, path: Path, root: Path, record: JobRecord, estimate_id: str) -> tuple[dict[str, Any], list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    summary: dict[str, Any] = {}
    line_items: list[dict[str, Any]] = []
    source_path = rel(path, root)

    if "Estimate" not in wb.sheetnames:
        return summary, line_items, [f"No 'Estimate' sheet found in {path.name}"]

    ws = wb["Estimate"]
    if getattr(ws, "merged_cells", None) and ws.merged_cells.ranges:
        warnings.append("merged cell ambiguity")

    summary.update(extract_summary_labels(ws))
    summary.update(extract_coating_and_warranty(ws))

    sections = find_sections(ws)
    if not sections:
        warnings.append("section not recognized")

    for index, section in enumerate(sections):
        next_row = sections[index + 1]["row"] if index + 1 < len(sections) else ws.max_row + 1
        section_items, section_warnings = extract_section_line_items(
            ws=ws,
            section=section,
            end_row=next_row,
            source_path=source_path,
            record=record,
            estimate_id=estimate_id,
        )
        line_items.extend(section_items)
        warnings.extend(section_warnings)
    existing_rows = {item["source_row"] for item in line_items if item.get("source_sheet") == ws.title}
    adder_items = extract_estimate_adders(ws, source_path, record, estimate_id, existing_rows)
    line_items.extend(adder_items)
    summary.update(adder_rollups(adder_items))
    return summary, line_items, warnings


def extract_summary_labels(ws: Any) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, labels in SUMMARY_LABELS.items():
        value = first_value_right_of_any(ws, labels)
        if value is not None:
            out[key] = value
    return out


def extract_coating_and_warranty(ws: Any) -> dict[str, Any]:
    text_values = [
        str(cell.value).strip()
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None and str(cell.value).strip()
    ]
    combined = "\n".join(text_values)
    out: dict[str, Any] = {"coating_required": bool(re.search(r"\b(coating|top coat|topcoat)\b", combined, flags=re.I))}
    if "coating_type" not in out:
        match = re.search(r"\b(?:coating type|coating)\s*[:\-]?\s*([A-Za-z0-9 /+\-]+)", combined, flags=re.I)
        if match:
            out["coating_type"] = match.group(1).strip()
    if "warranty_years" not in out:
        match = re.search(r"\b(\d{1,2})\s*(?:year|yr)\s*warranty\b", combined, flags=re.I)
        if match:
            out["warranty_years"] = int(match.group(1))
    return out


def first_value_right_of_any(ws: Any, labels: list[str]) -> Any:
    needles = [norm_label(label) for label in labels]
    for row in ws.iter_rows():
        cells = list(row)
        for index, cell in enumerate(cells):
            label = norm_label(cell.value)
            if any(needle in label for needle in needles):
                right_values = [c.value for c in cells[index + 1 :] if c.value not in (None, "")]
                if right_values:
                    numeric = first_numeric(right_values)
                    return numeric if numeric is not None else str(right_values[0]).strip()
    return None


def find_sections(ws: Any) -> list[dict[str, Any]]:
    sections: list[dict[str, Any]] = []
    for row in ws.iter_rows():
        for cell in row:
            section = section_name(cell.value)
            if section and is_likely_section_row(ws, cell.row, cell.column):
                sections.append({"section": section, "row": cell.row, "column": cell.column})
                break
    deduped: list[dict[str, Any]] = []
    for section in sections:
        if deduped and deduped[-1]["row"] == section["row"]:
            continue
        deduped.append(section)
    return deduped


def is_likely_section_row(ws: Any, row_num: int, section_col: int) -> bool:
    row_values = [cell.value for cell in ws[row_num]]
    non_blank = [value for value in row_values if value not in (None, "")]
    if len(non_blank) <= 2:
        return True
    right_values = [ws.cell(row=row_num, column=col).value for col in range(section_col + 1, min(ws.max_column, section_col + 5) + 1)]
    return not any(numeric_value(value) is not None for value in right_values)


def extract_section_line_items(
    *,
    ws: Any,
    section: dict[str, Any],
    end_row: int,
    source_path: str,
    record: JobRecord,
    estimate_id: str,
) -> tuple[list[dict[str, Any]], list[str]]:
    section_name_text = section["section"]
    warnings: list[str] = []
    items: list[dict[str, Any]] = []
    header = find_line_item_header(ws, section["row"] + 1, min(section["row"] + 12, end_row - 1))

    if section_name_text == "Labor / Subcontractor":
        labor_items, labor_warnings = extract_labor_line_items(ws, section, end_row, source_path, record, header, estimate_id)
        return labor_items, labor_warnings

    if not header:
        warnings.append(f"section not recognized: {section_name_text}")
        return items, warnings

    header_row, columns = header
    blank_streak = 0
    blank_warning_added = False
    for row_num in range(header_row + 1, end_row):
        row_values = [ws.cell(row=row_num, column=col).value for col in range(1, ws.max_column + 1)]
        if not any(value not in (None, "") for value in row_values):
            blank_streak += 1
            if not blank_warning_added:
                warnings.append(f"skipped blank row: {ws.title}!{row_num}")
                blank_warning_added = True
            if blank_streak >= 8:
                break
            continue
        blank_streak = 0
        if is_total_or_subtotal_row(row_values):
            continue
        name = value_at(ws, row_num, columns, "line_item_name") or first_text(row_values)
        if not name:
            continue
        quantity = numeric_value(value_at(ws, row_num, columns, "quantity"))
        unit_cost = numeric_value(value_at(ws, row_num, columns, "unit_cost"))
        extended_cost = numeric_value(value_at(ws, row_num, columns, "extended_cost"))
        row_warnings = []
        if quantity is None:
            row_warnings.append("missing quantity")
        if unit_cost is None:
            row_warnings.append("missing unit cost")
        if extended_cost is None:
            row_warnings.append("missing total")
        if row_warnings:
            warnings.append(f"{ws.title}!{row_num}: {', '.join(row_warnings)}")
        items.append(
            line_item_row(
                record=record,
                estimate_id=estimate_id,
                estimate_file=source_path,
                section=section_name_text,
                source_sheet=ws.title,
                source_row=row_num,
                line_item_name=str(name).strip(),
                description=text_value(value_at(ws, row_num, columns, "description")),
                quantity=quantity,
                unit=text_value(value_at(ws, row_num, columns, "unit")),
                unit_cost=unit_cost,
                unit_price=numeric_value(value_at(ws, row_num, columns, "unit_price")),
                extended_cost=extended_cost,
                markup_pct=numeric_value(value_at(ws, row_num, columns, "markup_pct")),
                vendor=text_value(value_at(ws, row_num, columns, "vendor")),
                notes="; ".join(row_warnings),
            )
        )
    return items, warnings


def extract_labor_line_items(
    ws: Any,
    section: dict[str, Any],
    end_row: int,
    source_path: str,
    record: JobRecord,
    header: tuple[int, dict[str, int]] | None,
    estimate_id: str,
) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    if not header:
        warnings.append("section not recognized: Labor / Subcontractor")
        return [], warnings
    header_row, columns = header
    items: list[dict[str, Any]] = []
    for row_num in range(header_row + 1, end_row):
        row_values = [ws.cell(row=row_num, column=col).value for col in range(1, ws.max_column + 1)]
        if not any(value not in (None, "") for value in row_values):
            warnings.append(f"skipped blank row: {ws.title}!{row_num}")
            continue
        if is_total_or_subtotal_row(row_values):
            break
        task = value_at(ws, row_num, columns, "line_item_name") or ws.cell(row=row_num, column=section["column"]).value
        labor_days = numeric_value(value_at(ws, row_num, columns, "labor_days"))
        crew_size = numeric_value(value_at(ws, row_num, columns, "crew_size"))
        labor_hours = numeric_value(value_at(ws, row_num, columns, "labor_hours"))
        if not task or labor_days is None or labor_hours is None:
            continue
        items.append(
            line_item_row(
                record=record,
                estimate_id=estimate_id,
                estimate_file=source_path,
                section="Labor / Subcontractor",
                source_sheet=ws.title,
                source_row=row_num,
                line_item_name=str(task).strip(),
                labor_days=labor_days,
                crew_size=crew_size,
                labor_hours=labor_hours,
            )
        )
    return items, warnings


def extract_estimate_adders(
    ws: Any,
    source_path: str,
    record: JobRecord,
    estimate_id: str,
    existing_rows: set[int],
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    start_row = max(1, ws.max_row - 80)
    for row_num in range(start_row, ws.max_row + 1):
        if row_num in existing_rows:
            continue
        row_cells = list(ws[row_num])
        text_parts = [str(cell.value).strip() for cell in row_cells if text_value(cell.value)]
        row_text = " ".join(text_parts).strip()
        if not meaningful_adder_text(row_text):
            continue
        amount = rightmost_amount(row_cells)
        if amount is None:
            continue
        category = infer_line_item_category(row_text)
        items.append(
            line_item_row(
                record=record,
                estimate_id=estimate_id,
                estimate_file=source_path,
                section="Estimate Adders",
                source_sheet=ws.title,
                source_row=row_num,
                line_item_name=adder_name(row_text),
                line_item_category=category,
                description=row_text,
                extended_cost=amount,
            )
        )
    return items


def adder_rollups(items: list[dict[str, Any]]) -> dict[str, Any]:
    rollups = {
        "adders_subtotal": 0,
        "warranty_amount": 0,
        "insurance_amount": 0,
        "rental_amount": 0,
        "subcontractor_amount": 0,
        "misc_materials_amount": 0,
    }
    for item in items:
        amount = numeric_value(item.get("extended_cost")) or 0
        rollups["adders_subtotal"] += amount
        category = item.get("line_item_category")
        if category == "Warranty":
            rollups["warranty_amount"] += amount
        elif category == "Insurance":
            rollups["insurance_amount"] += amount
        elif category in {"Rental / Site Services", "Equipment Rental"}:
            rollups["rental_amount"] += amount
        elif category == "Subcontractor":
            rollups["subcontractor_amount"] += amount
        elif category == "Materials":
            rollups["misc_materials_amount"] += amount
    return {key: (int(value) if float(value).is_integer() else value) for key, value in rollups.items() if value}


def find_line_item_header(ws: Any, start_row: int, end_row: int) -> tuple[int, dict[str, int]] | None:
    for row_num in range(start_row, max(start_row, end_row) + 1):
        columns: dict[str, int] = {}
        for cell in ws[row_num]:
            key = column_key(cell.value)
            if key:
                columns[key] = cell.column
        if "labor_days" in columns and "crew_size" in columns and "labor_hours" in columns:
            columns.setdefault("line_item_name", 1)
            return row_num, columns
        if "quantity" in columns or "unit_cost" in columns or "extended_cost" in columns:
            columns.setdefault("line_item_name", first_name_column(ws, row_num))
            return row_num, columns
    return None


def line_item_row(record: JobRecord, estimate_file: str, section: str, source_sheet: str, source_row: int, estimate_id: str = "", **values: Any) -> dict[str, Any]:
    row = {
        "estimate_id": estimate_id,
        "job_id": record.job_id,
        "estimate_file": estimate_file,
        "division": record.division,
        "pipeline_status": record.pipeline_status,
        "customer": record.customer,
        "job_name": record.job_name,
        "section": section,
        "line_item_name": None,
        "line_item_category": None,
        "description": None,
        "quantity": None,
        "unit": None,
        "unit_cost": None,
        "unit_price": None,
        "extended_cost": None,
        "markup_pct": None,
        "labor_days": None,
        "crew_size": None,
        "labor_hours": None,
        "vendor": None,
        "notes": None,
        "source_sheet": source_sheet,
        "source_row": source_row,
    }
    row.update(values)
    return {field: row.get(field) for field in ESTIMATE_LINE_ITEM_FIELDS}


def write_dataset_csv(rows: list[dict[str, Any]], fields: list[str], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows([{field: row.get(field) for field in fields} for row in rows])


def write_dataset_json(rows: list[dict[str, Any]], fields: list[str], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    normalized = [{field: row.get(field) for field in fields} for row in rows]
    path.write_text(json.dumps(normalized, indent=2, ensure_ascii=False), encoding="utf-8")


def norm_label(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def section_name(value: Any) -> str | None:
    label = norm_label(value)
    return SECTION_ALIASES.get(label)


def column_key(value: Any) -> str | None:
    label = norm_label(value)
    aliases = {
        "item": "line_item_name",
        "task": "line_item_name",
        "description": "description",
        "qty": "quantity",
        "quantity": "quantity",
        "unit": "unit",
        "uom": "unit",
        "unit cost": "unit_cost",
        "cost": "unit_cost",
        "unit price": "unit_price",
        "price": "unit_price",
        "extended cost": "extended_cost",
        "total cost": "extended_cost",
        "total": "extended_cost",
        "markup": "markup_pct",
        "markup %": "markup_pct",
        "days": "labor_days",
        "no. of people": "crew_size",
        "no of people": "crew_size",
        "number of people": "crew_size",
        "total hours": "labor_hours",
        "hours": "labor_hours",
        "vendor": "vendor",
        "supplier": "vendor",
        "notes": "notes",
    }
    return aliases.get(label)


def first_name_column(ws: Any, row_num: int) -> int:
    for cell in ws[row_num]:
        if norm_label(cell.value) in {"item", "task", "description"}:
            return cell.column
    return 1


def value_at(ws: Any, row_num: int, columns: dict[str, int], key: str) -> Any:
    column = columns.get(key)
    if not column:
        return None
    return ws.cell(row=row_num, column=column).value


def numeric_value(value: Any) -> float | int | None:
    parsed = money(value)
    if parsed is None:
        return None
    return int(parsed) if float(parsed).is_integer() else parsed


def rightmost_amount(cells: list[Any]) -> float | int | None:
    for cell in reversed(cells):
        value = getattr(cell, "value", cell)
        amount = amount_value(value)
        if amount is not None:
            return amount
    return None


def amount_value(value: Any) -> float | int | None:
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)
    if value is None:
        return None
    text = str(value)
    matches = re.findall(r"\$\s*(-?\(?[0-9][0-9,]*(?:\.\d{1,2})?\)?)", text)
    if not matches:
        return None
    raw = matches[-1].replace(",", "").replace("(", "-").replace(")", "").strip()
    try:
        parsed = float(raw)
    except ValueError:
        return None
    return int(parsed) if parsed.is_integer() else parsed


def meaningful_adder_text(text: str) -> bool:
    normalized = norm_label(text)
    if not normalized or len(normalized) < 3:
        return False
    if not re.search(r"[a-zA-Z]", normalized):
        return False
    return not any(re.search(pattern, normalized) for pattern in ADDER_SKIP_PATTERNS)


def adder_name(text: str) -> str:
    cleaned = re.sub(r"\s+\$?\s*-?\(?[0-9][0-9,]*(?:\.\d{1,2})?\)?\s*$", "", text).strip(" -")
    return cleaned or text.strip()


def infer_line_item_category(text: str) -> str:
    normalized = norm_label(text)
    if "warranty" in normalized:
        return "Warranty"
    if "insurance" in normalized:
        return "Insurance"
    if "porta" in normalized or "john" in normalized:
        return "Rental / Site Services"
    if "subcontractor" in normalized or re.search(r"\bsub\b", normalized):
        return "Subcontractor"
    if "lift" in normalized or "rental" in normalized:
        return "Equipment Rental"
    if any(term in normalized for term in ["material", "caulk", "brush grade", "rustnox"]):
        return "Materials"
    if "labor" in normalized:
        return "Labor"
    return "Misc"


def text_value(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def first_numeric(values: list[Any]) -> float | int | None:
    for value in values:
        parsed = numeric_value(value)
        if parsed is not None:
            return parsed
    return None


def first_text(values: list[Any]) -> str | None:
    for value in values:
        text = text_value(value)
        if text and not section_name(text):
            return text
    return None


def is_total_or_subtotal_row(values: list[Any]) -> bool:
    text = " ".join(str(value).lower() for value in values if value is not None)
    return "subtotal" in text or text.startswith("total ") or "total hours" in text or "total days" in text
