from __future__ import annotations

import hashlib
import json
import re
from datetime import date, datetime, timedelta
import os
from pathlib import Path
from typing import Any

from .estimate_selection import select_primary_estimate
from .job_tracking_extractor import apply_job_tracking_to_record, is_job_tracking_file
from .models import JobRecord, money, rel
from .schedule_extractor import apply_schedule_extraction

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".heic", ".webp", ".tif", ".tiff"}
SPREADSHEET_EXTS = {".xlsx", ".xlsm", ".xls"}
DOC_EXTS = {".doc", ".docx", ".pdf"}
IMAGE_MANIFEST_NAME = ".image_manifest.json"
NON_JOB_FOLDER_NAMES = {
    "pics",
    "photos",
    "images",
    "drone",
    "drones",
    "aerial",
    "aerials",
    "warranties",
    "warranty",
    "invoices",
    "invoice",
    "contracts",
    "contract",
    "old",
    "archive",
    "archives",
    "template",
    "templates",
}

AERIAL_KEYWORDS = [
    "aerial",
    "drone",
    "drone photos",
    "uav",
    "overhead",
    "satellite",
    "eagleview",
    "eagle view",
    "hover",
    "roof report",
]

def path_contains_any(path, keywords):
    text = str(path).lower()
    return any(keyword.lower() in text for keyword in keywords)

def path_has_keyword(path, keywords):
    text = str(path).lower()
    return any(keyword.lower() in text for keyword in keywords)

def slugify(value: str) -> str:
    value = value.upper()
    value = re.sub(r"[^A-Z0-9]+", "-", value).strip("-")
    return re.sub(r"-+", "-", value)[:80] or "UNKNOWN-JOB"


def split_city_state_zip(value: str | None) -> tuple[str | None, str | None, str | None]:
    if not value:
        return None, None, None
    # Handles "Jeffersonville, IN 47130".
    m = re.match(r"\s*(?P<city>.*?),\s*(?P<state>[A-Z]{2})\s*(?P<zip>\d{5}(?:-\d{4})?)?\s*$", value)
    if not m:
        return value.strip(), None, None
    return m.group("city").strip(), m.group("state"), m.group("zip")


def _folder_key(path: Path) -> str:
    return re.sub(r"\s+", " ", path.name.strip().lower())


def is_non_job_folder(path: Path) -> bool:
    key = _folder_key(path)
    return key in NON_JOB_FOLDER_NAMES or key.startswith(".")


def immediate_child_job_folders(root: Path) -> tuple[list[Path], int]:
    children = [p for p in sorted(root.iterdir()) if p.is_dir()]
    skipped_count = sum(1 for child in children if is_non_job_folder(child))
    return [child for child in children if not is_non_job_folder(child)], skipped_count


def find_job_folders(root: Path) -> list[Path]:
    """Return likely job folders from an exported SharePoint/OneDrive directory."""
    root = root.resolve()
    immediate_children, skipped_admin_count = immediate_child_job_folders(root)
    find_job_folders.last_immediate_child_count = len(immediate_children)  # type: ignore[attr-defined]
    find_job_folders.last_skipped_admin_count = skipped_admin_count  # type: ignore[attr-defined]
    if immediate_children:
        return immediate_children

    candidates: list[Path] = []

    for directory in [root, *[p for p in root.rglob("*") if p.is_dir()]]:
        files = [p for p in directory.iterdir() if p.is_file()]
        if not files:
            continue
        has_job_artifact = any(
            p.suffix.lower() in SPREADSHEET_EXTS | DOC_EXTS
            or p.suffix.lower() in IMAGE_EXTS
            or p.name == IMAGE_MANIFEST_NAME
            for p in files
        )
        # Avoid returning Pics as a separate job folder.
        if has_job_artifact and not is_non_job_folder(directory):
            candidates.append(directory)

    # Prefer folders that contain an estimate, invoice, contract, proposal, or subfolder pics.
    scored: list[tuple[int, Path]] = []
    for c in candidates:
        names = " ".join(p.name.lower() for p in c.rglob("*") if p.is_file())
        score = 0
        for term in ["estimate", "invoice", "contract", "proposal", "warranty", "job spec"]:
            if term in names:
                score += 1
        if any(p.name == IMAGE_MANIFEST_NAME for p in c.rglob("*") if p.is_file()):
            score += 1
        if any(p.is_dir() and p.name.lower() in {"pics", "photos", "images"} for p in c.iterdir()):
            score += 1
        scored.append((score, c))

    # Keep only top-level-ish winners: if a parent and child both match, keep the parent when it scores higher.
    winners = [p for score, p in scored if score > 0]
    final: list[Path] = []
    for p in sorted(winners, key=lambda x: len(x.parts)):
        if not any(str(p).startswith(str(existing) + os.sep) for existing in final):
            final.append(p)
    return final or [root]


def _load_image_manifest_entries(folder: Path) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for manifest in folder.rglob(IMAGE_MANIFEST_NAME):
        try:
            loaded = json.loads(manifest.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if isinstance(loaded, list):
            entries.extend(entry for entry in loaded if isinstance(entry, dict))
    return entries


def classify_files(folder: Path) -> dict[str, Any]:
    files = [p for p in folder.rglob("*") if p.is_file()]
    lower = {p: p.name.lower() for p in files}
    image_manifest_entries = _load_image_manifest_entries(folder)

    job_tracking_files = [p for p in files if p.suffix.lower() in SPREADSHEET_EXTS and is_job_tracking_file(p)]
    estimate_files = [
        p for p in files
        if p.suffix.lower() in SPREADSHEET_EXTS
        and "estimate" in lower[p]
    ]
    if not estimate_files:
        estimate_files = [p for p in files if p.suffix.lower() in SPREADSHEET_EXTS and p not in job_tracking_files]

    invoice_files = [p for p in files if "invoice" in lower[p] and p.suffix.lower() == ".pdf"]
    signed_contracts = [p for p in files if "signed" in lower[p] and "contract" in lower[p]]
    warranties = [p for p in files if "warranty" in lower[p]]
    proposals = [p for p in files if "proposal" in lower[p]]
    job_specs = [p for p in files if "job spec" in lower[p] or "scope" in lower[p]]
    # Check full paths, not just filenames, so folders like "Drone/" count as aerial evidence.
    all_paths = list(folder.rglob("*"))
    aerials = [p for p in all_paths if path_contains_any(p, AERIAL_KEYWORDS)]
    notes = [p for p in files if "note" in lower[p] or "handwritten" in lower[p]]
    photos = [p for p in files if p.suffix.lower() in IMAGE_EXTS]
    skipped_image_count = len(image_manifest_entries)

    duplicate_count: int | None
    if skipped_image_count:
        duplicate_count = None
    else:
        hashes: dict[str, int] = {}
        for photo in photos:
            try:
                digest = hashlib.sha1(photo.read_bytes()).hexdigest()
                hashes[digest] = hashes.get(digest, 0) + 1
            except OSError:
                pass
        duplicate_count = sum(count - 1 for count in hashes.values() if count > 1)

    return {
        "files": files,
        "estimate_files": estimate_files,
        "job_tracking_files": job_tracking_files,
        "invoice_files": invoice_files,
        "signed_contracts": signed_contracts,
        "warranties": warranties,
        "proposals": proposals,
        "job_specs": job_specs,
        "aerials": aerials,
        "notes": notes,
        "photos": photos,
        "photo_count": len(photos) + skipped_image_count,
        "duplicate_photo_count": duplicate_count,
        "image_files_cached": skipped_image_count == 0,
        "skipped_image_count": skipped_image_count,
        "warnings": [],
    }


def _coerce_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    # Excel serial date fallback. 45684 = 2025-02-26 in Excel's 1900 date system.
    if isinstance(value, (int, float)) and 30000 < value < 60000:
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=float(value))).date().isoformat()  # type: ignore[name-defined]
    return str(value)


def _norm_label(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _clean_number(value: float | None) -> int | float | None:
    if value is None:
        return None
    return int(value) if float(value).is_integer() else value


def _find_cell_containing(ws: Any, text: str) -> Any | None:
    needle = text.strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            if needle in _norm_label(cell.value):
                return cell
    return None


def _find_header_columns(ws: Any, start_row: int, max_row: int) -> tuple[int, dict[str, int]] | None:
    for row_num in range(start_row, max_row + 1):
        headers: dict[str, int] = {}
        for cell in ws[row_num]:
            label = _norm_label(cell.value)
            if label == "days":
                headers["days"] = cell.column
            elif label in {"no. of people", "no of people", "number of people"}:
                headers["crew_size"] = cell.column
            elif label == "total hours":
                headers["total_hours"] = cell.column
        if {"days", "crew_size", "total_hours"}.issubset(headers):
            return row_num, headers
    return None


def _find_total_row(ws: Any, start_row: int, max_row: int) -> int | None:
    for row_num in range(start_row, max_row + 1):
        labels = [_norm_label(cell.value) for cell in ws[row_num]]
        if "total hours" in labels and "total days" in labels:
            return row_num
    return None


def _most_common_crew_size(values: list[int]) -> int | None:
    if not values:
        return None
    counts: dict[int, int] = {}
    for value in values:
        counts[value] = counts.get(value, 0) + 1
    return sorted(counts, key=lambda value: (-counts[value], -value))[0]


def _value_right_of_label_in_row(ws: Any, row_num: int, label: str) -> Any:
    target = label.strip().lower()
    row = list(ws[row_num])
    for index, cell in enumerate(row):
        if _norm_label(cell.value) == target and index + 1 < len(row):
            return row[index + 1].value
    return None


def extract_labor_schedule(ws: Any) -> dict[str, Any]:
    section = _find_cell_containing(ws, "Labor / Subcontractor")
    if not section:
        return {}

    header = _find_header_columns(ws, section.row + 1, min(ws.max_row, section.row + 20))
    if not header:
        return {}
    header_row, columns = header

    total_row = _find_total_row(ws, header_row + 1, ws.max_row)
    if not total_row:
        return {}

    breakdown: list[dict[str, Any]] = []
    crew_sizes: list[int] = []
    task_col = section.column
    for row_num in range(header_row + 1, total_row):
        task = ws.cell(row=row_num, column=task_col).value
        days = _number(ws.cell(row=row_num, column=columns["days"]).value)
        crew_size = _number(ws.cell(row=row_num, column=columns["crew_size"]).value)
        total_hours = _number(ws.cell(row=row_num, column=columns["total_hours"]).value)
        if not task or days is None or total_hours is None:
            continue
        task_text = str(task).strip()
        if not task_text or "subtotal" in task_text.lower() or task_text.lower().startswith("total"):
            continue
        item = {
            "task": task_text,
            "days": _clean_number(days),
            "crew_size": _clean_number(crew_size),
            "total_hours": _clean_number(total_hours),
        }
        breakdown.append(item)
        if crew_size is not None:
            crew_sizes.append(int(crew_size))

    total_hours = _number(_value_right_of_label_in_row(ws, total_row, "Total Hours"))
    total_days = _number(_value_right_of_label_in_row(ws, total_row, "Total Days"))
    out: dict[str, Any] = {
        "estimated_labor_hours": _clean_number(total_hours),
        "estimated_duration_days": _clean_number(total_days),
        "estimated_crew_size": _most_common_crew_size(crew_sizes),
        "labor_duration_source": "Estimate sheet Labor / Subcontractor section",
        "labor_schedule_breakdown": breakdown,
    }
    return {key: value for key, value in out.items() if value not in (None, [])}


def extract_hours_per_day(wb: Any) -> float | int | None:
    if "People" not in wb.sheetnames:
        return None
    ws = wb["People"]
    cell = _find_cell_containing(ws, "Hours /Day")
    if not cell:
        return None
    return _clean_number(_number(ws.cell(row=cell.row, column=cell.column + 1).value))


def extract_estimate_xlsx(path: Path) -> dict[str, Any]:
    """Extract key fields from Spray-Tec-style estimate workbooks.

    This deliberately starts with label-based extraction instead of absolute cells so it can survive small template changes.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("Install openpyxl to read Excel estimate files: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "Estimate" not in wb.sheetnames:
        return {"warnings": [f"No 'Estimate' sheet found in {path.name}"]}
    ws = wb["Estimate"]

    def row_values(row: int) -> list[Any]:
        return [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]

    def value_right_of(label: str, occurrence: int = 1, preferred_col: int | None = None) -> Any:
        label_norm = label.strip().lower()
        matches = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().lower() == label_norm:
                    matches.append(cell)
        if preferred_col is not None:
            preferred = [c for c in matches if c.column == preferred_col]
            if len(preferred) >= occurrence:
                c = preferred[occurrence - 1]
                return ws.cell(row=c.row, column=c.column + 1).value
        if len(matches) >= occurrence:
            c = matches[occurrence - 1]
            return ws.cell(row=c.row, column=c.column + 1).value
        return None

    def numeric_to_right(label_contains: str, pick: str = "last") -> float | None:
        needle = label_contains.strip().lower()
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and needle in cell.value.strip().lower():
                    values = row_values(cell.row)[cell.column:]
                    nums = [v for v in values if isinstance(v, (int, float))]
                    if not nums:
                        return None
                    return money(nums[0] if pick == "first" else nums[-1])
        return None

    def estimated_sqft_value() -> float | None:
        # In this template, "Est. Square Feet:" is a label row and the actual base sqft is in the next row.
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "est. square feet" in cell.value.strip().lower():
                    next_values = row_values(cell.row + 1)
                    nums = [v for v in next_values if isinstance(v, (int, float)) and v > 10]
                    return money(nums[0]) if nums else None
        return None

    def row_number(label_contains: str) -> int | None:
        needle = label_contains.strip().lower()
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and needle in cell.value.strip().lower():
                    return cell.row
        return None

    city, state, zip_code = split_city_state_zip(value_right_of("City, State, Zip:", preferred_col=2))

    out: dict[str, Any] = {
        "estimate_file": str(path),
        "estimate_date": _coerce_date(value_right_of("Today's Date:", preferred_col=2)),
        "job_name": value_right_of("Job Name:", preferred_col=2),
        "job_type": str(value_right_of("Job Type:", preferred_col=2)).strip() if value_right_of("Job Type:", preferred_col=2) else None,
        "site_address": value_right_of("Site Address:", preferred_col=2),
        "city": city,
        "state": state,
        "zip_code": zip_code,
        "contact_name": value_right_of("Contact:", preferred_col=2),
        "contact_title": value_right_of("Title:", preferred_col=2),
        "contact_email": value_right_of("Email Address:", preferred_col=2),
        "contact_phone": value_right_of("Phone:", preferred_col=2),
        "estimated_sqft": estimated_sqft_value(),
        "material_subtotal": numeric_to_right("Subtotal Material"),
        "labor_subtotal": numeric_to_right("Subtotal Labor"),
        "warranty_bonding_insurance_subtotal": numeric_to_right("Subtotal Warranty"),
        "total_job_cost": numeric_to_right("Total Job Cost"),
        "overhead_amount": numeric_to_right("Estimated O/H"),
        "profit_amount": numeric_to_right("Profit"),
        "worksheet_price": numeric_to_right("Work Sheet Price"),
        "final_price": numeric_to_right("Work Sheet Price + Additional"),
        "price_per_sqft": numeric_to_right("Price / Sq. Ft", pick="first"),
        "warnings": [],
    }
    out.update(extract_labor_schedule(ws))
    hours_per_day = extract_hours_per_day(wb)
    if hours_per_day is not None:
        out["estimated_hours_per_day"] = hours_per_day

    # Pull percentage values from the percentage rows.
    for pct_label, pct_key in [("Estimated O/H", "overhead_pct"), ("Profit", "profit_pct")]:
        rn = row_number(pct_label)
        pct = None
        if rn:
            values = row_values(rn)
            for i, val in enumerate(values):
                if isinstance(val, str) and "percentage" in val.lower():
                    right = values[i + 1] if i + 1 < len(values) else None
                    pct = money(right)
                    break
        out[pct_key] = pct

    return out


def parse_invoice_filename(path: Path) -> dict[str, Any]:
    text = path.name
    out: dict[str, Any] = {"invoice_file": str(path)}

    m = re.search(r"invoice\s*no\.?\s*([A-Za-z0-9\-]+)", text, flags=re.I)
    if m:
        out["invoice_number"] = m.group(1)

    m = re.search(r"\$\s*([0-9,]+(?:\.\d{2})?)", text)
    if m:
        out["invoice_amount"] = money(m.group(1))

    m = re.search(r"\((\d{1,2})\.(\d{1,2})\.(\d{2,4})\)", text)
    if m:
        month, day, year = m.groups()
        year = "20" + year if len(year) == 2 else year
        out["invoice_date"] = f"{int(year):04d}-{int(month):02d}-{int(day):02d}"

    return out


def infer_customer_from_folder(folder_name: str, job_name: str | None) -> str | None:
    """Infer customer from noisy SharePoint folder names.

    Handles examples like:
    - "2026 ROOFING_COMPLETED_Andriot's 711 Main Street..."
    - "ACRE Derb E Cigs - 1710 E 10th Street"
    """
    raw = folder_name.replace("_", " ")
    raw = re.sub(r"\s+", " ", raw).strip()

    # Remove common parent-folder/status noise that can appear when Graph cache flattens paths.
    cleaned = re.sub(r"\b20\d{2}\b", " ", raw, flags=re.I)
    cleaned = re.sub(r"\broofing\b", " ", cleaned, flags=re.I)
    cleaned = re.sub(r"\b(completed|complete|active|open|closed|estimates?|jobs?)\b", " ", cleaned, flags=re.I)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -_")

    # If the job name appears in the folder, use the part before it.
    if job_name:
        idx = cleaned.lower().find(job_name.lower())
        if idx > 0:
            before = cleaned[:idx].strip(" -_")
            if before:
                return before.upper() if len(before) <= 6 else before.title()

    # Fallback: take everything before the first street-number-looking token.
    m = re.match(r"(.+?)\s+\d{2,6}\b", cleaned)
    if m:
        candidate = m.group(1).strip(" -_")
        if candidate:
            return candidate.upper() if len(candidate) <= 6 else candidate.title()

    # Last fallback: first few words, but avoid returning empty/noisy values.
    words = cleaned.split()
    if words:
        candidate = " ".join(words[:3]).strip(" -_")
        return candidate.upper() if len(candidate) <= 6 else candidate.title()

    return None


def infer_status(record: JobRecord, folder_context: str = "") -> str:
    """Infer job status from artifacts and SharePoint folder context."""
    context = folder_context.lower()

    if record.has_invoice:
        return "Invoiced"
    if record.has_signed_contract:
        return "Contracted"
    if "contracted repairs" in context or "contracted repair" in context:
        return "Contracted Repairs"
    if "contracted" in context:
        return "Contracted"
    if "proposed" in context:
        return "Proposed"
    if any(term in context for term in ["completed", "complete", "closed"]):
        return "Completed"
    if any(term in context for term in ["active", "in progress", "open"]):
        return "Active"
    if record.estimate_file:
        return "Estimated"
    return "Folder Created"


def scan_job_folder(folder: Path, root: Path | None = None, scan_context: str = "") -> JobRecord:
    root = root or folder.parent
    info = classify_files(folder)
    record = JobRecord(
        job_id=slugify(folder.name),
        folder_name=folder.name,
        folder_path=rel(folder, root),
        photo_count=info["photo_count"],
        duplicate_photo_count=info["duplicate_photo_count"],
        image_files_cached=info["image_files_cached"],
        skipped_image_count=info["skipped_image_count"],
        has_signed_contract=bool(info["signed_contracts"]),
        has_invoice=bool(info["invoice_files"]),
        has_warranty=bool(info["warranties"]),
        has_proposal=bool(info["proposals"]),
        has_job_spec=bool(info["job_specs"]),
        has_job_tracking_form=bool(info["job_tracking_files"]),
        has_aerial=bool(info["aerials"]),
        has_notes=bool(info["notes"]),
    )
    record.warnings.extend(info["warnings"])

    estimate_files = sorted(info["estimate_files"], key=lambda path: path.name.lower())
    record.estimate_file_count = len(estimate_files)
    record.estimate_files = [rel(path, root) for path in estimate_files]
    record.multiple_estimates_found = len(estimate_files) > 1

    if info["estimate_files"]:
        parsed_estimates: list[dict[str, Any]] = []
        for estimate_file in estimate_files:
            parsed: dict[str, Any] = {"path": estimate_file, "estimate_file": rel(estimate_file, root), "warnings": []}
            try:
                extracted = extract_estimate_xlsx(estimate_file)
                parsed.update(extracted)
                parsed["path"] = estimate_file
                parsed["estimate_file"] = rel(estimate_file, root)
            except Exception as exc:  # Keep the scanner running when one workbook is bad.
                parsed["warnings"] = [f"Estimate parse failed: {exc}"]
                record.warnings.append(f"Estimate parse failed for {rel(estimate_file, root)}: {exc}")
            parsed_estimates.append(parsed)

        primary, reason = select_primary_estimate(parsed_estimates)
        record.estimate_selection_reason = (
            f"Multiple estimate workbooks found; {reason}" if record.multiple_estimates_found else reason
        )
        if primary:
            primary_path = primary["path"]
            record.primary_estimate_file = rel(primary_path, root)
            record.estimate_file = record.primary_estimate_file
            record.supporting_estimate_files = [
                rel(path, root) for path in estimate_files if path != primary_path
            ]
            for key, value in primary.items():
                if key in {"path", "estimate_file"}:
                    continue
                if key == "warnings":
                    record.warnings.extend(value or [])
                elif hasattr(record, key):
                    setattr(record, key, value)

    if info["invoice_files"]:
        invoice = info["invoice_files"][0]
        parsed = parse_invoice_filename(invoice)
        for key, value in parsed.items():
            if hasattr(record, key):
                setattr(record, key, value)
        record.invoice_file = rel(invoice, root)

    record.customer = infer_customer_from_folder(record.folder_name, record.job_name)
    folder_context = f"{scan_context} {record.folder_path} {record.folder_name}"
    record.status = infer_status(record, folder_context)
    apply_job_tracking_to_record(record, folder, root, info["job_tracking_files"])
    apply_schedule_extraction(record, folder, root, info)

    completed_context = any(term in folder_context.lower() for term in ["completed", "complete", "closed"])
    if completed_context and not record.has_invoice:
        record.warnings.append("Completed job has no invoice")

    if record.final_price and record.invoice_amount and abs(record.final_price - record.invoice_amount) > 1:
        record.warnings.append(
            f"Final price {record.final_price} does not match invoice amount {record.invoice_amount}"
        )

    if (
        record.estimate_file
        and record.job_type
        and "roof" in record.job_type.lower()
        and record.labor_subtotal == 0
    ):
        record.warnings.append("Labor subtotal is zero for a roof job; verify estimate extraction or workbook values")

    if not record.estimate_file:
        record.warnings.append("No estimate workbook found")

    return record
