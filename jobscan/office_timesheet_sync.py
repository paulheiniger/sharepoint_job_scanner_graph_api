from __future__ import annotations

import argparse
import csv
import json
import re
import tempfile
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterable

from .graph_client import GraphClient, SharePointTarget
from .sharepoint_sync import _safe_name

WEEKDAYS = {"monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"}
HEADER_ALIASES = {
    "PROJECT": "project",
    "CODE": "code",
    "APPROX. TIME SPENT": "approx_time_spent",
    "APPROX TIME SPENT": "approx_time_spent",
    "START": "start_time",
    "END": "end_time",
    "HUBSPOT NOTES": "hubspot_notes",
    "ADDITIONAL NOTES": "additional_notes",
}
DETAIL_FIELDS = [
    "employee_folder",
    "employee_name",
    "work_date",
    "sheet_day",
    "year",
    "month_folder",
    "project",
    "code",
    "code_missing",
    "row_type",
    "duration_hours",
    "approx_time_spent",
    "start_time",
    "end_time",
    "day_start",
    "lunch",
    "day_end",
    "hubspot_notes",
    "additional_notes",
    "source_file",
    "source_path",
    "warnings",
]
EMPLOYEE_DAILY_SUMMARY_FIELDS = [
    "work_date",
    "employee_name",
    "total_hours",
    "line_count",
    "timed_entry_count",
    "activity_only_count",
    "missing_code_count",
    "project_count",
    "warning_count",
]
CODE_SUMMARY_FIELDS = [
    "code",
    "total_hours",
    "employee_count",
    "project_count",
    "line_count",
    "timed_entry_count",
    "activity_only_count",
    "missing_code_count",
    "date_min",
    "date_max",
]
PROJECT_TOUCH_SUMMARY_FIELDS = [
    "project",
    "code",
    "total_hours",
    "employee_count",
    "date_min",
    "date_max",
    "line_count",
    "timed_entry_count",
    "activity_only_count",
    "missing_code_count",
    "latest_notes",
]
ACTIONABLE_WARNING_PATTERNS = {
    "invalid duration",
    "invalid start/end time",
    "unusually large duration",
    "long office timesheet entry",
    "possible duplicate sheet",
    "missing code",
}
MONTH_RE = re.compile(r"\b(?:\d{1,2}\s+)?(january|february|march|april|may|june|july|august|september|october|november|december)\b", re.I)
YEAR_RE = re.compile(r"\b(20\d{2})\b")
DURATION_UNIT_RE = re.compile(
    r"(?P<amount>(?:\d+\s+)?\d+\s*/\s*\d+|\d+(?:\.\d+)?)\s*(?P<unit>hours?|hrs?|h|minutes?|mins?|m)\b",
    re.I,
)


@dataclass
class ScanStats:
    files_scanned: int = 0
    templates_skipped: int = 0


@dataclass
class ScanResult:
    records: list[dict[str, Any]]
    stats: ScanStats


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="minutes")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value).strip()


def _normal_header(value: Any) -> str:
    return re.sub(r"\s+", " ", _stringify(value)).upper()


def _parse_date(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            return from_excel(value).date().isoformat()
        except Exception:
            return ""
    text = _stringify(value)
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%B %d, %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text).date().isoformat()
    except ValueError:
        return text


def _parse_time(value: Any) -> time | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, (int, float)):
        if 0 <= float(value) < 1:
            total_minutes = round(float(value) * 24 * 60)
            return (datetime.min + timedelta(minutes=total_minutes)).time().replace(second=0, microsecond=0)
        return None
    text = _stringify(value).lower().replace(".", "")
    text = re.sub(r"\s+", " ", text)
    for fmt in ("%I:%M %p", "%I %p", "%H:%M", "%H"):
        try:
            return datetime.strptime(text.upper(), fmt).time().replace(second=0, microsecond=0)
        except ValueError:
            pass
    return None


def _format_time_value(value: Any) -> str:
    parsed = _parse_time(value)
    return parsed.strftime("%H:%M") if parsed else _stringify(value)


def _clean_duration_text(value: Any) -> str:
    return (
        _stringify(value)
        .lower()
        .replace("\u00a0", " ")
        .replace(",", "")
        .strip()
    )


def _round_duration(hours: float) -> float:
    return round(hours, 4)


def _parse_duration_amount(text: str) -> float | None:
    cleaned = re.sub(r"\s+", " ", text.strip())
    fraction_match = re.fullmatch(r"(?:(?P<whole>\d+)\s+)?(?P<numerator>\d+)\s*/\s*(?P<denominator>\d+)", cleaned)
    if fraction_match:
        denominator = int(fraction_match.group("denominator"))
        if denominator == 0:
            return None
        whole = int(fraction_match.group("whole") or 0)
        numerator = int(fraction_match.group("numerator"))
        return whole + numerator / denominator
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_duration_hours(value: Any, *, from_approx_time: bool = False) -> tuple[float | None, str | None]:
    """Parse an approximate duration value into hours.

    Returns (hours, warning). Numeric Excel time fractions are treated as fractions
    of a day only outside the approximate-duration column.
    Bare integer values are treated as minutes because office timesheets commonly use minute counts.
    """
    if value in (None, ""):
        return None, None
    if isinstance(value, timedelta):
        hours = value.total_seconds() / 3600
        return (_round_duration(hours), None) if hours > 0 else (None, None)
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        hours = value.hour + value.minute / 60 + value.second / 3600
        return (_round_duration(hours), None) if hours > 0 else (None, None)
    if isinstance(value, (int, float)):
        numeric = float(value)
        if numeric <= 0:
            return None, None
        if 0 < numeric < 1 and not from_approx_time:
            return _round_duration(numeric * 24), None
        if from_approx_time and 0 < numeric <= 12 and not numeric.is_integer():
            return _round_duration(numeric), None
        if numeric > 180:
            return None, "invalid duration"
        if isinstance(value, float) and not numeric.is_integer() and numeric <= 12:
            return _round_duration(numeric), None
        if 1 <= numeric <= 180:
            return _round_duration(numeric / 60), None
        return None, "invalid duration"

    text = _clean_duration_text(value)
    if not text:
        return None, None

    colon_match = re.fullmatch(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", text)
    if colon_match:
        hours = int(colon_match.group(1)) + int(colon_match.group(2)) / 60 + int(colon_match.group(3) or 0) / 3600
        return (_round_duration(hours), None) if 0 < hours <= 24 else (None, "invalid duration")

    unit_matches = list(DURATION_UNIT_RE.finditer(text))
    if unit_matches:
        remainder = DURATION_UNIT_RE.sub("", text)
        if re.sub(r"[\s:+-]+", "", remainder):
            return None, "invalid duration"
        hours = 0.0
        minutes = 0.0
        for match in unit_matches:
            amount = _parse_duration_amount(match.group("amount"))
            if amount is None:
                return None, "invalid duration"
            unit = match.group("unit").lower()
            if unit.startswith("h"):
                hours += amount
            else:
                minutes += amount
        duration = hours + minutes / 60
        return (_round_duration(duration), None) if 0 < duration <= 24 else (None, "invalid duration")

    decimal_match = re.fullmatch(r"\d{1,2}\.\d+", text)
    if decimal_match:
        numeric = float(text)
        return (_round_duration(numeric), None) if 0 < numeric <= 12 else (None, "invalid duration")

    numeric_match = re.fullmatch(r"\d+", text)
    if numeric_match:
        numeric = int(text)
        if numeric <= 0:
            return None, None
        if numeric <= 180:
            return _round_duration(numeric / 60), None
        return None, "invalid duration"
    return None, "invalid duration"


def _time_text_has_ampm(value: Any) -> bool:
    return bool(re.search(r"(^|[^a-z])[ap]\s*\.?\s*m\.?($|[^a-z])", _stringify(value).lower()))


def _duration_from_start_end(start: Any, end: Any) -> tuple[float | None, str | None]:
    start_time = _parse_time(start)
    end_time = _parse_time(end)
    if start_time and end_time:
        start_dt = datetime.combine(date.today(), start_time)
        end_dt = datetime.combine(date.today(), end_time)
        hours = (end_dt - start_dt).total_seconds() / 3600
        if hours <= 0 and not _time_text_has_ampm(start) and not _time_text_has_ampm(end):
            afternoon_end = end_dt + timedelta(hours=12)
            afternoon_hours = (afternoon_end - start_dt).total_seconds() / 3600
            if 0 < afternoon_hours <= 12:
                return _round_duration(afternoon_hours), None
        if hours <= 0:
            overnight_hours = (end_dt + timedelta(days=1) - start_dt).total_seconds() / 3600
            if 0 < overnight_hours <= 12:
                return _round_duration(overnight_hours), None
            return None, "invalid start/end time"
        if hours <= 12:
            return _round_duration(hours), None
        return None, "invalid start/end time"
    if start or end:
        return None, "invalid start/end time"
    return None, "missing duration"


def _has_time_intent(approx: Any, start: Any, end: Any) -> bool:
    return any(_stringify(value) for value in (approx, start, end))


def _duration_from_values(approx: Any, start: Any, end: Any) -> tuple[float | None, list[str]]:
    warnings: list[str] = []
    if not _has_time_intent(approx, start, end):
        return None, warnings
    approx_duration, approx_warning = parse_duration_hours(approx, from_approx_time=True)
    if approx_duration is not None:
        if approx_duration > 12:
            warnings.append("unusually large duration")
        return approx_duration, warnings
    if approx_warning:
        warnings.append(approx_warning)

    start_end_duration, start_end_warning = _duration_from_start_end(start, end)
    if start_end_duration is not None:
        if start_end_duration > 12:
            warnings.append("unusually large duration")
        return start_end_duration, warnings
    if start_end_warning == "invalid start/end time":
        warnings.append(start_end_warning)
    warnings.append("missing duration")
    return None, warnings


def _append_text(existing: str, extra: str) -> str:
    extra = extra.strip()
    if not extra:
        return existing
    return f"{existing}\n{extra}" if existing else extra


def _is_blank_row(values: Iterable[Any]) -> bool:
    return not any(_stringify(value) for value in values)


def _weekday_base(sheet_name: str) -> str:
    cleaned_name = re.sub(r"\s*\(\d+\)\s*$", "", sheet_name.strip())
    return re.sub(r"[^a-z]", "", cleaned_name.lower())


def _looks_like_weekday(sheet_name: str) -> bool:
    cleaned = _weekday_base(sheet_name)
    return cleaned in WEEKDAYS


def _is_possible_duplicate_sheet(sheet_name: str) -> bool:
    return bool(re.search(r"\(\d+\)\s*$", sheet_name.strip())) and _looks_like_weekday(sheet_name)


def _metadata_from_path(path: Path, root: Path) -> dict[str, str]:
    try:
        rel = path.relative_to(root)
    except ValueError:
        rel = path
    parts = rel.parts
    employee_index = 1 if parts and parts[0].lower() == "timesheets" else 0
    employee_folder = parts[employee_index] if len(parts) > employee_index + 1 else ""
    year = ""
    month_folder = ""
    for part in parts[:-1]:
        if not year:
            match = YEAR_RE.search(part)
            if match:
                year = match.group(1)
        if not month_folder and MONTH_RE.search(part):
            month_folder = part
    return {
        "employee_folder": employee_folder,
        "year": year,
        "month_folder": month_folder,
        "source_file": path.name,
        "source_path": str(path),
    }


def _find_header_map(sheet: Any) -> dict[str, int]:
    header_map: dict[str, int] = {}
    header_row = 5
    for col_idx, cell in enumerate(sheet[header_row], start=1):
        mapped = HEADER_ALIASES.get(_normal_header(cell.value))
        if mapped:
            header_map[mapped] = col_idx
    if header_map:
        return header_map
    return {
        "project": 1,
        "code": 2,
        "approx_time_spent": 3,
        "start_time": 4,
        "end_time": 5,
        "hubspot_notes": 6,
        "additional_notes": 7,
    }


def _row_value(row: tuple[Any, ...], col_idx: int | None) -> Any:
    if not col_idx or col_idx < 1 or col_idx > len(row):
        return None
    return row[col_idx - 1]


def _is_continuation(values: dict[str, Any]) -> bool:
    has_text = any(_stringify(values.get(field)) for field in ("project", "hubspot_notes", "additional_notes"))
    has_structured = any(_stringify(values.get(field)) for field in ("code", "approx_time_spent", "start_time", "end_time"))
    has_project = bool(_stringify(values.get("project")))
    return has_text and not has_project and not has_structured


def _has_activity_content(record: dict[str, Any]) -> bool:
    return any(_stringify(record.get(field)) for field in ("project", "code", "hubspot_notes", "additional_notes"))


def _row_type_for_record(record: dict[str, Any]) -> str:
    if record.get("duration_hours") is not None:
        return "timed_entry"
    if _has_activity_content(record):
        return "activity_only"
    return ""


def _finalize_record(record: dict[str, Any]) -> dict[str, Any]:
    warnings = list(record.get("_warnings", []))
    if not record.get("work_date"):
        warnings.append("missing work_date")
    if not record.get("project"):
        warnings.append("missing project")
    code_missing = not bool(_stringify(record.get("code")))
    duration, duration_warnings = _duration_from_values(record.get("_approx_raw"), record.get("_start_raw"), record.get("_end_raw"))
    record["duration_hours"] = duration
    record["row_type"] = _row_type_for_record(record)
    record["code_missing"] = code_missing
    warnings.extend(duration_warnings)
    if code_missing and (_is_timed_entry(record) or _has_time_intent(record.get("_approx_raw"), record.get("_start_raw"), record.get("_end_raw"))):
        warnings.append("missing code")
    record["warnings"] = "; ".join(dict.fromkeys(warnings))
    for hidden in ("_warnings", "_approx_raw", "_start_raw", "_end_raw"):
        record.pop(hidden, None)
    return record


def _scan_workbook(path: Path, root: Path) -> list[dict[str, Any]]:
    metadata = _metadata_from_path(path, root)
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return [{
            **{field: "" for field in DETAIL_FIELDS},
            **metadata,
            "warnings": f"unreadable workbook: {exc}",
        }]

    records: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            if not _looks_like_weekday(sheet.title):
                continue
            if sheet.max_row < 6:
                continue

            employee_name = _stringify(sheet["A2"].value) or metadata["employee_folder"]
            work_date = _parse_date(sheet["C2"].value)
            day_start = _format_time_value(sheet["A4"].value)
            lunch = _format_time_value(sheet["B4"].value)
            day_end = _format_time_value(sheet["C4"].value)
            header_map = _find_header_map(sheet)
            sheet_warnings = ["possible duplicate sheet"] if _is_possible_duplicate_sheet(sheet.title) else []
            current: dict[str, Any] | None = None

            for row in sheet.iter_rows(min_row=6, values_only=True):
                if _is_blank_row(row):
                    continue
                values = {field: _row_value(row, col_idx) for field, col_idx in header_map.items()}
                if _is_continuation(values) and current:
                    current["project"] = _append_text(current.get("project", ""), _stringify(values.get("project"))) if not current.get("project") else current["project"]
                    current["hubspot_notes"] = _append_text(current.get("hubspot_notes", ""), _stringify(values.get("hubspot_notes")))
                    for extra_note in (_stringify(values.get("project")), _stringify(values.get("additional_notes"))):
                        current["additional_notes"] = _append_text(current.get("additional_notes", ""), extra_note)
                    continue

                if current:
                    records.append(_finalize_record(current))

                current = {
                    **metadata,
                    "employee_name": employee_name,
                    "work_date": work_date,
                    "sheet_day": sheet.title,
                    "project": _stringify(values.get("project")),
                    "code": _stringify(values.get("code")),
                    "duration_hours": None,
                    "approx_time_spent": _stringify(values.get("approx_time_spent")),
                    "start_time": _format_time_value(values.get("start_time")),
                    "end_time": _format_time_value(values.get("end_time")),
                    "day_start": day_start,
                    "lunch": lunch,
                    "day_end": day_end,
                    "hubspot_notes": _stringify(values.get("hubspot_notes")),
                    "additional_notes": _stringify(values.get("additional_notes")),
                    "_approx_raw": values.get("approx_time_spent"),
                    "_start_raw": values.get("start_time"),
                    "_end_raw": values.get("end_time"),
                    "_warnings": list(sheet_warnings),
                }

            if current:
                records.append(_finalize_record(current))
    finally:
        workbook.close()
    return records


def _is_template_or_temp(path: Path) -> tuple[bool, str]:
    name = path.name.lower()
    if path.name.startswith("~$"):
        return True, "possible template skipped: temporary Excel file"
    if "template" in name:
        return True, "possible template skipped"
    return False, ""


def find_workbooks(root: Path) -> tuple[list[Path], int]:
    workbooks: list[Path] = []
    templates_skipped = 0
    for path in sorted(root.rglob("*.xlsx")):
        should_skip, _warning = _is_template_or_temp(path)
        if should_skip:
            templates_skipped += 1
            continue
        workbooks.append(path)
    return workbooks, templates_skipped


def _record_in_filters(
    record: dict[str, Any],
    start_date: str,
    end_date: str,
    employee: str,
    code: str,
    project: str,
) -> bool:
    work_date = record.get("work_date") or ""
    if start_date and work_date and work_date < start_date:
        return False
    if end_date and work_date and work_date > end_date:
        return False
    if employee:
        needle = employee.lower()
        haystack = f"{record.get('employee_name', '')} {record.get('employee_folder', '')}".lower()
        if needle not in haystack:
            return False
    if code and code.lower() not in (record.get("code") or "").lower():
        return False
    if project and project.lower() not in (record.get("project") or "").lower():
        return False
    return True


def scan_office_timesheets(
    root: Path,
    start_date: str = "",
    end_date: str = "",
    employee: str = "",
    code: str = "",
    project: str = "",
    include_skipped: bool = False,
) -> ScanResult:
    _ = include_skipped
    workbooks, templates_skipped = find_workbooks(root)
    records: list[dict[str, Any]] = []
    for workbook in workbooks:
        records.extend(_scan_workbook(workbook, root))
    filtered = [record for record in records if _record_in_filters(record, start_date, end_date, employee, code, project)]
    return ScanResult(records=filtered, stats=ScanStats(files_scanned=len(workbooks), templates_skipped=templates_skipped))


def _is_line_item_record(record: dict[str, Any]) -> bool:
    return bool(record.get("employee_name") or record.get("project") or record.get("code") or record.get("duration_hours"))


def _record_notes(record: dict[str, Any]) -> str:
    notes = [record.get("hubspot_notes", ""), record.get("additional_notes", "")]
    return "\n".join(note for note in notes if note)


def _is_timed_entry(record: dict[str, Any]) -> bool:
    return record.get("row_type") == "timed_entry"


def _is_activity_only(record: dict[str, Any]) -> bool:
    return record.get("row_type") == "activity_only"


def build_employee_daily_summary(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str], dict[str, Any]] = {}
    projects: defaultdict[tuple[str, str], set[str]] = defaultdict(set)
    for record in records:
        if not _is_line_item_record(record):
            continue
        key = (record.get("work_date", ""), record.get("employee_name", ""))
        group = groups.setdefault(key, {
            "work_date": key[0],
            "employee_name": key[1],
            "total_hours": 0.0,
            "line_count": 0,
            "timed_entry_count": 0,
            "activity_only_count": 0,
            "missing_code_count": 0,
            "project_count": 0,
            "warning_count": 0,
        })
        if _is_timed_entry(record):
            group["total_hours"] += float(record.get("duration_hours") or 0)
            group["timed_entry_count"] += 1
        if _is_activity_only(record):
            group["activity_only_count"] += 1
        if record.get("code_missing"):
            group["missing_code_count"] += 1
        group["line_count"] += 1
        if record.get("warnings"):
            group["warning_count"] += 1
        if record.get("project"):
            projects[key].add(record["project"])
    for key, group in groups.items():
        group["total_hours"] = round(group["total_hours"], 4)
        group["project_count"] = len(projects[key])
    return sorted(groups.values(), key=lambda row: (row["work_date"], row["employee_name"]))


def build_code_summary(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[str, dict[str, Any]] = {}
    employees: defaultdict[str, set[str]] = defaultdict(set)
    projects: defaultdict[str, set[str]] = defaultdict(set)
    dates: defaultdict[str, list[str]] = defaultdict(list)
    for record in records:
        if not _is_line_item_record(record):
            continue
        key = record.get("code", "")
        group = groups.setdefault(key, {
            "code": key,
            "total_hours": 0.0,
            "employee_count": 0,
            "project_count": 0,
            "line_count": 0,
            "timed_entry_count": 0,
            "activity_only_count": 0,
            "missing_code_count": 0,
            "date_min": "",
            "date_max": "",
        })
        if _is_timed_entry(record):
            group["total_hours"] += float(record.get("duration_hours") or 0)
            group["timed_entry_count"] += 1
        if _is_activity_only(record):
            group["activity_only_count"] += 1
        if record.get("code_missing"):
            group["missing_code_count"] += 1
        group["line_count"] += 1
        if record.get("employee_name"):
            employees[key].add(record["employee_name"])
        if record.get("project"):
            projects[key].add(record["project"])
        if record.get("work_date"):
            dates[key].append(record["work_date"])
    for key, group in groups.items():
        group["total_hours"] = round(group["total_hours"], 4)
        group["employee_count"] = len(employees[key])
        group["project_count"] = len(projects[key])
        if dates[key]:
            group["date_min"] = min(dates[key])
            group["date_max"] = max(dates[key])
    return sorted(groups.values(), key=lambda row: row["code"])


def build_project_touch_summary(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str], dict[str, Any]] = {}
    employees: defaultdict[tuple[str, str], set[str]] = defaultdict(set)
    dates: defaultdict[tuple[str, str], list[str]] = defaultdict(list)
    latest_note_dates: dict[tuple[str, str], str] = {}
    for record in records:
        if not _is_line_item_record(record):
            continue
        key = (record.get("project", ""), record.get("code", ""))
        group = groups.setdefault(key, {
            "project": key[0],
            "code": key[1],
            "total_hours": 0.0,
            "employee_count": 0,
            "date_min": "",
            "date_max": "",
            "line_count": 0,
            "timed_entry_count": 0,
            "activity_only_count": 0,
            "missing_code_count": 0,
            "latest_notes": "",
        })
        if _is_timed_entry(record):
            group["total_hours"] += float(record.get("duration_hours") or 0)
            group["timed_entry_count"] += 1
        if _is_activity_only(record):
            group["activity_only_count"] += 1
        if record.get("code_missing"):
            group["missing_code_count"] += 1
        group["line_count"] += 1
        if record.get("employee_name"):
            employees[key].add(record["employee_name"])
        if record.get("work_date"):
            dates[key].append(record["work_date"])
        notes = _record_notes(record)
        note_date = record.get("work_date") or ""
        if notes and note_date >= latest_note_dates.get(key, ""):
            latest_note_dates[key] = note_date
            group["latest_notes"] = notes
    for key, group in groups.items():
        group["total_hours"] = round(group["total_hours"], 4)
        group["employee_count"] = len(employees[key])
        if dates[key]:
            group["date_min"] = min(dates[key])
            group["date_max"] = max(dates[key])
    return sorted(groups.values(), key=lambda row: (row["project"], row["code"]))


def warning_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [record for record in records if _has_actionable_warning(record)]


def _has_actionable_warning(record: dict[str, Any]) -> bool:
    warnings = record.get("warnings") or ""
    return any(pattern in warnings for pattern in ACTIONABLE_WARNING_PATTERNS)


def build_console_summary(records: list[dict[str, Any]], stats: ScanStats) -> dict[str, Any]:
    return {
        "files_scanned": stats.files_scanned,
        "templates_skipped": stats.templates_skipped,
        "entries_written": len(records),
        "rows_with_missing_duration": sum("missing duration" in (record.get("warnings") or "") for record in records),
        "rows_with_invalid_duration": sum("invalid duration" in (record.get("warnings") or "") for record in records),
        "rows_with_unusually_large_duration": sum("unusually large duration" in (record.get("warnings") or "") for record in records),
        "total_parsed_hours": round(sum(float(record.get("duration_hours") or 0) for record in records), 4),
    }


def print_console_summary(summary: dict[str, Any]) -> None:
    print(f"Files scanned: {summary['files_scanned']}")
    print(f"Templates skipped: {summary['templates_skipped']}")
    print(f"Entries written: {summary['entries_written']}")
    print(f"Rows with missing duration: {summary['rows_with_missing_duration']}")
    print(f"Rows with invalid duration: {summary['rows_with_invalid_duration']}")
    print(f"Rows with unusually large duration: {summary['rows_with_unusually_large_duration']}")
    print(f"Total parsed hours: {summary['total_parsed_hours']}")


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(rows, indent=2, default=str), encoding="utf-8")


def _extract_zip(zip_path: Path) -> tempfile.TemporaryDirectory[str]:
    temp_dir = tempfile.TemporaryDirectory(prefix="office_timesheet_zip_")
    with zipfile.ZipFile(zip_path) as archive:
        archive.extractall(temp_dir.name)
    return temp_dir


def _scan_root_from_extracted_zip(temp_dir: tempfile.TemporaryDirectory[str]) -> Path:
    root = Path(temp_dir.name)
    children = [child for child in root.iterdir() if child.is_dir()]
    if len(children) == 1 and children[0].name.lower() == "timesheets":
        return children[0]
    return root


def sync_sharepoint_office_timesheets(
    *,
    sharepoint_url: str,
    library: str,
    folder: str,
    cache: Path,
    max_depth: int = 8,
    force: bool = False,
) -> Path:
    target = SharePointTarget.from_url(sharepoint_url, library=library, folder_path=folder)
    client = GraphClient()
    site = client.get_site(target.hostname, target.site_path)
    drive = client.get_drive_by_name(site["id"], target.library)
    root_item = client.get_root_or_path_item(drive["id"], target.folder_path)
    sync_root = cache / _safe_name(site.get("name") or target.site_path.strip("/").replace("/", "_")) / _safe_name(target.folder_path or "root")

    def walk(item_id: str, local_dir: Path, depth: int) -> None:
        if depth > max_depth:
            return
        local_dir.mkdir(parents=True, exist_ok=True)
        for child in client.list_children(drive["id"], item_id):
            name = child.get("name", "")
            if not name:
                continue
            if child.get("folder") is not None:
                walk(child["id"], local_dir / _safe_name(name), depth + 1)
                continue
            if Path(name).suffix.lower() != ".xlsx":
                continue
            destination = local_dir / _safe_name(name)
            if destination.exists() and not force:
                continue
            client.download_item(drive["id"], child["id"], destination)

    walk(root_item["id"], sync_root, 0)
    return sync_root


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Scan office/admin/sales timesheet workbooks and summarize employee time, codes, project touches, and warnings."
    )
    parser.add_argument("--root", type=Path, default=None, help="Local office timesheets folder or exported ZIP")
    parser.add_argument("--sharepoint-url", default="", help="Site URL, e.g. https://contoso.sharepoint.com/sites/Operations")
    parser.add_argument("--library", default="Documents", help="SharePoint document library name. Default: Documents")
    parser.add_argument("--folder", default="", help="Folder path inside the library, e.g. Timesheets")
    parser.add_argument("--cache", type=Path, default=Path(".cache/sharepoint_office_timesheets"), help="Local cache folder for SharePoint downloads")
    parser.add_argument("--force", action="store_true", help="Redownload SharePoint workbooks even when cached")
    parser.add_argument("--out", type=Path, default=Path("output/office_timesheet_entries.csv"), help="Detail CSV output path")
    parser.add_argument(
        "--employee-daily-summary",
        type=Path,
        default=Path("output/office_timesheet_employee_daily_summary.csv"),
        help="Employee/day summary CSV output path",
    )
    parser.add_argument(
        "--code-summary",
        type=Path,
        default=Path("output/office_timesheet_code_summary.csv"),
        help="Code summary CSV output path",
    )
    parser.add_argument(
        "--project-touch-summary",
        type=Path,
        default=Path("output/office_timesheet_project_touch_summary.csv"),
        help="Project/code touch summary CSV output path",
    )
    parser.add_argument(
        "--warnings-out",
        type=Path,
        default=Path("output/office_timesheet_warnings.csv"),
        help="Warnings-only CSV output path",
    )
    parser.add_argument("--json", type=Path, default=None, help="Optional detail JSON output path")
    parser.add_argument("--start-date", default="", help="Inclusive ISO date filter, e.g. 2026-06-01")
    parser.add_argument("--end-date", default="", help="Inclusive ISO date filter, e.g. 2026-06-30")
    parser.add_argument("--employee", default="", help="Employee name or folder substring filter")
    parser.add_argument("--code", default="", help="Code substring filter")
    parser.add_argument("--project", default="", help="Project substring filter")
    parser.add_argument("--dry-run", action="store_true", help="Scan and print counts without writing output files")
    args = parser.parse_args()

    temp_dir: tempfile.TemporaryDirectory[str] | None = None
    if args.sharepoint_url:
        scan_root = sync_sharepoint_office_timesheets(
            sharepoint_url=args.sharepoint_url,
            library=args.library,
            folder=args.folder,
            cache=args.cache,
            force=args.force,
        )
        print(f"SharePoint cache: {scan_root}")
    elif args.root:
        if args.root.suffix.lower() == ".zip":
            temp_dir = _extract_zip(args.root)
            scan_root = _scan_root_from_extracted_zip(temp_dir)
        else:
            scan_root = args.root
    else:
        parser.error("Provide --root for local scanning or --sharepoint-url for Graph scanning.")

    try:
        scan_result = scan_office_timesheets(
            scan_root,
            start_date=args.start_date,
            end_date=args.end_date,
            employee=args.employee,
            code=args.code,
            project=args.project,
        )
        records = scan_result.records
        employee_daily_summary = build_employee_daily_summary(records)
        code_summary = build_code_summary(records)
        project_touch_summary = build_project_touch_summary(records)
        warnings = warning_records(records)
        console_summary = build_console_summary(records, scan_result.stats)

        if args.dry_run:
            print(f"Dry run: scanned root {scan_root}")
            print(f"Employee daily summary rows: {len(employee_daily_summary)}")
            print(f"Code summary rows: {len(code_summary)}")
            print(f"Project touch summary rows: {len(project_touch_summary)}")
            print(f"Warning rows: {len(warnings)}")
            print_console_summary(console_summary)
            return

        write_csv(args.out, records, DETAIL_FIELDS)
        write_csv(args.employee_daily_summary, employee_daily_summary, EMPLOYEE_DAILY_SUMMARY_FIELDS)
        write_csv(args.code_summary, code_summary, CODE_SUMMARY_FIELDS)
        write_csv(args.project_touch_summary, project_touch_summary, PROJECT_TOUCH_SUMMARY_FIELDS)
        write_csv(args.warnings_out, warnings, DETAIL_FIELDS)
        if args.json:
            write_json(args.json, records)

        print_console_summary(console_summary)
        print(f"CSV: {args.out}")
        print(f"Employee daily summary CSV: {args.employee_daily_summary}")
        print(f"Code summary CSV: {args.code_summary}")
        print(f"Project touch summary CSV: {args.project_touch_summary}")
        print(f"Warnings CSV: {args.warnings_out}")
        if args.json:
            print(f"JSON: {args.json}")
    finally:
        if temp_dir:
            temp_dir.cleanup()


if __name__ == "__main__":
    main()
