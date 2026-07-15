from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import subprocess
import time
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from dotenv import load_dotenv
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Connection, Engine
from sqlalchemy.exc import DBAPIError, InterfaceError, OperationalError, PendingRollbackError, SQLAlchemyError

from .graph_client import GraphClient, GraphError, SharePointTarget
from .job_search import first_nonblank, normalize_search_text, tokenize_search_text

SUPPORTED_EXTENSIONS = {".pdf", ".doc", ".docx", ".xlsx", ".xlsm", ".txt", ".csv"}
JOB_SPEC_CANDIDATE_PATTERN = r"(job[ _-]*spec|jobspec|job specification|spec form|scope of work|work scope|(^|\\s)spec(\\s|$)|job tracking|tracking form|field notes|site notes|inspection notes|estimator notes)"
JOB_SPEC_EXCLUDE_PATTERN = r"(submittal|submittals|sds|pds|tds|technical data|data sheet|sales sheet|brochure|certificate of liability|cert tracking)"
ESTIMATOR_RELEVANT_DOCUMENT_TYPES = [
    "estimate",
    "proposal",
    "contract",
    "warranty",
    "job_tracking",
    "specification",
    "field_notes",
    "site_notes",
]
ESTIMATOR_RELEVANT_CANDIDATE_PATTERN = (
    r"(estimate|proposal|quote|scope of work|work scope|job[ _-]*spec|jobspec|job specification|"
    r"job tracking|tracking form|field notes|site notes|inspection notes|estimator notes|"
    r"contract|agreement|warranty)"
)
ESTIMATOR_RELEVANT_EXCLUDE_PATTERN = (
    r"(submittal|submittals|sds|pds|tds|technical data|data sheet|sales sheet|brochure|"
    r"certificate of liability|cert tracking|aerial|eagleview|drone|photo|photos|picture|image|"
    r"bid package|bid documents)"
)
TEXT_EMPTY_THRESHOLD = 20
MAX_XLSX_ROWS_PER_SHEET = 5000
MAX_XLSX_COLUMNS_PER_SHEET = 120


class DocumentAcquisitionError(RuntimeError):
    pass


class TransientDocumentDatabaseError(RuntimeError):
    pass


@dataclass
class ExtractedContent:
    content_type: str
    source_locator: str
    text_content: str
    page_number: int | None = None
    sheet_name: str | None = None
    cell_range: str | None = None
    row_number: int | None = None
    section_name: str | None = None


@dataclass
class ExtractionResult:
    rows: list[ExtractedContent]
    extraction_method: str
    requires_ocr: bool = False


def text_or_none(value: Any) -> str | None:
    text_value = str(value or "").strip()
    return text_value or None


def is_transient_database_error(exc: Exception) -> bool:
    if isinstance(exc, PendingRollbackError):
        return True
    if isinstance(exc, (OperationalError, InterfaceError, DBAPIError)) and getattr(exc, "connection_invalidated", False):
        return True
    message = str(exc).lower()
    return isinstance(exc, (OperationalError, InterfaceError, DBAPIError, PendingRollbackError)) and any(
        marker in message
        for marker in (
            "ssl connection has been closed unexpectedly",
            "server closed the connection unexpectedly",
            "connection already closed",
            "can't reconnect until invalid transaction is rolled back",
            "connection not open",
            "connection is closed",
        )
    )


def rollback_connection(connection: Connection) -> None:
    try:
        connection.rollback()
    except Exception:
        pass


def safe_filename(value: Any) -> str:
    name = str(value or "document").strip() or "document"
    return re.sub(r"[^A-Za-z0-9._-]+", "_", name)[:180]


def file_sha1(path: Path) -> str:
    digest = hashlib.sha1()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def content_id_for(document_id: str, row: ExtractedContent) -> str:
    key = "||".join(
        str(part or "")
        for part in [
            document_id,
            row.content_type,
            row.source_locator,
            row.page_number,
            row.sheet_name,
            row.cell_range,
            row.row_number,
            row.text_content,
        ]
    )
    return f"content-{hashlib.sha1(key.encode('utf-8')).hexdigest()[:28]}"


def normalized_content(text_value: str) -> str:
    return normalize_search_text(text_value)


def postgres_safe_text(value: Any) -> str | None:
    if value is None:
        return None
    return str(value).replace("\x00", "")


def stable_cache_path(document: dict[str, Any], cache_root: Path) -> Path:
    document_id = first_nonblank(document.get("document_id")) or "document"
    extension = first_nonblank(document.get("file_extension")) or Path(first_nonblank(document.get("file_name"))).suffix
    if extension and not extension.startswith("."):
        extension = "." + extension
    return cache_root / "_document_content_cache" / f"{safe_filename(document_id)}{extension.lower()}"


def is_nonempty_file(path: Path) -> bool:
    return path.exists() and path.is_file() and path.stat().st_size > 0


def looks_like_html_file(path: Path) -> bool:
    try:
        prefix = path.read_bytes()[:512].lstrip().lower()
    except OSError:
        return False
    return prefix.startswith(b"<!doctype html") or prefix.startswith(b"<html") or b"<title>sign in" in prefix


def find_existing_cached_file(document: dict[str, Any], cache_root: Path) -> Path | None:
    cached = text_or_none(document.get("cached_file_path"))
    if cached and is_nonempty_file(Path(cached)):
        return Path(cached)

    relative_path = text_or_none(document.get("relative_path"))
    if relative_path:
        direct = cache_root / relative_path.strip("/")
        if is_nonempty_file(direct):
            return direct
        relative_parts = Path(relative_path).parts
        file_name = Path(relative_path).name
    else:
        relative_parts = ()
        file_name = first_nonblank(document.get("file_name"))

    if not file_name:
        return None
    for candidate in cache_root.rglob(file_name):
        if not is_nonempty_file(candidate):
            continue
        if relative_parts and tuple(candidate.parts[-len(relative_parts) :]) != relative_parts:
            continue
        return candidate
    return None


def ensure_local_document(document: dict[str, Any], cache_root: Path, force_download: bool = False) -> Path:
    if not force_download:
        existing = find_existing_cached_file(document, cache_root)
        if existing:
            return existing

    drive_id = first_nonblank(document.get("drive_id"))
    drive_item_id = first_nonblank(document.get("drive_item_id"))
    if not drive_id or not drive_item_id:
        raise DocumentAcquisitionError("No cached file found and document row does not include drive_id plus drive_item_id for download.")

    destination = stable_cache_path(document, cache_root)
    try:
        GraphClient(max_retries=2).download_item(drive_id, drive_item_id, destination)
    except GraphError as exc:
        raise DocumentAcquisitionError(str(exc)) from exc
    if not is_nonempty_file(destination):
        raise DocumentAcquisitionError("Downloaded file is missing or empty.")
    if looks_like_html_file(destination):
        try:
            destination.unlink()
        except OSError:
            pass
        raise DocumentAcquisitionError("Downloaded file appears to be HTML, not document content.")
    return destination


def planned_acquisition_method(document: dict[str, Any], cache_root: Path, force_download: bool = False) -> str:
    if not force_download and find_existing_cached_file(document, cache_root):
        return "cached_file"
    if first_nonblank(document.get("drive_id")) and first_nonblank(document.get("drive_item_id")):
        return "graph_content_download"
    return "unavailable"


def manifest_metadata_rows(cache_root: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for manifest_path in cache_root.rglob(".jobscan_manifest.json"):
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if not isinstance(manifest, dict):
            continue
        manifest_drive_id = manifest.get("drive_id")
        docs = manifest.get("documents") if isinstance(manifest.get("documents"), list) else []
        items = manifest.get("items") if isinstance(manifest.get("items"), dict) else {}
        for item_id, item in items.items():
            if not isinstance(item, dict):
                continue
            parent = item.get("parentReference") if isinstance(item.get("parentReference"), dict) else {}
            rows.append(
                {
                    "drive_id": item.get("drive_id") or parent.get("driveId") or manifest_drive_id,
                    "drive_item_id": item.get("drive_item_id") or item.get("graph_item_id") or item.get("id") or item_id,
                    "sharepoint_url": item.get("webUrl") or item.get("web_url"),
                    "relative_path": item.get("relative_path") or item.get("local_path"),
                    "file_name": item.get("name"),
                }
            )
        for doc in docs:
            if not isinstance(doc, dict):
                continue
            parent = doc.get("parentReference") if isinstance(doc.get("parentReference"), dict) else {}
            rows.append(
                {
                    "drive_id": doc.get("drive_id") or parent.get("driveId") or manifest_drive_id,
                    "drive_item_id": doc.get("drive_item_id") or doc.get("graph_item_id") or doc.get("id"),
                    "sharepoint_url": doc.get("webUrl") or doc.get("web_url") or doc.get("sharepoint_url"),
                    "relative_path": doc.get("relative_path") or doc.get("local_path"),
                    "file_name": doc.get("name") or doc.get("file_name"),
                }
            )
    out: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for row in rows:
        if not row.get("drive_id") or not row.get("drive_item_id"):
            continue
        key = (
            str(row.get("drive_id") or ""),
            str(row.get("drive_item_id") or ""),
            str(row.get("sharepoint_url") or ""),
            str(row.get("relative_path") or ""),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def backfill_document_drive_metadata(
    connection: Connection,
    cache_root: Path,
    *,
    limit: int | None = None,
    job_id: str | None = None,
    progress_every: int = 0,
) -> int:
    candidates = manifest_metadata_rows(cache_root)
    if limit is not None and limit > 0:
        candidates = candidates[:limit]
    total = len(candidates)
    if progress_every and progress_every > 0:
        print(f"Cached manifest metadata candidates: {total}", flush=True)
    updated = 0
    for index, candidate in enumerate(candidates, start=1):
        params = {
            "drive_id": candidate.get("drive_id"),
            "drive_item_id": candidate.get("drive_item_id"),
            "sharepoint_url": candidate.get("sharepoint_url"),
            "relative_path": candidate.get("relative_path"),
            "file_name": candidate.get("file_name"),
            "job_id": job_id,
        }
        result = connection.execute(
            text(
                """
                UPDATE documents
                SET drive_id = COALESCE(NULLIF(drive_id, ''), :drive_id),
                    drive_item_id = COALESCE(NULLIF(drive_item_id, ''), :drive_item_id),
                    updated_at = NOW()
                WHERE (:job_id IS NULL OR job_id = :job_id)
                  AND (drive_id IS NULL OR drive_id = '' OR drive_item_id IS NULL OR drive_item_id = '')
                  AND (
                    (sharepoint_url IS NOT NULL AND sharepoint_url <> '' AND sharepoint_url = :sharepoint_url)
                    OR (relative_path IS NOT NULL AND relative_path <> '' AND relative_path = :relative_path)
                  )
                """
            ),
            params,
        )
        updated += int(result.rowcount or 0)
        if progress_every and progress_every > 0 and (index % progress_every == 0 or index == total):
            print(f"Backfill metadata progress: {index}/{total} candidates, {updated} rows updated", flush=True)
    return updated


def resolve_graph_metadata_for_document(
    client: GraphClient,
    document: dict[str, Any],
    *,
    site_url: str,
    library: str,
    root_folder: str = "",
) -> dict[str, Any]:
    target = SharePointTarget.from_url(site_url, library=library, folder_path=root_folder)
    site = client.get_site(target.hostname, target.site_path)
    drive = client.get_drive_by_name(site["id"], library)
    relative_path = first_nonblank(document.get("relative_path"))
    file_name = first_nonblank(document.get("file_name"))
    if not relative_path and not file_name:
        raise DocumentAcquisitionError("Document row has no relative_path or file_name for Graph path resolution.")
    path_parts = [part.strip("/") for part in [root_folder, relative_path or file_name] if part and part.strip("/")]
    graph_path = "/".join(path_parts)
    item = client.get_root_or_path_item(drive["id"], graph_path)
    return {
        "document_id": document.get("document_id"),
        "drive_id": drive["id"],
        "drive_item_id": item.get("id"),
        "sharepoint_url": item.get("webUrl") or document.get("sharepoint_url"),
        "file_name": item.get("name") or document.get("file_name"),
    }


def update_document_drive_metadata(connection: Connection, metadata: dict[str, Any]) -> int:
    result = connection.execute(
        text(
            """
            UPDATE documents
            SET drive_id = :drive_id,
                drive_item_id = :drive_item_id,
                sharepoint_url = COALESCE(NULLIF(sharepoint_url, ''), :sharepoint_url),
                updated_at = NOW()
            WHERE document_id = :document_id
            """
        ),
        metadata,
    )
    return int(result.rowcount or 0)


def extract_pdf(path: Path) -> ExtractionResult:
    try:
        try:
            from pypdf import PdfReader
        except ImportError:
            from PyPDF2 import PdfReader  # type: ignore[no-redef]
    except ImportError as exc:
        raise RuntimeError("Install pypdf to extract PDF text: pip install pypdf") from exc

    reader = PdfReader(str(path))
    rows: list[ExtractedContent] = []
    total_chars = 0
    for index, page in enumerate(reader.pages, start=1):
        page_text = (page.extract_text() or "").strip()
        if not page_text:
            continue
        total_chars += len(page_text)
        rows.append(
            ExtractedContent(
                content_type="pdf_page",
                source_locator=f"page {index}",
                page_number=index,
                text_content=page_text,
            )
        )
    return ExtractionResult(rows=rows, extraction_method="pypdf", requires_ocr=total_chars < TEXT_EMPTY_THRESHOLD)


def _docx_text(node: ET.Element, ns: dict[str, str]) -> str:
    return "".join(text_node.text or "" for text_node in node.findall(".//w:t", ns)).strip()


def _docx_paragraph_style(node: ET.Element, ns: dict[str, str]) -> str:
    style = node.find("./w:pPr/w:pStyle", ns)
    if style is None:
        return ""
    return style.attrib.get(f"{{{ns['w']}}}val", "")


def extract_docx(path: Path) -> ExtractionResult:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with zipfile.ZipFile(path) as package:
        xml = package.read("word/document.xml")
    root = ET.fromstring(xml)
    body = root.find("w:body", ns)
    if body is None:
        return ExtractionResult(rows=[], extraction_method="docx-xml")

    rows: list[ExtractedContent] = []
    current_section: str | None = None
    paragraph_count = 0
    table_count = 0
    for child in list(body):
        tag = child.tag.rsplit("}", 1)[-1]
        if tag == "p":
            content = _docx_text(child, ns)
            if not content:
                continue
            paragraph_count += 1
            style = _docx_paragraph_style(child, ns)
            if style.lower().startswith("heading"):
                current_section = content
                content_type = "docx_heading"
            else:
                content_type = "docx_paragraph"
            rows.append(
                ExtractedContent(
                    content_type=content_type,
                    source_locator=f"paragraph {paragraph_count}",
                    section_name=current_section,
                    text_content=content,
                )
            )
        elif tag == "tbl":
            table_count += 1
            for row_number, row in enumerate(child.findall(".//w:tr", ns), start=1):
                cell_values = [_docx_text(cell, ns) for cell in row.findall("./w:tc", ns)]
                cell_values = [value for value in cell_values if value]
                if not cell_values:
                    continue
                rows.append(
                    ExtractedContent(
                        content_type="docx_table_row",
                        source_locator=f"table {table_count} row {row_number}",
                        row_number=row_number,
                        section_name=current_section,
                        text_content=" | ".join(cell_values),
                    )
                )
    return ExtractionResult(rows=rows, extraction_method="docx-xml")


def extract_legacy_doc(path: Path) -> ExtractionResult:
    converters = [
        ("textutil", ["textutil", "-convert", "txt", "-stdout", str(path)]),
        ("antiword", ["antiword", str(path)]),
        ("catdoc", ["catdoc", str(path)]),
    ]
    errors: list[str] = []
    for name, command in converters:
        if not shutil.which(name):
            continue
        try:
            result = subprocess.run(
                command,
                check=False,
                capture_output=True,
                text=True,
                timeout=45,
            )
        except Exception as exc:
            errors.append(f"{name}: {exc}")
            continue
        text_content = (result.stdout or "").strip()
        if result.returncode == 0 and len(text_content) >= TEXT_EMPTY_THRESHOLD:
            return ExtractionResult(
                rows=[
                    ExtractedContent(
                        content_type="legacy_doc_text",
                        source_locator="file",
                        text_content=text_content,
                    )
                ],
                extraction_method=name,
            )
        error = (result.stderr or result.stdout or "").strip()
        errors.append(f"{name}: {error or 'no text extracted'}")
    detail = "; ".join(errors) if errors else "no supported converter found"
    raise RuntimeError(
        "Legacy .doc extraction requires textutil, antiword, or catdoc on the runtime. "
        f"Attempted converters: {detail}"
    )


def _xlsx_cell_text(cell: Any) -> str | None:
    return text_or_none(getattr(cell, "value", None))


def extract_xlsx(path: Path) -> ExtractionResult:
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Install openpyxl to extract Excel content: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    formula_wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    rows: list[ExtractedContent] = []
    try:
        for ws_index, ws in enumerate(wb.worksheets):
            if getattr(ws, "sheet_state", "visible") != "visible":
                continue
            formula_ws = formula_wb.worksheets[ws_index]
            max_row = min(ws.max_row or 0, MAX_XLSX_ROWS_PER_SHEET)
            max_col = min(ws.max_column or 0, MAX_XLSX_COLUMNS_PER_SHEET)
            if max_row <= 0 or max_col <= 0:
                continue
            formula_rows = formula_ws.iter_rows(max_row=max_row, max_col=max_col)
            value_rows = ws.iter_rows(max_row=max_row, max_col=max_col)
            for row_number, (row, formula_row) in enumerate(zip(value_rows, formula_rows), start=1):
                values: list[tuple[int, str]] = []
                for column_number, (cell, formula_cell) in enumerate(zip(row, formula_row), start=1):
                    value = _xlsx_cell_text(cell)
                    if value is None:
                        value = _xlsx_cell_text(formula_cell)
                    if value:
                        values.append((column_number, value))
                if not values:
                    continue
                start_col = get_column_letter(values[0][0])
                end_col = get_column_letter(values[-1][0])
                cell_range = f"{start_col}{row_number}:{end_col}{row_number}"
                text_content = " | ".join(f"{get_column_letter(col)}{row_number}: {value}" for col, value in values)
                rows.append(
                    ExtractedContent(
                        content_type="xlsx_row",
                        source_locator=f"{ws.title}!{cell_range}",
                        sheet_name=ws.title,
                        cell_range=cell_range,
                        row_number=row_number,
                        text_content=text_content,
                    )
                )
    finally:
        wb.close()
        formula_wb.close()
    return ExtractionResult(rows=rows, extraction_method="openpyxl")


def extract_text_file(path: Path) -> ExtractionResult:
    text_content = path.read_text(encoding="utf-8", errors="replace").strip()
    rows = [ExtractedContent(content_type="text_file", source_locator="file", text_content=text_content)] if text_content else []
    return ExtractionResult(rows=rows, extraction_method="plain-text")


def extract_document_file(path: Path, document: dict[str, Any]) -> ExtractionResult:
    extension = (first_nonblank(document.get("file_extension")) or path.suffix).lower()
    if extension == ".pdf":
        return extract_pdf(path)
    if extension == ".doc":
        return extract_legacy_doc(path)
    if extension == ".docx":
        return extract_docx(path)
    if extension in {".xlsx", ".xlsm"}:
        return extract_xlsx(path)
    if extension in {".txt", ".csv"}:
        return extract_text_file(path)
    raise RuntimeError(f"Unsupported document type for extraction: {extension or path.suffix or 'unknown'}")


def document_content_table_available(connection: Connection) -> bool:
    return bool(
        connection.execute(
            text(
                """
                SELECT EXISTS (
                    SELECT 1 FROM information_schema.tables
                    WHERE table_schema = 'public' AND table_name = 'document_content'
                )
                """
            )
        ).scalar()
    )


def documents_table_available(connection: Connection) -> bool:
    try:
        return bool(
            connection.execute(
                text(
                    """
                    SELECT EXISTS (
                        SELECT 1 FROM information_schema.tables
                        WHERE table_schema = 'public' AND table_name = 'documents'
                    )
                    """
                )
            ).scalar()
        )
    except SQLAlchemyError:
        try:
            connection.execute(text("SELECT 1 FROM documents LIMIT 1"))
            return True
        except SQLAlchemyError:
            return False


def update_document_failure(connection: Connection, document_id: str, error: str, cached_file_path: str | None = None) -> None:
    connection.execute(
        text(
            """
            UPDATE documents
            SET extraction_status = 'failed',
                extraction_error = :error,
                cached_file_path = COALESCE(:cached_file_path, cached_file_path),
                updated_at = NOW()
            WHERE document_id = :document_id
            """
        ),
        {"document_id": document_id, "error": error[:1000], "cached_file_path": cached_file_path},
    )


def should_skip_extraction(document: dict[str, Any], current_hash: str, force: bool = False) -> bool:
    if force:
        return False
    status = first_nonblank(document.get("extraction_status"))
    previous_hash = first_nonblank(document.get("content_hash"))
    return bool(previous_hash and previous_hash == current_hash and status in {"succeeded", "failed", "ocr_required"})


def replace_document_content(
    connection: Connection,
    document: dict[str, Any],
    path: Path,
    result: ExtractionResult,
    current_hash: str,
) -> int:
    document_id = first_nonblank(document.get("document_id"))
    job_id = text_or_none(document.get("job_id"))
    now = datetime.now(timezone.utc)
    connection.execute(text("DELETE FROM document_content WHERE document_id = :document_id"), {"document_id": document_id})
    for row in result.rows:
        text_content = postgres_safe_text(row.text_content) or ""
        source_locator = postgres_safe_text(row.source_locator)
        sheet_name = postgres_safe_text(row.sheet_name)
        cell_range = postgres_safe_text(row.cell_range)
        section_name = postgres_safe_text(row.section_name)
        connection.execute(
            text(
                """
                INSERT INTO document_content (
                    content_id, document_id, job_id, content_type, source_locator, page_number,
                    sheet_name, cell_range, row_number, section_name, text_content, normalized_text,
                    extraction_method, content_hash, created_at, updated_at
                )
                VALUES (
                    :content_id, :document_id, :job_id, :content_type, :source_locator, :page_number,
                    :sheet_name, :cell_range, :row_number, :section_name, :text_content, :normalized_text,
                    :extraction_method, :content_hash, :created_at, :updated_at
                )
                ON CONFLICT (content_id) DO UPDATE SET
                    text_content = EXCLUDED.text_content,
                    normalized_text = EXCLUDED.normalized_text,
                    updated_at = EXCLUDED.updated_at
                """
            ),
            {
                "content_id": content_id_for(document_id, row),
                "document_id": document_id,
                "job_id": job_id,
                "content_type": row.content_type,
                "source_locator": source_locator,
                "page_number": row.page_number,
                "sheet_name": sheet_name,
                "cell_range": cell_range,
                "row_number": row.row_number,
                "section_name": section_name,
                "text_content": text_content,
                "normalized_text": normalized_content(text_content),
                "extraction_method": result.extraction_method,
                "content_hash": current_hash,
                "created_at": now,
                "updated_at": now,
            },
        )
    connection.execute(
        text(
            """
            UPDATE documents
            SET extraction_status = 'succeeded',
                extraction_method = :extraction_method,
                extraction_error = NULL,
                extracted_at = :extracted_at,
                content_hash = :content_hash,
                cached_file_path = :cached_file_path,
                requires_ocr = :requires_ocr,
                updated_at = NOW()
            WHERE document_id = :document_id
            """
        ),
        {
            "document_id": document_id,
            "extraction_method": result.extraction_method,
            "extracted_at": now,
            "content_hash": current_hash,
            "cached_file_path": str(path),
            "requires_ocr": bool(result.requires_ocr),
        },
    )
    return len(result.rows)


def mark_ocr_required(connection: Connection, document: dict[str, Any], path: Path, current_hash: str, result: ExtractionResult) -> None:
    connection.execute(
        text(
            """
            UPDATE documents
            SET extraction_status = 'ocr_required',
                extraction_method = :extraction_method,
                extraction_error = 'No extractable text found; OCR is required.',
                extracted_at = :extracted_at,
                content_hash = :content_hash,
                cached_file_path = :cached_file_path,
                requires_ocr = TRUE,
                updated_at = NOW()
            WHERE document_id = :document_id
            """
        ),
        {
            "document_id": document["document_id"],
            "extraction_method": result.extraction_method,
            "extracted_at": datetime.now(timezone.utc),
            "content_hash": current_hash,
            "cached_file_path": str(path),
        },
    )


def extract_one_document(connection: Connection, document: dict[str, Any], cache_root: Path, *, force: bool = False) -> tuple[str, int]:
    document_id = first_nonblank(document.get("document_id"))
    if not document_id:
        raise RuntimeError("Document row is missing document_id.")
    try:
        path = ensure_local_document(document, cache_root, force_download=force)
        current_hash = file_sha1(path)
        if should_skip_extraction(document, current_hash, force=force):
            return "skipped", 0
        result = extract_document_file(path, document)
        if result.requires_ocr and not result.rows:
            mark_ocr_required(connection, document, path, current_hash, result)
            return "ocr_required", 0
        row_count = replace_document_content(connection, document, path, result, current_hash)
        return "extracted", row_count
    except Exception as exc:
        if is_transient_database_error(exc):
            rollback_connection(connection)
            raise TransientDocumentDatabaseError(
                "Database connection dropped during document extraction; rolled back and will retry/resume safely."
            ) from exc
        try:
            update_document_failure(connection, document_id, str(exc))
        except Exception as failure_exc:
            if is_transient_database_error(failure_exc):
                rollback_connection(connection)
                raise TransientDocumentDatabaseError(
                    "Database connection dropped while recording document extraction failure; "
                    "rolled back and will retry/resume safely."
                ) from failure_exc
            raise
        return "failed", 0


def prepare_document_extraction(document: dict[str, Any], cache_root: Path, *, force: bool = False) -> tuple[str, Path, str, ExtractionResult | None]:
    document_id = first_nonblank(document.get("document_id"))
    if not document_id:
        raise RuntimeError("Document row is missing document_id.")
    path = ensure_local_document(document, cache_root, force_download=force)
    current_hash = file_sha1(path)
    if should_skip_extraction(document, current_hash, force=force):
        return "skipped", path, current_hash, None
    result = extract_document_file(path, document)
    return "prepared", path, current_hash, result


def write_prepared_document_extraction(
    connection: Connection,
    document: dict[str, Any],
    path: Path,
    current_hash: str,
    result: ExtractionResult,
) -> tuple[str, int]:
    if result.requires_ocr and not result.rows:
        mark_ocr_required(connection, document, path, current_hash, result)
        return "ocr_required", 0
    row_count = replace_document_content(connection, document, path, result, current_hash)
    return "extracted", row_count


def record_document_failure_with_retry(
    engine: Engine,
    document: dict[str, Any],
    error: str,
    *,
    cached_file_path: str | None = None,
    retries: int = 2,
    backoff_seconds: float = 0.25,
) -> None:
    document_id = first_nonblank(document.get("document_id"))
    max_attempts = max(1, retries + 1)
    for attempt in range(1, max_attempts + 1):
        try:
            with engine.begin() as connection:
                update_document_failure(connection, document_id, error, cached_file_path=cached_file_path)
            return
        except Exception as exc:
            if not is_transient_database_error(exc) or attempt >= max_attempts:
                raise
            try:
                engine.dispose()
            except Exception:
                pass
            time.sleep(min(1.0, backoff_seconds * (2 ** (attempt - 1))))


def extract_one_document_with_retry(
    engine: Engine,
    document: dict[str, Any],
    cache_root: Path,
    *,
    force: bool = False,
    retries: int = 2,
    backoff_seconds: float = 0.25,
) -> tuple[str, int]:
    try:
        status, path, current_hash, result = prepare_document_extraction(document, cache_root, force=force)
    except Exception as exc:
        if is_transient_database_error(exc):
            raise TransientDocumentDatabaseError(
                "Database connection dropped during document acquisition; retry/resume safely."
            ) from exc
        cached_file_path = None
        try:
            path = stable_cache_path(document, cache_root)
            if path.exists():
                cached_file_path = str(path)
        except Exception:
            cached_file_path = None
        record_document_failure_with_retry(
            engine,
            document,
            str(exc),
            cached_file_path=cached_file_path,
            retries=retries,
            backoff_seconds=backoff_seconds,
        )
        return "failed", 0
    if status == "skipped" or result is None:
        return status, 0

    max_attempts = max(1, retries + 1)
    for attempt in range(1, max_attempts + 1):
        connection: Connection | None = None
        try:
            with engine.begin() as connection:
                return write_prepared_document_extraction(connection, document, path, current_hash, result)
        except TransientDocumentDatabaseError:
            print(
                "Database connection dropped during document extraction; rolled back and will retry/resume safely.",
                flush=True,
            )
            try:
                engine.dispose()
            except Exception:
                pass
            if attempt >= max_attempts:
                raise
            time.sleep(min(1.0, backoff_seconds * (2 ** (attempt - 1))))
        except Exception as exc:
            if not is_transient_database_error(exc):
                raise
            if connection is not None:
                rollback_connection(connection)
            print(
                "Database connection dropped during document extraction write; rolled back and will retry/resume safely.",
                flush=True,
            )
            try:
                engine.dispose()
            except Exception:
                pass
            if attempt >= max_attempts:
                raise TransientDocumentDatabaseError("Database connection dropped during document extraction write.") from exc
            time.sleep(min(1.0, backoff_seconds * (2 ** (attempt - 1))))
    raise TransientDocumentDatabaseError("Database connection dropped during document extraction.")


def document_selection_sql(
    *,
    document_id: str | None,
    job_id: str | None,
    pending: bool,
    failed: bool = False,
    document_type: str | None = None,
    job_spec_candidates: bool = False,
    estimator_relevant: bool = False,
) -> tuple[str, dict[str, Any]]:
    where: list[str] = []
    params: dict[str, Any] = {}
    if document_id:
        where.append("document_id = :document_id")
        params["document_id"] = document_id
    if job_id:
        where.append("job_id = :job_id")
        params["job_id"] = job_id
    if document_type:
        where.append("document_type = :document_type")
        params["document_type"] = document_type
    if job_spec_candidates:
        where.append(
            """
            (
                document_type = ANY(:job_spec_note_document_types)
                OR LOWER(COALESCE(file_name, '') || ' ' || COALESCE(relative_path, '')) ~ :job_spec_candidate_pattern
                OR (
                    document_type = 'specification'
                    AND LOWER(COALESCE(file_name, '') || ' ' || COALESCE(relative_path, '')) !~ :job_spec_exclude_pattern
                )
            )
            """
        )
        params["job_spec_note_document_types"] = ["field_notes", "job_tracking", "site_notes"]
        params["job_spec_candidate_pattern"] = JOB_SPEC_CANDIDATE_PATTERN
        params["job_spec_exclude_pattern"] = JOB_SPEC_EXCLUDE_PATTERN
    if estimator_relevant:
        where.append(
            """
            (
                document_type = ANY(:estimator_relevant_document_types)
                OR (
                    LOWER(COALESCE(file_name, '') || ' ' || COALESCE(relative_path, '')) ~ :estimator_relevant_candidate_pattern
                    AND LOWER(COALESCE(file_name, '') || ' ' || COALESCE(relative_path, '')) !~ :estimator_relevant_exclude_pattern
                )
            )
            """
        )
        params["estimator_relevant_document_types"] = ESTIMATOR_RELEVANT_DOCUMENT_TYPES
        params["estimator_relevant_candidate_pattern"] = ESTIMATOR_RELEVANT_CANDIDATE_PATTERN
        params["estimator_relevant_exclude_pattern"] = ESTIMATOR_RELEVANT_EXCLUDE_PATTERN
    if pending:
        where.append("(extraction_status IS NULL OR extraction_status IN ('not_started', 'pending'))")
    if failed:
        where.append("extraction_status = 'failed'")
    where.append("LOWER(COALESCE(file_extension, '')) = ANY(:extensions)")
    params["extensions"] = sorted(SUPPORTED_EXTENSIONS)
    where_sql = "WHERE " + " AND ".join(where) if where else ""
    return where_sql, params


def load_documents_for_extraction(
    connection: Connection,
    *,
    document_id: str | None = None,
    job_id: str | None = None,
    pending: bool = False,
    failed: bool = False,
    document_type: str | None = None,
    job_spec_candidates: bool = False,
    estimator_relevant: bool = False,
    limit: int = 10,
) -> list[dict[str, Any]]:
    where_sql, params = document_selection_sql(
        document_id=document_id,
        job_id=job_id,
        pending=pending,
        failed=failed,
        document_type=document_type,
        job_spec_candidates=job_spec_candidates,
        estimator_relevant=estimator_relevant,
    )
    limit_sql = ""
    if limit and limit > 0:
        params["limit"] = limit
        limit_sql = "\n        LIMIT :limit"
    sql = f"""
        SELECT *
        FROM documents
        {where_sql}
        ORDER BY updated_at NULLS FIRST, file_name{limit_sql}
    """
    return [dict(row) for row in connection.execute(text(sql), params).mappings().all()]


def list_document_content(connection: Connection | Engine, document_id: str) -> list[dict[str, Any]]:
    manager = connection.connect() if isinstance(connection, Engine) else None
    conn = manager.__enter__() if manager else connection
    try:
        if not document_content_table_available(conn):
            return []
        rows = conn.execute(
            text(
                """
                SELECT *
                FROM document_content
                WHERE document_id = :document_id
                ORDER BY page_number NULLS LAST, sheet_name NULLS LAST, row_number NULLS LAST, source_locator
                """
            ),
            {"document_id": document_id},
        ).mappings()
        return [dict(row) for row in rows]
    finally:
        if manager:
            manager.__exit__(None, None, None)


def excerpt(text_value: str, tokens: list[str], width: int = 240) -> str:
    text_value = " ".join(str(text_value or "").split())
    if not text_value:
        return ""
    normalized = normalize_search_text(text_value)
    position = -1
    for token in tokens:
        position = normalized.find(token.lower())
        if position >= 0:
            break
    if position < 0:
        return text_value[:width]
    start = max(0, position - width // 3)
    return text_value[start : start + width]


def search_extracted_text(
    connection: Connection | Engine,
    query: str,
    job_id: str | None = None,
    document_type: str | None = None,
    limit: int = 20,
) -> list[dict[str, Any]]:
    manager = connection.connect() if isinstance(connection, Engine) else None
    conn = manager.__enter__() if manager else connection
    try:
        if not document_content_table_available(conn):
            return []
        tokens = tokenize_search_text(query)
        params: dict[str, Any] = {"limit": limit}
        where: list[str] = []
        if job_id:
            where.append("c.job_id = :job_id")
            params["job_id"] = str(job_id)
        if document_type and document_type != "all":
            where.append("d.document_type = :document_type")
            params["document_type"] = document_type
        for index, token in enumerate(tokens):
            key = f"token_{index}"
            where.append("c.normalized_text LIKE :" + key)
            params[key] = f"%{token.lower()}%"
        where_sql = "WHERE " + " AND ".join(where) if where else ""
        rows = conn.execute(
            text(
                f"""
                SELECT c.document_id, c.job_id, d.file_name, d.document_type, d.sharepoint_url,
                       c.content_type, c.source_locator, c.page_number, c.sheet_name, c.row_number,
                       c.text_content
                FROM document_content c
                JOIN documents d ON d.document_id = c.document_id
                {where_sql}
                ORDER BY d.file_name, c.page_number NULLS LAST, c.sheet_name NULLS LAST, c.row_number NULLS LAST
                LIMIT :limit
                """
            ),
            params,
        ).mappings()
        return [{**dict(row), "excerpt": excerpt(str(row.get("text_content") or ""), tokens)} for row in rows]
    finally:
        if manager:
            manager.__exit__(None, None, None)


def print_status(connection: Connection) -> None:
    if not documents_table_available(connection):
        print("Documents table not found. Apply db/add_documents_table.sql first.")
        return
    rows = connection.execute(
        text(
            """
            SELECT COALESCE(extraction_status, 'not_started') AS status, COUNT(*) AS count
            FROM documents
            GROUP BY COALESCE(extraction_status, 'not_started')
            ORDER BY status
            """
        )
    ).fetchall()
    print("Document extraction status:")
    for status, count in rows:
        print(f"  {status}: {count}")


def print_identifier_status(connection: Connection) -> None:
    if not documents_table_available(connection):
        print("Documents table not found. Apply db/add_documents_table.sql first.")
        return
    row = connection.execute(
        text(
            """
            SELECT
                COUNT(*) AS total,
                COUNT(*) FILTER (WHERE drive_id IS NOT NULL AND drive_id <> '' AND drive_item_id IS NOT NULL AND drive_item_id <> '') AS with_identifiers,
                COUNT(*) FILTER (WHERE drive_id IS NULL OR drive_id = '' OR drive_item_id IS NULL OR drive_item_id = '') AS missing_identifiers
            FROM documents
            """
        )
    ).mappings().one()
    print("Document acquisition identifiers:")
    print(f"  documents_total: {row['total']}")
    print(f"  documents_with_drive_identifiers: {row['with_identifiers']}")
    print(f"  documents_missing_drive_identifiers: {row['missing_identifiers']}")


def load_documents_missing_drive_metadata(
    connection: Connection,
    *,
    job_id: str | None = None,
    limit: int = 10,
    document_type: str | None = None,
    job_spec_candidates: bool = False,
) -> list[dict[str, Any]]:
    params: dict[str, Any] = {"limit": limit}
    where = ["(drive_id IS NULL OR drive_id = '' OR drive_item_id IS NULL OR drive_item_id = '')"]
    if job_id:
        where.append("job_id = :job_id")
        params["job_id"] = job_id
    if document_type:
        where.append("document_type = :document_type")
        params["document_type"] = document_type
    if job_spec_candidates:
        where.append(
            """
            (
                document_type = ANY(:job_spec_note_document_types)
                OR LOWER(COALESCE(file_name, '') || ' ' || COALESCE(relative_path, '')) ~ :job_spec_candidate_pattern
                OR (
                    document_type = 'specification'
                    AND LOWER(COALESCE(file_name, '') || ' ' || COALESCE(relative_path, '')) !~ :job_spec_exclude_pattern
                )
            )
            """
        )
        params["job_spec_note_document_types"] = ["field_notes", "job_tracking", "site_notes"]
        params["job_spec_candidate_pattern"] = JOB_SPEC_CANDIDATE_PATTERN
        params["job_spec_exclude_pattern"] = JOB_SPEC_EXCLUDE_PATTERN
    sql = f"""
        SELECT *
        FROM documents
        WHERE {' AND '.join(where)}
        ORDER BY updated_at NULLS FIRST, file_name
        LIMIT :limit
    """
    return [dict(row) for row in connection.execute(text(sql), params).mappings().all()]


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    load_dotenv()
    parser = argparse.ArgumentParser(description="Extract source-aware searchable text from indexed Spray-Tec documents.")
    parser.add_argument("--document-id")
    parser.add_argument("--job-id")
    parser.add_argument("--pending", action="store_true")
    parser.add_argument("--failed", action="store_true", help="Retry documents currently marked failed.")
    parser.add_argument("--status", action="store_true")
    parser.add_argument("--identifier-status", action="store_true", help="Show document rows with and without Graph drive identifiers.")
    parser.add_argument("--backfill-metadata", action="store_true", help="Backfill drive identifiers from cached SharePoint manifests.")
    parser.add_argument("--resolve-metadata", action="store_true", help="Resolve missing drive identifiers from Graph by site/library/path without downloading content.")
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--document-type")
    parser.add_argument(
        "--job-spec-candidates",
        action="store_true",
        help="Limit extraction to job spec, scope, job tracking, and field-note candidate documents.",
    )
    parser.add_argument(
        "--estimator-relevant",
        action="store_true",
        help="Limit extraction to estimate/proposal/spec/job tracking/note/warranty/contract documents and matching filenames.",
    )
    parser.add_argument("--limit", type=int, default=10)
    parser.add_argument("--cache-root", type=Path, default=Path(".cache/sharepoint"))
    parser.add_argument("--site-url", help="SharePoint site URL for --resolve-metadata.")
    parser.add_argument("--library", default="Documents", help="SharePoint library name for --resolve-metadata.")
    parser.add_argument("--root-folder", default="", help="Optional library-relative root folder prepended to document relative_path for --resolve-metadata.")
    parser.add_argument("--database-url", default=os.getenv("DATABASE_URL") or os.getenv("NEON_DATABASE_URL"))
    parser.add_argument("--debug", action="store_true")
    parser.add_argument("--progress-every", type=int, default=1000, help="Print progress every N records for metadata backfill/extraction loops. Use 0 to silence progress.")
    parser.add_argument("--fail-fast", action="store_true", help="Stop the extraction batch on the first document-level transient failure.")
    parser.add_argument("--max-document-failures", type=int, default=0, help="Stop after N document-level transient failures. Use 0 for no failure cap.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    if not args.database_url:
        raise SystemExit("Set --database-url, DATABASE_URL, or NEON_DATABASE_URL.")
    engine = create_engine(args.database_url, future=True)
    if args.status:
        with engine.connect() as conn:
            print_status(conn)
        return 0
    if args.identifier_status:
        with engine.connect() as conn:
            print_identifier_status(conn)
        return 0
    if args.backfill_metadata:
        with engine.begin() as conn:
            updated = backfill_document_drive_metadata(
                conn,
                args.cache_root,
                limit=args.limit,
                job_id=args.job_id,
                progress_every=args.progress_every,
            )
        print(f"Drive identifiers backfilled from cached manifests: {updated}")
        return 0
    if args.resolve_metadata:
        if not args.site_url:
            raise SystemExit("--resolve-metadata requires --site-url.")
        client = GraphClient(max_retries=2)
        with engine.connect() as conn:
            documents = load_documents_missing_drive_metadata(
                conn,
                job_id=args.job_id,
                limit=args.limit,
                document_type=args.document_type,
                job_spec_candidates=args.job_spec_candidates,
            )
        resolved = 0
        failed = 0
        total = len(documents)
        for index, document in enumerate(documents, start=1):
            label = first_nonblank(document.get("file_name"), document.get("document_id"))
            try:
                metadata = resolve_graph_metadata_for_document(
                    client,
                    document,
                    site_url=args.site_url,
                    library=args.library,
                    root_folder=args.root_folder,
                )
                with engine.begin() as conn:
                    resolved += update_document_drive_metadata(conn, metadata)
                print(f"[{index}/{total}] {label} — identifiers resolved")
            except Exception as exc:
                failed += 1
                print(f"[{index}/{total}] {label} — metadata resolution failed: {str(exc)[:240]}")
        print(f"Identifiers resolved: {resolved}")
        print(f"Identifier resolution failures: {failed}")
        return 0
    if not any([args.document_id, args.job_id, args.pending, args.failed, args.job_spec_candidates, args.estimator_relevant]):
        raise SystemExit("Choose --document-id, --job-id, --pending, --failed, --job-spec-candidates, --estimator-relevant, or --status.")

    with engine.connect() as conn:
        documents = load_documents_for_extraction(
            conn,
            document_id=args.document_id,
            job_id=args.job_id,
            pending=args.pending,
            failed=args.failed,
            document_type=args.document_type,
            job_spec_candidates=args.job_spec_candidates,
            estimator_relevant=args.estimator_relevant,
            limit=args.limit,
        )
    total = len(documents)
    status_counts: dict[str, int] = {}
    transient_failures: list[str] = []
    for index, document in enumerate(documents, start=1):
        label = first_nonblank(document.get("file_name"), document.get("document_id"))
        try:
            status, count = extract_one_document_with_retry(engine, document, args.cache_root, force=args.force or args.failed)
        except TransientDocumentDatabaseError as exc:
            transient_failures.append(str(label))
            print(f"[{index}/{total}] {label} — database connection failure: {str(exc)[:240]}", flush=True)
            if args.fail_fast or (args.max_document_failures > 0 and len(transient_failures) >= args.max_document_failures):
                print("Stopping extraction because document failure limit was reached.", flush=True)
                return 1
            continue
        status_counts[status] = status_counts.get(status, 0) + 1
        acquisition_method = planned_acquisition_method(document, args.cache_root, force_download=args.force)
        print(f"[{index}/{total}] {label} — {status} {count} content rows — acquisition: {acquisition_method}")
        if args.debug:
            print(f"  document_id: {document.get('document_id')}")
            print(f"  sharepoint_url: {document.get('sharepoint_url') or '-'}")
    print("Document extraction run summary:")
    print(f"  documents_selected: {total}")
    for status, count in sorted(status_counts.items()):
        print(f"  {status}: {count}")
    print(f"  transient_document_failures: {len(transient_failures)}")
    if transient_failures:
        print("  first_failed_documents:")
        for label in transient_failures[:10]:
            print(f"    - {label}")
    return 1 if transient_failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
