from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from jobscan.estimator.decision_graph import (
    build_decision_graph,
    build_graphs_from_files,
    main as decision_graph_main,
    write_decision_graph_outputs,
)
from jobscan.estimator.template_intelligence import extract_template_intelligence


INSULATION_TEMPLATE = Path("templates/Estimate Insulation - McCall Residence (Walls Only) (1).xlsx")
ROOFING_TEMPLATE = Path("templates/Estimate FINAL- Recoat 15 YR.xlsx")


def _graphs() -> tuple[dict, dict]:
    insulation = build_decision_graph(extract_template_intelligence(INSULATION_TEMPLATE, template_type="insulation"))
    roofing = build_decision_graph(extract_template_intelligence(ROOFING_TEMPLATE, template_type="roofing"))
    return insulation, roofing


def _node(graph: dict, decision_id: str) -> dict:
    return next(row for row in graph["decision_nodes"] if row["decision_id"] == decision_id)


def test_decision_graph_collapses_repeated_material_rows() -> None:
    insulation, roofing = _graphs()

    foam = _node(insulation, "insulation_foam_system")
    assert foam["rows_controlled"] == [19, 20, 21]
    assert foam["category"] == "product_selection"
    assert {"foam_product_selector", "manufacturer", "density", "area_sqft", "thickness_inches", "yield_or_coverage", "unit_price"}.issubset(
        set(foam["input_fields"])
    )
    assert {"estimated_units", "estimated_sets", "estimated_cost"}.issubset(set(foam["computed_fields"]))

    thermal_barrier = _node(insulation, "insulation_thermal_barrier")
    assert thermal_barrier["rows_controlled"] == [30, 31, 32]
    assert thermal_barrier["category"] == "product_selection"
    assert {"coating_product_selector", "gal_per_100_sqft", "unit_price"}.issubset(set(thermal_barrier["input_fields"]))
    assert {"estimated_gallons", "estimated_cost"}.issubset(set(thermal_barrier["computed_fields"]))

    surface_areas = _node(insulation, "insulation_surface_areas")
    assert surface_areas["category"] == "scope_geometry"
    assert {"surface_type", "gross_area_sqft", "deduction_area_sqft", "net_area_sqft"}.issubset(
        set(surface_areas["input_fields"])
    )

    r_targets = _node(insulation, "insulation_r_value_targets")
    assert r_targets["category"] == "scope_requirement"
    assert {"surface_type", "target_r_value", "source_text"}.issubset(set(r_targets["input_fields"]))

    thickness = _node(insulation, "insulation_thickness_calculation")
    assert thickness["category"] == "formula_model"
    assert {"required_thickness_inches", "rounded_thickness_inches"}.issubset(set(thickness["computed_fields"]))

    coating = _node(roofing, "roofing_coating_system")
    assert coating["rows_controlled"] == [26, 27, 28]
    assert coating["category"] == "product_selection"
    assert {"coating_product_selector", "area_sqft", "gal_per_100_sqft", "wet_mils_estimate", "waste_factor_pct"}.issubset(
        set(coating["input_fields"])
    )
    assert {"estimated_gallons", "estimated_cost"}.issubset(set(coating["computed_fields"]))


def test_decision_graph_labor_nodes_preserve_crew_rate_decisions() -> None:
    insulation, roofing = _graphs()

    foam_labor = _node(insulation, "insulation_labor_foam")
    assert foam_labor["category"] == "labor_planning"
    assert {"labor_task", "days", "crew_size", "selected_daily_rate_cell", "hourly_rate"}.issubset(set(foam_labor["input_fields"]))
    assert {"total_hours", "daily_rate", "calculated_cost"}.issubset(set(foam_labor["computed_fields"]))
    assert "People!B11" in foam_labor["dependencies"]

    prep_labor = _node(roofing, "roofing_labor_prep")
    assert prep_labor["category"] == "labor_planning"
    assert {"crew_size", "selected_daily_rate_cell", "blended_rate"}.issubset(set(prep_labor["input_fields"]))
    assert {"total_hours", "daily_rate", "calculated_cost"}.issubset(set(prep_labor["computed_fields"]))


def test_decision_graph_preserves_selector_options_and_crew_size_choices() -> None:
    insulation, roofing = _graphs()
    selectors = insulation["selector_options"] + roofing["selector_options"]

    assert any(row["decision_id"] == "insulation_foam_system" and row["resolved_item_name"] == "Gaco 2.0 lb." for row in selectors)
    assert any(row["decision_id"] == "roofing_coating_system" and row["resolved_item_name"] == "Gaco Silicone" for row in selectors)
    assert any(row["decision_id"] == "insulation_thinner" and row["resolved_item_name"] == "Xylene" for row in selectors)
    assert any(row["decision_id"] == "insulation_caulk_sealant" and row["resolved_item_name"] == "Liquid Flashing" for row in selectors)
    assert any(row["decision_id"] == "insulation_lift_equipment" and row["resolved_item_name"] == "Forklift" for row in selectors)
    assert any(row["source_type"] == "people_daily_rate_selector" and row["selector_code"] == 3 for row in selectors)


def test_decision_graph_row_traceability_and_area_models() -> None:
    insulation, roofing = _graphs()

    trace = insulation["row_traceability"] + roofing["row_traceability"]
    assert any(row["decision_id"] == "insulation_foam_system" and row["sheet_name"] == "Estimate" and row["row_number"] == 19 for row in trace)
    assert any(row["decision_id"] == "roofing_coating_system" and row["sheet_name"] == "Estimate" and row["row_number"] == 26 for row in trace)
    assert any(row["decision_id"] == "insulation_area_takeoff" and row["sheet_name"] == "Sq Ft Calculation" for row in trace)

    area_models = insulation["area_calculation_models"]
    assert any(row["area_type"] == "wall" and row["model"] == "height_width_area" for row in area_models)
    assert any(row["area_type"] == "gable" and row["model"] == "triangle_height_width_area" for row in area_models)


def test_decision_graph_json_and_excel_outputs_write_successfully(tmp_path) -> None:
    insulation, roofing = _graphs()

    json_paths, xlsx_path = write_decision_graph_outputs(insulation, roofing, tmp_path)

    assert {path.name for path in json_paths} == {"template_decision_graph_insulation.json", "template_decision_graph_roofing.json"}
    assert all(path.exists() for path in json_paths)
    assert xlsx_path.exists()
    assert json.loads((tmp_path / "template_decision_graph_insulation.json").read_text())["template_type"] == "insulation"

    workbook = load_workbook(xlsx_path, read_only=True)
    assert {
        "decision_nodes",
        "selector_options",
        "formula_models",
        "row_traceability",
        "crew_rate_options",
        "area_calculation_models",
    }.issubset(set(workbook.sheetnames))


def test_decision_graph_cli_resolves_common_input_paths(tmp_path) -> None:
    insulation_intelligence = extract_template_intelligence(INSULATION_TEMPLATE, template_type="insulation")
    roofing_intelligence = extract_template_intelligence(ROOFING_TEMPLATE, template_type="roofing")
    insulation_path = tmp_path / "template_intelligence_insulation.json"
    roofing_path = tmp_path / "template_intelligence_roofing.json"
    insulation_path.write_text(json.dumps(insulation_intelligence, default=str), encoding="utf-8")
    roofing_path.write_text(json.dumps(roofing_intelligence, default=str), encoding="utf-8")

    insulation_graph, roofing_graph = build_graphs_from_files(insulation_path, roofing_path)
    assert _node(insulation_graph, "insulation_foam_system")["rows_controlled"] == [19, 20, 21]
    assert _node(roofing_graph, "roofing_coating_system")["rows_controlled"] == [26, 27, 28]

    result = decision_graph_main(
        [
            "--insulation",
            str(insulation_path),
            "--roofing",
            str(roofing_path),
            "--out-dir",
            str(tmp_path / "out"),
        ]
    )
    assert result == 0
    assert (tmp_path / "out" / "template_decision_graph_insulation.json").exists()
    assert (tmp_path / "out" / "template_decision_graph_roofing.json").exists()
    assert (tmp_path / "out" / "template_decision_graph_summary.xlsx").exists()
