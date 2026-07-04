from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from jobscan.estimator import template_rows as tr
from jobscan.estimator.schemas import EstimateRecommendation, EstimatorData
from jobscan.estimator.template_intelligence import (
    extract_template_intelligence,
    main as template_intelligence_main,
)
from jobscan.estimator.workbench import build_estimating_workbench


TEMPLATE_PATH = Path("templates/Estimate Insulation - McCall Residence (Walls Only) (1).xlsx")
ROOFING_TEMPLATE_PATH = Path("templates/Estimate FINAL- Recoat 15 YR.xlsx")


def test_insulation_template_intelligence_extracts_selector_maps_and_formula_models() -> None:
    intelligence = extract_template_intelligence(TEMPLATE_PATH)
    selector_rows = intelligence["selector_maps"]
    row19_choices = {
        str(row["selector_code"]): row["resolved_item_name"]
        for row in selector_rows
        if row["row_number"] == 19
    }

    assert row19_choices == {
        "11": "Gaco 2.0 lb.",
        "12": "Gaco 0.5 lb.",
        "21": "NCFI 2.0 lb.",
        "22": "NCFI 0.5 lb.",
        "31": "BASF 2.0 lb.",
        "32": "BASF 0.5 lb.",
        "41": "PSI 2.0 lb.",
        "42": "PSI 0.5 lb.",
        "51": "Demilec 2.0 lb.",
        "52": "Demilec 0.5 lb.",
    }

    materials = {row["row_number"]: row for row in intelligence["materials_rows"]}
    foam = materials[19]
    assert foam["selector_code"] == 11
    assert foam["selector_cell_role"] == "selector_code"
    assert foam["area_sqft"] == 2800
    assert foam["thickness_inches"] == 4.25
    assert foam["yield_or_coverage"] == 13500
    assert foam["yield_factor"] == 13500
    assert round(foam["estimated_units"], 6) == 881.481481
    assert round(foam["estimated_sets"], 6) == 0.881481
    assert foam["foam_brand"] == "Gaco"
    assert foam["foam_density_lb"] == 2.0
    assert foam["formula_model"] == "foam_sets_from_area_thickness_yield"

    formula_models = {(row["row_number"], row["template_bucket"]): row for row in intelligence["formula_models"]}
    assert formula_models[(19, "foam")]["formula_model"] == "foam_sets_from_area_thickness_yield"
    assert formula_models[(30, "thermal_barrier_coating")]["formula_model"] == "coating_gallons_from_area_rate_waste"
    assert formula_models[(30, "thermal_barrier_coating")]["waste_margin_cell"] == "A34"

    coating = materials[30]
    assert coating["gal_per_100_sqft"] == 1.5
    assert coating["formula_model"] == "coating_gallons_from_area_rate_waste"

    thinner_choices = {row["selector_code"]: row["resolved_item_name"] for row in selector_rows if row["row_number"] == 37}
    assert thinner_choices == {"1": "Naphtha VM&P", "2": "Mineral Spirits", "3": "Xylene"}
    assert {row["selector_code"] for row in selector_rows if row["row_number"] == 41} == {"1", "2", "3"}
    assert {row["selector_code"] for row in selector_rows if row["row_number"] == 47} == {"1", "2", "3", "4"}

    sqft_rows = {row["row_number"]: row for row in intelligence["sq_ft_calculation"]}
    assert sqft_rows[4]["model"] == "height_width_area"
    assert sqft_rows[9]["area_type"] == "ceiling"
    assert sqft_rows[13]["model"] == "triangle_height_width_area"
    assert sqft_rows[15]["final_total_cell"] == "F15"

    labor_rows = {row["row_number"]: row for row in intelligence["labor_rows"]}
    foam_labor = labor_rows[86]
    assert foam_labor["template_bucket"] == "labor_foam"
    assert foam_labor["formula_model"] == "labor_cost_from_days_crew_rate"
    assert foam_labor["formula_mode"] == "mixed_formula"
    assert foam_labor["days_cell"] == "B86"
    assert foam_labor["crew_selector_cell"] == "C86"
    assert foam_labor["crew_person_selector_code"] == 3
    assert foam_labor["selected_daily_rate_cell"] == "People!F12"
    assert foam_labor["daily_rate_cell"] == "J86"
    assert foam_labor["total_hours_cell"] == "D86"
    assert foam_labor["hourly_rate_cell"] == "G86"
    assert foam_labor["cost_formula"] == "=IF(G86=0, B86*J86, D86*G86)"
    assert "People!B11" in foam_labor["formula_dependencies"]

    loading_labor = labor_rows[95]
    assert loading_labor["template_bucket"] == "labor_loading"
    assert loading_labor["formula_model"] == "loading_cost_from_hours_people_rate_trip_count"
    assert loading_labor["formula_mode"] == "hours_based"
    assert loading_labor["total_hours_cell"] == "C95"
    assert loading_labor["cost_formula"].startswith("=IF(E95=1")

    meals_lodging = labor_rows[100]
    assert meals_lodging["template_bucket"] == "meals_lodging"
    assert meals_lodging["formula_model"] == "meals_lodging_cost_from_days_people_daily_amount"
    assert meals_lodging["formula_mode"] == "days_based"

    selector_three = next(row for row in intelligence["people_rate_table"] if row["selector_code"] == 3)
    assert selector_three["daily_rate_cell"] == "People!F12"
    assert selector_three["daily_rate_formula"] == "=SUM(F3:F10)*B11"

    people_refs = intelligence["people_labor_references"]
    assert any(row.get("source_type") == "people_daily_rate_selector" and row.get("selector_code") == 3 for row in people_refs)
    assert any(
        row.get("source_type") == "labor_row_formula_reference"
        and row.get("row_number") == 86
        and row.get("formula_mode") == "mixed_formula"
        and row.get("selected_daily_rate_cell") == "People!F12"
        for row in people_refs
    )


def test_template_intelligence_cli_writes_json_and_workbook(tmp_path) -> None:
    output = tmp_path / "insulation_template_intelligence.json"

    result = template_intelligence_main(["--template", str(TEMPLATE_PATH), "--out", str(output)])

    assert result == 0
    assert output.exists()
    assert output.with_suffix(".xlsx").exists()


def test_roofing_template_intelligence_extracts_workbook_engine() -> None:
    intelligence = extract_template_intelligence(ROOFING_TEMPLATE_PATH, template_type="roofing")

    assert intelligence["template_type"] == "roofing"
    assert {row["sheet_name"] for row in intelligence["workbook_sheets"]} >= {"Estimate", "People", "Materials"}

    selector_rows = intelligence["selector_maps"]
    coating_choices = {
        str(row["selector_code"]): row["resolved_item_name"]
        for row in selector_rows
        if row["row_number"] == 26
    }
    assert coating_choices["11"] == "Gaco Silicone"
    assert coating_choices["33"] == "AW Urethane"

    material_rows = {row["row_number"]: row for row in intelligence["material_rows"]}
    coating = material_rows[26]
    assert coating["template_bucket"] == "coating"
    assert coating["resolved_item_name"] == "Gaco Silicone"
    assert coating["formula_model"] == "coating_gallons_from_area_rate_waste"
    assert coating["gal_per_100_sqft"] == 1
    assert coating["waste_factor_cell"] == "A30"
    assert coating["waste_factor_pct"] == 15
    assert coating["wet_mils_estimate"] == 16
    assert "A30" in coating["formula_dependencies"]

    foam = material_rows[19]
    assert foam["formula_model"] == "foam_sets_from_area_thickness_yield"
    assert foam["selector_code"] == 11
    assert foam["area_sqft"] == 865
    assert foam["thickness_inches"] == 1.5
    assert foam["yield_or_coverage"] == 2600

    lookup_names = {row["table_name"] for row in intelligence["lookup_tables"]}
    assert {"fabric", "board", "fasteners", "crew_rate_matrix", "workers"}.issubset(lookup_names)

    labor_rows = {row["row_number"]: row for row in intelligence["labor_rows"]}
    setup = labor_rows[116]
    assert setup["template_bucket"] == "labor_prep"
    assert setup["formula_model"] == "labor_cost_from_days_crew_rate"
    assert setup["formula_mode"] == "mixed_formula"
    assert setup["days_cell"] == "B116"
    assert setup["crew_selector_cell"] == "C116"
    assert setup["crew_person_selector_code"] == 5
    assert setup["selected_daily_rate_cell"] == "People!H12"
    assert setup["daily_rate_cell"] == "J116"
    assert setup["total_hours_cell"] == "D116"
    assert setup["hourly_rate_cell"] == "G116"
    assert setup["cost_cell"] == "H116"
    assert setup["cost_formula"] == "=IF(G116=0, B116*J116, D116*G116)"
    assert "People!B11" in setup["formula_dependencies"]

    people_rate_rows = intelligence["people_rate_table"]
    selector_five = next(row for row in people_rate_rows if row["selector_code"] == 5)
    assert selector_five["daily_rate_cell"] == "People!H12"
    assert selector_five["daily_rate_formula"] == "=SUM(H3:H10)*B11"
    assert "H3:H10" in selector_five["formula_dependencies"]

    people_refs = intelligence["people_labor_references"]
    assert any(row.get("source_type") == "people_daily_rate_selector" and row.get("selector_code") == 5 for row in people_refs)
    assert any(
        row.get("source_type") == "labor_row_formula_reference"
        and row.get("row_number") == 116
        and row.get("formula_mode") == "mixed_formula"
        for row in people_refs
    )

    totals = {row["row_number"]: row for row in intelligence["totals_markups"]}
    assert totals[165]["template_bucket"] == "overhead"
    assert totals[165]["percentage"] == 35
    assert totals[167]["template_bucket"] == "profit"

    named_ranges = {row["name"]: row["attr_text"] for row in intelligence["named_ranges"]}
    assert named_ranges["Coating_Databas"] == "Estimate!#REF!"
    assert any(row["hidden_type"] == "column" and row["address"] == "J" for row in intelligence["hidden_tables"])
    assert any(row["product_name"] == "Gaco Silicone" for row in intelligence["pricing_product_references"])
    assert any(row["lookup_key"] == "Foreman / Sprayer" for row in intelligence["people_labor_references"])


def test_roofing_template_intelligence_cli_writes_requested_sheets(tmp_path) -> None:
    output = tmp_path / "roofing_template_intelligence.json"

    result = template_intelligence_main(
        ["--template", str(ROOFING_TEMPLATE_PATH), "--template-type", "roofing", "--out", str(output)]
    )

    assert result == 0
    assert output.exists()
    workbook = load_workbook(output.with_suffix(".xlsx"), read_only=True)
    assert workbook.sheetnames[:13] == [
        "Workbook Sheets",
        "Selector Maps",
        "Lookup Tables",
        "Material Rows",
        "Labor Rows",
        "Formula Models",
        "Hidden Tables",
        "Named Ranges",
        "Workbook Row Catalog",
        "Pricing Product References",
        "People Labor References",
        "People Rate Table",
        "Totals Markups",
    ]
    assert {"Coating Decisions", "Labor Decisions", "Markup Decisions"}.issubset(set(workbook.sheetnames))
    assert "People Rate Table" in workbook.sheetnames


def test_insulation_template_parser_preserves_formula_roles() -> None:
    parsed = tr.parse_document_content_row(
        {
            "document_id": "DOCINTEL",
            "job_id": "JOBINTEL",
            "source_file": "Estimate Insulation - Test.xlsx",
            "sheet_name": "Estimate",
            "row_number": 19,
            "cell_range": "A19:H19",
            "text_content": "A19: 11 | B19: Gaco 2.0 lb. | C19: 2800 | D19: 4.25 | E19: 1.63 | F19: 13500 | G19: 881.48 | H19: 1436.81",
        },
        template_type="insulation",
    )

    assert parsed is not None
    assert parsed["template_bucket"] == "foam"
    assert parsed["selector_code"] == 11
    assert parsed["quantity"] == 2800
    assert parsed["quantity_cell_role"] == "area_sqft"
    assert parsed["area_sqft"] == 2800
    assert parsed["thickness_inches"] == 4.25
    assert parsed["yield_or_coverage"] == 13500
    assert parsed["yield_factor"] == 13500
    assert parsed["estimated_units"] == 881.48
    assert parsed["estimated_sets"] == 0.88148
    assert parsed["foam_brand"] == "Gaco"
    assert parsed["foam_density_lb"] == 2.0
    assert parsed["estimated_cost"] == 1436.81
    assert parsed["formula_model"] == "foam_sets_from_area_thickness_yield"


def test_insulation_labor_parser_preserves_decision_formula_fields() -> None:
    parsed = tr.parse_document_content_row(
        {
            "document_id": "DOCLABORINTEL",
            "job_id": "JOBLABORINTEL",
            "source_file": "Estimate Insulation - Test.xlsx",
            "sheet_name": "Estimate",
            "row_number": 86,
            "cell_range": "A86:J86",
            "text_content": (
                "A86: Foam | B86: 1.5 | C86: 3 | "
                "D86: 45 | G86: 84.6 | H86: 3807 | J86: 2538"
            ),
        },
        template_type="insulation",
    )

    assert parsed is not None
    assert parsed["template_bucket"] == "labor_foam"
    assert parsed["days"] == 1.5
    assert parsed["crew_size"] == 3
    assert parsed["crew_selector_code"] == 3
    assert parsed["total_hours"] == 45
    assert parsed["hourly_rate"] == 84.6
    assert parsed["estimated_cost"] == 3807
    assert parsed["calculated_cost"] == 3807
    assert parsed["daily_rate"] == 2538
    assert parsed["formula_mode"] == "mixed_formula"


def test_insulation_dc315_parser_preserves_gallon_model() -> None:
    parsed = tr.parse_document_content_row(
        {
            "document_id": "DOCDC315",
            "job_id": "JOBDC315",
            "source_file": "Estimate Insulation - Test.xlsx",
            "sheet_name": "Estimate",
            "row_number": 30,
            "cell_range": "A30:H30",
            "text_content": "A30: 1 | B30: DC 315 TB | C30: 2400 | D30: 1.5 | E30: 52 | G30: 45 | H30: 2340",
        },
        template_type="insulation",
    )

    assert parsed is not None
    assert parsed["template_bucket"] == "thermal_barrier_coating"
    assert parsed["selector_code"] == 1
    assert parsed["area_sqft"] == 2400
    assert parsed["gal_per_100_sqft"] == 1.5
    assert parsed["estimated_gallons"] == 45
    assert parsed["waste_margin_cell"] == "A34"
    assert parsed["formula_model"] == "coating_gallons_from_area_rate_waste"


def test_insulation_workbench_uses_foam_sets_area_thickness_model() -> None:
    recommendation = EstimateRecommendation(
        parsed_fields={
            "division": "Insulation",
            "template_type": "insulation",
            "project_type": "spray foam insulation",
            "estimated_sqft": 1000,
            "net_insulation_area_sqft": 1000,
            "foam_thickness_inches": 3,
            "notes": "Spray foam insulation for walls.",
        },
        recommended_scope=[],
        material_plan=[{"category": "foam", "package": "foam", "included_in_total": True}],
        labor_plan=[],
        travel_plan={},
        historical_calibration={},
        similar_examples=[],
        estimate_low=None,
        estimate_target=None,
        estimate_high=None,
        review_flags=[],
        human_review_required=True,
        draft_workbook_inputs={"header": {"C12_estimated_sqft": 1000}},
    )
    data = EstimatorData(
        template_rows=pd.DataFrame(
            [
                {
                    "job_id": "F1",
                    "source_file": "insulation.xlsx",
                    "template_type": "insulation",
                    "template_bucket": "foam",
                    "row_number": 19,
                    "area_sqft": 1000,
                    "thickness_inches": 2,
                    "estimated_units": 4,
                    "estimated_cost": 400,
                    "yield_or_coverage": 500000,
                }
            ]
        )
    )

    workbench = build_estimating_workbench(recommendation, data)
    foam = workbench["insulation_foam_template_decisions"][0]

    assert "materials" not in workbench
    assert foam["formula_model"] == "foam_sets_from_area_thickness_yield"
    assert foam["workbook_row"] == "19-21"
    assert foam["basis_sqft"] == 1000
    assert foam["calculated_output_summary"].startswith("units=")
