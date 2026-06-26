from __future__ import annotations

import json
import math
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4

import pandas as pd
from openpyxl import load_workbook

from jobscan.estimator.calibration_audit import (
    AUDIT_SHEETS,
    build_calibration_audit,
    write_calibration_audit,
)
from jobscan.estimator.field_estimator import estimate_from_field_notes
from jobscan.estimator.schemas import EstimatorData
from test_field_estimator import field_data


SAMPLE_NOTE = (
    "Roof coating estimate. Metal roof in Louisville KY. Main roof is 120 ft by 80 ft. "
    "Deduct two skylight areas, each 4 ft by 8 ft. Roof condition is fair with some rusted fasteners. "
    "Customer wants a 10-year silicone coating system. Access is easy. Few penetrations."
)


def test_calibration_audit_writes_json_and_xlsx(tmp_path: Path) -> None:
    data = field_data()
    recommendation = estimate_from_field_notes(SAMPLE_NOTE, {"estimated_sqft": 0}, data=data)

    audit = build_calibration_audit(recommendation, data, notes=SAMPLE_NOTE, case_id="sample_case")
    paths = write_calibration_audit(audit, tmp_path, case_id="sample_case")

    assert paths["json"] == tmp_path / "sample_case_audit.json"
    assert paths["xlsx"] == tmp_path / "sample_case_audit.xlsx"
    parsed = json.loads(paths["json"].read_text(encoding="utf-8"))
    assert parsed["case_id"] == "sample_case"
    assert parsed["sheets"]["material_audit"]
    assert parsed["sheets"]["labor_audit"]
    workbook = load_workbook(paths["xlsx"], read_only=True)
    assert set(AUDIT_SHEETS).issubset(set(workbook.sheetnames))


def test_material_audit_flags_historical_cost_ratio_when_physical_quantity_and_current_price_exist() -> None:
    data = EstimatorData(
        jobs=pd.DataFrame([{"job_id": "J1", "estimated_sqft": 10000}]),
        template_rows=pd.DataFrame(
            [
                {
                    "job_id": "J1",
                    "template_bucket": "primer",
                    "line_item_kind": "material",
                    "row_label": "Epoxy Primer",
                    "quantity": 20,
                    "unit": "pail",
                    "estimated_cost": 3000,
                }
            ]
        ),
        pricing_catalog=pd.DataFrame(
            [
                {
                    "pricing_item_id": "P1",
                    "product_name": "Epoxy Primer",
                    "category": "Primer",
                    "unit_price": 150,
                    "is_current": True,
                    "needs_review": False,
                }
            ]
        ),
    )
    recommendation = SimpleNamespace(
        parsed_fields={"estimated_sqft": 10000, "project_type": "roof coating"},
        recommended_scope=[],
        material_plan=[
            {
                "item": "Primer allowance",
                "category": "primer",
                "quantity": 10000,
                "unit": "sqft",
                "unit_price": 0.35,
                "estimated_cost": 3500,
                "selected_price_source": "historical_cost_ratio_fallback",
                "calibration_method": "historical_cost_ratio_fallback",
                "needs_review": True,
            }
        ],
        labor_plan=[],
        travel_plan={},
        historical_calibration={},
        similar_examples=[],
        estimate_low=0,
        estimate_target=0,
        estimate_high=0,
        review_flags=[],
        human_review_required=True,
        draft_workbook_inputs={"header": {"C12_estimated_sqft": 10000}},
        debug={},
    )

    audit = build_calibration_audit(recommendation, data, notes="test", case_id="primer_case")
    primer_audit = [row for row in audit["sheets"]["material_audit"] if row.get("package") == "primer"]

    assert primer_audit
    assert primer_audit[0]["status"] == "FAIL"
    assert primer_audit[0]["issue"] == "historical_cost_ratio_used_despite_physical_quantity_and_current_pricing"
    assert primer_audit[0]["physical_quantity_evidence_count"] >= 1
    assert primer_audit[0]["current_pricing_match_count"] >= 1


def test_labor_audit_flags_rule_based_fallback_when_historical_rows_exist() -> None:
    data = EstimatorData(
        jobs=pd.DataFrame([{"job_id": "J1", "estimated_sqft": 9536, "project_type": "roof coating"}]),
        template_rows=pd.DataFrame(
            [
                {
                    "job_id": "J1",
                    "template_bucket": "labor_prep",
                    "line_item_kind": "labor",
                    "days": 2,
                    "crew_size": 4,
                    "total_hours": 64,
                    "estimated_cost": 4200,
                }
            ]
        ),
    )
    recommendation = SimpleNamespace(
        parsed_fields={"estimated_sqft": 9536, "surface_area_sqft": 9536, "project_type": "roof coating", "substrate": "metal", "coating_type": "silicone"},
        recommended_scope=["roof coating"],
        material_plan=[],
        labor_plan=[
            {
                "task": "labor_prep",
                "crew_size": 4,
                "total_hours": 40,
                "estimated_cost": 2880,
                "evidence_count": 0,
                "calibration_method": "rule_based_fallback",
                "needs_review": True,
            }
        ],
        travel_plan={},
        historical_calibration={},
        similar_examples=[],
        estimate_low=0,
        estimate_target=0,
        estimate_high=0,
        review_flags=[],
        human_review_required=True,
        draft_workbook_inputs={"header": {"C12_estimated_sqft": 9536}},
        debug={},
    )

    audit = build_calibration_audit(recommendation, data, notes="test", case_id="labor_case")
    prep_audit = [row for row in audit["sheets"]["labor_audit"] if row.get("task") == "labor_prep"]

    assert prep_audit
    assert prep_audit[0]["status"] == "FAIL"
    assert prep_audit[0]["issue"] == "fallback_used_despite_valid_historical_evidence"
    assert prep_audit[0]["valid_historical_evidence_count"] >= 1


def test_calibration_audit_export_sanitizes_complex_values(tmp_path: Path) -> None:
    aware_datetime = datetime(2026, 6, 26, 14, 0, tzinfo=UTC)
    row_uuid = uuid4()
    recommendation = SimpleNamespace(
        parsed_fields={
            "estimated_sqft": 10000,
            "project_type": "roof coating",
            "generated_at": aware_datetime,
            "timestamp": pd.Timestamp("2026-06-26T09:15:00", tz="America/New_York"),
            "uuid": row_uuid,
            "bad_number": math.nan,
            "nested": {"amount": Decimal("12.50"), "items": [aware_datetime, math.inf]},
        },
        recommended_scope=[],
        material_plan=[{"item": "Coating", "category": "coating", "estimated_cost": Decimal("1200.25"), "created_at": aware_datetime}],
        labor_plan=[],
        travel_plan={},
        historical_calibration={},
        similar_examples=[],
        estimate_low=0,
        estimate_target=0,
        estimate_high=0,
        review_flags=[],
        human_review_required=False,
        draft_workbook_inputs={"header": {"C12_estimated_sqft": 10000, "created_at": aware_datetime}},
        debug={},
    )

    audit = build_calibration_audit(recommendation, EstimatorData(), notes="test", case_id="timezone_case")
    paths = write_calibration_audit(audit, tmp_path, case_id="timezone_case")

    parsed = json.loads(paths["json"].read_text(encoding="utf-8"))
    assert parsed["sheets"]["parsed_scope"][0]["bad_number"] is None
    assert parsed["sheets"]["parsed_scope"][0]["nested"]["items"][1] is None
    workbook = load_workbook(paths["xlsx"], read_only=True, data_only=True)
    assert "parsed_scope" in workbook.sheetnames


def test_similar_jobs_audit_classifies_weak_strong_and_outliers() -> None:
    recommendation = SimpleNamespace(
        parsed_fields={"estimated_sqft": 10000, "project_type": "roof coating"},
        recommended_scope=[],
        material_plan=[],
        labor_plan=[],
        travel_plan={},
        historical_calibration={},
        similar_examples=[
            {
                "job_id": "strong",
                "job_name": "Metal roof coating",
                "job_type": "roof coating",
                "price_per_sqft": 8,
                "reason_matched": "same substrate and coating package",
            },
            {
                "job_id": "weak",
                "job_name": "Same city",
                "division": "Roofing",
                "price_per_sqft": 90,
                "reason_matched": "same division and city",
            },
        ],
        estimate_low=0,
        estimate_target=0,
        estimate_high=0,
        review_flags=[],
        human_review_required=False,
        draft_workbook_inputs={"header": {"C12_estimated_sqft": 10000}},
        debug={},
    )

    audit = build_calibration_audit(recommendation, EstimatorData(), notes="test", case_id="similar_case")
    rows = audit["sheets"]["similar_jobs_audit"]
    by_job = {row.get("job_id"): row for row in rows}

    assert by_job["strong"]["match_strength"] == "strong"
    assert by_job["weak"]["match_strength"] == "weak"
    assert by_job["weak"]["outlier_flag"] is True
