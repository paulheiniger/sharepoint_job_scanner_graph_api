from __future__ import annotations

import json

import pandas as pd
from sqlalchemy import create_engine, text

from jobscan.estimator import template_rows as tr


def content_row(row_number: int, text_content: str, **kwargs):
    row = {
        "document_id": "DOC1",
        "job_id": "JOB1",
        "source_file": "Estimate.xlsx",
        "sheet_name": "Estimate",
        "row_number": row_number,
        "cell_range": f"A{row_number}:J{row_number}",
        "text_content": text_content,
    }
    row.update(kwargs)
    return row


def create_sqlite_schema(engine) -> None:
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE documents (
                    document_id TEXT PRIMARY KEY,
                    job_id TEXT,
                    file_name TEXT,
                    file_extension TEXT,
                    document_type TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE document_content (
                    content_id TEXT PRIMARY KEY,
                    document_id TEXT,
                    job_id TEXT,
                    sheet_name TEXT,
                    row_number INTEGER,
                    cell_range TEXT,
                    text_content TEXT
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE estimate_template_rows (
                    template_row_id TEXT PRIMARY KEY,
                    document_id TEXT NOT NULL,
                    job_id TEXT,
                    source_file TEXT,
                    template_type TEXT,
                    sheet_name TEXT,
                    row_number INTEGER,
                    cell_range TEXT,
                    template_bucket TEXT,
                    template_section TEXT,
                    line_item_kind TEXT,
                    row_label TEXT,
                    raw_text TEXT,
                    cell_values TEXT,
                    formula_cells TEXT,
                    selected_item_name TEXT,
                    quantity NUMERIC,
                    unit TEXT,
                    unit_price NUMERIC,
                    estimated_units NUMERIC,
                    estimated_cost NUMERIC,
                    selector_code NUMERIC,
                    resolved_item_name TEXT,
                    area_sqft NUMERIC,
                    thickness_inches NUMERIC,
                    yield_or_coverage NUMERIC,
                    yield_factor NUMERIC,
                    estimated_sets NUMERIC,
                    foam_brand TEXT,
                    foam_density_lb NUMERIC,
                    units_per_sqft_per_inch NUMERIC,
                    sets_per_sqft_per_inch NUMERIC,
                    cost_per_sqft_per_inch NUMERIC,
                    gal_per_100_sqft NUMERIC,
                    gal_per_sqft NUMERIC,
                    estimated_gallons NUMERIC,
                    linear_ft NUMERIC,
                    ft_per_unit NUMERIC,
                    margin_pct NUMERIC,
                    waste_margin_cell TEXT,
                    quantity_cell_role TEXT,
                    formula_model TEXT,
                    days NUMERIC,
                    crew_size NUMERIC,
                    total_hours NUMERIC,
                    daily_rate NUMERIC,
                    crew_selector_code NUMERIC,
                    hourly_rate NUMERIC,
                    calculated_cost NUMERIC,
                    formula_mode TEXT,
                    trips NUMERIC,
                    round_trip_miles NUMERIC,
                    cost_per_mile NUMERIC,
                    warranty_years NUMERIC,
                    overhead_pct NUMERIC,
                    profit_pct NUMERIC,
                    parsed_confidence NUMERIC,
                    needs_review BOOLEAN DEFAULT FALSE,
                    parser_version TEXT,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        )


def test_parse_cell_labeled_text_numeric_and_formulas() -> None:
    cell_values, formula_cells, malformed_count = tr.parse_cell_labeled_text(
        "A26: 11 | B26: Gaco Silicone | D26: 1 | E26: 42 | G26: =C26*D26 | bad fragment"
    )

    assert cell_values == {"A26": 11, "B26": "Gaco Silicone", "D26": 1, "E26": 42}
    assert formula_cells == {"G26": "=C26*D26"}
    assert malformed_count == 1


def test_row_26_maps_to_coating_and_extracts_material_fields() -> None:
    parsed = tr.parse_document_content_row(
        content_row(26, "A26: 11 | B26: Gaco Silicone | C26: 150 | E26: 42 | G26: =formula | H26: 6300")
    )

    assert parsed["template_bucket"] == "coating"
    assert parsed["line_item_kind"] == "material"
    assert parsed["selected_item_name"] == "Gaco Silicone"
    assert parsed["quantity"] == 150
    assert parsed["unit_price"] == 42
    assert parsed["estimated_units"] is None
    assert parsed["estimated_cost"] == 6300
    assert parsed["formula_cells"] == {"G26": "=formula"}


def test_row_39_maps_to_primer() -> None:
    parsed = tr.parse_document_content_row(content_row(39, "A39: Primer | B39: Bleed Block | C39: 2 | E39: 125"))

    assert parsed["template_bucket"] == "primer"
    assert parsed["selected_item_name"] == "Bleed Block"


def test_flooring_material_rows_preserve_flooring_formula_inputs() -> None:
    parsed = tr.parse_document_content_row(
        content_row(
            26,
            "A26: 11 | B26: NPI Epoxy | C26: 2400 | D26: 1 | E26: 45 | F26: 707 - Black | G26: 26.4 | H26: 1188",
            source_file="Estimate Flooring - Lee Sporting Shop.xlsx",
        ),
        template_type=tr.TEMPLATE_TYPE_FLOORING,
    )

    assert parsed["template_type"] == "flooring"
    assert parsed["template_bucket"] == "floor_base_coat"
    assert parsed["line_item_kind"] == "material"
    assert parsed["area_sqft"] == 2400
    assert parsed["gal_per_100_sqft"] == 1
    assert parsed["estimated_gallons"] == 26.4
    assert parsed["estimated_units"] == 26.4
    assert parsed["formula_model"] == "floor_coating_gallons_from_area_rate_margin"


def test_flooring_labor_rows_preserve_mixed_formula_inputs() -> None:
    parsed = tr.parse_document_content_row(
        content_row(
            120,
            "A120: Prep & Base 707 | B120: 0.5 | C120: 3 | D120: 12 | H120: 2528.66 | J120: 1685.78",
            source_file="Estimate Flooring - Lee Sporting Shop.xlsx",
        ),
        template_type=tr.TEMPLATE_TYPE_FLOORING,
    )

    assert parsed["template_bucket"] == "labor_floor_prep_base"
    assert parsed["line_item_kind"] == "labor"
    assert parsed["days"] == 0.5
    assert parsed["crew_size"] == 3
    assert parsed["total_hours"] == 12
    assert parsed["daily_rate"] == 1685.78
    assert parsed["formula_mode"] == "mixed_formula"


def test_flooring_flake_adder_classifies_as_material() -> None:
    parsed = tr.parse_document_content_row(
        content_row(
            177,
            "A177: Additional Amount w/o Markup | G177: Flake (10) | H177: 1320",
            source_file="Estimate Flooring - Lee Sporting Shop.xlsx",
        ),
        template_type=tr.TEMPLATE_TYPE_FLOORING,
    )

    assert parsed["template_bucket"] == "floor_flake"
    assert parsed["line_item_kind"] == "material"
    assert parsed["estimated_cost"] == 1320


def test_row_106_sales_inspection_extracts_travel_fields() -> None:
    parsed = tr.parse_document_content_row(content_row(106, "A106: Sales/Inspection | B106: 2 | C106: 180 | E106: .75 | H106: 270"))

    assert parsed["template_bucket"] == "sales_inspection_trips"
    assert parsed["line_item_kind"] == "travel"
    assert parsed["trips"] == 2
    assert parsed["round_trip_miles"] == 180
    assert parsed["cost_per_mile"] == 0.75
    assert parsed["estimated_cost"] == 270


def test_row_108_truck_expense_extracts_travel_fields() -> None:
    parsed = tr.parse_document_content_row(content_row(108, "A108: Truck | B108: 3 | C108: 90 | E108: 0.75 | H108: 202.5"))

    assert parsed["template_bucket"] == "truck_expense"
    assert parsed["trips"] == 3
    assert parsed["estimated_cost"] == 202.5


def test_row_116_labor_prep_extracts_labor_fields() -> None:
    parsed = tr.parse_document_content_row(
        content_row(116, "A116: Pwash/Prep | B116: 4 | C116: 5 | D116: 220 | H116: 7607.6 | J116: 1901.9")
    )

    assert parsed["template_bucket"] == "labor_prep"
    assert parsed["line_item_kind"] == "labor"
    assert parsed["days"] == 4
    assert parsed["crew_size"] == 5
    assert parsed["total_hours"] == 220
    assert parsed["estimated_cost"] == 7607.6
    assert parsed["daily_rate"] == 1901.9


def test_normalize_roofing_labor_bucket_common_labels() -> None:
    cases = {
        "Set Up/Safety": "labor_loading",
        "Setup/Safety": "labor_loading",
        "Set-Up": "labor_loading",
        "PW/Prep": "labor_prep",
        "Pwash/Prep": "labor_prep",
        "Clean/Prep": "labor_prep",
        "Prep/Clean": "labor_prep",
        "Prep/Prime": "labor_prep",
        "PW/Prep/Prime": "labor_prep",
        "Prime": "labor_prime",
        "Primer": "labor_prime",
        "Prime Coat": "labor_prime",
        "Flash curbs": "labor_details",
        "Pitch Pockets": "labor_details",
        "Expansion Joints": "labor_details",
        "Touch/Clean Up": "labor_cleanup",
        "Foam/Base": "labor_base",
        "TO/Foam/Base": "labor_base",
        "Seam Sealer": "labor_seam_sealer",
        "Top Coat": "labor_top_coat",
        "Caulk": "labor_caulk",
    }

    for label, expected in cases.items():
        result = tr.normalize_roofing_labor_bucket(label)
        assert result["primary_bucket"] == expected, label


def test_roofing_labor_setup_safety_does_not_map_to_prep() -> None:
    parsed = tr.parse_document_content_row(
        content_row(116, "A116: Set Up/Safety | B116: 1 | C116: 4 | D116: 32 | H116: 2400")
    )

    assert parsed["template_bucket"] == "labor_loading"
    assert parsed["template_bucket"] != "labor_prep"
    assert parsed["line_item_kind"] == "labor"
    assert parsed["days"] == 1
    assert parsed["total_hours"] == 32


def test_roofing_composite_labor_labels_preserve_secondary_tags() -> None:
    prep_prime = tr.parse_document_content_row(content_row(116, "A116: PW/Prep/Prime | B116: 2 | C116: 4 | D116: 64 | H116: 4000"))
    foam_base = tr.parse_document_content_row(content_row(122, "A122: TO/Foam/Base | B122: 3 | C122: 4 | D122: 96 | H122: 7200"))

    assert prep_prime["template_bucket"] == "labor_prep"
    assert "labor_prime" in prep_prime["package_tags"]
    assert prep_prime["is_composite_label"] is True
    assert foam_base["template_bucket"] == "labor_base"


def test_row_139_labor_traveling() -> None:
    parsed = tr.parse_document_content_row(content_row(139, "A139: Traveling | C139: 16 | E139: 3 | G139: 72 | H139: 3456"))

    assert parsed["template_bucket"] == "labor_traveling"
    assert parsed["line_item_kind"] == "travel"
    assert parsed["days"] is None
    assert parsed["total_hours"] == 16
    assert parsed["crew_size"] == 3
    assert parsed["unit_price"] == 72


def test_insulation_loading_row_is_hours_based_not_days() -> None:
    parsed = tr.parse_document_content_row(
        content_row(95, "A95: Loading | C95: 0.5 | E95: 1 | G95: 25.5 | H95: 12.75", template_type="insulation")
    )

    assert parsed["template_bucket"] == "labor_loading"
    assert parsed["days"] is None
    assert parsed["total_hours"] == 0.5
    assert parsed["crew_size"] == 1
    assert parsed["unit_price"] == 25.5


def test_row_154_warranty_extracts_years() -> None:
    parsed = tr.parse_document_content_row(content_row(154, "A154: Warranty | C154: 20 | E154: 12000 | H154: 2400"))

    assert parsed["template_bucket"] == "warranty"
    assert parsed["warranty_years"] == 20
    assert parsed["quantity"] == 12000
    assert parsed["estimated_cost"] == 2400


def test_overhead_profit_and_worksheet_price_rows() -> None:
    overhead = tr.parse_document_content_row(content_row(165, "A165: Overhead | F165: 10% | H165: 5000"))
    profit = tr.parse_document_content_row(content_row(167, "A167: Profit | F167: 25 | H167: 12000"))
    price = tr.parse_document_content_row(content_row(169, "A169: Price | H169: 85000"))

    assert overhead["template_bucket"] == "overhead"
    assert overhead["overhead_pct"] == 10
    assert profit["template_bucket"] == "profit"
    assert profit["profit_pct"] == 25
    assert price["template_bucket"] == "worksheet_price"
    assert price["estimated_cost"] == 85000


def test_row_173_misc_materials_adder_uses_column_f_amount() -> None:
    parsed = tr.parse_document_content_row(content_row(173, "A173: Misc. Materials | F173: 3500"))

    assert parsed["template_bucket"] == "misc_materials"
    assert parsed["template_section"] == "estimate_adders"
    assert parsed["line_item_kind"] == "material"
    assert parsed["selected_item_name"] == "Misc. Materials"
    assert parsed["estimated_cost"] == 3500
    assert parsed["needs_review"] is False


def test_row_174_misc_insurance_adder() -> None:
    parsed = tr.parse_document_content_row(content_row(174, "A174: Misc. Insurance | F174: 1600"))

    assert parsed["template_bucket"] == "misc_insurance"
    assert parsed["template_section"] == "estimate_adders"
    assert parsed["line_item_kind"] == "insurance"
    assert parsed["estimated_cost"] == 1600
    assert parsed["needs_review"] is False


def test_row_175_lift_rental_adder() -> None:
    parsed = tr.parse_document_content_row(content_row(175, "A175: Lift Rental | F175: 3500"))

    assert parsed["template_bucket"] == "lift"
    assert parsed["template_section"] == "estimate_adders"
    assert parsed["line_item_kind"] == "equipment"
    assert parsed["estimated_cost"] == 3500


def test_placeholder_adder_without_amount_is_skipped() -> None:
    parsed = tr.parse_document_content_row(content_row(176, "A176: Additional Amount w/o Markup"))

    assert parsed is None


def test_placeholder_adder_with_h_cell_but_no_numeric_amount_is_skipped() -> None:
    parsed = tr.parse_document_content_row(content_row(176, "A176: Additional Amount w/o Markup | H176: =F176"))

    assert parsed is None


def test_placeholder_adder_label_in_other_cell_is_skipped() -> None:
    parsed = tr.parse_document_content_row(content_row(176, "B176: Additional Amount w/o Markup | F176: "))

    assert parsed is None


def test_placeholder_adder_rows_do_not_parse_as_unknown_review_rows() -> None:
    rows = tr.parse_document_content_rows(
        [
            content_row(173, "A173: Additional Amount w/o Markup"),
            content_row(174, "B174: Additional Amount w/o Markup | H174: =F174"),
            content_row(175, "A175: Misc. Materials | F175: 3500"),
        ]
    )

    assert len(rows) == 1
    assert rows[0]["template_bucket"] == "misc_materials"
    assert all(row["template_bucket"] != "unknown" for row in rows)
    assert all(row["needs_review"] is False for row in rows)


def test_custom_adder_with_amount_maps_to_generic_adder() -> None:
    parsed = tr.parse_document_content_row(content_row(177, "A177: Crane coordination | F177: 900"))

    assert parsed["template_bucket"] == "estimate_adder"
    assert parsed["template_section"] == "estimate_adders"
    assert parsed["line_item_kind"] == "other"
    assert parsed["estimated_cost"] == 900
    assert parsed["needs_review"] is False


def test_adder_formula_in_column_f_is_preserved() -> None:
    parsed = tr.parse_document_content_row(content_row(178, "A178: Misc. Equipment | F178: =SUM(F173:F177)"))

    assert parsed["template_bucket"] == "misc_equipment"
    assert parsed["line_item_kind"] == "equipment"
    assert parsed["estimated_cost"] is None
    assert parsed["formula_cells"] == {"F178": "=SUM(F173:F177)"}
    assert parsed["needs_review"] is True


def test_missing_and_malformed_cells_do_not_crash() -> None:
    parsed = tr.parse_document_content_row(content_row(999, "malformed only | A999: Mystery"))

    assert parsed["template_bucket"] == "unknown"
    assert parsed["needs_review"] is True
    assert parsed["cell_values"] == {"A999": "Mystery"}


def test_idempotent_upsert_preserves_document_content() -> None:
    engine = create_engine("sqlite:///:memory:", future=True)
    create_sqlite_schema(engine)
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO documents (document_id, job_id, file_name, file_extension, document_type)
                VALUES ('DOC1', 'JOB1', 'Estimate.xlsx', '.xlsx', 'estimate')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO document_content (
                    content_id, document_id, job_id, sheet_name, row_number, cell_range, text_content
                )
                VALUES (
                    'CONTENT1', 'DOC1', 'JOB1', 'Estimate', 116, 'A116:J116',
                    'A116: Pwash/Prep | B116: 4 | C116: 5 | D116: 220 | H116: 7607.6 | J116: 1901.9'
                )
                """
            )
        )

    first = tr.parse_existing_document_content(engine)
    second = tr.parse_existing_document_content(engine)

    assert first["rows_upserted"] == 1
    assert second["rows_upserted"] == 1
    with engine.connect() as conn:
        content_count = conn.execute(text("SELECT COUNT(*) FROM document_content")).scalar_one()
        template_count = conn.execute(text("SELECT COUNT(*) FROM estimate_template_rows")).scalar_one()
        raw_text = conn.execute(text("SELECT text_content FROM document_content WHERE content_id = 'CONTENT1'")).scalar_one()
    assert content_count == 1
    assert template_count == 1
    assert "Pwash/Prep" in raw_text


def test_parse_existing_persists_insulation_foam_formula_fields() -> None:
    engine = create_engine("sqlite:///:memory:", future=True)
    create_sqlite_schema(engine)
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO documents (document_id, job_id, file_name, file_extension, document_type)
                VALUES ('DOC1', 'JOB1', 'Estimate Insulation.xlsx', '.xlsx', 'estimate')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO document_content (
                    content_id, document_id, job_id, sheet_name, row_number, cell_range, text_content
                )
                VALUES (
                    'CONTENT1', 'DOC1', 'JOB1', 'Estimate', 19, 'A19:H19',
                    'A19: 12 | B19: Gaco 0.5 lb. Open Cell Spray Foam | C19: 2226 | D19: 3.68 | E19: 1.6 | F19: 4500 | G19: 1820.16 | H19: 2912.26'
                )
                """
            )
        )

    summary = tr.parse_existing_document_content(engine)

    assert summary["rows_upserted"] == 1
    with engine.connect() as conn:
        row = conn.execute(
            text(
                """
                SELECT template_type, template_bucket, area_sqft, thickness_inches, yield_or_coverage,
                       yield_factor, estimated_units, estimated_sets, unit_price, formula_model
                FROM estimate_template_rows
                WHERE document_id = 'DOC1'
                """
            )
        ).mappings().one()

    assert row["template_type"] == "insulation"
    assert row["template_bucket"] == "foam"
    assert row["area_sqft"] == 2226
    assert row["thickness_inches"] == 3.68
    assert row["yield_or_coverage"] == 4500
    assert row["yield_factor"] == 4500
    assert row["estimated_units"] == 1820.16
    assert row["estimated_sets"] == 1.82016
    assert row["unit_price"] == 1.6
    assert row["formula_model"] == "foam_sets_from_area_thickness_yield"


def test_parse_existing_persists_insulation_waste_margin_cell_reference() -> None:
    engine = create_engine("sqlite:///:memory:", future=True)
    create_sqlite_schema(engine)
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO documents (document_id, job_id, file_name, file_extension, document_type)
                VALUES ('DOC1', 'JOB1', 'Estimate Insulation.xlsx', '.xlsx', 'estimate')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO document_content (
                    content_id, document_id, job_id, sheet_name, row_number, cell_range, text_content
                )
                VALUES (
                    'CONTENT1', 'DOC1', 'JOB1', 'Estimate', 30, 'A30:H30',
                    'A30: 11 | B30: DC315 | C30: 1000 | D30: 1.25 | E30: 42 | G30: 12.5 | H30: 525'
                )
                """
            )
        )

    summary = tr.parse_existing_document_content(engine)

    assert summary["rows_upserted"] == 1
    with engine.connect() as conn:
        row = conn.execute(
            text(
                """
                SELECT template_bucket, waste_margin_cell, quantity_cell_role, formula_model
                FROM estimate_template_rows
                WHERE document_id = 'DOC1'
                """
            )
        ).mappings().one()

    assert row["template_bucket"] == "thermal_barrier_coating"
    assert row["waste_margin_cell"] == "A34"
    assert row["quantity_cell_role"] == "area_sqft"
    assert row["formula_model"] == "coating_gallons_from_area_rate_waste"


def test_repair_flooring_template_type_overwrites_misclassified_rows() -> None:
    engine = create_engine("sqlite:///:memory:", future=True)
    create_sqlite_schema(engine)
    row_id = tr.stable_template_row_id("DOCF", "Estimate", 26, "A26:H26")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO documents (document_id, job_id, file_name, file_extension, document_type)
                VALUES ('DOCF', 'JOBF', 'Estimate Flooring - Lee Sporting Shop.xlsx', '.xlsx', 'estimate')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO document_content (
                    content_id, document_id, job_id, sheet_name, row_number, cell_range, text_content
                )
                VALUES
                    (
                        'CONTENT1', 'DOCF', 'JOBF', 'Estimate', 3, 'A3:C3',
                        'A3: Job Type | C3: Floor System'
                    ),
                    (
                        'CONTENT2', 'DOCF', 'JOBF', 'Estimate', 26, 'A26:H26',
                        'A26: 11 | B26: NPI Epoxy | C26: 2400 | D26: 1 | E26: 45 | G26: 26.4 | H26: 1188'
                    ),
                    (
                        'CONTENT3', 'DOCF', 'JOBF', 'Estimate', 120, 'A120:J120',
                        'A120: Prep & Base 707 | B120: 0.5 | C120: 3 | D120: 12 | H120: 2528.66 | J120: 1685.78'
                    )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO estimate_template_rows (
                    template_row_id, document_id, job_id, source_file, template_type,
                    sheet_name, row_number, cell_range, template_bucket, line_item_kind,
                    selected_item_name, needs_review, parser_version
                )
                VALUES (
                    :row_id, 'DOCF', 'JOBF', 'Estimate Flooring - Lee Sporting Shop.xlsx', 'roofing',
                    'Estimate', 26, 'A26:H26', 'coating', 'material',
                    'NPI Epoxy', false, :parser_version
                )
                """
            ),
            {"row_id": row_id, "parser_version": tr.PARSER_VERSION},
        )

    summary = tr.repair_flooring_template_type(engine)

    assert summary["documents_considered"] == 1
    assert summary["rows_upserted"] == 3
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT row_number, template_type, template_bucket, line_item_kind, estimated_units, formula_model
                FROM estimate_template_rows
                WHERE document_id = 'DOCF'
                ORDER BY row_number
                """
            )
        ).mappings().all()
        count = conn.execute(
            text("SELECT COUNT(*) FROM estimate_template_rows WHERE template_row_id = :row_id"),
            {"row_id": row_id},
        ).scalar_one()

    assert count == 1
    by_row = {row["row_number"]: row for row in rows}
    assert by_row[26]["template_type"] == "flooring"
    assert by_row[26]["template_bucket"] == "floor_base_coat"
    assert by_row[26]["estimated_units"] == 26.4
    assert by_row[26]["formula_model"] == "floor_coating_gallons_from_area_rate_margin"
    assert by_row[120]["template_type"] == "flooring"
    assert by_row[120]["template_bucket"] == "labor_floor_prep_base"
    assert by_row[120]["line_item_kind"] == "labor"


def test_parse_existing_is_bounded_visible_and_xlsx_only(capsys) -> None:
    engine = create_engine("sqlite:///:memory:", future=True)
    create_sqlite_schema(engine)
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO documents (document_id, job_id, file_name, file_extension, document_type)
                VALUES
                    ('DOC1', 'JOB1', 'Estimate 1.xlsx', '.xlsx', 'estimate'),
                    ('DOC2', 'JOB2', 'Proposal.pdf', '.pdf', 'proposal'),
                    ('DOC3', 'JOB3', 'Estimate 3.xlsx', '.xlsx', 'estimate')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO document_content (
                    content_id, document_id, job_id, sheet_name, row_number, cell_range, text_content
                )
                VALUES
                    (
                        'CONTENT1', 'DOC1', 'JOB1', 'Estimate', 116, 'A116:J116',
                        'A116: Pwash/Prep | B116: 4 | C116: 5 | D116: 220 | H116: 7607.6'
                    ),
                    (
                        'CONTENT2', 'DOC2', 'JOB2', NULL, NULL, NULL,
                        'PDF paragraph that should not be scanned A116: fake'
                    ),
                    (
                        'CONTENT3', 'DOC3', 'JOB3', 'Estimate', 26, 'A26:H26',
                        'A26: 11 | B26: Silicone | C26: 10 | E26: 40 | H26: 400'
                    )
                """
            )
        )

    summary = tr.parse_existing_document_content(engine, limit_documents=1, progress=True)

    assert summary["documents_considered"] == 1
    assert summary["rows_read"] == 1
    assert summary["rows_upserted"] == 1
    captured = capsys.readouterr().out
    assert "Template row parse: documents considered: 1" in captured
    assert "[1/1] Estimate 1.xlsx" in captured
    with engine.connect() as conn:
        parsed_documents = conn.execute(text("SELECT DISTINCT document_id FROM estimate_template_rows")).scalars().all()
    assert parsed_documents == ["DOC1"]


def test_parse_existing_only_unparsed_skips_current_parser_rows() -> None:
    engine = create_engine("sqlite:///:memory:", future=True)
    create_sqlite_schema(engine)
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO documents (document_id, job_id, file_name, file_extension, document_type)
                VALUES
                    ('DOC1', 'JOB1', 'Estimate 1.xlsx', '.xlsx', 'estimate'),
                    ('DOC2', 'JOB2', 'Estimate 2.xlsx', '.xlsx', 'estimate')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO document_content (
                    content_id, document_id, job_id, sheet_name, row_number, cell_range, text_content
                )
                VALUES
                    (
                        'CONTENT1', 'DOC1', 'JOB1', 'Estimate', 116, 'A116:J116',
                        'A116: Pwash/Prep | B116: 4 | C116: 5 | D116: 220 | H116: 7607.6'
                    ),
                    (
                        'CONTENT2', 'DOC2', 'JOB2', 'Estimate', 26, 'A26:H26',
                        'A26: 11 | B26: Silicone | C26: 10 | E26: 40 | H26: 400'
                    )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO estimate_template_rows (
                    template_row_id, document_id, job_id, sheet_name, row_number,
                    cell_range, template_bucket, line_item_kind, needs_review, parser_version
                )
                VALUES ('existing', 'DOC1', 'JOB1', 'Estimate', 116, 'A116:J116', 'labor_prep', 'labor', false, :parser_version)
                """
            ),
            {"parser_version": tr.PARSER_VERSION},
        )

    summary = tr.parse_existing_document_content(engine, only_unparsed=True)

    assert summary["documents_considered"] == 1
    with engine.connect() as conn:
        parsed_documents = conn.execute(
            text("SELECT DISTINCT document_id FROM estimate_template_rows ORDER BY document_id")
        ).scalars().all()
    assert parsed_documents == ["DOC1", "DOC2"]


def test_unused_placeholder_adder_deletes_stale_template_row() -> None:
    engine = create_engine("sqlite:///:memory:", future=True)
    create_sqlite_schema(engine)
    stale_id = tr.stable_template_row_id("DOC1", "Estimate", 176, "A176:J176")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO documents (document_id, job_id, file_name, file_extension, document_type)
                VALUES ('DOC1', 'JOB1', 'Estimate.xlsx', '.xlsx', 'estimate')
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO document_content (
                    content_id, document_id, job_id, sheet_name, row_number, cell_range, text_content
                )
                VALUES (
                    'CONTENT1', 'DOC1', 'JOB1', 'Estimate', 176, 'A176:J176',
                    'A176: Additional Amount w/o Markup | H176: =F176'
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO estimate_template_rows (
                    template_row_id, document_id, job_id, sheet_name, row_number,
                    cell_range, template_bucket, line_item_kind, needs_review
                )
                VALUES (:template_row_id, 'DOC1', 'JOB1', 'Estimate', 176, 'A176:J176', 'unknown', 'unknown', true)
                """
            ),
            {"template_row_id": stale_id},
        )

    summary = tr.parse_existing_document_content(engine)

    assert summary["placeholder_rows_deleted"] == 1
    with engine.connect() as conn:
        count = conn.execute(text("SELECT COUNT(*) FROM estimate_template_rows")).scalar_one()
    assert count == 0


def test_query_helpers_and_summaries() -> None:
    rows = pd.DataFrame(
        [
            tr.parse_document_content_row(content_row(116, "A116: Prep | B116: 2 | C116: 4 | D116: 64 | H116: 2000")),
            tr.parse_document_content_row(content_row(26, "A26: 11 | B26: Silicone | C26: 10 | E26: 40 | H26: 400")),
            tr.parse_document_content_row(content_row(169, "A169: Price | H169: 10000")),
        ]
    )

    summary = tr.bucket_summary(rows)
    labor = tr.labor_task_summary(rows)
    met = tr.material_equipment_travel_summary(rows)
    totals = tr.totals_for_document(rows)

    assert set(summary["template_bucket"]) == {"labor_prep", "coating", "worksheet_price"}
    assert labor.iloc[0]["median_total_hours"] == 64
    assert met.iloc[0]["median_unit_price"] == 40
    assert totals["worksheet_price"] == 10000


def test_detect_workbook_template_type_insulation(tmp_path) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Estimate"
    ws["C3"] = "Insulation - Walls Only"
    ws["A116"] = "Total Job Cost"
    workbook.create_sheet("People")
    workbook.create_sheet("Materials")
    workbook.create_sheet("General")
    workbook.create_sheet("Sq Ft Calculation")
    workbook.create_sheet("Performance & Payment Bonds")
    path = tmp_path / "Estimate Insulation - Test.xlsx"
    workbook.save(path)

    assert tr.detect_workbook_template_type(path) == "insulation"


def test_detect_workbook_template_type_flooring(tmp_path) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Estimate"
    ws["C3"] = "Floor System"
    ws["A116"] = "Grind/Patch/"
    ws["A120"] = "Prep & Base 707"
    ws["A130"] = "Trip #3 Top Coat"
    workbook.create_sheet("People")
    workbook.create_sheet("Materials")
    workbook.create_sheet("General")
    workbook.create_sheet("Performance & Payment Bonds")
    path = tmp_path / "Estimate Flooring - Test.xlsx"
    workbook.save(path)

    assert tr.detect_workbook_template_type(path) == "flooring"


def test_insulation_document_rows_use_insulation_template_map() -> None:
    rows = tr.parse_document_content_rows(
        [
            content_row(3, "A3: Job Type: | C3: Insulation - Walls Only", document_id="DOCINS", source_file="Estimate Insulation - Test.xlsx"),
            content_row(19, "A19: 11 | B19: Gaco 2.0 lb. | C19: 2800 | D19: 4.25 | E19: 1.63 | G19: 740 | H19: 1206.2", document_id="DOCINS", source_file="Estimate Insulation - Test.xlsx"),
            content_row(26, "A26: Primer | C26: 100 | E26: 30 | G26: 0.4 | H26: 12", document_id="DOCINS", source_file="Estimate Insulation - Test.xlsx"),
            content_row(78, "A78: Set Up | B78: 0.1 | C78: 3 | D78: 2.4 | H78: 250", document_id="DOCINS", source_file="Estimate Insulation - Test.xlsx"),
            content_row(86, "A86: Foam | B86: 1.5 | C86: 3 | D86: 36 | H86: 1200", document_id="DOCINS", source_file="Estimate Insulation - Test.xlsx"),
            content_row(116, "A116: Total Job Cost | H116: 10000", document_id="DOCINS", source_file="Estimate Insulation - Test.xlsx"),
            content_row(118, "A118: Estimated O/H | F118: 30 | H118: 3000", document_id="DOCINS", source_file="Estimate Insulation - Test.xlsx"),
            content_row(120, "A120: Profit | F120: 10 | H120: 1300", document_id="DOCINS", source_file="Estimate Insulation - Test.xlsx"),
            content_row(122, "A122: Work Sheet Price | H122: 14300", document_id="DOCINS", source_file="Estimate Insulation - Test.xlsx"),
            content_row(123, "A123: Work Sheet Price + Additional Amount w/o Markup | F123: 500 | H123: 14800", document_id="DOCINS", source_file="Estimate Insulation - Test.xlsx"),
            content_row(137, "A137: Price / Sq. Ft: | B137: 5.42 | C137: Est. Sets: | D137: 0.74", document_id="DOCINS", source_file="Estimate Insulation - Test.xlsx"),
        ]
    )

    by_row = {row["row_number"]: row for row in rows}
    assert by_row[19]["template_type"] == "insulation"
    assert by_row[19]["template_bucket"] == "foam"
    assert by_row[19]["line_item_kind"] == "material"
    assert by_row[26]["template_bucket"] == "primer"
    assert by_row[78]["template_bucket"] == "labor_set_up"
    assert by_row[78]["line_item_kind"] == "labor"
    assert by_row[86]["template_bucket"] == "labor_foam"
    assert by_row[116]["template_bucket"] == "total_job_cost"
    assert by_row[116]["line_item_kind"] == "total"
    assert by_row[118]["overhead_pct"] == 30
    assert by_row[120]["profit_pct"] == 10
    assert by_row[122]["estimated_cost"] == 14300
    assert by_row[123]["estimated_cost"] == 14800
    assert by_row[137]["unit_price"] == 5.42
    assert by_row[137]["estimated_units"] == 0.74
    assert not any(row["template_bucket"] == "labor_prep" for row in rows)
