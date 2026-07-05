from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from sqlalchemy import create_engine, text

from jobscan.estimator.unknown_rows import (
    apply_mapping,
    build_unknown_clusters,
    export_unknown_review,
    extract_highlighted_template_row_roles,
    read_template_rows,
    summarize_unknown_rows,
)


def create_template_row_db():
    engine = create_engine("sqlite:///:memory:")
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE estimate_template_rows (
                    template_row_id TEXT PRIMARY KEY,
                    document_id TEXT,
                    job_id TEXT,
                    source_file TEXT,
                    template_type TEXT,
                    sheet_name TEXT,
                    row_number INTEGER,
                    template_bucket TEXT,
                    line_item_kind TEXT,
                    row_label TEXT,
                    selected_item_name TEXT,
                    raw_text TEXT,
                    estimated_cost NUMERIC
                )
                """
            )
        )
        rows = [
            ("r1", "d1", "j1", "Job 1001 Estimate.xlsx", "roofing", "Estimate", 115, "coating", "material", "Coating", "Silicone", "", 1000),
            ("r2", "d1", "j1", "Job 1001 Estimate.xlsx", "roofing", "Estimate", 116, "unknown", "unknown", "Pwash/Prep", "", "", None),
            ("r3", "d1", "j1", "Job 1001 Estimate.xlsx", "roofing", "Estimate", 117, "labor_base", "labor", "Base Coat", "", "", None),
            ("r4", "d2", "j2", "Job 1002 Estimate.xlsx", "roofing", "Estimate", 116, "unknown", "unknown", "Pwash/Prep", "", "", None),
            ("r5", "d2", "j2", "Job 1002 Estimate.xlsx", "roofing", "Estimate", 120, "unknown", "unknown", "Additional Amount w/o Markup", "", "", None),
            ("r6", "d3", "j3", "Insulation 1003 Estimate.xlsx", "insulation", "Estimate", 86, "unknown", "unknown", "Foam", "Gaco 2.0 lb.", "", None),
        ]
        conn.execute(
            text(
                """
                INSERT INTO estimate_template_rows (
                    template_row_id, document_id, job_id, source_file, template_type, sheet_name,
                    row_number, template_bucket, line_item_kind, row_label, selected_item_name,
                    raw_text, estimated_cost
                )
                VALUES (
                    :template_row_id, :document_id, :job_id, :source_file, :template_type, :sheet_name,
                    :row_number, :template_bucket, :line_item_kind, :row_label, :selected_item_name,
                    :raw_text, :estimated_cost
                )
                """
            ),
            [
                {
                    "template_row_id": row[0],
                    "document_id": row[1],
                    "job_id": row[2],
                    "source_file": row[3],
                    "template_type": row[4],
                    "sheet_name": row[5],
                    "row_number": row[6],
                    "template_bucket": row[7],
                    "line_item_kind": row[8],
                    "row_label": row[9],
                    "selected_item_name": row[10],
                    "raw_text": row[11],
                    "estimated_cost": row[12],
                }
                for row in rows
            ],
        )
    return engine


def test_unknown_row_clustering_groups_and_suggests_buckets() -> None:
    engine = create_template_row_db()
    with engine.connect() as conn:
        unknown = read_template_rows(conn, unknown_only=True)
        all_rows = read_template_rows(conn, unknown_only=False)

    clusters, samples, mapping = build_unknown_clusters(unknown, all_rows, limit=10)
    prep = clusters[clusters["row_label"].eq("Pwash/Prep")].iloc[0]

    assert prep["row_count"] == 2
    assert prep["suggested_bucket"] == "labor_prep"
    assert prep["suggested_line_item_kind"] == "labor"
    assert "coating/material" in prep["nearby_known_buckets_above"]
    assert "labor_base/labor" in prep["nearby_known_buckets_below"]
    assert not samples.empty
    assert set(mapping.columns) >= {"cluster_id", "target_template_bucket", "approved"}


def test_summary_reports_unknown_counts() -> None:
    engine = create_template_row_db()
    with engine.connect() as conn:
        rows = read_template_rows(conn, unknown_only=False)
    summary = summarize_unknown_rows(rows)

    assert summary["total_unknown_rows"] == 4
    assert any(row["value"] == "roofing" and row["count"] == 3 for row in summary["unknown_by_template_type"])
    assert any(row["value"] == "Pwash/Prep" and row["count"] == 2 for row in summary["top_unknown_row_labels"])


def test_export_unknown_review_writes_expected_csvs(tmp_path: Path) -> None:
    engine = create_template_row_db()
    with engine.begin() as conn:
        paths = export_unknown_review(conn, tmp_path, limit=5)

    assert paths["clusters"].exists()
    assert paths["samples"].exists()
    assert paths["mapping"].exists()
    clusters = pd.read_csv(paths["clusters"])
    assert {"cluster_id", "row_count", "suggested_bucket", "nearby_known_buckets_above"}.issubset(clusters.columns)
    assert paths["actionable_clusters"].exists()
    assert paths["non_informational_clusters"].exists()
    assert paths["metadata_clusters"].exists()


def test_highlighted_template_rows_split_non_informational_unknown_clusters(tmp_path: Path) -> None:
    engine = create_template_row_db()
    row_role_hints = {
        ("roofing", "Estimate", 120): {
            "template_type": "roofing",
            "sheet_name": "Estimate",
            "row_number": 120,
            "template_row_role": "template_header_or_instruction",
            "template_row_role_source": "test-template.xlsx",
            "template_row_role_detail": "yellow_cols=1,2,3",
        }
    }

    with engine.begin() as conn:
        paths = export_unknown_review(conn, tmp_path, limit=10, row_role_hints=row_role_hints)

    all_clusters = pd.read_csv(paths["clusters"])
    non_info = pd.read_csv(paths["non_informational_clusters"])
    actionable = pd.read_csv(paths["actionable_clusters"])
    header_row = all_clusters[all_clusters["row_number"].eq(120)].iloc[0]

    assert header_row["template_row_role"] == "template_header_or_instruction"
    assert header_row["suggested_bucket"] == "template_scaffolding"
    assert header_row["suggested_line_item_kind"] == "header"
    assert 120 in set(non_info["row_number"])
    assert 120 not in set(actionable["row_number"])
    assert paths["row_role_hints"].exists()


def test_text_scaffold_rows_split_without_highlight_hints(tmp_path: Path) -> None:
    engine = create_template_row_db()

    with engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO estimate_template_rows (
                    template_row_id, document_id, job_id, source_file, template_type, sheet_name,
                    row_number, template_bucket, line_item_kind, row_label, selected_item_name,
                    raw_text, estimated_cost
                )
                VALUES (
                    'r7', 'd4', 'j4', 'Job 1004 Estimate.xlsx', 'roofing', 'Estimate',
                    7, 'unknown', 'unknown', '', 'Title:', '', NULL
                )
                """
            )
        )
        paths = export_unknown_review(conn, tmp_path, limit=10)

    non_info = pd.read_csv(paths["non_informational_clusters"])
    actionable = pd.read_csv(paths["actionable_clusters"])
    metadata = pd.read_csv(paths["metadata_clusters"])

    assert 7 in set(metadata["row_number"])
    assert 7 not in set(non_info["row_number"])
    assert 7 not in set(actionable["row_number"])


def test_extract_highlighted_template_row_roles_reads_yellow_rows(tmp_path: Path) -> None:
    workbook_path = tmp_path / "Estimate Roofing Test.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Estimate"
    yellow = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    worksheet["A14"] = "Type"
    worksheet["B14"] = "Sq. Ft."
    worksheet["A14"].fill = yellow
    worksheet["B14"].fill = yellow
    worksheet["A15"] = "Actionable"
    workbook.save(workbook_path)

    hints = extract_highlighted_template_row_roles(workbook_path)

    assert ("roofing", "Estimate", 14) in hints
    assert ("roofing", "Estimate", 15) not in hints
    hint = hints[("roofing", "Estimate", 14)]
    assert hint["template_row_role"] == "template_header_or_instruction"
    assert "A14=Type" in hint["template_row_role_detail"]


def test_apply_mapping_dry_run_and_apply_preserves_original_bucket(tmp_path: Path) -> None:
    engine = create_template_row_db()
    mapping_path = tmp_path / "unknown_mapping_template.csv"
    pd.DataFrame(
        [
            {
                "cluster_id": "c1",
                "match_template_type": "roofing",
                "match_sheet_name": "Estimate",
                "match_row_number": 116,
                "match_row_label_pattern": "Pwash/Prep",
                "match_selected_item_pattern": "",
                "target_template_bucket": "labor_prep",
                "target_line_item_kind": "labor",
                "notes": "Approved test mapping",
                "approved": True,
            },
            {
                "cluster_id": "c2",
                "match_template_type": "roofing",
                "match_sheet_name": "Estimate",
                "match_row_number": 120,
                "match_row_label_pattern": "Additional Amount",
                "match_selected_item_pattern": "",
                "target_template_bucket": "estimate_adder",
                "target_line_item_kind": "other",
                "notes": "Not approved",
                "approved": False,
            },
        ]
    ).to_csv(mapping_path, index=False)

    with engine.begin() as conn:
        dry = apply_mapping(conn, mapping_path, dry_run=True, output_dir=tmp_path)
        unchanged = conn.execute(text("SELECT COUNT(*) FROM estimate_template_rows WHERE template_bucket = 'unknown'")).scalar_one()
    assert dry["matched_rows"].sum() == 2
    assert unchanged == 4
    assert (tmp_path / "approved_unknown_row_mappings.json").exists()
    assert (tmp_path / "approved_unknown_row_mappings.py").exists()

    with engine.begin() as conn:
        applied = apply_mapping(conn, mapping_path, dry_run=False, output_dir=tmp_path)
        remapped = conn.execute(text("SELECT COUNT(*) FROM estimate_template_rows WHERE template_bucket = 'labor_prep'")).scalar_one()
        original = conn.execute(text("SELECT DISTINCT original_template_bucket FROM estimate_template_rows WHERE row_number = 116")).scalar_one()
        unapproved = conn.execute(text("SELECT template_bucket FROM estimate_template_rows WHERE template_row_id = 'r5'")).scalar_one()

    assert applied["matched_rows"].sum() == 2
    assert remapped == 2
    assert original == "unknown"
    assert unapproved == "unknown"
