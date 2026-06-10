from pathlib import Path
import tempfile

from jobscan.batch_sharepoint_sync import BatchScanRoot, add_batch_context, crew_schedule_rows
from jobscan.extractors import extract_estimate_xlsx
from jobscan.scan import records_as_dicts, scan_root
from jobscan.schedule_extractor import add_business_days, parse_duration_days, parse_start_date


def write_labor_schedule_workbook(path: Path) -> None:
    import openpyxl

    tasks = [
        ("Set Up/Safety", 1, 6, 66),
        ("Tear Off/Rock", 50, 6, 3300),
        ("Board", 15, 6, 990),
        ("Wall/EM Work", 7, 6, 462),
        ("Foam/Base", 18, 6, 1188),
        ("Caulk/SF", 7, 6, 462),
        ("Details", 2, 5, 110),
        ("Top Coat/Gran", 8, 5, 440),
        ("Touch/Clean Up", 2, 5, 110),
        ("Lower Roofs (Dock + Awning)", None, None, None),
        ("Add Labor for Lower Roofs", 6, 6, 396),
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estimate"
    people = wb.create_sheet("People")
    ws["A112"] = "Labor / Subcontractor"
    ws["B114"] = "Days"
    ws["C114"] = "No. of People"
    ws["D114"] = "Total Hours"
    for offset, (task, days, crew_size, total_hours) in enumerate(tasks, start=115):
        ws.cell(row=offset, column=1).value = task
        ws.cell(row=offset, column=2).value = days
        ws.cell(row=offset, column=3).value = crew_size
        ws.cell(row=offset, column=4).value = total_hours
    ws["A148"] = "Total Hours"
    ws["B148"] = 10930
    ws["C148"] = "Total Days"
    ws["D148"] = 116
    people["A11"] = "Hours /Day"
    people["B11"] = 11
    wb.save(path)


def test_parse_schedule_dates_and_durations() -> None:
    assert parse_start_date("scheduled start: 06.15.26") == "2026-06-15"
    assert parse_start_date("mobilize 6/15 in 2026") == "2026-06-15"
    assert parse_duration_days("install days: 2") == 2
    assert parse_duration_days("estimated duration 1 week") == 5
    assert add_business_days("2026-06-12", 3) == "2026-06-16"


def test_scan_root_does_not_assign_crew_from_job_spec_text() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        job = root / "Pence UofL"
        job.mkdir()
        (job / "Job Spec.txt").write_text(
            "Crew Leader: Santos\nstart date: 6/15/26\nproduction days: 7\n",
            encoding="utf-8",
        )

        record = scan_root(root, scan_context="2026 Roofing/Contracted")[0]
        add_batch_context(record, BatchScanRoot(folder="2026 Roofing/Contracted", pipeline_status="Contracted"))
        row = records_as_dicts([record])[0]

    assert row["crew_leader"] is None
    assert row["assigned_crew_leader"] is None
    assert row["estimated_start_date"] is None
    assert row["estimated_duration_days"] is None
    assert row["schedule_status"] == "Not Ready"
    assert row["ready_to_schedule"] is False
    assert row["blocking_issue"] == "Missing estimated duration"
    assert row["suggested_crew_type"] is None
    assert row["suggested_crew_reason"] == "manual_needed"


def test_crew_schedule_rows_surface_missing_schedule_info() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        job = root / "Folder Only Job"
        job.mkdir()

        record = scan_root(root, scan_context="2026 Roofing/Folder Created")[0]
        add_batch_context(record, BatchScanRoot(folder="2026 Roofing/Folder Created", pipeline_status="Folder Created"))
        row = crew_schedule_rows([record])[0]

    assert row["schedule_status"] == "Not Ready"
    assert row["ready_to_schedule"] is False
    assert "Missing estimated duration" in row["blocking_issue"]
    assert "Missing estimated start date" not in row["blocking_issue"]
    assert "Missing crew leader" not in row["blocking_issue"]
    assert "Missing job spec" not in row["blocking_issue"]


def test_estimate_parser_extracts_labor_schedule_duration() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "Estimate Roofing (2026) - ACRE Bens Bargain (+ Loading Dock + Awning).xlsx"
        write_labor_schedule_workbook(path)

        extracted = extract_estimate_xlsx(path)

    assert extracted["estimated_duration_days"] == 116
    assert extracted["estimated_labor_hours"] == 10930
    assert extracted["estimated_hours_per_day"] == 11
    assert extracted["estimated_crew_size"] == 6
    assert extracted["labor_duration_source"] == "Estimate sheet Labor / Subcontractor section"
    task_names = [item["task"] for item in extracted["labor_schedule_breakdown"]]
    assert "Lower Roofs (Dock + Awning)" not in task_names
    assert task_names == [
        "Set Up/Safety",
        "Tear Off/Rock",
        "Board",
        "Wall/EM Work",
        "Foam/Base",
        "Caulk/SF",
        "Details",
        "Top Coat/Gran",
        "Touch/Clean Up",
        "Add Labor for Lower Roofs",
    ]


def test_scan_root_uses_estimate_duration_for_schedule_fields() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        job = root / "ACRE Bens Bargain"
        job.mkdir()
        write_labor_schedule_workbook(job / "Estimate Roofing (2026) - ACRE Bens Bargain (+ Loading Dock + Awning).xlsx")

        record = scan_root(root, scan_context="2026 Roofing/Contracted")[0]
        add_batch_context(record, BatchScanRoot(folder="2026 Roofing/Contracted", pipeline_status="Contracted"))
        row = records_as_dicts([record])[0]

    assert row["estimated_duration_days"] == 116
    assert row["estimated_labor_hours"] == 10930
    assert row["estimated_hours_per_day"] == 11
    assert row["estimated_crew_size"] == 6
    assert row["labor_schedule_breakdown"][0]["task"] == "Set Up/Safety"
    assert row["schedule_status"] == "Needs Assignment"
    assert row["ready_to_schedule"] is True
    assert row["schedule_confidence"] == "medium"
    assert "Missing estimated duration" not in row["blocking_issue"]
    assert row["blocking_issue"] == "Needs crew assignment"
    assert row["suggested_crew_type"] is None
    assert row["suggested_crew_reason"] == "manual_needed"


def test_manual_assignment_changes_schedule_status() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        job = root / "ACRE Bens Bargain"
        job.mkdir()
        write_labor_schedule_workbook(job / "Estimate Roofing (2026) - ACRE Bens Bargain.xlsx")

        record = scan_root(root, scan_context="2026 Roofing/Contracted")[0]
        record.assigned_crew_leader = "Mariano"
        add_batch_context(record, BatchScanRoot(folder="2026 Roofing/Contracted", pipeline_status="Contracted"))
        row = records_as_dicts([record])[0]

    assert row["schedule_status"] == "Needs Start Date"
    assert row["ready_to_schedule"] is True
    assert row["blocking_issue"] == "Needs start date"
