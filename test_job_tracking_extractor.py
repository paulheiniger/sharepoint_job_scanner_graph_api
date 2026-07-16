from pathlib import Path
import tempfile

from jobscan.job_tracking_extractor import (
    JOB_TRACKING_DAILY_FIELDS,
    JOB_TRACKING_SUMMARY_FIELDS,
    extract_job_tracking_file,
    scan_job_tracking_for_records,
)
from jobscan.scan import records_as_dicts, scan_root


def write_tracking_workbook(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Actual Amounts"
    headers = [
        "Meredith, Scott - Residence (Roofing 2026)",
        "Labor Hours",
        "Travel Hours",
        "Load Hours",
        "OS Hours",
        "Mileage",
        "OS Mileage",
        "Foam Strokes",
        "Thickness (in.)",
        "Sq. Ft. (Foam)",
        "Foam Yield",
        "A-Side Lot #",
        "B-Side Lot #",
        "Base Coat",
        "Sq. Ft. (Base)",
        "Gal/Sq. (Base)",
        "Top Coat",
        "Sq. Ft. (Top)",
        "Gal/Sq. (Top)",
        "Granules",
        "AF Buttergrade",
        "Caulk",
        "Primer",
        "SF",
        "Crew",
        "Notes",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=2, column=col).value = value

    dates = ["05.07.26", "05/08/26", "05-09-2026", "05.10.26", "05.11.26", "05.12.26", "05.13.26", "05.14.26"]
    for index, value in enumerate(dates, start=3):
        ws.cell(row=index, column=1).value = value
        ws.cell(row=index, column=2).value = 1
        ws.cell(row=index, column=3).value = 1
        ws.cell(row=index, column=4).value = 0.5
        ws.cell(row=index, column=5).value = 0.25
        ws.cell(row=index, column=6).value = 5
        ws.cell(row=index, column=7).value = 6
        ws.cell(row=index, column=8).value = 0.5
        ws.cell(row=index, column=9).value = 1.5
        ws.cell(row=index, column=10).value = 100
        ws.cell(row=index, column=11).value = 3000
        ws.cell(row=index, column=12).value = "A1"
        ws.cell(row=index, column=13).value = "B1"
        ws.cell(row=index, column=14).value = 0.5
        ws.cell(row=index, column=20).value = 3
        ws.cell(row=index, column=21).value = 1
        ws.cell(row=index, column=22).value = 2
        ws.cell(row=index, column=23).value = 4
        ws.cell(row=index, column=24).value = 5
        ws.cell(row=index, column=25).value = "Crew"
        ws.cell(row=index, column=26).value = "Notes"

    ws.cell(row=11, column=1).value = "Daily Totals"
    totals = {
        2: 72.41,
        3: 9.72,
        4: 4.34,
        5: 8.25,
        6: 67,
        7: 77,
        8: 6.5,
        9: 1.5,
        10: 800,
        11: 3000,
        14: 6.5,
        17: 0,
        20: 24,
        21: 9.25,
        22: 19,
        23: 4,
        24: 5,
    }
    for col, value in totals.items():
        ws.cell(row=11, column=col).value = value

    ws["A22"] = "Estimated Amounts"
    estimated_headers = [
        None,
        "Labor Hours",
        "Travel Hours",
        "Load Hours",
        "Overhead",
        "Mileage",
        "OS Mileage",
        "Foam Strokes",
        "Thickness (in.)",
        "Sq. Ft. (Foam)",
        "Foam Yield",
        "Base Coat",
        "Sq. Ft. (Base)",
        "Gal/Sq. (Base)",
        "Top Coat",
        "Sq. Ft. (Top)",
        "Gal/Sq. (Top)",
        "Granules",
        "AF Buttergrade",
        "Caulk",
        "Primer",
        "SF",
    ]
    estimated_values = [None, 175, 50, 8, 12, 160, 240, 10, 1.5, 1000, 2800, 119, None, None, 119, None, None, 30, 15, 80, 6, 7]
    for col, value in enumerate(estimated_headers, start=1):
        ws.cell(row=23, column=col).value = value
    for col, value in enumerate(estimated_values, start=1):
        ws.cell(row=24, column=col).value = value

    ws["A25"] = "Over/Under"
    for col, value in enumerate(estimated_headers, start=1):
        ws.cell(row=26, column=col).value = value
    ws.cell(row=27, column=2).value = -102.59
    ws.cell(row=27, column=3).value = -40.28
    ws.cell(row=27, column=4).value = -3.66
    ws.cell(row=27, column=8).value = -3.5
    ws.cell(row=27, column=10).value = -200
    ws.cell(row=27, column=12).value = -112.5
    ws.cell(row=27, column=15).value = -119
    ws.cell(row=27, column=18).value = -6
    ws.cell(row=27, column=19).value = -5.75
    ws.cell(row=27, column=20).value = -61
    ws.cell(row=27, column=21).value = -2
    ws.cell(row=27, column=22).value = -2
    wb.save(path)


def write_direct_header_tracking_workbook(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = [
        "Pegasus 39 Pearce - Various Repairs 2026",
        "Labor Hours",
        "Travel Hours",
        "Load Hours",
        "OS Hours",
        "Mileage",
        "OS Mileage",
        "Top Coat",
        "Sq. Ft. (Top)",
        "Gal/Sq. (Top)",
        "Granules",
        "Caulk",
        "Primer",
        "SF",
        "Crew",
        "Notes",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=2, column=col).value = value

    rows = [
        ("07.06.26", 34.85, 3, 1.5, None, 10, None, None, None, None, None, None, None, None, "Santos", "Power washed 7000 sq ft."),
        ("07.07.26", 53.65, 3.7, 2.75, None, 13, None, None, None, None, None, 48, 6, 10, "Santos", "Sprayed primer, caulk, and SF."),
    ]
    for row_num, row in enumerate(rows, start=3):
        for col, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col).value = value

    ws["A18"] = "Insert Additional Lines Here"
    ws["A20"] = "Daily Totals"
    ws["B20"] = 88.5
    ws["C20"] = 6.7
    ws["D20"] = 4.25
    ws["F20"] = 23
    ws["L20"] = 48
    ws["M20"] = 6
    ws["N20"] = 10

    ws["A22"] = "Estimated Amounts"
    for col, value in enumerate(headers[1:14], start=2):
        ws.cell(row=23, column=col).value = value
    ws["B24"] = 880
    ws["C24"] = 127.5
    ws["D24"] = 34
    ws["F24"] = 340
    ws["H24"] = 350
    ws["I24"] = 15250
    ws["J24"] = 1.5
    ws["L24"] = 160
    ws["M24"] = 61
    ws["N24"] = 40

    ws["A25"] = "Over/Under"
    for col, value in enumerate(headers[1:14], start=2):
        ws.cell(row=26, column=col).value = value
    ws["B27"] = 791.5
    ws["C27"] = 120.8
    ws["D27"] = 29.75
    ws["L27"] = 112
    ws["M27"] = 55
    ws["N27"] = 30
    wb.save(path)


def test_job_tracking_extractor_reads_daily_summary_and_estimates() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        job = root / "Meredith Residence"
        job.mkdir()
        path = job / "Job Tracking Form - Meredith Residence (Roofing 2026).xlsx"
        write_tracking_workbook(path)

        records = scan_root(root, scan_context="2026 Roofing/Completed")
        row = records_as_dicts(records)[0]
        summaries, daily = extract_job_tracking_file(path, root, records[0])

    summary = summaries[0]
    assert len(daily) == 8
    assert set(summary) == set(JOB_TRACKING_SUMMARY_FIELDS)
    assert set(daily[0]) == set(JOB_TRACKING_DAILY_FIELDS)
    assert daily[0]["work_date"] == "2026-05-07"
    assert summary["actual_first_work_date"] == "2026-05-07"
    assert summary["actual_last_work_date"] == "2026-05-14"
    assert summary["actual_work_day_count"] == 8
    assert summary["actual_labor_hours"] == 72.41
    assert summary["actual_travel_hours"] == 9.72
    assert summary["actual_load_hours"] == 4.34
    assert summary["actual_mileage"] == 67
    assert summary["actual_os_mileage"] == 77
    assert summary["actual_foam_strokes"] == 6.5
    assert summary["actual_foam_thickness_inches"] == 1.5
    assert summary["actual_foam_sqft"] == 800
    assert summary["actual_foam_yield"] == 3000
    assert summary["actual_base_coat_1"] == 6.5
    assert summary["actual_granules"] == 24
    assert summary["actual_af_buttergrade"] == 9.25
    assert summary["actual_caulk"] == 19
    assert summary["actual_primer"] == 4
    assert summary["actual_sf"] == 5
    assert summary["estimated_labor_hours"] == 175
    assert summary["estimated_travel_hours"] == 50
    assert summary["estimated_load_hours"] == 8
    assert summary["estimated_mileage"] == 160
    assert summary["estimated_os_mileage"] == 240
    assert summary["estimated_foam_strokes"] == 10
    assert summary["estimated_foam_thickness_inches"] == 1.5
    assert summary["estimated_foam_sqft"] == 1000
    assert summary["estimated_foam_yield"] == 2800
    assert summary["estimated_base_coat_1"] == 119
    assert summary["estimated_base_coat_2"] == 119
    assert summary["estimated_granules"] == 30
    assert summary["estimated_af_buttergrade"] == 15
    assert summary["estimated_caulk"] == 80
    assert summary["estimated_primer"] == 6
    assert summary["estimated_sf"] == 7
    assert summary["labor_hours_variance"] == -102.59
    assert summary["foam_strokes_variance"] == -3.5
    assert summary["foam_sqft_variance"] == -200
    assert summary["granules_variance"] == -6
    assert summary["primer_variance"] == -2
    assert summary["sf_variance"] == -2
    assert daily[0]["foam_strokes"] == 0.5
    assert daily[0]["foam_thickness_inches"] == 1.5
    assert daily[0]["foam_sqft"] == 100
    assert daily[0]["foam_yield"] == 3000
    assert daily[0]["a_side_lot"] == "A1"
    assert daily[0]["b_side_lot"] == "B1"
    assert daily[0]["granules"] == 3
    assert daily[0]["primer"] == 4
    assert daily[0]["sf"] == 5

    assert row["has_job_tracking_form"] is True
    assert row["job_tracking_file"] == "Meredith Residence/Job Tracking Form - Meredith Residence (Roofing 2026).xlsx"
    assert row["actual_labor_hours"] == 72.41
    assert row["labor_hours_variance"] == -102.59


def test_job_tracking_extractor_reads_foam_lbs_estimate_header() -> None:
    import openpyxl

    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        job = root / "Graves County Athletic Multipurpose Facility"
        job.mkdir()
        path = job / "Job Tracking Form - Graves County Athletic Multipurpose Facility.xlsx"
        write_tracking_workbook(path)

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        ws["H23"] = "Foam (lbs.)"
        ws["H24"] = 3343.24
        wb.save(path)

        records = scan_root(root, scan_context="2026 Roofing/Completed")
        summaries, _daily = extract_job_tracking_file(path, root, records[0])

    assert summaries[0]["estimated_foam_lbs"] == 3343.24
    assert summaries[0]["estimated_foam_strokes"] is None


def test_job_tracking_extractor_combines_multiple_tracking_sheets_per_file() -> None:
    import openpyxl

    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        job = root / "WT Young"
        job.mkdir()
        path = job / "Job Tracking - Phase 2 WT Young.xlsx"
        write_tracking_workbook(path)

        wb = openpyxl.load_workbook(path)
        ws = wb.create_sheet("Gregg")
        ws["A1"] = "Actual Amounts"
        headers = [
            "UK WT Phase 2",
            "Labor Hours",
            "Travel Hours",
            "Load Hours",
            "OS Hours",
            "Mileage",
            "OS Mileage",
            "Foam Strokes",
            "Thickness (in.)",
            "Sq. Ft. (Foam)",
            "Foam Yield",
            "Crew",
            "Notes",
        ]
        for col, value in enumerate(headers, start=1):
            ws.cell(row=2, column=col).value = value
        ws["A3"] = "05.15.26"
        ws["B3"] = 10
        ws["C3"] = 2
        ws["L3"] = "Gregg"
        ws["M3"] = "Supplemental labor sheet."
        ws["A8"] = "Daily Totals"
        ws["B8"] = 10
        ws["C8"] = 2
        ws["A10"] = "Estimated Amounts"
        ws["B11"] = "Labor Hours"
        ws["C11"] = "Travel Hours"
        ws["B12"] = 175
        ws["C12"] = 50
        ws["A13"] = "Over/Under"
        wb.save(path)

        record = scan_root(root, scan_context="2026 Roofing/Contracted")[0]
        summaries, daily = extract_job_tracking_file(path, root, record)

    assert len(summaries) == 1
    summary = summaries[0]
    assert len(daily) == 9
    assert summary["actual_work_day_count"] == 9
    assert summary["actual_labor_hours"] == 82.41
    assert summary["actual_travel_hours"] == 11.72
    assert summary["actual_foam_strokes"] == 6.5
    assert summary["actual_foam_sqft"] == 800
    assert summary["estimated_labor_hours"] == 175
    assert summary["estimated_foam_sqft"] == 1000
    assert summary["labor_hours_variance"] == 92.59
    assert "Supplemental labor sheet." in summary["tracking_notes"]


def test_job_tracking_extractor_reads_direct_header_tracking_form() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        job = root / "Pegasus 39 Pearce"
        job.mkdir()
        path = job / "Job Tracking Form - Pegasus 39 Pearce Repairs (2026).xlsx"
        write_direct_header_tracking_workbook(path)
        record = scan_root(root, scan_context="2026 Roofing/Contracted")[0]

        summaries, daily = extract_job_tracking_file(path, root, record)

    summary = summaries[0]
    assert len(daily) == 2
    assert daily[0]["work_date"] == "2026-07-06"
    assert daily[0]["labor_hours"] == 34.85
    assert daily[1]["caulk"] == 48
    assert summary["actual_first_work_date"] == "2026-07-06"
    assert summary["actual_last_work_date"] == "2026-07-07"
    assert summary["actual_work_day_count"] == 2
    assert summary["actual_labor_hours"] == 88.5
    assert summary["actual_travel_hours"] == 6.7
    assert summary["actual_load_hours"] == 4.25
    assert summary["actual_caulk"] == 48
    assert summary["actual_primer"] == 6
    assert summary["actual_sf"] == 10
    assert summary["estimated_labor_hours"] == 880
    assert summary["estimated_caulk"] == 160
    assert summary["labor_hours_variance"] == 791.5
    assert not summary["tracking_warnings"]


def test_scan_job_tracking_for_records_finds_tracking_workbook() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        job = root / "Meredith Residence"
        job.mkdir()
        write_tracking_workbook(job / "Job Tracking Form - Meredith Residence (Roofing 2026).xlsx")
        record = scan_root(root, scan_context="2026 Roofing/Completed")[0]

        summaries, daily = scan_job_tracking_for_records(root, [record])

    assert len(summaries) == 1
    assert len(daily) == 8


def test_job_tracking_extractor_skips_empty_tracking_workbook() -> None:
    import openpyxl

    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        job = root / "Empty Tracking"
        job.mkdir()
        path = job / "Job Tracking Form - Empty.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.delete_rows(1, ws.max_row)
        wb.save(path)
        record = scan_root(root, scan_context="2026 Roofing/Contracted")[0]

        summaries, daily = extract_job_tracking_file(path, root, record)

    assert summaries == []
    assert daily == []
