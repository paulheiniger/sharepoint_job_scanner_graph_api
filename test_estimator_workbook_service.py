from __future__ import annotations

import hashlib
import shutil
from copy import deepcopy
from pathlib import Path

import pytest

pytest.importorskip("openpyxl")
import openpyxl

from jobscan.estimator.workbook_profile import (
    discover_flooring_workbook_profile,
    discover_insulation_workbook_profile,
    discover_roofing_workbook_profile,
)
from jobscan.estimator.workbook_recommendations import (
    normalize_template_material_pricing,
)
from jobscan.estimator.workbook_service import (
    EstimateWorkbookArtifact,
    EstimateWorkbookInputError,
    EstimateWorkbookOutputError,
    create_estimate_workbook,
    create_estimate_workbook_options,
    estimate_template_path,
    validate_semantic_workbook_payload,
)


TEMPLATE_PATH = Path("templates/Estimate + Spec - Roofing.xlsx")
INSULATION_TEMPLATE_PATH = Path("templates/Estimate + Spec - Insulation.xlsx")
FLOORING_TEMPLATE_PATH = Path("templates/Estimate + Spec - Flooring.xlsx")


def test_checked_in_templates_are_the_service_defaults(monkeypatch) -> None:
    monkeypatch.delenv("ESTIMATOR_ROOFING_TEMPLATE_PATH", raising=False)
    monkeypatch.delenv("ESTIMATOR_INSULATION_TEMPLATE_PATH", raising=False)

    assert estimate_template_path("roofing", base_dir=Path.cwd()) == (
        TEMPLATE_PATH.resolve()
    )
    assert estimate_template_path("insulation", base_dir=Path.cwd()) == (
        INSULATION_TEMPLATE_PATH.resolve()
    )
    assert estimate_template_path("flooring", base_dir=Path.cwd()) == (
        FLOORING_TEMPLATE_PATH.resolve()
    )


def complete_payload() -> dict:
    return {
        "confirmed": True,
        "template_type": "roofing",
        "header": {
            "job_name": "Complete Roofing Workbook",
            "job_type": "Roof restoration",
            "site_address": "100 Test Street",
            "estimated_sqft": 5000,
            "mobilizations": 3,
        },
        "pricing": {"overhead_pct": 35, "profit_pct": 15},
        "materials": [
            {
                "category": "coating",
                "item": "Gaco Silicone",
                "include": True,
                "selector_code": 11,
                "area_sqft": 5000,
                "gal_per_100_sqft": 1.5,
                "unit_price": 36,
            }
        ],
        "labor": [
            {
                "task": "labor_prep",
                "include": True,
                "days": 2,
                "crew_size": 4,
            },
            {
                "task": "labor_loading",
                "include": True,
                "hours_per_trip": 1.5,
                "crew_size": 2,
            },
            {
                "task": "labor_traveling",
                "include": True,
                "hours_per_trip": 3,
                "crew_size": 4,
            },
        ],
        "logistics": [
            {
                "category": "sales_inspection_trips",
                "item": "Sales and inspection mileage",
                "include": True,
                "trip_count": 2,
                "round_trip_miles": 180,
            },
            {
                "category": "truck_expense",
                "item": "Production truck mileage",
                "include": True,
                "trip_count": 3,
                "round_trip_miles": 180,
            },
        ],
        "adders": [],
        "scope_of_work": ["QA estimate workbook generation."],
        "spec_notes": ["Estimator review required."],
    }


def grossman_integrity_payload() -> dict:
    return {
        "confirmed": True,
        "template_type": "roofing",
        "structured_scope": {
            "template_type": "roofing",
            "declared_total_area_sqft": 5136,
            "area_reconciliation": {
                "declared_total_area_sqft": 5136,
                "calculated_total_area_sqft": 5456,
            },
            "area_scopes": [
                {
                    "scope_id": "tearoff",
                    "scope_role": "exclusive_area",
                    "label": "Full-removal section",
                    "area_sqft": 3120,
                    "action": (
                        "Full removal down to wood decking; replace deteriorated "
                        "decking within the nested allowance"
                    ),
                    "proposed_assembly": "2 inch Resista ISO and 1.5 inch coated foam roof",
                },
                {
                    "scope_id": "deck-repair",
                    "parent_scope_id": "tearoff",
                    "scope_role": "nested_sub_scope",
                    "label": "Deteriorated decking",
                    "area_sqft": 320,
                    "action": "Remove and replace deteriorated decking",
                },
                {
                    "scope_id": "recover",
                    "scope_role": "exclusive_area",
                    "label": "Foam over existing",
                    "area_sqft": 2016,
                    "proposed_assembly": "1.5 inch coated foam over existing roof",
                },
            ],
        },
        "header": {
            "job_name": "Grossman Tuning Integrity Regression",
            "job_type": "Roof replacement and coated foam",
            "site_address": "830 South 1st Street",
            "city_state_zip": "Louisville, KY 40203",
            "estimated_sqft": 5136,
            "mobilizations": 12,
            "estimated_days": 5,
            "estimated_crew_size": 6,
            "repair_area_description": "320 sq. ft. nested deck-repair allowance",
        },
        "pricing": {"overhead_pct": 35, "profit_pct": 15},
        "materials": [
            {
                "category": "roofing_foam",
                "item": "Gaco Roof 2.7 SPF",
                "area_sqft": 5136,
                "basis_sqft": 5250,
                "quantity_adjustment_reason": "Production allowance over measured roof area",
                "thickness_inches": 1.5,
                "yield_factor": 2700,
                "unit_price": 2.15,
            },
            {
                "category": "coating",
                "item": "Gaco Silicone base coat",
                "area_sqft": 5136,
                "basis_sqft": 5200,
                "quantity_adjustment_reason": "Rounded coating production area",
                "gal_per_100_sqft": 1.15,
                "unit_price": 32,
            },
            {
                "category": "coating",
                "item": "Gaco Silicone top coat",
                "area_sqft": 5136,
                "basis_sqft": 5200,
                "quantity_adjustment_reason": "Rounded coating production area",
                "gal_per_100_sqft": 1.15,
                "unit_price": 32,
            },
            {
                "category": "board_stock",
                "item": "2 inch Resista ISO",
                "area_sqft": 3120,
                "basis_sqft": 3136,
                "quantity_adjustment_reason": "98 full 4x8 sheets",
                "thickness_inches": 2,
                "price_per_square": 105,
            },
            {
                "category": "fasteners",
                "item": "Roof fasteners",
                "quantity": 1176,
                "unit_price": 100,
            },
            {
                "category": "plates",
                "item": "Insulation plates",
                "quantity": 1176,
                "unit_price": 80,
            },
            {
                "category": "granules",
                "item": "3M granules",
                "area_sqft": 5136,
                "quantity": 25.68,
                "unit_price": 25,
            },
            {"category": "edge_metal", "linear_ft": 150, "unit_price": 8},
            {"category": "gutter", "linear_ft": 52, "unit_price": 10},
            {"category": "downspouts", "linear_ft": 20, "unit_price": 9},
        ],
        "labor": [
            {"task": "labor_prep", "days": 1, "crew_size": 5},
            {"task": "labor_tearoff", "days": 2, "crew_size": 6},
            {"task": "labor_board", "days": 1, "crew_size": 6},
            {"task": "labor_base", "days": 1, "crew_size": 5},
            {"task": "labor_top_coat", "days": 1, "crew_size": 5},
            {
                "task": "labor_loading",
                "hours_per_trip": 1,
                "crew_size": 2,
            },
            {
                "task": "labor_traveling",
                "hours_per_trip": 2,
                "crew_size": 5,
            },
        ],
        "logistics": [
            {
                "category": "sales_inspection_trips",
                "trip_count": 3,
                "round_trip_miles": 65,
            },
            {
                "category": "truck_expense",
                "trip_count": 12,
                "round_trip_miles": 65,
            },
            {
                "category": "dumpster",
                "item": "20-yard dumpster",
                "area_sqft": 3120,
                "basis_sqft": 3200,
                "quantity_adjustment_reason": "Rounded debris allowance",
                "debris_thickness_inches": 2.5,
                "unit_price": 600,
            },
        ],
        "adders": [],
        "scope_of_work": [
            "Remove the 3,120 sq. ft. section to wood deck and install 2-inch ISO.",
            "Treat 320 sq. ft. as nested deteriorated-deck repair, not additional roof area.",
            "Install a 1.5-inch coated SPF system over the complete 5,136 sq. ft. roof.",
        ],
        "spec_notes": ["All quantities remain subject to estimator review."],
    }


def complete_insulation_payload() -> dict:
    return {
        "confirmed": True,
        "template_type": "insulation",
        "header": {
            "job_name": "Handwritten Field Notes Test",
            "job_type": "Spray foam insulation",
            "estimated_sqft": 4611,
            "sqft_calculation_rows": [
                {"description": "Second-floor apartment gross walls", "area_sqft": 974},
                {"description": "Second-floor openings", "area_sqft": -81},
                {"description": "First-floor apartment gross walls", "area_sqft": 1449},
                {"description": "First-floor openings", "area_sqft": -324},
                {"description": "Shop gross walls", "area_sqft": 3304},
                {"description": "Shop openings", "area_sqft": -711},
            ],
        },
        "pricing": {"overhead_pct": 35, "profit_pct": 15},
        "materials": [
            {
                "category": "foam",
                "item": "Gaco 2.0 lb. closed-cell foam",
                "include": True,
                "selector_code": 11,
                "area_sqft": 4611,
                "thickness_inches": 2,
                "unit_price": 2.2,
                "yield_factor": 3500,
            }
        ],
        "labor": [
            {"task": "labor_set_up", "include": True, "days": 0.25, "crew_size": 3},
            {"task": "labor_mask", "include": True, "days": 0.25, "crew_size": 3},
            {"task": "labor_foam", "include": True, "days": 5, "crew_size": 3},
            {"task": "labor_cleanup", "include": True, "days": 0.25, "crew_size": 3},
            {"task": "labor_loading", "include": False},
            {"task": "labor_traveling", "include": False},
        ],
        "logistics": [
            {"category": "sales_inspection_trips", "include": False},
            {"category": "truck_expense", "include": False},
        ],
        "adders": [
            {
                "label": "Gable insulation - scope/area review",
                "include": True,
                "needs_review": True,
            },
            {
                "label": "Stair-wall insulation - area review",
                "include": True,
                "needs_review": True,
            },
        ],
        "scope_of_work": [
            "Apply 2 inches of closed-cell spray foam to 4,611 sq. ft. of apartment and shop walls."
        ],
        "spec_notes": ["Gables and stair area require estimator confirmation."],
    }


def complete_flooring_payload() -> dict:
    return {
        "confirmed": True,
        "template_type": "flooring",
        "header": {
            "job_name": "Robert Gardiner Stables - Garage Floor",
            "job_type": "Floor system repair",
            "site_address": "2849 Mt. Eden Rd.",
            "city_state_zip": "Shelbyville, KY",
            "contact": "Cornell",
            "estimated_sqft": 484,
            "mobilizations": 4,
        },
        "pricing": {"overhead_pct": 30, "profit_pct": 10},
        "materials": [
            {
                "category": "coating",
                "item": "NPI 707 Epoxy",
                "include": True,
                "selector_code": 11,
                "area_sqft": 500,
                "gal_per_100_sqft": 1,
                "unit_price": 40.5,
                "waste_factor_pct": 10,
            },
            {
                "category": "coating",
                "item": "NPI Polyaspartic",
                "include": True,
                "selector_code": 11,
                "area_sqft": 500,
                "gal_per_100_sqft": 0.75,
                "unit_price": 77.1,
                "waste_factor_pct": 10,
            },
        ],
        "labor": [
            {"task": "labor_floor_grind_patch", "include": True, "days": 1, "crew_size": 2},
            {"task": "labor_floor_corner_repair", "include": True, "days": 0.5, "crew_size": 2},
            {"task": "labor_floor_prep_base_flake", "include": True, "days": 0.5, "crew_size": 2},
            {"task": "labor_floor_top_coat", "include": True, "days": 0.25, "crew_size": 2},
            {"task": "labor_floor_cleanup", "include": True, "days": 0.25, "crew_size": 2},
            {"task": "labor_loading", "include": True, "hours_per_trip": 0.5, "crew_size": 1},
            {"task": "labor_traveling", "include": True, "hours_per_trip": 0.5, "crew_size": 2},
        ],
        "logistics": [
            {
                "category": "sales_inspection_trips",
                "item": "Sales and inspection mileage",
                "include": True,
                "trip_count": 2,
                "round_trip_miles": 20,
            },
            {
                "category": "truck_expense",
                "item": "Production truck mileage",
                "include": True,
                "trip_count": 4,
                "round_trip_miles": 20,
            },
            {
                "category": "generator",
                "item": "Generator",
                "include": True,
                "days": 2,
                "unit_price": 30,
            },
        ],
        "adders": [
            {"label": "Misc. pucks and supplies", "include": True, "amount": 400},
            {
                "label": "Patch materials package",
                "include": True,
                "amount": 1800,
                "notes": "Mortar, joint tubes, Quick Fix, flake, and freight.",
            },
            {
                "label": "5x5 corner concrete replacement scope",
                "include": True,
                "needs_review": True,
                "notes": "Confirm tear-out and new concrete responsibility before pricing.",
            },
        ],
        "scope_of_work": [
            "Grind and patch approximately 100 linear feet of thin cracks.",
            "Prepare 484 sq. ft. and apply NPI 707, flake, and polyaspartic top coat.",
        ],
        "spec_notes": [
            "The 5x5 corner concrete tear-out and replacement remains unpriced pending confirmation."
        ],
    }


def complete_roofing_repair_payload() -> dict:
    return {
        "confirmed": True,
        "template_type": "roofing",
        "header": {
            "job_name": "Elsby Property - Statewide Mortgage",
            "job_type": "Corner Leak Repair",
            "site_address": "752 Highlander Pointe Dr.",
            "city_state_zip": "Floyds Knobs, IN 47119",
            "contact": "Rick Elsby",
            "title": "Owner",
            "email": "rick@example.invalid",
            "phone": "502-555-0100",
            "estimator": "Anthony Palmer",
            "estimated_sqft": 8184,
            "mobilizations": 1,
            "estimated_days": 1,
            "estimated_hours": 20,
            "estimated_crew_size": 2,
            "repair_area_description": "50' x 12'",
            "warranty_description": "2 Yr. Workmanship",
        },
        "pricing": {"overhead_pct": 35, "profit_pct": 15},
        "warranty": {
            "include": True,
            "years": 2,
            "warranty_type": "Workmanship",
            "area_sqft": 8184,
            "unit_cost": 0,
        },
        "materials": [
            {
                "category": "coating",
                "item": "Gaco Silicone",
                "selector_code": 11,
                "area_sqft": 250,
                "gal_per_100_sqft": 2.5,
                "unit_price": 32,
                "waste_factor_pct": 20,
            },
            {
                "category": "caulk_sealant",
                "item": "Silicone Sausage",
                "selector_code": 2,
                "quantity": 16,
                "unit_price": 11,
            },
            {
                "category": "caulk_sealant",
                "item": "Gaco SF-2000",
                "selector_code": 5,
                "quantity": 5,
                "unit_price": 35,
            },
        ],
        "labor": [
            {
                "task": "labor_setup_safety",
                "label": "Setup/Safety",
                "days": 0.25,
                "crew_size": 2,
            },
            {"task": "labor_prep", "days": 0.25, "crew_size": 2},
            {"task": "labor_caulk", "days": 0.5, "crew_size": 2},
            {
                "task": "labor_loading",
                "hours_per_trip": 1,
                "crew_size": 1,
                "hourly_rate": 25.5,
            },
            {
                "task": "labor_traveling",
                "hours_per_trip": 2,
                "crew_size": 2,
                "hourly_rate": 34,
            },
        ],
        "logistics": [
            {
                "category": "sales_inspection_trips",
                "trip_count": 1,
                "round_trip_miles": 85,
                "unit_price": 0.75,
            },
            {
                "category": "truck_expense",
                "trip_count": 1,
                "round_trip_miles": 85,
                "unit_price": 1.25,
            },
            {"category": "generator", "days": 1, "unit_price": 50},
        ],
        "adders": [
            {"label": "Warranty - 2-yr Workmanship", "amount": 750},
            {
                "label": "Misc. Materials (Fabric, Fasteners, Solvent, Etc.)",
                "amount": 250,
            },
        ],
        "scope_of_work": [
            "Set up jobsite safety around the localized repair area.",
            "Inspect inside the leaking area before starting; the building was locked during the site visit.",
            "Clean the 50 ft x 12 ft repair area.",
            "Tighten or replace loose or missing fasteners.",
            "Three-course 50 linear ft of wall-to-roof counter flashing seam and 12 linear ft of horizontal lap seam.",
            "Seal minor damage, all fasteners, and four vent or pipe penetrations with silicone caulk and SF-2000.",
            "Complete touch-up and remove work-related debris.",
        ],
        "spec_notes": [
            "Repair footprint is approximately 600 sq. ft.; the coating allowance is 250 sq. ft.",
            "Inspect parapet wall and coping cap seams and confirm substrate conditions before work.",
        ],
    }


def test_template_profile_discovers_cost_dependencies() -> None:
    profile = discover_roofing_workbook_profile(TEMPLATE_PATH)

    assert profile.sales_inspection.trips_cell == "B106"
    assert profile.sales_inspection.miles_cell == "C106"
    assert profile.sales_inspection.cost_cell == "H106"
    assert profile.truck_expense.cost_cell == "H108"
    assert profile.per_trip_labor_rows["labor_loading"].cost_cell == "H137"
    assert profile.crew_daily_rate_cells[4] == "G12"
    assert profile.labor_subtotal_cell == "H148"
    assert profile.final_price_cell == "H170"
    assert profile.warranty is not None
    assert profile.warranty.years_cell == "C154"
    assert profile.warranty.cost_cell == "H154"
    assert profile.labor_rows["labor_setup_safety"].row == 116
    assert profile.labor_rows["labor_setup_safety"].cost_cell == "H116"


def test_insulation_template_profile_discovers_cost_dependencies() -> None:
    profile = discover_insulation_workbook_profile(INSULATION_TEMPLATE_PATH)

    assert profile.sales_inspection.cost_cell == "H68"
    assert profile.truck_expense.cost_cell == "H70"
    assert profile.labor_rows["labor_foam"].cost_cell == "H86"
    assert profile.per_trip_labor_rows["labor_loading"].cost_cell == "H95"
    assert profile.crew_daily_rate_cells[3] == "F12"
    assert profile.material_subtotal_cell == "H72"
    assert profile.labor_subtotal_cell == "H103"
    assert profile.final_price_cell == "H123"
    assert profile.warranty is None


def test_flooring_template_profile_discovers_cost_dependencies() -> None:
    profile = discover_flooring_workbook_profile(FLOORING_TEMPLATE_PATH)

    assert profile.job_spec_sheet is None
    assert profile.sales_inspection.cost_cell == "H106"
    assert profile.truck_expense.cost_cell == "H108"
    assert profile.labor_rows["labor_floor_grind_patch"].cost_cell == "H116"
    assert profile.labor_rows["labor_floor_top_coat"].cost_cell == "H130"
    assert profile.per_trip_labor_rows["labor_loading"].cost_cell == "H137"
    assert profile.crew_daily_rate_cells[2] == "E12"
    assert profile.material_subtotal_cell == "H110"
    assert profile.labor_subtotal_cell == "H148"
    assert profile.final_price_cell == "H170"
    assert profile.warranty is not None
    assert profile.warranty.area_cell == "E154"


def test_semantic_validation_requires_explicit_travel_and_per_trip_labor() -> None:
    profile = discover_roofing_workbook_profile(TEMPLATE_PATH)
    payload = complete_payload()
    payload["logistics"] = []
    payload["labor"] = [payload["labor"][0]]

    with pytest.raises(EstimateWorkbookInputError) as exc_info:
        validate_semantic_workbook_payload(payload, profile)

    message = str(exc_info.value)
    assert "Sales/inspection travel" in message
    assert "Truck travel" in message
    assert "Loading labor" in message
    assert "Traveling labor" in message


def test_grossman_integrity_payload_reconciles_measured_and_purchase_areas() -> None:
    profile = discover_roofing_workbook_profile(TEMPLATE_PATH)

    validate_semantic_workbook_payload(grossman_integrity_payload(), profile)


def test_grossman_integrity_blocks_320_sqft_as_full_iso_basis() -> None:
    profile = discover_roofing_workbook_profile(TEMPLATE_PATH)
    payload = grossman_integrity_payload()
    board = next(
        item for item in payload["materials"] if item["category"] == "board_stock"
    )
    board["area_sqft"] = 320
    board["basis_sqft"] = 320
    board["quantity_adjustment_reason"] = ""

    with pytest.raises(EstimateWorkbookInputError) as exc_info:
        validate_semantic_workbook_payload(payload, profile)

    assert "3120 sq. ft. basis" in str(exc_info.value)


def test_grossman_integrity_requires_reason_for_purchase_allowance() -> None:
    profile = discover_roofing_workbook_profile(TEMPLATE_PATH)
    payload = grossman_integrity_payload()
    foam = next(
        item for item in payload["materials"] if item["category"] == "roofing_foam"
    )
    foam["quantity_adjustment_reason"] = ""

    with pytest.raises(EstimateWorkbookInputError) as exc_info:
        validate_semantic_workbook_payload(payload, profile)

    assert "quantity_adjustment_reason" in str(exc_info.value)


def test_semantic_validation_rejects_warranty_for_template_without_warranty_row() -> None:
    profile = discover_insulation_workbook_profile(INSULATION_TEMPLATE_PATH)
    payload = complete_insulation_payload()
    payload["warranty"] = {
        "include": True,
        "manufacturer": "Example",
        "years": 10,
        "unit_cost": 0.1,
    }

    with pytest.raises(EstimateWorkbookInputError) as exc_info:
        validate_semantic_workbook_payload(payload, profile)

    assert "does not support warranty inputs" in str(exc_info.value)


def test_semantic_validation_rejects_flat_amount_in_warranty_unit_cost() -> None:
    profile = discover_roofing_workbook_profile(TEMPLATE_PATH)
    payload = complete_payload()
    payload["warranty"] = {
        "include": True,
        "years": 10,
        "warranty_type": "material",
        "area_sqft": 5000,
        "unit_cost": 250,
    }

    with pytest.raises(EstimateWorkbookInputError) as exc_info:
        validate_semantic_workbook_payload(payload, profile)

    assert "dollars per square foot" in str(exc_info.value)


def test_semantic_validation_rejects_conflicting_warranty_years() -> None:
    profile = discover_roofing_workbook_profile(TEMPLATE_PATH)
    payload = complete_payload()
    payload["header"]["warranty_description"] = "2 Yr. Workmanship"
    payload["warranty"] = {
        "include": True,
        "years": 15,
        "warranty_type": "material",
        "area_sqft": 5000,
        "unit_cost": 0.2,
    }

    with pytest.raises(EstimateWorkbookInputError) as exc_info:
        validate_semantic_workbook_payload(payload, profile)

    assert "conflicts with warranty.years" in str(exc_info.value)


def test_semantic_validation_rejects_warranty_area_that_template_will_ignore() -> None:
    profile = discover_roofing_workbook_profile(TEMPLATE_PATH)
    payload = complete_payload()
    payload["warranty"] = {
        "include": True,
        "years": 10,
        "warranty_type": "material",
        "area_sqft": 600,
        "unit_cost": 0.2,
    }

    with pytest.raises(EstimateWorkbookInputError) as exc_info:
        validate_semantic_workbook_payload(payload, profile)

    assert "must match header.estimated_sqft" in str(exc_info.value)


def test_semantic_validation_rejects_two_tasks_for_shared_repair_row() -> None:
    profile = discover_roofing_workbook_profile(TEMPLATE_PATH)
    payload = complete_roofing_repair_payload()
    payload["labor"].append(
        {"task": "labor_full_repair", "days": 0.5, "crew_size": 2}
    )

    with pytest.raises(EstimateWorkbookInputError) as exc_info:
        validate_semantic_workbook_payload(payload, profile)

    assert "same template activity row" in str(exc_info.value)


@pytest.mark.skipif(
    not (shutil.which("soffice") or shutil.which("libreoffice")),
    reason="A spreadsheet recalculation engine is not installed.",
)
def test_create_workbook_persists_calculated_costs(monkeypatch, tmp_path: Path) -> None:
    source_hash = hashlib.sha256(TEMPLATE_PATH.read_bytes()).hexdigest()
    monkeypatch.setenv("ESTIMATOR_API_ARTIFACT_DIR", str(tmp_path))

    artifact = create_estimate_workbook(complete_payload(), base_dir=Path.cwd())

    values = openpyxl.load_workbook(artifact.path, data_only=True)
    formulas = openpyxl.load_workbook(artifact.path, data_only=False)
    estimate_values = values["Estimate"]
    estimate_formulas = formulas["Estimate"]
    people_values = values["People"]

    assert estimate_values["H106"].value == pytest.approx(270)
    assert estimate_values["H108"].value == pytest.approx(675)
    assert estimate_values["H137"].value == pytest.approx(229.5)
    assert estimate_values["H139"].value == pytest.approx(513)
    assert people_values["G12"].value == pytest.approx(1890)
    assert estimate_values["H118"].value == pytest.approx(3780)
    assert estimate_values["H148"].value > 0
    assert estimate_values["H170"].value > 0
    assert estimate_formulas["H106"].value == "=B106*C106*E106"
    assert estimate_formulas["H118"].value.startswith("=IF(")
    assert artifact.calculated_outputs["sales_inspection_cost"] == pytest.approx(270)
    assert artifact.calculated_outputs["truck_expense_cost"] == pytest.approx(675)
    assert artifact.calculated_outputs["labor_prep_daily_rate"] == pytest.approx(1890)
    assert hashlib.sha256(TEMPLATE_PATH.read_bytes()).hexdigest() == source_hash


@pytest.mark.skipif(
    not (shutil.which("soffice") or shutil.which("libreoffice")),
    reason="A spreadsheet recalculation engine is not installed.",
)
def test_create_grossman_integrity_workbook_uses_purchase_bases(
    monkeypatch,
    tmp_path: Path,
) -> None:
    monkeypatch.setenv("ESTIMATOR_API_ARTIFACT_DIR", str(tmp_path))

    artifact = create_estimate_workbook(
        grossman_integrity_payload(),
        base_dir=Path.cwd(),
    )

    values = openpyxl.load_workbook(artifact.path, data_only=True)
    formulas = openpyxl.load_workbook(artifact.path, data_only=False)
    estimate_values = values["Estimate"]
    estimate_formulas = formulas["Estimate"]

    assert estimate_values["C12"].value == 5136
    assert estimate_values["C19"].value == 5250
    assert estimate_values["C26"].value == 5200
    assert estimate_values["C27"].value == 5200
    assert estimate_values["C58"].value == 3136
    assert estimate_values["B106"].value == 3
    assert estimate_values["C106"].value == 65
    assert estimate_values["B108"].value == 12
    assert estimate_values["C108"].value == 65
    assert estimate_values["H148"].value > 0
    assert estimate_values["H170"].value > 0
    assert "Measured scope: 5,136.00 sq ft" in estimate_formulas["A19"].comment.text
    assert "98 full 4x8 sheets" in estimate_formulas["A58"].comment.text
    assert artifact.calculated_outputs["total_job_cost"] > 0
    assert artifact.calculated_outputs["worksheet_price"] > 0


@pytest.mark.skipif(
    not (shutil.which("soffice") or shutil.which("libreoffice")),
    reason="A spreadsheet recalculation engine is not installed.",
)
def test_roofing_foam_set_price_cannot_create_a_thousand_fold_cost(
    monkeypatch,
    tmp_path: Path,
) -> None:
    monkeypatch.setenv("ESTIMATOR_API_ARTIFACT_DIR", str(tmp_path))
    payload = grossman_integrity_payload()
    foam = next(
        item for item in payload["materials"] if item["category"] == "roofing_foam"
    )
    foam["unit_price"] = 2_100
    prepared, warnings = normalize_template_material_pricing(payload)

    artifact = create_estimate_workbook(prepared, base_dir=Path.cwd())

    values = openpyxl.load_workbook(artifact.path, data_only=True)["Estimate"]
    assert values["E19"].value == pytest.approx(2.1)
    assert values["H19"].value < 10_000
    assert artifact.calculated_outputs["material_subtotal"] < 100_000
    assert artifact.calculated_outputs["material_cost_per_sqft"] < 25
    assert any("$2,100.00/set" in warning for warning in warnings)


@pytest.mark.skipif(
    not (shutil.which("soffice") or shutil.which("libreoffice")),
    reason="A spreadsheet recalculation engine is not installed.",
)
def test_output_validation_rejects_economically_impossible_material_cost(
    monkeypatch,
    tmp_path: Path,
) -> None:
    monkeypatch.setenv("ESTIMATOR_API_ARTIFACT_DIR", str(tmp_path))
    payload = grossman_integrity_payload()
    foam = next(
        item for item in payload["materials"] if item["category"] == "roofing_foam"
    )
    foam["unit_price"] = 2_100

    with pytest.raises(EstimateWorkbookOutputError, match=r"\$250"):
        create_estimate_workbook(payload, base_dir=Path.cwd())

    assert list(tmp_path.glob("*.xlsx")) == []


@pytest.mark.skipif(
    not (shutil.which("soffice") or shutil.which("libreoffice")),
    reason="A spreadsheet recalculation engine is not installed.",
)
def test_create_localized_roof_repair_on_standard_template(
    monkeypatch,
    tmp_path: Path,
) -> None:
    source_hash = hashlib.sha256(TEMPLATE_PATH.read_bytes()).hexdigest()
    monkeypatch.setenv("ESTIMATOR_API_ARTIFACT_DIR", str(tmp_path))

    artifact = create_estimate_workbook(
        complete_roofing_repair_payload(),
        base_dir=Path.cwd(),
    )

    values = openpyxl.load_workbook(artifact.path, data_only=True)
    formulas = openpyxl.load_workbook(artifact.path, data_only=False)
    estimate_values = values["Estimate"]
    estimate_formulas = formulas["Estimate"]
    job_spec = values["Job Spec"]

    assert estimate_values["C12"].value == 8184
    assert estimate_values["H26"].value == pytest.approx(240)
    assert estimate_values["H43"].value == pytest.approx(176)
    assert estimate_values["H45"].value == pytest.approx(175)
    assert estimate_values["H99"].value == pytest.approx(50)
    assert estimate_values["H106"].value == pytest.approx(63.75)
    assert estimate_values["H108"].value == pytest.approx(106.25)
    assert estimate_values["A116"].value == "Setup/Safety"
    assert estimate_values["H116"].value > 0
    assert estimate_values["H118"].value > 0
    assert estimate_values["H126"].value > 0
    assert estimate_values["H148"].value > 0
    assert estimate_values["H170"].value > 0
    assert estimate_formulas["H116"].value.startswith("=IF(")
    assert job_spec["B8"].value == 1
    assert job_spec["B9"].value == pytest.approx(24)
    assert job_spec["B10"].value == 2
    assert job_spec["G9"].value == "50' x 12'"
    assert job_spec["G4"].value == "2 Yr. Workmanship"
    assert formulas["TEST Job Spec"].sheet_state == "hidden"
    assert formulas["Tracking"]["I21"].value.startswith("=IFERROR(")
    assert "C154" in formulas["Warranty"]["F5"].value
    assert "50 ft x 12 ft repair area" in str(job_spec["A13"].value)
    assert hashlib.sha256(TEMPLATE_PATH.read_bytes()).hexdigest() == source_hash


@pytest.mark.skipif(
    not (shutil.which("soffice") or shutil.which("libreoffice")),
    reason="A spreadsheet recalculation engine is not installed.",
)
def test_create_insulation_workbook_persists_calculated_costs(
    monkeypatch,
    tmp_path: Path,
) -> None:
    source_hash = hashlib.sha256(INSULATION_TEMPLATE_PATH.read_bytes()).hexdigest()
    monkeypatch.setenv("ESTIMATOR_API_ARTIFACT_DIR", str(tmp_path))

    artifact = create_estimate_workbook(
        complete_insulation_payload(),
        base_dir=Path.cwd(),
    )

    values = openpyxl.load_workbook(artifact.path, data_only=True)
    formulas = openpyxl.load_workbook(artifact.path, data_only=False)
    estimate_values = values["Estimate"]
    estimate_formulas = formulas["Estimate"]
    sqft_values = values["Sq Ft Calculation"]
    job_spec_values = values["Job Spec"]

    assert sqft_values["F15"].value == pytest.approx(4611)
    assert estimate_values["C12"].value == pytest.approx(4611)
    assert estimate_values["D12"].value == pytest.approx(4611)
    assert estimate_values["H19"].value == pytest.approx(5796.685714, rel=1e-6)
    assert estimate_values["H86"].value == pytest.approx(7245)
    assert estimate_values["H103"].value == pytest.approx(8331.75)
    assert estimate_values["H116"].value > 0
    assert estimate_values["H123"].value > 0
    assert estimate_values["B137"].value > 0
    assert estimate_values["A126"].value == "Gable insulation - scope/area review"
    assert estimate_values["A127"].value == "Stair-wall insulation - area review"
    assert job_spec_values["G10"].value == pytest.approx(4611)
    assert "Gables and stair area" in job_spec_values["A13"].value
    assert estimate_formulas["H19"].value == "=E19*G19"
    assert estimate_formulas["H86"].value.startswith("=IF(")
    assert artifact.template_profile["profile_version"] == (
        "spraytec.insulation_template_profile.v1"
    )
    assert artifact.calculated_outputs["labor_foam_cost"] == pytest.approx(7245)
    assert hashlib.sha256(INSULATION_TEMPLATE_PATH.read_bytes()).hexdigest() == source_hash


@pytest.mark.skipif(
    not (shutil.which("soffice") or shutil.which("libreoffice")),
    reason="A spreadsheet recalculation engine is not installed.",
)
def test_create_flooring_workbook_persists_calculated_costs(
    monkeypatch,
    tmp_path: Path,
) -> None:
    source_hash = hashlib.sha256(FLOORING_TEMPLATE_PATH.read_bytes()).hexdigest()
    monkeypatch.setenv("ESTIMATOR_API_ARTIFACT_DIR", str(tmp_path))

    artifact = create_estimate_workbook(
        complete_flooring_payload(),
        base_dir=Path.cwd(),
    )

    values = openpyxl.load_workbook(artifact.path, data_only=True)
    formulas = openpyxl.load_workbook(artifact.path, data_only=False)
    estimate_values = values["Estimate"]
    estimate_formulas = formulas["Estimate"]

    assert estimate_values["C12"].value == pytest.approx(484)
    assert estimate_values["H26"].value == pytest.approx(222.75)
    assert estimate_values["H27"].value == pytest.approx(318.0375)
    assert estimate_values["H106"].value == pytest.approx(26)
    assert estimate_values["H108"].value == pytest.approx(92)
    assert estimate_values["H116"].value == pytest.approx(708)
    assert estimate_values["H130"].value == pytest.approx(177)
    assert estimate_values["H137"].value == pytest.approx(66)
    assert estimate_values["H139"].value == pytest.approx(52)
    assert estimate_values["H148"].value == pytest.approx(1888)
    assert estimate_values["H170"].value == pytest.approx(5970.833375)
    assert "REVIEW" in str(estimate_values["A174"].value)
    assert estimate_values["F174"].value in (None, 0)
    assert estimate_formulas["H26"].value == "=E26*G26"
    assert estimate_formulas["H116"].value.startswith("=IF(")
    assert artifact.template_profile["profile_version"] == (
        "spraytec.flooring_template_profile.v1"
    )
    assert artifact.calculated_outputs["labor_floor_grind_patch_cost"] == pytest.approx(708)
    assert hashlib.sha256(FLOORING_TEMPLATE_PATH.read_bytes()).hexdigest() == source_hash


@pytest.mark.skipif(
    not (shutil.which("soffice") or shutil.which("libreoffice")),
    reason="A spreadsheet recalculation engine is not installed.",
)
def test_create_workbook_options_generates_independent_calculated_files(
    monkeypatch,
    tmp_path: Path,
) -> None:
    monkeypatch.setenv("ESTIMATOR_API_ARTIFACT_DIR", str(tmp_path))
    base = complete_flooring_payload()
    base["warranty"] = {
        "include": True,
        "manufacturer": "NPI",
        "years": 10,
        "warranty_type": "material",
        "unit_cost": 0.1,
    }
    alternate = deepcopy(base)
    alternate["header"]["estimated_sqft"] = 600
    alternate["warranty"]["years"] = 15
    alternate["warranty"]["unit_cost"] = 0.2
    for material in alternate["materials"]:
        material["area_sqft"] = 600

    artifacts = create_estimate_workbook_options(
        [("Base 484 SF", base), ("Alternate 600 SF", alternate)],
        base_dir=Path.cwd(),
    )

    assert len(artifacts) == 2
    assert artifacts[0][1].file_name.endswith("Base_484_SF.xlsx")
    assert artifacts[1][1].file_name.endswith("Alternate_600_SF.xlsx")
    base_values = openpyxl.load_workbook(artifacts[0][1].path, data_only=True)[
        "Estimate"
    ]
    alternate_values = openpyxl.load_workbook(
        artifacts[1][1].path,
        data_only=True,
    )["Estimate"]
    assert base_values["C12"].value == pytest.approx(484)
    assert alternate_values["C12"].value == pytest.approx(600)
    assert base_values["C154"].value == 10
    assert base_values["H154"].value == pytest.approx(48.4)
    assert alternate_values["C154"].value == 15
    assert alternate_values["H154"].value == pytest.approx(120)
    assert alternate_values["H26"].value > base_values["H26"].value
    assert (
        artifacts[1][1].calculated_outputs["worksheet_price"]
        > artifacts[0][1].calculated_outputs["worksheet_price"]
    )


def test_create_workbook_options_removes_prior_artifacts_on_failure(
    monkeypatch,
    tmp_path: Path,
) -> None:
    first_path = tmp_path / "first.xlsx"
    first_path.write_bytes(b"first")
    first_artifact = EstimateWorkbookArtifact(
        artifact_id="a" * 32,
        file_name="first.xlsx",
        path=first_path,
        calculated_outputs={},
        template_profile={},
    )
    calls = iter(
        [
            first_artifact,
            EstimateWorkbookInputError(["missing warranty selection"]),
        ]
    )

    def fake_create(*_args, **_kwargs):
        result = next(calls)
        if isinstance(result, Exception):
            raise result
        return result

    monkeypatch.setattr(
        "jobscan.estimator.workbook_service.create_estimate_workbook",
        fake_create,
    )

    with pytest.raises(EstimateWorkbookInputError) as exc_info:
        create_estimate_workbook_options(
            [("10-year", {}), ("15-year", {})],
            base_dir=Path.cwd(),
        )

    assert "15-year: missing warranty selection" in str(exc_info.value)
    assert not first_path.exists()
