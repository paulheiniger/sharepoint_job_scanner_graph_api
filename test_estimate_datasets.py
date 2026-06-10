from pathlib import Path
import tempfile

from jobscan.estimate_datasets import (
    ESTIMATE_LINE_ITEM_FIELDS,
    ESTIMATE_SUMMARY_FIELDS,
    extract_estimate_dataset,
    scan_estimate_datasets_for_records,
    write_dataset_csv,
    write_dataset_json,
)
from jobscan.models import JobRecord
from jobscan.scan import records_as_dicts, scan_root


def write_detail_workbook(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estimate"
    people = wb.create_sheet("People")

    ws["A1"] = "Job Name:"
    ws["B1"] = "ACRE Bens Bargain"
    ws["A3"] = "Materials"
    ws["A4"] = "Item"
    ws["B4"] = "Quantity"
    ws["C4"] = "Unit"
    ws["D4"] = "Unit Cost"
    ws["E4"] = "Extended Cost"
    ws["A5"] = "Foam Kit"
    ws["B5"] = 12
    ws["C5"] = "set"
    ws["D5"] = 100
    ws["E5"] = 1200
    ws["A6"] = "Fasteners"
    ws["B6"] = 500
    ws["C6"] = "ea"
    ws["D6"] = 0.25
    ws["E6"] = 125

    ws["A112"] = "Labor / Subcontractor"
    ws["B114"] = "Days"
    ws["C114"] = "No. of People"
    ws["D114"] = "Total Hours"
    ws["A115"] = "Set Up/Safety"
    ws["B115"] = 1
    ws["C115"] = 6
    ws["D115"] = 66
    ws["A116"] = "Details"
    ws["B116"] = 2
    ws["C116"] = 5
    ws["D116"] = 110
    ws["A148"] = "Total Hours"
    ws["B148"] = 10930
    ws["C148"] = "Total Days"
    ws["D148"] = 116
    people["A11"] = "Hours /Day"
    people["B11"] = 11
    wb.save(path)


def write_duration_only_workbook(path: Path, *, job_name: str, total_days: float, total_hours: float) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estimate"
    ws["B2"] = "Job Name:"
    ws["C2"] = job_name
    ws["A10"] = "Labor / Subcontractor"
    ws["B12"] = "Days"
    ws["C12"] = "No. of People"
    ws["D12"] = "Total Hours"
    ws["A13"] = "Work"
    ws["B13"] = total_days
    ws["C13"] = 2
    ws["D13"] = total_hours
    ws["A20"] = "Total Hours"
    ws["B20"] = total_hours
    ws["C20"] = "Total Days"
    ws["D20"] = total_days
    wb.save(path)


def write_adders_workbook(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estimate"
    rows = [
        ("Warranty - 10-yr", 750),
        ("Misc. Insurance", 240),
        ("Porta John", 150),
        ("Misc. Materials", 500),
        ("BEN PW SUBCONTRACTOR - $1,500.00 (Wash/Clean/Brush Algae)", 1800),
        ("Lift Rental", 3500),
        ("Caulk(80)/Prem Brush Grade (15)", 1500),
        ("Ashby Labor + RustNox Materials- 2 Coats & 1 RustNox ($9,260.00)", 10400),
        ("Total Job Cost", 20000),
        ("Overhead", 1000),
        ("Profit", 2500),
    ]
    for row_num, (label, amount) in enumerate(rows, start=80):
        ws.cell(row=row_num, column=1).value = label
        ws.cell(row=row_num, column=6).value = amount
    wb.save(path)


def test_extract_estimate_dataset_summary_and_line_items() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        job = root / "ACRE Bens Bargain"
        job.mkdir()
        path = job / "Estimate Roofing (2026) - ACRE Bens Bargain.xlsx"
        write_detail_workbook(path)
        record = JobRecord(
            job_id="ACRE-BENS-BARGAIN",
            folder_name=job.name,
            folder_path=job.name,
            division="Roofing",
            pipeline_status="Contracted",
            customer="ACRE",
        )

        summary, line_items = extract_estimate_dataset(path, root, record)

    assert summary["job_id"] == "ACRE-BENS-BARGAIN"
    assert summary["estimate_file"] == "ACRE Bens Bargain/Estimate Roofing (2026) - ACRE Bens Bargain.xlsx"
    assert summary["estimated_labor_hours"] == 10930
    assert summary["estimated_duration_days"] == 116
    assert summary["estimated_crew_size"] == 6
    assert summary["estimated_hours_per_day"] == 11
    assert set(summary) == set(ESTIMATE_SUMMARY_FIELDS)

    names = [item["line_item_name"] for item in line_items]
    assert "Foam Kit" in names
    assert "Fasteners" in names
    assert "Set Up/Safety" in names
    assert "Details" in names
    details = next(item for item in line_items if item["line_item_name"] == "Details")
    assert details["section"] == "Labor / Subcontractor"
    assert details["labor_days"] == 2
    assert details["crew_size"] == 5
    assert details["labor_hours"] == 110
    assert set(line_items[0]) == set(ESTIMATE_LINE_ITEM_FIELDS)


def test_scan_estimate_datasets_for_records_and_writers() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        job = root / "ACRE Bens Bargain"
        job.mkdir()
        write_detail_workbook(job / "Estimate Roofing (2026) - ACRE Bens Bargain.xlsx")
        record = JobRecord(
            job_id="ACRE-BENS-BARGAIN",
            folder_name=job.name,
            folder_path=job.name,
            division="Roofing",
            pipeline_status="Contracted",
            customer="ACRE",
        )

        summaries, line_items = scan_estimate_datasets_for_records(root, [record])
        csv_path = root / "estimate_summary.csv"
        json_path = root / "estimate_line_items.json"
        write_dataset_csv(summaries, ESTIMATE_SUMMARY_FIELDS, csv_path)
        write_dataset_json(line_items, ESTIMATE_LINE_ITEM_FIELDS, json_path)

        assert csv_path.exists()
        assert json_path.exists()

    assert len(summaries) == 1
    assert len(line_items) >= 4


def test_multiple_estimates_select_primary_but_emit_all_summaries() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        job = root / "Owenton First Baptist Church"
        job.mkdir()
        write_duration_only_workbook(job / "Estimate - STAMP.xlsx", job_name=job.name, total_days=0.25, total_hours=16)
        write_duration_only_workbook(job / "Estimate - IR Scan.xlsx", job_name=job.name, total_days=0, total_hours=10)
        write_duration_only_workbook(job / "Estimate - Coated Polyurethane Foam Roof.xlsx", job_name=job.name, total_days=3.5, total_hours=278.5)

        record = scan_root(root, scan_context="2026 Roofing/Contracted")[0]
        row = records_as_dicts([record])[0]
        summaries, line_items = scan_estimate_datasets_for_records(root, [record])

    assert row["estimate_file_count"] == 3
    assert row["multiple_estimates_found"] is True
    assert row["primary_estimate_file"] == "Owenton First Baptist Church/Estimate - Coated Polyurethane Foam Roof.xlsx"
    assert row["estimate_file"] == row["primary_estimate_file"]
    assert row["estimated_duration_days"] == 3.5
    assert row["estimated_labor_hours"] == 278.5
    assert "Estimate - STAMP.xlsx" in " ".join(row["supporting_estimate_files"])
    assert "Estimate - IR Scan.xlsx" in " ".join(row["supporting_estimate_files"])
    assert "Multiple estimate workbooks found" not in row["warnings"]
    assert row["estimate_selection_reason"].startswith("Multiple estimate workbooks found")
    assert "selected largest estimated_duration_days" in row["estimate_selection_reason"]

    assert len(summaries) == 3
    assert {summary["estimate_role"] for summary in summaries} == {"primary", "supporting"}
    assert len({summary["estimate_id"] for summary in summaries}) == 3
    primary = next(summary for summary in summaries if summary["estimate_role"] == "primary")
    assert primary["estimate_scope_type"] == "Coated Polyurethane Foam Roof"
    assert primary["estimated_duration_days"] == 3.5
    assert all(item["estimate_id"] for item in line_items)


def test_bottom_estimate_adders_are_captured_and_rolled_up() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        job = root / "Adder Job"
        job.mkdir()
        path = job / "Estimate Roofing - Adders.xlsx"
        write_adders_workbook(path)
        record = JobRecord(job_id="ADDER-JOB", folder_name=job.name, folder_path=job.name)

        summary, line_items = extract_estimate_dataset(path, root, record)

    adders = [item for item in line_items if item["section"] == "Estimate Adders"]
    assert len(adders) == 8
    by_name = {item["line_item_name"]: item for item in adders}
    assert by_name["Warranty - 10-yr"]["extended_cost"] == 750
    assert by_name["Warranty - 10-yr"]["line_item_category"] == "Warranty"
    assert by_name["Misc. Insurance"]["line_item_category"] == "Insurance"
    assert by_name["Porta John"]["line_item_category"] == "Rental / Site Services"
    assert by_name["BEN PW SUBCONTRACTOR - $1,500.00 (Wash/Clean/Brush Algae)"]["line_item_category"] == "Subcontractor"
    assert by_name["Lift Rental"]["line_item_category"] == "Equipment Rental"
    assert by_name["Caulk(80)/Prem Brush Grade (15)"]["line_item_category"] == "Materials"
    assert by_name["Ashby Labor + RustNox Materials- 2 Coats & 1 RustNox ($9,260.00)"]["extended_cost"] == 10400
    assert "Total Job Cost" not in by_name
    assert summary["adders_subtotal"] == 18840
    assert summary["warranty_amount"] == 750
    assert summary["insurance_amount"] == 240
    assert summary["rental_amount"] == 3650
    assert summary["subcontractor_amount"] == 1800
    assert summary["misc_materials_amount"] == 12400
