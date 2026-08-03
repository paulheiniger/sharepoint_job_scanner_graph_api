from __future__ import annotations

import hashlib
from pathlib import Path

import pytest

from jobscan.estimator.workbook_writer import generate_estimate_workbook, resolve_default_template_path


pytest.importorskip("openpyxl")
import openpyxl


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def minimal_template(path: Path, *, template_type: str = "roofing") -> Path:
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Estimate"
    if template_type == "insulation":
        workbook.create_sheet("Sq Ft Calculation")
    workbook.save(path)
    return path


def test_generate_estimate_workbook_accepts_decision_native_payload(tmp_path: Path) -> None:
    template_path = resolve_default_template_path()
    original_hash = file_hash(template_path)
    inputs = {
        "template_type": "roofing",
        "header": {
            "C2_job_name": "Decision Only Roof",
            "C3_job_type": "roof coating",
            "C12_estimated_sqft": 10000,
        },
        "workbook_decisions": [
            {
                "row_type": "material",
                "decision_id": "roofing_coating_system_row_26",
                "template_bucket": "coating",
                "category": "coating",
                "item": "Gaco Silicone",
                "workbook_row": "26",
                "selector_code": "11",
                "area_sqft": 10000,
                "gal_per_100_sqft": 1.5,
                "waste_factor_pct": 10,
                "unit_price": 38,
                "estimated_gallons": 166.67,
                "estimated_cost": 6333.46,
            }
        ],
    }

    output_path = generate_estimate_workbook(inputs, template_path, tmp_path)

    assert output_path.exists()
    assert output_path.suffix == ".xlsx"
    assert file_hash(template_path) == original_hash
    openpyxl.load_workbook(output_path, data_only=False)


def test_generate_roofing_workbook_writes_decision_input_cells(tmp_path: Path) -> None:
    template_path = minimal_template(tmp_path / "roofing_template.xlsx")
    inputs = {
        "template_type": "roofing",
        "header": {
            "C2_job_name": "Roofing Selector Draft",
            "C3_job_type": "roof coating",
            "C4_site_address": "123 Roof St",
            "C12_estimated_sqft": 10000,
        },
        "pricing": {
            "overhead_pct": 35,
            "profit_pct": 25.5,
        },
        "workbook_decisions": [
            {
                "row_type": "material",
                "decision_id": "roofing_coating_system_row_26",
                "template_bucket": "coating",
                "category": "coating",
                "item": "Gaco Silicone",
                "workbook_row": "26",
                "selector_code": "11",
                "area_sqft": 10000,
                "gal_per_100_sqft": 1.5,
                "waste_factor_pct": 10,
                "unit_price": 38,
                "estimated_gallons": 166.67,
                "estimated_cost": 6333.46,
            },
            {
                "row_type": "labor",
                "decision_id": "roofing_labor_base",
                "template_bucket": "labor_base",
                "task": "labor_base",
                "workbook_row": "122",
                "adjusted_days": 2,
                "crew_size": 4,
                "hourly_rate": 72,
                "total_hours": 64,
                "estimated_cost": 4608,
            },
        ],
    }

    output_path = generate_estimate_workbook(inputs, template_path, tmp_path, "roofing_decisions.xlsx")
    workbook = openpyxl.load_workbook(output_path, data_only=False)
    ws = workbook["Estimate"]

    assert ws["C2"].value == "Roofing Selector Draft"
    assert ws["C3"].value == "roof coating"
    assert ws["C4"].value == "123 Roof St"
    assert ws["C12"].value == 10000
    assert ws["A26"].value == 11
    assert ws["C26"].value == 10000
    assert ws["D26"].value == 1.5
    assert ws["E26"].value == 38
    assert ws["A30"].value == 10
    assert ws["B122"].value == 2
    assert ws["C122"].value == 4
    assert ws["D122"].value == 72
    assert ws["G122"].value == 64
    assert ws["F165"].value == 35
    assert ws["F167"].value == 25.5


def test_generate_roofing_spec_template_maps_labor_by_label_and_preserves_special_adders(
    tmp_path: Path,
) -> None:
    template_path = Path("templates/Estimate + Spec - Roofing.xlsx")
    original_hash = file_hash(template_path)
    inputs = {
        "template_type": "roofing",
        "header": {
            "C2_job_name": "Template Aware Roof",
            "C3_job_type": "roof coating",
            "C4_site_address": "100 Main Street",
            "C5_city_state_zip": "Louisville, KY 40202",
            "C6_contact": "Taylor Customer",
            "C8_email": "taylor@example.com",
            "C9_phone": "555-0100",
            "G9_estimator": "Spray-Tec Estimator",
            "C12_estimated_sqft": 5000,
            "mobilizations": 2,
        },
        "scope_of_work": [
            "Power wash and prepare the existing roof surface.",
            "Apply the selected base and top-coat system.",
        ],
        "spec_notes": ["Draft generated for estimator review."],
        "workbook_decisions": [
            {
                "row_type": "material",
                "category": "roofing_foam",
                "item": "Roof foam area one",
                "selector_code": 11,
                "area_sqft": 3000,
                "thickness_inches": 1.5,
                "yield_factor": 2600,
                "unit_price": 2.15,
            },
            {
                "row_type": "material",
                "category": "roofing_foam",
                "item": "Roof foam area two",
                "selector_code": 11,
                "area_sqft": 2000,
                "thickness_inches": 1.5,
                "yield_factor": 2600,
                "unit_price": 2.15,
            },
            {
                "row_type": "labor",
                "task": "labor_prep",
                "adjusted_days": 1.5,
                "crew_size": 4,
            },
            {
                "row_type": "labor",
                "task": "labor_base",
                "adjusted_days": 2,
                "crew_size": 4,
            },
            {
                "row_type": "labor",
                "task": "labor_top_coat",
                "adjusted_days": 1,
                "crew_size": 4,
            },
            {
                "row_type": "adder",
                "item": "Estimator review allowance",
                "estimated_cost": 750,
            },
            {
                "row_type": "material",
                "category": "edge_metal",
                "item": "24 gauge edge metal",
                "linear_ft": 180,
                "unit_price": 8.5,
            },
            {
                "row_type": "material",
                "category": "penetrations",
                "item": "Roof penetrations",
                "quantity": 6,
            },
        ],
    }

    output_path = generate_estimate_workbook(
        inputs,
        template_path,
        tmp_path,
        "template_aware_roof.xlsx",
    )
    workbook = openpyxl.load_workbook(output_path, data_only=False)
    estimate = workbook["Estimate"]
    job_spec = workbook["Job Spec"]

    assert estimate["A116"].value == "FULL REPAIR "
    assert estimate["C19"].value == 3000
    assert estimate["C20"].value == 2000
    assert estimate["B118"].value == 1.5
    assert estimate["C118"].value == 4
    assert estimate["B124"].value == 2
    assert estimate["C124"].value == 4
    assert estimate["B130"].value == 1
    assert estimate["C130"].value == 4
    assert estimate["A173"].value == "Warranty - 10yr"
    assert estimate["A177"].value == "Estimator review allowance"
    assert estimate["F177"].value == 750
    assert estimate["C82"].value == 180
    assert estimate["E82"].value == 8.5
    assert estimate["D49"].value == 6
    assert "Power wash and prepare" in job_spec["A13"].value
    assert job_spec["B7"].value == 2
    assert "Draft generated" in job_spec["A51"].value
    assert file_hash(template_path) == original_hash


def test_generate_insulation_workbook_writes_decision_input_cells(tmp_path: Path) -> None:
    template_path = minimal_template(tmp_path / "insulation_template.xlsx", template_type="insulation")
    inputs = {
        "template_type": "insulation",
        "header": {
            "C2_job_name": "Insulation Decision Draft",
            "C3_job_type": "spray foam insulation",
            "net_area_sqft": 2388,
        },
        "pricing": {
            "overhead_pct": 30,
            "profit_pct": 20,
        },
        "workbook_decisions": [
            {
                "row_type": "material",
                "decision_id": "insulation_foam_template_selector",
                "template_bucket": "foam",
                "category": "foam",
                "item": "Gaco 2.0 lb.",
                "workbook_row": "19-21",
                "selector_code": "11",
                "area_sqft": 2388,
                "thickness_inches": 3,
                "yield_factor": 12000,
                "unit_price": 2.4,
                "estimated_units": 597,
                "estimated_cost": 1432.8,
            },
            {
                "row_type": "labor",
                "decision_id": "insulation_labor_foam",
                "template_bucket": "labor_foam",
                "task": "labor_foam",
                "workbook_row": "86",
                "adjusted_days": 1.5,
                "crew_size": 3,
                "hourly_rate": 72,
                "total_hours": 36,
                "estimated_cost": 2592,
            },
        ],
    }

    output_path = generate_estimate_workbook(inputs, template_path, tmp_path, "insulation_decisions.xlsx")
    workbook = openpyxl.load_workbook(output_path, data_only=False)
    ws = workbook["Estimate"]

    assert ws["C2"].value == "Insulation Decision Draft"
    assert ws["C3"].value == "spray foam insulation"
    assert ws["A19"].value == 11
    assert ws["C19"].value == 2388
    assert ws["D19"].value == 3
    assert ws["E19"].value == 2.4
    assert ws["F19"].value == 12000
    assert ws["B86"].value == 1.5
    assert ws["C86"].value == 3
    assert ws["D86"].value == 72
    assert ws["G86"].value == 36
    assert ws["F118"].value == 30
    assert ws["F120"].value == 20
